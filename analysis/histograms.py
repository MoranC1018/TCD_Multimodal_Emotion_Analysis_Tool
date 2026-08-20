#!/usr/bin/env python3
"""Shared histogram, descriptor, and statistics engine for analysis.

Source-specific modules such as ``imotions.py`` and ``audio.py`` should only
discover and parse their own input format. Once they have built ``ParsedExport``
objects, they hand them here so histogram, descriptor, Spearman, and
chi-squared outputs are identical across sources.
"""

from __future__ import annotations

import csv
import hashlib
import html
import json
import math
import os
import re
import shutil
import stat
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from itertools import combinations
from pathlib import Path
from typing import Iterable, Iterator, Mapping, Sequence

from procurement.input_limits import read_control_json
from spreadsheet_safety import SpreadsheetSafeWriter, neutralize_spreadsheet_value


BIN_WIDTH = 5
SCORE_RANGE = (0, 100)
SIGNED_RANGE = (-100, 100)
BIN_VALUES_0_TO_100 = list(range(SCORE_RANGE[0], SCORE_RANGE[1], BIN_WIDTH))
BIN_VALUES_MINUS100_TO_100 = list(range(SIGNED_RANGE[0], SIGNED_RANGE[1], BIN_WIDTH))
CORE_EMOTION_ORDER = [
    "Anger",
    "Contempt",
    "Disgust",
    "Fear",
    "Joy",
    "Sadness",
    "Surprise",
]
DESCRIPTOR_METRICS = [
    "count",
    "missing",
    "mean",
    "stddev",
    "min",
    "q1",
    "median",
    "q3",
    "max",
    "kurtosis",
    "nonzero_count",
    "nonzero_percent",
]

STRUCTURAL_NAME_PATTERNS = [
    r"^row$",
    r"^timestamp$",
    r"^duration$",
    r"^eventsource(\.\d+)?$",
    r"^sample ?number(\.\d+)?$",
    r"^slideevent$",
    r"^stimtype$",
    r"^collectionphase$",
    r"^sourcestimuliname$",
]

STRUCTURAL_CATEGORY_PATTERNS = [
    r"timestamp",
    r"event source",
    r"counter",
    r"slideshow",
]

LANDMARK_NAME_PATTERNS = [
    r"^feature id_\d+$",
    r"^feature-x_\d+$",
    r"^feature-y_\d+$",
]

GEOMETRY_NAME_PATTERNS = [
    r"^width$",
    r"^height$",
    r"^interocular distance$",
]

REGION_KEYWORDS = OrderedDict(
    [
        ("brow", ["brow"]),
        ("eye", ["eye", "lid", "blink"]),
        ("mouth_lips_jaw", ["mouth", "lip", "jaw", "smile", "smirk", "dimpler", "chin"]),
        ("cheek_nose", ["cheek", "nose"]),
        ("head_rotation", ["pitch", "yaw", "roll", "head rotation", "attention"]),
    ]
)

GRAPH_COLORS = [
    "#2f6fbb",
    "#c44e52",
    "#55a868",
    "#8172b2",
    "#ccb974",
    "#64b5cd",
    "#dd8452",
    "#4c72b0",
]
OUTPUT_OWNER_FILE = ".meap_output_owner.json"
MAX_OUTPUT_OWNER_BYTES = 64 * 1024
MAX_OUTPUT_OWNER_ITEMS = 8192
MAX_OWNED_DIRECTORIES = 4096


@dataclass(frozen=True)
class ColumnInfo:
    unique_name: str
    original_name: str
    display_name: str = ""
    category: str = ""
    group: str = ""
    unit: str = ""
    description: str = ""
    device: str = ""
    provided_by: str = ""
    channel_identifier: str = ""
    scale_hint: str = ""

    @property
    def label(self) -> str:
        """Return the human-facing statistic name used in output files."""

        base = (self.display_name or self.original_name or self.unique_name).strip()
        base = re.sub(r"\.\d+$", "", base)
        return re.sub(r"\s+", " ", base) or "Unnamed statistic"


@dataclass
class ParsedExport:
    source: str
    path: Path
    header: list[str]
    info: dict[str, ColumnInfo]
    rows: Sequence[dict[str, str]]
    speaker: str = ""
    video: str = ""


@dataclass
class HistogramBucket:
    classification: str
    statistic: str
    source: str
    bin_start: str
    bin_end: str
    count: int
    percent: float


@dataclass
class HistogramTable:
    classification: str
    statistic: str
    sources: list[str]
    bins: list[str]
    rows: dict[str, list[int]]


@dataclass(frozen=True)
class ChiSquareResult:
    table: HistogramTable
    observed: dict[str, list[int]]
    expected: dict[str, list[float]]
    residuals: dict[str, list[float]]
    x_squared: float
    df: int
    p_value: float


@dataclass(frozen=True)
class SpearmanResult:
    rho: float
    s: float
    p_value: float
    n: int


@dataclass
class AnalysisResult:
    input_dir: Path
    output_dir: Path
    other_findings_dir: Path
    histogram_paths: list[Path] = field(default_factory=list)
    graph_paths: list[Path] = field(default_factory=list)
    descriptive_path: Path | None = None
    correlation_path: Path | None = None
    statistical_paths: list[Path] = field(default_factory=list)
    domain_output_dirs: dict[str, Path] = field(default_factory=dict)


def analyse_parsed_exports(
    *,
    input_dir: Path,
    output_dir: Path,
    exports: Sequence[ParsedExport],
    discovery_log: Sequence[str],
    write_graphs: bool = True,
    include_logscale: bool = False,
    include_landmarks: bool = False,
    include_timing: bool = False,
    exclude_geometry: bool = False,
    report_domain: str | None = None,
) -> AnalysisResult:
    """Run shared table/statistical outputs for already parsed input exports."""

    other_findings_dir = output_dir / "other_findings"
    graph_dir = other_findings_dir / "histogram_graphs"

    output_dir.mkdir(parents=True, exist_ok=True)
    other_findings_dir.mkdir(parents=True, exist_ok=True)
    graph_dir.mkdir(parents=True, exist_ok=True)
    clean_previous_outputs(output_dir, other_findings_dir, graph_dir)

    source_order = [export.source for export in exports]

    log_lines = [
        f"Run started: {datetime.now().astimezone().isoformat(timespec='seconds')}",
        f"Input folder: {input_dir}",
        f"Output folder: {output_dir}",
        "",
        "CSV discovery:",
        *discovery_log,
        "",
    ]

    descriptors: list[dict[str, object]] = []
    manifest_rows: list[dict[str, object]] = []
    histogram_tables: dict[str, OrderedDict[str, HistogramTable]] = {
        "emotion_0_to_100": OrderedDict(),
        "valence_minus100_to_100": OrderedDict(),
        "other_0_to_100": OrderedDict(),
        "other_numeric": OrderedDict(),
    }
    region_series_by_source: dict[str, dict[str, list[float | None]]] = {}
    region_column_rows: list[dict[str, object]] = []

    for export in exports:
        numeric_values = collect_numeric_values(export)
        classifications = classify_histogram_columns(
            export,
            numeric_values,
            include_landmarks=include_landmarks,
            include_timing=include_timing,
            exclude_geometry=exclude_geometry,
        )
        numeric_values, classifications = filter_report_domain(export, numeric_values, classifications, report_domain)
        report_export = export_with_header(export, [column for column in export.header if column in numeric_values])

        descriptors.extend(build_descriptor_rows(report_export, numeric_values, classifications))
        manifest_rows.extend(build_manifest_rows(report_export, numeric_values, classifications))

        for column_name, values in numeric_values.items():
            classification = classifications.get(column_name, "descriptor_only")
            if classification == "descriptor_only":
                continue
            info = report_export.info[column_name]
            add_histogram_counts(
                histogram_tables,
                classification,
                info.label,
                report_export.source,
                values,
                scale_hint=info.scale_hint,
            )

        source_regions, source_region_columns = build_region_series(report_export, numeric_values)
        region_series_by_source[report_export.source] = source_regions
        region_column_rows.extend(source_region_columns)

    histogram_paths = write_histogram_csvs(output_dir, source_order, histogram_tables)
    if include_logscale:
        histogram_paths.extend(write_logscale_histogram_csvs(other_findings_dir, source_order, histogram_tables))

    descriptive_path = other_findings_dir / "descriptive_statistics.csv"
    write_descriptive_statistics_csv(descriptive_path, source_order, descriptors)

    manifest_path = other_findings_dir / "column_manifest.csv"
    write_dict_rows(manifest_path, manifest_rows)

    region_map_path = other_findings_dir / "facial_region_column_map.csv"
    write_dict_rows(region_map_path, region_column_rows)

    correlation_path = other_findings_dir / "facial_region_correlations.csv"
    write_dict_rows(correlation_path, build_region_correlation_rows(region_series_by_source, source_order))

    graph_paths: list[Path] = []
    if write_graphs:
        graph_paths = write_histogram_graphs(graph_dir, histogram_tables)
        if include_logscale:
            logscale_graph_dir = other_findings_dir / "logscale_histogram_graphs"
            logscale_graph_dir.mkdir(parents=True, exist_ok=True)
            graph_paths.extend(write_histogram_graphs(logscale_graph_dir, histogram_tables, logscale=True))

    statistical_paths = write_statistical_csvs(output_dir, source_order, histogram_tables)

    log_lines.extend(
        [
            f"Descriptor rows: {len(descriptors)}",
            f"Histogram CSVs: {len(histogram_paths)}",
            f"Histogram graphs: {len(graph_paths)}",
            f"Logscale histograms: {'enabled' if include_logscale else 'disabled'}",
            f"Chi-squared/Spearman CSVs: {len(statistical_paths)}",
            f"Run finished: {datetime.now().astimezone().isoformat(timespec='seconds')}",
        ]
    )
    (other_findings_dir / "run_log.txt").write_text("\n".join(log_lines) + "\n", encoding="utf-8")

    return AnalysisResult(
        input_dir=input_dir,
        output_dir=output_dir,
        other_findings_dir=other_findings_dir,
        histogram_paths=histogram_paths,
        graph_paths=graph_paths,
        descriptive_path=descriptive_path,
        correlation_path=correlation_path,
        statistical_paths=statistical_paths,
    )


def analysis_root() -> Path:
    """Return the analysis package folder that owns local inputs/outputs."""

    return Path(__file__).resolve().parent


def default_imotions_root() -> Path:
    """Return the analysis-local iMotions input folder."""

    return analysis_root() / "iMotions_Output"


def default_output_root() -> Path:
    """Return the analysis-local report output folder."""

    return analysis_root() / "output"


def resolve_input_folder(input_folder: str | Path, imotions_root: str | Path | None = None) -> Path:
    """Resolve direct paths and bare names under analysis/iMotions_Output."""

    candidate = Path(input_folder)
    if candidate.exists():
        return candidate.resolve()

    if not candidate.is_absolute():
        root = Path(imotions_root).resolve() if imotions_root else default_imotions_root()
        rooted_candidate = root / candidate
        if rooted_candidate.exists():
            return rooted_candidate.resolve()

    return candidate.resolve()


def resolve_output_folder(input_dir: str | Path, output_root: str | Path | None = None) -> Path:
    """Resolve the folder-named report destination under analysis/output."""

    input_path = Path(input_dir)
    root = Path(output_root).resolve() if output_root else default_output_root()
    return source_bound_output(root / safe_filename(input_path.name), input_path)


def resolve_typed_output_folder(
    input_dir: str | Path,
    output_root: str | Path | None,
    source_type: str,
) -> Path:
    """Resolve output/<source_type>/<input-folder> without duplicating the type folder."""

    input_path = Path(input_dir)
    root = Path(output_root).resolve() if output_root else default_output_root()
    typed_root = root if root.name.casefold() == source_type.casefold() else root / source_type
    return source_bound_output(typed_root / safe_filename(input_path.name), input_path)


def source_bound_output(candidate: Path, input_path: Path) -> Path:
    """Keep same-named sources from sharing or deleting one report folder."""

    candidate = candidate.resolve()
    source = str(input_path.expanduser().resolve())
    if not candidate.exists() or not any(candidate.iterdir()):
        return candidate
    owner = read_output_owner(candidate)
    if owner.get("source") == source:
        return candidate
    suffix = hashlib.sha256(source.encode("utf-8")).hexdigest()[:10]
    fallback = candidate.with_name(f"{candidate.name}_{suffix}")
    if not fallback.exists() or not any(fallback.iterdir()):
        return fallback
    fallback_owner = read_output_owner(fallback)
    if fallback_owner.get("source") == source:
        return fallback
    raise ValueError(f"Analysis output is already owned by another source: {fallback}")


def analyse_domain_split_parsed_exports(
    *,
    input_dir: Path,
    output_root: str | Path | None,
    exports: Sequence[ParsedExport],
    discovery_log: Sequence[str],
    write_graphs: bool = True,
    include_logscale: bool = False,
    include_landmarks: bool = False,
    include_timing: bool = False,
    exclude_geometry: bool = False,
) -> AnalysisResult:
    """Write the two analysis parent folders: emotion and raw."""

    if not any(collect_numeric_values(export) for export in exports):
        raise ValueError("No non-empty numeric data rows were found in the selected exports.")
    domain_results: dict[str, AnalysisResult] = {}
    for domain in ("emotion", "raw"):
        domain_results[domain] = analyse_grouped_parsed_exports(
            input_dir=input_dir,
            output_dir=resolve_typed_output_folder(input_dir, output_root, domain),
            exports=exports,
            discovery_log=discovery_log,
            write_graphs=write_graphs,
            include_logscale=include_logscale,
            include_landmarks=include_landmarks,
            include_timing=include_timing,
            exclude_geometry=exclude_geometry,
            report_domain=domain,
        )

    emotion_result = domain_results["emotion"]
    aggregate = AnalysisResult(
        input_dir=input_dir,
        output_dir=emotion_result.output_dir,
        other_findings_dir=emotion_result.other_findings_dir,
        domain_output_dirs={domain: result.output_dir for domain, result in domain_results.items()},
    )
    for result in domain_results.values():
        merge_analysis_result(aggregate, result)
    return aggregate


def analyse_grouped_parsed_exports(
    *,
    input_dir: Path,
    output_dir: Path,
    exports: Sequence[ParsedExport],
    discovery_log: Sequence[str],
    write_graphs: bool = True,
    include_logscale: bool = False,
    include_landmarks: bool = False,
    include_timing: bool = False,
    exclude_geometry: bool = False,
    report_domain: str | None = None,
) -> AnalysisResult:
    """Run the shared report writer per video and per speaker-combined group.

    The source-specific feeders parse files once. This helper then preserves a
    research-friendly folder contract: source run, speaker, individual videos,
    and one combined speaker folder using the same report format.
    """

    output_dir = output_dir.resolve()
    ensure_disjoint_paths(input_dir, output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    clean_legacy_flat_outputs(output_dir, input_dir)

    exports = filter_exports_for_report_domain(
        exports,
        report_domain,
        include_landmarks=include_landmarks,
        include_timing=include_timing,
        exclude_geometry=exclude_geometry,
    )

    grouped_exports: OrderedDict[str, list[ParsedExport]] = OrderedDict()
    for export in exports:
        speaker_label = speaker_label_for_export(export, input_dir)
        grouped_exports.setdefault(speaker_label, []).append(export)

    aggregate = AnalysisResult(input_dir=input_dir, output_dir=output_dir, other_findings_dir=output_dir)
    owned_directories: set[str] = set()
    speaker_batches: list[tuple[Path, list[ParsedExport]]] = []
    for speaker_label, speaker_exports in grouped_exports.items():
        speaker_dir = grouped_speaker_dir(output_dir, input_dir, speaker_label, len(grouped_exports))
        labelled_exports = labelled_exports_for_speaker(speaker_exports)
        if speaker_dir == output_dir:
            owned_directories.update(export.source for export in labelled_exports)
            owned_directories.add("combined")
        else:
            owned_directories.add(speaker_dir.name)
        speaker_batches.append((speaker_dir, labelled_exports))

    # Claim the destination before writing reports. If the process is
    # interrupted, the next run can identify and safely clean its own partial
    # directories instead of treating them as foreign output.
    write_output_owner(output_dir, input_dir, owned_directories)

    for speaker_dir, labelled_exports in speaker_batches:
        for labelled_export in labelled_exports:
            report_dir = speaker_dir / safe_filename(labelled_export.source)
            result = analyse_parsed_exports(
                input_dir=input_dir,
                output_dir=report_dir,
                exports=[labelled_export],
                discovery_log=grouped_discovery_log(input_dir, [labelled_export], discovery_log),
                write_graphs=write_graphs,
                include_logscale=include_logscale,
                include_landmarks=include_landmarks,
                include_timing=include_timing,
                exclude_geometry=exclude_geometry,
                report_domain=report_domain,
            )
            merge_analysis_result(aggregate, result)

        combined_result = analyse_parsed_exports(
            input_dir=input_dir,
            output_dir=speaker_dir / "combined",
            exports=labelled_exports,
            discovery_log=grouped_discovery_log(input_dir, labelled_exports, discovery_log),
            write_graphs=write_graphs,
            include_logscale=include_logscale,
            include_landmarks=include_landmarks,
            include_timing=include_timing,
            exclude_geometry=exclude_geometry,
            report_domain=report_domain,
        )
        merge_analysis_result(aggregate, combined_result)

    return aggregate


def filter_exports_for_report_domain(
    exports: Sequence[ParsedExport],
    report_domain: str | None,
    *,
    include_landmarks: bool,
    include_timing: bool,
    exclude_geometry: bool,
) -> list[ParsedExport]:
    """Drop exports that have no numeric columns for the requested parent folder."""

    if report_domain is None:
        return list(exports)

    matching_exports: list[ParsedExport] = []
    for export in exports:
        numeric_values = collect_numeric_values(export)
        classifications = classify_histogram_columns(
            export,
            numeric_values,
            include_landmarks=include_landmarks,
            include_timing=include_timing,
            exclude_geometry=exclude_geometry,
        )
        filtered_values, _filtered_classifications = filter_report_domain(
            export,
            numeric_values,
            classifications,
            report_domain,
        )
        if filtered_values:
            matching_exports.append(export)
    return matching_exports


def clean_legacy_flat_outputs(output_dir: Path, input_dir: Path) -> None:
    """Remove only directories declared as generated by this same source."""

    owner = read_output_owner(output_dir)
    if owner.get("source") != str(input_dir.expanduser().resolve()):
        return

    generated = owner.get("generated_directories")
    if isinstance(generated, list):
        for name in generated:
            child = owned_output_directory(output_dir, name)
            if child is not None:
                remove_generated_directory(child)

    for file_name in [
        "histograms.csv",
        "histograms.xlsx",
        "emotion_histograms_0_to_100.csv",
        "valence_histograms_minus100_to_100.csv",
        "other_histograms.csv",
        "chi_squared_results.csv",
        "spearman_results.csv",
    ]:
        path = output_dir / file_name
        if path.exists():
            path.unlink()

    legacy_other_findings = owned_output_directory(output_dir, "other_findings")
    if legacy_other_findings is not None:
        remove_generated_directory(legacy_other_findings)


def owned_output_directory(output_dir: Path, name: object) -> Path | None:
    """Return an existing ordinary child directory that is safe to delete."""

    if not isinstance(name, str) or not name or name in {".", ".."}:
        return None
    if any(separator in name for separator in ("/", "\\")) or ":" in name:
        return None
    relative = Path(name)
    if relative.is_absolute() or relative.drive or relative.root or relative.name != name:
        return None

    output = output_dir.expanduser().resolve()
    child = output / name
    try:
        metadata = child.lstat()
    except OSError:
        return None
    is_junction = getattr(child, "is_junction", None)
    if child.is_symlink() or (callable(is_junction) and is_junction()):
        return None
    reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    if reparse_flag and getattr(metadata, "st_file_attributes", 0) & reparse_flag:
        return None
    try:
        resolved = child.resolve(strict=True)
    except OSError:
        return None
    if resolved.parent != output or not resolved.is_dir():
        return None
    return resolved


def read_output_owner(output_dir: Path) -> dict[str, object]:
    path = output_dir / OUTPUT_OWNER_FILE
    try:
        payload = read_control_json(
            path,
            label="output owner",
            max_bytes=MAX_OUTPUT_OWNER_BYTES,
            max_items=MAX_OUTPUT_OWNER_ITEMS,
        )
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(payload, dict):
        return {}
    generated = payload.get("generated_directories")
    if isinstance(generated, list) and len(generated) > MAX_OWNED_DIRECTORIES:
        raise ValueError(f"output owner JSON may declare at most {MAX_OWNED_DIRECTORIES} directories")
    return payload


def write_output_owner(output_dir: Path, input_dir: Path, generated_directories: set[str]) -> None:
    payload = {
        "source": str(input_dir.expanduser().resolve()),
        "generated_directories": sorted(generated_directories),
    }
    owner_path = output_dir / OUTPUT_OWNER_FILE
    temporary = owner_path.with_name(f".{owner_path.name}.{os.getpid()}.tmp")
    try:
        with temporary.open("w", encoding="utf-8", newline="\n") as handle:
            handle.write(json.dumps(payload, indent=2) + "\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, owner_path)
    finally:
        temporary.unlink(missing_ok=True)


def ensure_disjoint_paths(input_dir: Path, output_dir: Path) -> None:
    """Refuse any report destination inside, above, or equal to its input."""

    source = input_dir.expanduser().resolve()
    target = output_dir.expanduser().resolve()
    if source == target or source in target.parents or target in source.parents:
        raise ValueError("Analysis input and output directories must not overlap.")


def speaker_label_for_export(export: ParsedExport, input_dir: Path) -> str:
    if export.speaker:
        return safe_filename(export.speaker)
    if export.path.parent.parent != export.path.parent and export.path.parent.parent != input_dir.parent:
        return safe_filename(export.path.parent.parent.name)
    return safe_filename(input_dir.name)


def video_label_for_export(export: ParsedExport) -> str:
    label = safe_filename(export.video or export.source or export.path.stem)
    if label.casefold() == "combined":
        return "combined_video"
    return label


def grouped_speaker_dir(output_dir: Path, input_dir: Path, speaker_label: str, speaker_count: int) -> Path:
    if speaker_count == 1 and speaker_label == safe_filename(input_dir.name):
        return output_dir
    return output_dir / speaker_label


def labelled_exports_for_speaker(exports: Sequence[ParsedExport]) -> list[ParsedExport]:
    used_labels: dict[str, int] = {}
    labelled: list[ParsedExport] = []
    for export in exports:
        base_label = video_label_for_export(export)
        count = used_labels.get(base_label, 0)
        used_labels[base_label] = count + 1
        source_label = base_label if count == 0 else f"{base_label}_{count + 1}"
        labelled.append(
            ParsedExport(
                source=source_label,
                path=export.path,
                header=export.header,
                info=export.info,
                rows=export.rows,
                speaker=export.speaker,
                video=export.video,
            )
        )
    return labelled


def grouped_discovery_log(
    input_dir: Path,
    exports: Sequence[ParsedExport],
    original_discovery_log: Sequence[str],
) -> list[str]:
    lines = list(original_discovery_log)
    lines.append("")
    lines.append("Grouped report inputs:")
    for export in exports:
        try:
            relative_path = export.path.relative_to(input_dir)
        except ValueError:
            relative_path = export.path
        lines.append(f"Selected {relative_path} as {export.source}.")
    return lines


def merge_analysis_result(target: AnalysisResult, source: AnalysisResult) -> None:
    target.histogram_paths.extend(source.histogram_paths)
    target.graph_paths.extend(source.graph_paths)
    if source.descriptive_path:
        target.descriptive_path = source.descriptive_path
    if source.correlation_path:
        target.correlation_path = source.correlation_path
    target.statistical_paths.extend(source.statistical_paths)
    target.domain_output_dirs.update(source.domain_output_dirs)


def clean_previous_outputs(
    output_dir: Path,
    other_findings_dir: Path,
    graph_dir: Path,
) -> None:
    """Remove prior generated files so reruns cannot leave stale findings behind."""

    for path in output_dir.glob("histogram_counts_*.csv"):
        path.unlink()
    for file_name in [
        "histograms.csv",
        "histograms.xlsx",
        "emotion_histograms_0_to_100.csv",
        "valence_histograms_minus100_to_100.csv",
        "other_histograms.csv",
        "chi_squared_results.csv",
        "spearman_results.csv",
    ]:
        path = output_dir / file_name
        if path.exists():
            path.unlink()

    for file_name in [
        "descriptive_statistics.csv",
        "descriptor_statistics.csv",  # Legacy name removed during migration.
        "column_manifest.csv",
        "facial_region_column_map.csv",
        "facial_region_correlations.csv",
        "run_log.txt",
        "logscale_histograms.csv",
    ]:
        path = other_findings_dir / file_name
        if path.exists():
            path.unlink()

    for folder in (graph_dir,):
        for path in folder.glob("*"):
            if path.is_file():
                path.unlink()
    logscale_graph_dir = other_findings_dir / "logscale_histogram_graphs"
    if logscale_graph_dir.exists():
        remove_generated_directory(logscale_graph_dir)
    legacy_report_dir = other_findings_dir / ("r" + "_convert_reports")
    if legacy_report_dir.exists():
        remove_generated_directory(legacy_report_dir)
    legacy_summary = other_findings_dir / ("r" + "_convert_summary.csv")
    if legacy_summary.exists():
        legacy_summary.unlink()


def remove_generated_directory(path: Path) -> None:
    """Remove generated directories even when synced folders mark them read-only."""

    def make_writable(function, failed_path, _exc_info):
        os.chmod(failed_path, stat.S_IREAD | stat.S_IWRITE)
        function(failed_path)

    try:
        shutil.rmtree(path, onerror=make_writable)
    except PermissionError:
        os.chmod(path, stat.S_IREAD | stat.S_IWRITE)
        path.rmdir()


def collect_numeric_values(export: ParsedExport) -> dict[str, list[float]]:
    """Collect numeric values for every column in one parsed export."""

    values: dict[str, list[float]] = {column: [] for column in export.header}
    for row in export.rows:
        for column in export.header:
            parsed = parse_float(row.get(column))
            if parsed is not None:
                values[column].append(parsed)
    return {column: column_values for column, column_values in values.items() if column_values}


def parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def classify_histogram_columns(
    export: ParsedExport,
    numeric_values: Mapping[str, Sequence[float]],
    *,
    include_landmarks: bool,
    include_timing: bool,
    exclude_geometry: bool,
) -> dict[str, str]:
    """Classify numeric columns into fixed histogram ranges or descriptor-only."""

    classifications: dict[str, str] = {}
    for column, values in numeric_values.items():
        info = export.info[column]
        exclude, _reason = should_exclude_column(
            column,
            info,
            include_timing=include_timing,
            include_landmarks=include_landmarks,
            exclude_geometry=exclude_geometry,
        )
        if exclude:
            classifications[column] = "descriptor_only"
        elif is_valence_column(info):
            classifications[column] = "valence_minus100_to_100"
        elif is_score_like(info, values):
            classifications[column] = "emotion_0_to_100" if is_emotion_column(info) else "other_0_to_100"
        else:
            classifications[column] = "other_numeric"
    return classifications


def should_exclude_column(
    column: str,
    info: ColumnInfo,
    *,
    include_timing: bool,
    include_landmarks: bool,
    exclude_geometry: bool,
) -> tuple[bool, str]:
    label = info.label
    fields = [column, info.original_name, label]
    categories = [info.category, info.group, info.channel_identifier]

    if not include_timing:
        if any(matches_any(field, STRUCTURAL_NAME_PATTERNS) for field in fields if field):
            return True, "structural/timing/counter column"
        if any(matches_any(field, STRUCTURAL_CATEGORY_PATTERNS) for field in categories if field):
            return True, "structural/timing/counter category"

    if not include_landmarks and any(matches_any(field, LANDMARK_NAME_PATTERNS) for field in fields if field):
        return True, "raw landmark feature column"

    if exclude_geometry and any(matches_any(field, GEOMETRY_NAME_PATTERNS) for field in fields if field):
        return True, "geometry/dimension column"

    return False, ""


def matches_any(text: str, patterns: Iterable[str]) -> bool:
    return any(re.search(pattern, text.strip(), flags=re.IGNORECASE) for pattern in patterns)


def is_valence_column(info: ColumnInfo) -> bool:
    text = f"{info.label} {info.original_name} {info.channel_identifier}".lower()
    return "valence" in text


def is_emotion_column(info: ColumnInfo) -> bool:
    """Return True for FEA emotion-category columns except Valence."""

    text = f"{info.category} {info.group} {info.channel_identifier} {info.label}".lower()
    return "emotion" in text and not is_valence_column(info)


def is_score_like(info: ColumnInfo, values: Sequence[float]) -> bool:
    if info.scale_hint == "raw_acoustic":
        return False

    if not values:
        return False
    finite = finite_values(values)
    if not finite:
        return False

    mn = min(finite)
    mx = max(finite)
    unit = (info.unit or "").strip().lower()
    text = f"{info.category} {info.group} {info.channel_identifier} {info.label}".lower()

    if unit == "index" and mn >= 0 and mx <= 100:
        return True
    if mn >= 0 and mx <= 1:
        return True

    score_keywords = [
        "emotion",
        "facial expression",
        "action unit",
        "attention",
        "engagement",
        "sentimentality",
        "confusion",
        "neutral",
        "blink",
        "speaking",
        "eye",
        "mouth",
        "brow",
        "lip",
        "smile",
        "smirk",
        "cheek",
        "jaw",
        "nose",
    ]
    return mn >= 0 and mx <= 100 and any(keyword in text for keyword in score_keywords)


def finite_values(values: Sequence[float]) -> list[float]:
    return [float(value) for value in values if math.isfinite(float(value))]


def filter_report_domain(
    export: ParsedExport,
    numeric_values: Mapping[str, Sequence[float]],
    classifications: Mapping[str, str],
    report_domain: str | None,
) -> tuple[dict[str, Sequence[float]], dict[str, str]]:
    if report_domain is None:
        return dict(numeric_values), dict(classifications)

    filtered_values: dict[str, Sequence[float]] = {}
    filtered_classifications: dict[str, str] = {}
    for column, values in numeric_values.items():
        classification = classifications.get(column, "descriptor_only")
        is_emotion_column_for_report = is_emotion_report_column(export.info[column], classification)
        if report_domain == "emotion" and is_emotion_column_for_report:
            filtered_values[column] = values
            filtered_classifications[column] = classification
        elif report_domain == "raw" and not is_emotion_column_for_report:
            filtered_values[column] = values
            filtered_classifications[column] = classification
    return filtered_values, filtered_classifications


def is_emotion_report_column(info: ColumnInfo, classification: str) -> bool:
    text = f"{info.category} {info.group} {info.channel_identifier} {info.label}".lower()
    return (
        classification in {"emotion_0_to_100", "valence_minus100_to_100"}
        or "categorical emotion" in text
        or "dimensional affect" in text
    )


def export_with_header(export: ParsedExport, header: Sequence[str]) -> ParsedExport:
    return ParsedExport(
        source=export.source,
        path=export.path,
        header=list(header),
        info=export.info,
        rows=export.rows,
        speaker=export.speaker,
        video=export.video,
    )


def normalise_values(values: Sequence[float], classification: str, *, scale_hint: str = "") -> list[float]:
    finite = finite_values(values)
    if not finite:
        return []
    if scale_hint in {"0_to_100", "minus100_to_100", "already_normalized"}:
        return finite
    if classification in {"emotion_0_to_100", "other_0_to_100"} and min(finite) >= 0 and max(finite) <= 1:
        return [value * 100.0 for value in finite]
    if classification == "valence_minus100_to_100" and min(finite) >= -1 and max(finite) <= 1:
        return [value * 100.0 for value in finite]
    return finite


def add_histogram_counts(
    histogram_tables: dict[str, OrderedDict[str, HistogramTable]],
    classification: str,
    statistic: str,
    source: str,
    values: Sequence[float],
    scale_hint: str = "",
) -> None:
    """Add one source column to the folder-level histogram table for a statistic."""

    normalised = normalise_values(values, classification, scale_hint=scale_hint)
    if classification in {"emotion_0_to_100", "other_0_to_100"}:
        bins = [str(value) for value in BIN_VALUES_0_TO_100]
        counts = fixed_range_counts(normalised, 0, 100, BIN_WIDTH)
    elif classification == "valence_minus100_to_100":
        bins = [str(value) for value in BIN_VALUES_MINUS100_TO_100]
        counts = fixed_range_counts(normalised, -100, 100, BIN_WIDTH)
    elif classification == "other_numeric":
        counts = auto_range_counts(normalised, BIN_WIDTH)
        bins = list(counts)
    else:
        return

    table_group = histogram_tables[classification]
    table = table_group.get(statistic)
    if table is None:
        table = HistogramTable(classification=classification, statistic=statistic, sources=[], bins=bins, rows={})
        table_group[statistic] = table
    else:
        table.bins = sorted_merged_bins(table.bins, bins)

    if source not in table.sources:
        table.sources.append(source)
    for bin_label in table.bins:
        table.rows.setdefault(bin_label, [0] * len(table.sources))

    source_index = table.sources.index(source)
    for row_counts in table.rows.values():
        while len(row_counts) < len(table.sources):
            row_counts.append(0)

    for bin_label, count in counts.items():
        table.rows.setdefault(bin_label, [0] * len(table.sources))
        table.rows[bin_label][source_index] = int(count)


def fixed_range_counts(values: Sequence[float], start: int, stop: int, width: int) -> OrderedDict[str, int]:
    counts: OrderedDict[str, int] = OrderedDict((str(bin_start), 0) for bin_start in range(start, stop, width))
    for value in values:
        if not math.isfinite(value):
            continue
        if value < start or value > stop:
            continue
        bin_start = int(math.floor(value / width) * width)
        if bin_start >= stop:
            bin_start = stop - width
        if bin_start < start:
            bin_start = start
        counts[str(bin_start)] += 1
    return counts


def auto_range_counts(values: Sequence[float], width: int) -> OrderedDict[str, int]:
    finite = finite_values(values)
    if not finite:
        return OrderedDict()
    start = int(math.floor(min(finite) / width) * width)
    stop = int(math.ceil(max(finite) / width) * width)
    if stop <= start:
        stop = start + width
    counts: OrderedDict[str, int] = OrderedDict((str(bin_start), 0) for bin_start in range(start, stop, width))
    max_value = max(finite)
    for value in finite:
        bin_start = int(math.floor(value / width) * width)
        if value == max_value and bin_start == stop:
            bin_start = stop - width
        counts.setdefault(str(bin_start), 0)
        counts[str(bin_start)] += 1
    return counts


def sorted_merged_bins(left: Sequence[str], right: Sequence[str]) -> list[str]:
    return [bin_label(value) for value in sorted({float(item) for item in [*left, *right]})]


def bin_label(value: float) -> str:
    """Keep bin labels stable so "0" and "0.0" do not split the same bucket."""

    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:g}"


def build_descriptor_rows(
    export: ParsedExport,
    numeric_values: Mapping[str, Sequence[float]],
    classifications: Mapping[str, str],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    total_rows = len(export.rows)
    for column, values in numeric_values.items():
        info = export.info[column]
        stats = describe(values)
        rows.append(
            {
                "source": export.source,
                "statistic": info.label,
                "source_column": info.original_name,
                "classification": classifications.get(column, "descriptor_only"),
                "category": info.category,
                "group": info.group,
                "unit": info.unit,
                "count": stats["count"],
                "missing": max(0, total_rows - int(stats["count"])),
                "mean": stats["mean"],
                "stddev": stats["stddev"],
                "min": stats["min"],
                "q1": stats["q1"],
                "median": stats["median"],
                "q3": stats["q3"],
                "max": stats["max"],
                "kurtosis": stats["kurtosis"],
                "nonzero_count": stats["nonzero_count"],
                "nonzero_percent": stats["nonzero_percent"],
            }
        )
    return rows


def describe(values: Sequence[float]) -> dict[str, object]:
    finite = sorted(finite_values(values))
    count = len(finite)
    if count == 0:
        return {
            "count": 0,
            "mean": "",
            "stddev": "",
            "min": "",
            "q1": "",
            "median": "",
            "q3": "",
            "max": "",
            "kurtosis": "",
            "nonzero_count": 0,
            "nonzero_percent": "",
        }

    mean = sum(finite) / count
    stddev = math.sqrt(sum((value - mean) ** 2 for value in finite) / (count - 1)) if count > 1 else 0.0
    nonzero_count = sum(1 for value in finite if value != 0)
    kurtosis = excess_kurtosis(finite, mean)
    return {
        "count": count,
        "mean": format_number(mean),
        "stddev": format_number(stddev),
        "min": format_number(finite[0]),
        "q1": format_number(percentile(finite, 0.25)),
        "median": format_number(percentile(finite, 0.50)),
        "q3": format_number(percentile(finite, 0.75)),
        "max": format_number(finite[-1]),
        "kurtosis": format_number(kurtosis),
        "nonzero_count": nonzero_count,
        "nonzero_percent": format_number(nonzero_count / count * 100.0),
    }


def excess_kurtosis(values: Sequence[float], mean: float | None = None) -> float:
    finite = finite_values(values)
    if not finite:
        return math.nan
    centre = sum(finite) / len(finite) if mean is None else mean
    variance = sum((value - centre) ** 2 for value in finite) / len(finite)
    if variance == 0:
        return 0.0
    fourth_moment = sum((value - centre) ** 4 for value in finite) / len(finite)
    return fourth_moment / (variance * variance) - 3.0


def write_descriptive_statistics_csv(
    path: Path,
    source_order: Sequence[str],
    descriptor_rows: Sequence[Mapping[str, object]],
) -> None:
    """Write descriptive statistics with each numeric descriptor compared across videos."""

    by_statistic: OrderedDict[str, dict[str, Mapping[str, object]]] = OrderedDict()
    for row in descriptor_rows:
        statistic = str(row.get("statistic", "") or "Unnamed statistic")
        source = str(row.get("source", "") or "Unknown source")
        by_statistic.setdefault(statistic, {})[source] = row

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = SpreadsheetSafeWriter(csv.writer(handle))
        for statistic, rows_by_source in by_statistic.items():
            ordered_sources = [source for source in source_order if source in rows_by_source]
            ordered_sources.extend(source for source in rows_by_source if source not in ordered_sources)
            first_row = next(iter(rows_by_source.values()))
            writer.writerow([statistic])
            writer.writerow(
                [
                    "classification",
                    first_row.get("classification", ""),
                    "category",
                    first_row.get("category", ""),
                    "unit",
                    first_row.get("unit", ""),
                ]
            )
            writer.writerow(["metric", *ordered_sources])
            for metric in DESCRIPTOR_METRICS:
                writer.writerow([metric, *[rows_by_source[source].get(metric, "") for source in ordered_sources]])
            writer.writerow([])


def percentile(sorted_values: Sequence[float], fraction: float) -> float:
    if not sorted_values:
        return math.nan
    if len(sorted_values) == 1:
        return float(sorted_values[0])
    position = (len(sorted_values) - 1) * fraction
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return float(sorted_values[int(position)])
    weight = position - lower
    return float(sorted_values[lower]) * (1.0 - weight) + float(sorted_values[upper]) * weight


def build_manifest_rows(
    export: ParsedExport,
    numeric_values: Mapping[str, Sequence[float]],
    classifications: Mapping[str, str],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for column in export.header:
        info = export.info[column]
        rows.append(
            {
                "source": export.source,
                "source_file": export.path.name,
                "source_column": info.original_name,
                "statistic": info.label,
                "classification": classifications.get(column, "not_numeric"),
                "category": info.category,
                "group": info.group,
                "unit": info.unit,
                "numeric_values": len(numeric_values.get(column, [])),
                "channel_identifier": info.channel_identifier,
                "provided_by": info.provided_by,
                "scale_hint": info.scale_hint,
                "description": info.description,
            }
        )
    return rows


def build_region_series(
    export: ParsedExport,
    numeric_values: Mapping[str, Sequence[float]],
) -> tuple[dict[str, list[float | None]], list[dict[str, object]]]:
    """Group facial action columns into region scores per row."""

    columns_by_region: OrderedDict[str, list[str]] = OrderedDict((region, []) for region in REGION_KEYWORDS)
    column_rows: list[dict[str, object]] = []
    for column in numeric_values:
        info = export.info[column]
        region = region_for_info(info)
        if not region:
            continue
        columns_by_region[region].append(column)
        column_rows.append(
            {
                "source": export.source,
                "region": region,
                "statistic": info.label,
                "source_column": info.original_name,
                "category": info.category,
                "group": info.group,
                "unit": info.unit,
            }
        )

    series_by_region: dict[str, list[float | None]] = {region: [] for region in columns_by_region}
    for row in export.rows:
        for region, columns in columns_by_region.items():
            row_values = [parse_float(row.get(column)) for column in columns]
            finite = [value for value in row_values if value is not None and math.isfinite(value)]
            series_by_region[region].append(sum(finite) / len(finite) if finite else None)

    return series_by_region, column_rows


def region_for_info(info: ColumnInfo) -> str | None:
    text = f"{info.label} {info.original_name} {info.category} {info.group} {info.channel_identifier}".lower()
    for region, keywords in REGION_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            return region
    return None


def build_region_correlation_rows(
    region_series_by_source: Mapping[str, Mapping[str, Sequence[float | None]]],
    source_order: Sequence[str],
) -> list[dict[str, object]]:
    """Return facial-region correlations as source-comparison rows.

    Each region pair is represented once per metric, with the input videos laid
    out across columns. This keeps the file short enough to scan while still
    leaving a stable row structure for downstream scripts.
    """

    per_source_rows: dict[str, dict[tuple[str, str], dict[str, object]]] = {}
    for source, region_series in region_series_by_source.items():
        per_source_rows[source] = {
            (row["region_a"], row["region_b"]): row
            for row in correlation_rows_for_source(source, region_series)
        }

    ordered_sources = [source for source in source_order if source in per_source_rows]
    ordered_sources.extend(source for source in per_source_rows if source not in ordered_sources)
    pair_order: list[tuple[str, str]] = []
    for left, right in combinations(REGION_KEYWORDS.keys(), 2):
        if any((left, right) in per_source_rows[source] for source in ordered_sources):
            pair_order.append((left, right))

    rows: list[dict[str, object]] = []
    for left, right in pair_order:
        pair_label = f"{left} vs {right}"
        for metric in ("pearson_r", "n"):
            row: dict[str, object] = {"region_pair": pair_label, "metric": metric}
            for source in ordered_sources:
                value = per_source_rows[source].get((left, right), {}).get(metric, "")
                row[source] = value
            rows.append(row)

    combined: dict[str, list[float | None]] = {region: [] for region in REGION_KEYWORDS}
    for region_series in region_series_by_source.values():
        for region in REGION_KEYWORDS:
            combined[region].extend(region_series.get(region, []))
    combined_rows = {
        (row["region_a"], row["region_b"]): row
        for row in correlation_rows_for_source("all_sources", combined)
    }
    for left, right in pair_order:
        combined_row = combined_rows.get((left, right), {})
        rows.append(
            {
                "region_pair": f"{left} vs {right}",
                "metric": "all_sources_pearson_r",
                "all_sources": combined_row.get("pearson_r", ""),
            }
        )
        rows.append(
            {
                "region_pair": f"{left} vs {right}",
                "metric": "all_sources_n",
                "all_sources": combined_row.get("n", ""),
            }
        )
    return rows


def correlation_rows_for_source(source: str, region_series: Mapping[str, Sequence[float | None]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    active_regions = [region for region, series in region_series.items() if any(value is not None for value in series)]
    for left, right in combinations(active_regions, 2):
        paired = [
            (a, b)
            for a, b in zip(region_series[left], region_series[right])
            if a is not None and b is not None and math.isfinite(a) and math.isfinite(b)
        ]
        r_value = pearson_correlation([a for a, _ in paired], [b for _, b in paired]) if len(paired) >= 3 else math.nan
        rows.append(
            {
                "source": source,
                "region_a": left,
                "region_b": right,
                "n": len(paired),
                "pearson_r": format_number(r_value),
            }
        )
    return rows


def write_histogram_csvs(
    output_dir: Path,
    source_order: Sequence[str],
    histogram_tables: Mapping[str, Mapping[str, HistogramTable]],
) -> list[Path]:
    """Write one readable histogram CSV plus a matching Excel workbook.

    The CSV uses clear sections with a blank line between statistic tables. The
    workbook mirrors the same layout across three sheets: Core emotions, Other
    0-100 findings, and Valence.
    """

    sections = histogram_output_sections(source_order, histogram_tables)
    csv_path = output_dir / "histograms.csv"
    write_sectioned_csv(csv_path, sections)
    paths = [csv_path]

    xlsx_path = output_dir / "histograms.xlsx"
    if write_xlsx_workbook(xlsx_path, sections):
        paths.append(xlsx_path)
    return paths


def write_logscale_histogram_csvs(
    other_findings_dir: Path,
    source_order: Sequence[str],
    histogram_tables: Mapping[str, Mapping[str, HistogramTable]],
) -> list[Path]:
    """Write optional log10(count + 1) histogram tables for skewed counts."""

    sections = histogram_output_sections(source_order, histogram_tables)
    csv_path = other_findings_dir / "logscale_histograms.csv"
    write_sectioned_csv(
        csv_path,
        sections,
        scale_label="log10(count + 1)",
        count_transform=lambda count: math.log10(count + 1),
        total_transform=lambda counts: math.log10(sum(counts) + 1),
    )
    return [csv_path]


def histogram_output_sections(
    source_order: Sequence[str],
    histogram_tables: Mapping[str, Mapping[str, HistogramTable]],
) -> OrderedDict[str, list[HistogramTable]]:
    return OrderedDict(
        [
            (
                "Core emotions (0-100)",
                sorted_histogram_tables(histogram_tables.get("emotion_0_to_100", {}).values(), source_order),
            ),
            (
                "Other 0-100 findings",
                sorted_histogram_tables(histogram_tables.get("other_0_to_100", {}).values(), source_order),
            ),
            (
                "Valence (-100 to 100)",
                sorted_histogram_tables(histogram_tables.get("valence_minus100_to_100", {}).values(), source_order),
            ),
        ]
    )


def sorted_histogram_tables(tables: Iterable[HistogramTable], source_order: Sequence[str]) -> list[HistogramTable]:
    def sort_key(table: HistogramTable) -> tuple[int, str]:
        try:
            core_index = CORE_EMOTION_ORDER.index(table.statistic)
        except ValueError:
            core_index = len(CORE_EMOTION_ORDER)
        return core_index, table.statistic.lower()

    return [reorder_table_sources(table, source_order) for table in sorted(tables, key=sort_key)]


def write_sectioned_csv(
    path: Path,
    sections: Mapping[str, Sequence[HistogramTable]],
    *,
    scale_label: str | None = None,
    count_transform=None,
    total_transform=None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = SpreadsheetSafeWriter(csv.writer(handle))
        for section_name, tables in sections.items():
            writer.writerow([section_name])
            if scale_label:
                writer.writerow(["scale", scale_label])
            if not tables:
                writer.writerow(["No matching columns found"])
                writer.writerow([])
                continue
            for table in tables:
                writer.writerow([table.statistic])
                writer.writerow(["bin_start", "bin_end", *table.sources, "total"])
                for bin_label in table.bins:
                    counts = table.rows.get(bin_label, [0] * len(table.sources))
                    if count_transform:
                        values = [format_cell(count_transform(count)) for count in counts]
                        total_value = format_cell(total_transform(counts) if total_transform else count_transform(sum(counts)))
                    else:
                        values = counts
                        total_value = sum(counts)
                    writer.writerow([bin_label, bin_end_label(bin_label, BIN_WIDTH), *values, total_value])
                writer.writerow([])


def write_xlsx_workbook(path: Path, sections: Mapping[str, Sequence[HistogramTable]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return False

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for section_name, tables in sections.items():
        sheet = workbook.create_sheet(title=excel_sheet_name(section_name))
        row_index = 1
        sheet.cell(
            row=row_index,
            column=1,
            value=neutralize_spreadsheet_value(section_name),
        ).font = Font(bold=True, size=14)
        row_index += 2
        if not tables:
            sheet.cell(row=row_index, column=1, value="No matching columns found")
            continue
        for table in tables:
            sheet.cell(
                row=row_index,
                column=1,
                value=neutralize_spreadsheet_value(table.statistic),
            ).font = Font(bold=True)
            row_index += 1
            headers = ["bin_start", "bin_end", *table.sources, "total"]
            for column_index, header in enumerate(headers, start=1):
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=neutralize_spreadsheet_value(header),
                ).font = Font(bold=True)
            row_index += 1
            for bin_label in table.bins:
                counts = table.rows.get(bin_label, [0] * len(table.sources))
                values = [bin_label, bin_end_label(bin_label, BIN_WIDTH), *counts, sum(counts)]
                for column_index, value in enumerate(values, start=1):
                    sheet.cell(
                        row=row_index,
                        column=column_index,
                        value=neutralize_spreadsheet_value(value),
                    )
                row_index += 1
            row_index += 1

        for column_cells in sheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = min(42, max([len(value) for value in values] + [10]) + 2)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return True


def excel_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[][\\/*?:]", "", name)
    return cleaned[:31] or "Sheet"


def bin_end_label(bin_start: str, width: int) -> str:
    value = float(bin_start)
    end = value + width
    return format_number(end)


def write_histogram_graphs(
    graph_dir: Path,
    histogram_tables: Mapping[str, Mapping[str, HistogramTable]],
    *,
    logscale: bool = False,
) -> list[Path]:
    graph_paths: list[Path] = []
    for classification, tables in histogram_tables.items():
        if classification == "other_numeric":
            continue
        for table in tables.values():
            if not any(sum(counts) for counts in table.rows.values()):
                continue
            suffix = "__logscale" if logscale else ""
            graph_path = graph_dir / f"{safe_filename(table.statistic).lower()}__{classification}{suffix}.svg"
            graph_path.write_text(render_histogram_svg(table, logscale=logscale), encoding="utf-8")
            graph_paths.append(graph_path)
    return graph_paths


def render_histogram_svg(table: HistogramTable, *, logscale: bool = False) -> str:
    """Render a dependency-free grouped bar chart as SVG."""

    bins = table.bins
    sources = table.sources
    max_raw_count = max((max(counts) if counts else 0 for counts in table.rows.values()), default=0)
    max_value = math.log10(max_raw_count + 1) if logscale else max_raw_count
    max_value = max(max_value, 1)

    left_margin = 82
    right_margin = 24
    top_margin = 62
    bottom_margin = 72
    bin_width = max(24, min(42, 760 // max(1, len(bins))))
    width = left_margin + right_margin + len(bins) * bin_width
    height = 420
    chart_height = height - top_margin - bottom_margin
    zero_y = top_margin + chart_height
    bar_width = max(2, (bin_width - 6) / max(1, len(sources)))

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        "<style>text{font-family:Arial,sans-serif;font-size:12px;fill:#202124}.axis{stroke:#555;stroke-width:1}.grid{stroke:#ddd;stroke-width:1}.title{font-size:17px;font-weight:700}.legend{font-size:11px}</style>",
        f'<rect width="{width}" height="{height}" fill="#fff"/>',
        f'<text class="title" x="{left_margin}" y="28">{html.escape(table.statistic)} ({html.escape(table.classification)}{"; log10(count + 1)" if logscale else ""})</text>',
    ]

    for tick in range(0, 5):
        ratio = tick / 4
        y = zero_y - ratio * chart_height
        value = (10 ** (max_value * ratio) - 1) if logscale else max_value * ratio
        parts.append(f'<line class="grid" x1="{left_margin}" y1="{y:.2f}" x2="{width - right_margin}" y2="{y:.2f}"/>')
        parts.append(f'<text x="{left_margin - 8}" y="{y + 4:.2f}" text-anchor="end">{format_number(value)}</text>')

    parts.append(f'<line class="axis" x1="{left_margin}" y1="{top_margin}" x2="{left_margin}" y2="{zero_y}"/>')
    parts.append(f'<line class="axis" x1="{left_margin}" y1="{zero_y}" x2="{width - right_margin}" y2="{zero_y}"/>')

    for bin_index, bin_label in enumerate(bins):
        x_base = left_margin + bin_index * bin_width + 3
        counts = table.rows.get(bin_label, [0] * len(sources))
        for source_index, source in enumerate(sources):
            count = counts[source_index] if source_index < len(counts) else 0
            scaled_count = math.log10(count + 1) if logscale else count
            bar_height = scaled_count / max_value * chart_height
            x = x_base + source_index * bar_width
            y = zero_y - bar_height
            color = GRAPH_COLORS[source_index % len(GRAPH_COLORS)]
            tooltip = f"{source} bin {bin_label}: {count}"
            if logscale:
                tooltip += f" (log10 count + 1 = {format_number(scaled_count)})"
            parts.append(
                f'<rect x="{x:.2f}" y="{y:.2f}" width="{bar_width:.2f}" height="{bar_height:.2f}" fill="{color}">'
                f'<title>{html.escape(tooltip)}</title></rect>'
            )
        if bin_index % max(1, math.ceil(len(bins) / 20)) == 0:
            label_x = left_margin + bin_index * bin_width + bin_width / 2
            parts.append(f'<text x="{label_x:.2f}" y="{zero_y + 18}" text-anchor="middle">{html.escape(bin_label)}</text>')

    legend_x = left_margin
    legend_y = height - 28
    for index, source in enumerate(sources):
        color = GRAPH_COLORS[index % len(GRAPH_COLORS)]
        x = legend_x + index * 150
        parts.append(f'<rect x="{x}" y="{legend_y - 10}" width="10" height="10" fill="{color}"/>')
        parts.append(f'<text class="legend" x="{x + 15}" y="{legend_y}">{html.escape(source[:22])}</text>')

    parts.append("</svg>")
    return "\n".join(parts)


def write_statistical_csvs(
    output_dir: Path,
    source_order: Sequence[str],
    histogram_tables: Mapping[str, Mapping[str, HistogramTable]],
) -> list[Path]:
    """Write Pearson chi-squared and Spearman outputs as comparison tables."""

    chi_path = output_dir / "chi_squared_results.csv"
    spearman_path = output_dir / "spearman_results.csv"

    with chi_path.open("w", encoding="utf-8", newline="") as handle:
        writer = SpreadsheetSafeWriter(csv.writer(handle))
        for table in iter_statistical_tables(source_order, histogram_tables):
            if len(table.sources) < 2 or sum(sum(counts) for counts in table.rows.values()) == 0:
                continue
            try:
                result = chi_square_test(table)
            except ValueError:
                continue
            pairwise_results = pairwise_chi_square(table)
            writer.writerow([f"Pearson chi-square: {table.statistic}"])
            writer.writerow(["classification", table.classification])
            writer.writerow(["source_count", len(result.table.sources), "bin_count", len(result.table.bins)])
            writer.writerow(["x_squared", format_number(result.x_squared), "df", result.df, "p_value", raw_p_value(result.p_value)])
            writer.writerow([])
            write_pairwise_matrix(writer, "Pairwise X-squared matrix", result.table.sources, pairwise_results, lambda item: format_cell(item.x_squared))
            write_pairwise_matrix(writer, "Pairwise df matrix", result.table.sources, pairwise_results, lambda item: item.df)
            write_pairwise_matrix(writer, "Pairwise p-value matrix", result.table.sources, pairwise_results, lambda item: raw_p_value(item.p_value))
            writer.writerow(["Observed vs expected"])
            writer.writerow(["bin_start", "source", "observed", "expected", "observed_minus_expected", "pearson_residual"])
            for bin_label in result.table.bins:
                for source_index, source in enumerate(result.table.sources):
                    observed = result.observed[bin_label][source_index]
                    expected = result.expected[bin_label][source_index]
                    residual = result.residuals[bin_label][source_index]
                    writer.writerow([bin_label, source, observed, format_cell(expected), format_cell(observed - expected), format_cell(residual)])
            writer.writerow([])

    with spearman_path.open("w", encoding="utf-8", newline="") as handle:
        writer = SpreadsheetSafeWriter(csv.writer(handle))
        for table in iter_statistical_tables(source_order, histogram_tables):
            if len(table.sources) < 2:
                continue
            pairwise_spearman = pairwise_spearman_results(table)
            if not pairwise_spearman:
                continue
            writer.writerow([f"Spearman rank correlation: {table.statistic}"])
            writer.writerow(["classification", table.classification])
            writer.writerow([])
            write_pairwise_matrix(writer, "Spearman rho matrix", table.sources, pairwise_spearman, lambda item: format_cell(item.rho))
            write_pairwise_matrix(writer, "Spearman p-value matrix", table.sources, pairwise_spearman, lambda item: raw_p_value(item.p_value))
            write_pairwise_matrix(writer, "Spearman n matrix", table.sources, pairwise_spearman, lambda item: item.n)
            writer.writerow([])
    return [chi_path, spearman_path]


def write_pairwise_matrix(writer: csv.writer, title: str, sources: Sequence[str], results: Mapping[tuple[str, str], object], formatter) -> None:
    writer.writerow([title])
    writer.writerow(["source", *sources])
    for row_source in sources:
        row = [row_source]
        for column_source in sources:
            result = results.get((row_source, column_source)) or results.get((column_source, row_source))
            row.append("" if row_source == column_source or result is None else formatter(result))
        writer.writerow(row)
    writer.writerow([])


def pairwise_spearman_results(table: HistogramTable) -> dict[tuple[str, str], SpearmanResult]:
    results: dict[tuple[str, str], SpearmanResult] = {}
    for left_index, right_index in combinations(range(len(table.sources)), 2):
        left = table.sources[left_index]
        right = table.sources[right_index]
        left_counts = [table.rows[bin_label][left_index] for bin_label in table.bins]
        right_counts = [table.rows[bin_label][right_index] for bin_label in table.bins]
        if len(left_counts) < 3:
            continue
        results[(left, right)] = spearman_test(left_counts, right_counts)
    return results


def iter_statistical_tables(
    source_order: Sequence[str],
    histogram_tables: Mapping[str, Mapping[str, HistogramTable]],
) -> Iterator[HistogramTable]:
    """Yield tables in the same human-facing order as the generated CSVs."""

    for classification in ("emotion_0_to_100", "other_0_to_100", "valence_minus100_to_100"):
        for table in histogram_tables.get(classification, {}).values():
            yield reorder_table_sources(table, source_order)


def reorder_table_sources(table: HistogramTable, source_order: Sequence[str]) -> HistogramTable:
    ordered_sources = [source for source in source_order if source in table.sources]
    ordered_sources.extend(source for source in table.sources if source not in ordered_sources)
    old_indexes = [table.sources.index(source) for source in ordered_sources]
    rows = {
        bin_label: [table.rows[bin_label][old_index] for old_index in old_indexes]
        for bin_label in table.bins
        if bin_label in table.rows
    }
    return HistogramTable(table.classification, table.statistic, ordered_sources, list(table.bins), rows)


def chi_square_test(table: HistogramTable) -> ChiSquareResult:
    observed = {bin_label: list(counts) for bin_label, counts in table.rows.items() if sum(counts) > 0}
    bins = list(observed)
    sources = list(table.sources)
    total = sum(sum(counts) for counts in observed.values())
    if total == 0:
        raise ValueError("Cannot run chi-square test on an empty table.")

    col_totals = [sum(observed[bin_label][index] for bin_label in bins) for index in range(len(sources))]
    expected: dict[str, list[float]] = {}
    residuals: dict[str, list[float]] = {}
    x_squared = 0.0
    for bin_label in bins:
        row_total = sum(observed[bin_label])
        expected_row: list[float] = []
        residual_row: list[float] = []
        for index, col_total in enumerate(col_totals):
            expected_value = row_total * col_total / total
            expected_row.append(expected_value)
            residual = 0.0 if expected_value == 0 else (observed[bin_label][index] - expected_value) / math.sqrt(expected_value)
            residual_row.append(residual)
            x_squared += residual * residual
        expected[bin_label] = expected_row
        residuals[bin_label] = residual_row

    df = (len(bins) - 1) * (len(sources) - 1)
    p_value = chi_square_sf(x_squared, df) if df > 0 else math.nan
    trimmed = HistogramTable(table.classification, table.statistic, sources, bins, observed)
    return ChiSquareResult(trimmed, observed, expected, residuals, x_squared, df, p_value)


def pairwise_chi_square(table: HistogramTable) -> dict[tuple[str, str], ChiSquareResult]:
    results: dict[tuple[str, str], ChiSquareResult] = {}
    for left_index, right_index in combinations(range(len(table.sources)), 2):
        left = table.sources[left_index]
        right = table.sources[right_index]
        rows: dict[str, list[int]] = {}
        for bin_label in table.bins:
            counts = [table.rows[bin_label][left_index], table.rows[bin_label][right_index]]
            if sum(counts) > 0:
                rows[bin_label] = counts
        if rows:
            pair_table = HistogramTable(table.classification, table.statistic, [left, right], list(rows), rows)
            results[(left, right)] = chi_square_test(pair_table)
    return results


def append_observed_expected_rows(rows: list[dict[str, object]], result: ChiSquareResult) -> None:
    for bin_label in result.table.bins:
        for source_index, source in enumerate(result.table.sources):
            observed = result.observed[bin_label][source_index]
            expected = result.expected[bin_label][source_index]
            residual = result.residuals[bin_label][source_index]
            rows.append(
                {
                    "section": "Observed vs expected",
                    "classification": result.table.classification,
                    "statistic": result.table.statistic,
                    "bin_start": bin_label,
                    "source": source,
                    "observed": observed,
                    "expected": format_cell(expected),
                    "observed_minus_expected": format_cell(observed - expected),
                    "pearson_residual": format_cell(residual),
                }
            )


def append_pairwise_chi_rows(
    rows: list[dict[str, object]],
    result: ChiSquareResult,
    pairwise_results: Mapping[tuple[str, str], ChiSquareResult],
) -> None:
    for (left, right), pair_result in pairwise_results.items():
        rows.append(
            {
                "section": "Pairwise Pearson chi-square",
                "classification": result.table.classification,
                "statistic": result.table.statistic,
                "source_a": left,
                "source_b": right,
                "x_squared": format_number(pair_result.x_squared),
                "df": pair_result.df,
                "p_value": raw_p_value(pair_result.p_value),
            }
        )


def append_pairwise_matrix_rows(
    rows: list[dict[str, object]],
    sources: Sequence[str],
    classification: str,
    statistic: str,
    pairwise_results: Mapping[tuple[str, str], ChiSquareResult],
) -> None:
    matrix_specs = [
        ("Pairwise X-squared matrix", lambda result: format_cell(result.x_squared)),
        ("Pairwise df matrix", lambda result: result.df),
        ("Pairwise p-value matrix", lambda result: raw_p_value(result.p_value)),
    ]
    for section, formatter in matrix_specs:
        for row_source in sources:
            for col_source in sources:
                result = pairwise_results.get((row_source, col_source)) or pairwise_results.get((col_source, row_source))
                rows.append(
                    {
                        "section": section,
                        "classification": classification,
                        "statistic": statistic,
                        "matrix_row_source": row_source,
                        "matrix_column_source": col_source,
                        "value": "" if row_source == col_source or result is None else formatter(result),
                    }
                )


def append_spearman_rows(
    rows: list[dict[str, object]],
    table: HistogramTable,
    pairwise_results: Mapping[tuple[str, str], ChiSquareResult],
) -> None:
    for left_index, right_index in combinations(range(len(table.sources)), 2):
        left = table.sources[left_index]
        right = table.sources[right_index]
        pair_result = pairwise_results.get((left, right))
        if pair_result is None or len(pair_result.table.bins) < 3:
            continue
        left_counts = [pair_result.observed[bin_label][0] for bin_label in pair_result.table.bins]
        right_counts = [pair_result.observed[bin_label][1] for bin_label in pair_result.table.bins]
        spearman = spearman_test(left_counts, right_counts)
        rows.append(
            {
                "section": "Spearman rank correlation",
                "classification": table.classification,
                "statistic": table.statistic,
                "source_a": left,
                "source_b": right,
                "S": format_number(spearman.s),
                "rho": format_number(spearman.rho),
                "p_value": raw_p_value(spearman.p_value),
                "n": spearman.n,
            }
        )


def spearman_test(left: Sequence[float], right: Sequence[float]) -> SpearmanResult:
    if len(left) != len(right):
        raise ValueError("Spearman inputs must have the same length.")
    if len(left) < 3:
        return SpearmanResult(math.nan, math.nan, math.nan, len(left))
    left_ranks = average_ranks(left)
    right_ranks = average_ranks(right)
    rho = pearson_correlation(left_ranks, right_ranks)
    if math.isnan(rho):
        return SpearmanResult(math.nan, math.nan, math.nan, len(left))
    s_value = (1.0 - rho) * len(left) * (len(left) * len(left) - 1.0) / 6.0
    if abs(rho) >= 1:
        p_value = 0.0
    else:
        df = len(left) - 2
        t_stat = abs(rho) * math.sqrt(df / (1 - rho * rho))
        p_value = student_t_two_tailed_p(t_stat, df)
    return SpearmanResult(rho, s_value, p_value, len(left))


def average_ranks(values: Sequence[float]) -> list[float]:
    indexed = sorted(enumerate(values), key=lambda item: item[1])
    ranks = [0.0] * len(values)
    i = 0
    while i < len(indexed):
        j = i + 1
        while j < len(indexed) and indexed[j][1] == indexed[i][1]:
            j += 1
        rank = (i + 1 + j) / 2.0
        for k in range(i, j):
            ranks[indexed[k][0]] = rank
        i = j
    return ranks


def pearson_correlation(left: Sequence[float], right: Sequence[float]) -> float:
    if len(left) != len(right) or len(left) < 2:
        return math.nan
    left_mean = sum(left) / len(left)
    right_mean = sum(right) / len(right)
    numerator = sum((a - left_mean) * (b - right_mean) for a, b in zip(left, right))
    left_ss = sum((a - left_mean) ** 2 for a in left)
    right_ss = sum((b - right_mean) ** 2 for b in right)
    denominator = math.sqrt(left_ss * right_ss)
    return numerator / denominator if denominator else math.nan


def chi_square_sf(x: float, df: int) -> float:
    return regularized_gamma_q(df / 2.0, x / 2.0)


def student_t_two_tailed_p(t_stat: float, df: int) -> float:
    x = df / (df + t_stat * t_stat)
    return regularized_beta(x, df / 2.0, 0.5)


def regularized_gamma_q(a: float, x: float) -> float:
    if x < 0 or a <= 0:
        raise ValueError("regularized_gamma_q requires x >= 0 and a > 0.")
    if x == 0:
        return 1.0
    if x < a + 1.0:
        return max(0.0, 1.0 - regularized_gamma_p_series(a, x))
    return regularized_gamma_q_fraction(a, x)


def regularized_beta(x: float, a: float, b: float) -> float:
    if x <= 0.0:
        return 0.0
    if x >= 1.0:
        return 1.0
    log_bt = math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log1p(-x)
    bt = math.exp(log_bt)
    if x < (a + 1.0) / (a + b + 2.0):
        return bt * beta_continued_fraction(a, b, x) / a
    return 1.0 - bt * beta_continued_fraction(b, a, 1.0 - x) / b


def regularized_gamma_p_series(a: float, x: float) -> float:
    eps = 3e-14
    ap = a
    delta = 1.0 / a
    total = delta
    for _ in range(1000):
        ap += 1.0
        delta *= x / ap
        total += delta
        if abs(delta) < abs(total) * eps:
            return total * math.exp(-x + a * math.log(x) - math.lgamma(a))
    raise RuntimeError("regularized gamma series did not converge.")


def regularized_gamma_q_fraction(a: float, x: float) -> float:
    eps = 3e-14
    fpmin = 1e-300
    b = x + 1.0 - a
    c = 1.0 / fpmin
    d = 1.0 / max(b, fpmin)
    h = d
    for i in range(1, 1000):
        an = -i * (i - a)
        b += 2.0
        d = an * d + b
        if abs(d) < fpmin:
            d = fpmin
        c = b + an / c
        if abs(c) < fpmin:
            c = fpmin
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < eps:
            return math.exp(-x + a * math.log(x) - math.lgamma(a)) * h
    raise RuntimeError("regularized gamma fraction did not converge.")


def beta_continued_fraction(a: float, b: float, x: float) -> float:
    eps = 3e-14
    fpmin = 1e-300
    qab = a + b
    qap = a + 1.0
    qam = a - 1.0
    c = 1.0
    d = 1.0 - qab * x / qap
    if abs(d) < fpmin:
        d = fpmin
    d = 1.0 / d
    h = d
    for m in range(1, 1000):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < fpmin:
            d = fpmin
        c = 1.0 + aa / c
        if abs(c) < fpmin:
            c = fpmin
        d = 1.0 / d
        h *= d * c

        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < fpmin:
            d = fpmin
        c = 1.0 + aa / c
        if abs(c) < fpmin:
            c = fpmin
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < eps:
            return h
    raise RuntimeError("regularized beta fraction did not converge.")


def write_dict_rows(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = SpreadsheetSafeWriter(csv.DictWriter(handle, fieldnames=fieldnames))
        writer.writeheader()
        writer.writerows(rows)


def format_cell(value: object) -> object:
    if isinstance(value, float):
        return format_number(value)
    return value


def format_number(value: object) -> str:
    if value is None:
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if math.isnan(number):
        return "NA"
    if math.isinf(number):
        return "Inf" if number > 0 else "-Inf"
    if abs(number - round(number)) < 1e-12:
        return str(int(round(number)))
    formatted = f"{number:.2f}"
    if formatted in {"-0.00", "0.00"}:
        return "0"
    return formatted.rstrip("0").rstrip(".")


def format_p_value(value: float) -> str:
    if math.isnan(value):
        return "NA"
    if value < 2.2e-16:
        return "< 2.2e-16"
    if value < 0.001:
        return f"= {value:.2e}"
    return f"= {format_number(value)}"


def raw_p_value(value: float) -> str:
    if math.isnan(value):
        return "NA"
    if value < 2.2e-16:
        return format_p_value(value)
    if value < 0.001:
        return f"{value:.2e}"
    return format_number(value)


def safe_filename(text: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "", str(text))
    cleaned = re.sub(r"\s+", "_", cleaned.strip())
    cleaned = re.sub(r"_+", "_", cleaned)
    cleaned = cleaned.strip("._ ")
    return cleaned or "output"


