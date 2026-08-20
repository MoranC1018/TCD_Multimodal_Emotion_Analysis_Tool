from __future__ import annotations

import csv
import json
import re
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from analysis.text_pipeline.postprocess import (
    DEFAULT_SEGMENT_SAMPLE_COUNTS,
    analyse_text_segments_folder,
    build_whisper_index,
    detect_text_variant,
    evenly_spaced_indexes,
    format_graph_number,
    describe,
    infer_default_whisper_root,
    normalize_segment_sample_counts,
    parse_segment_sample_counts,
    parse_video_name,
    resolve_segment_output_folder,
    render_grouped_bar_svg,
    validate_text_asset_path,
)
from analysis.text_pipeline.contracts import Category
from analysis.text_pipeline.constructs import CONSTRUCTS, construct_row
from analysis.text_pipeline.provenance import inspect_upstream_provenance
from analysis.text_pipeline.reports import SpeakerQuickRow, render_speaker_quick_table
from analysis.text_pipeline.transaction import replace_output_dir
from processing.text_analysis.contracts import file_sha256, inventory_digest
from processing.text_analysis.prepare_input.whisper_to_rocksteady import (
    PreparedSegment,
    replace_segment_directory,
)


ROCKSTEADY_ROWS = (
    ("10", "1", "0", "2", "0", "3", "2", "1"),
    ("5", "0", "1", "0", "0", "0", "1", "2"),
    ("0", "0", "0", "0", "0", "0", "0", "0"),
)


def write_rocksteady_csv(path: Path, *, include_moral: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "Title",
        "Terms",
        "Active",
        "Negativ",
        "Research Theme",
        "Passive",
        "Positiv",
        "Strong",
        "Weak",
    ]
    if include_moral:
        headers.append("Moral")
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for segment_id, values in enumerate(ROCKSTEADY_ROWS, start=1):
            row = [f"{path.stem}__segment_{segment_id:06d}", *values]
            if include_moral:
                row.append("1" if segment_id == 1 else "0")
            writer.writerow(row)


def write_valid_rocksteady_csv(path: Path, segment_count: int) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Title", "Terms", "Active", "Negativ", "Research Theme", "Passive", "Positiv", "Strong", "Weak"])
        for segment_id in range(1, segment_count + 1):
            writer.writerow(
                [
                    f"{path.stem}__segment_{segment_id:06d}",
                    "10",
                    str(segment_id % 2),
                    "0",
                    str(segment_id % 3),
                    "0",
                    str(segment_id % 4),
                    "1",
                    "0",
                ]
            )


def write_whisper_json(
    whisper_root: Path,
    video_stem: str,
    *,
    country: str = "France",
    speaker: str = "Test Speaker",
    segment_count: int = 3,
) -> Path:
    path = whisper_root / country / speaker / f"{video_stem}.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    segments = [
        {
            "id": index,
            "start": float((index - 1) * 2),
            "end": float(index * 2),
            "text": f"Test segment {index} text",
        }
        for index in range(1, segment_count + 1)
    ]
    path.write_text(
        json.dumps({"language": "en", "segments": segments}),
        encoding="utf-8",
    )
    return path


def write_custom_rocksteady_csv(
    path: Path,
    categories: list[str],
    *,
    segment_count: int = 1,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Title", "Terms", *categories])
        for segment_id in range(1, segment_count + 1):
            writer.writerow(
                [
                    f"{path.stem}__segment_{segment_id:06d}",
                    "100",
                    *("1" for _ in categories),
                ]
            )


def write_prepare_mapping(
    prepare_root: Path,
    *,
    country: str,
    speaker: str,
    video: str,
    segments: list[PreparedSegment],
) -> Path:
    video_dir = prepare_root / country / speaker / video
    replace_segment_directory(
        video_dir,
        video,
        segments,
        video_identity=f"{country}/{speaker}/{video}",
    )
    return video_dir / ".prepare_manifest.json"


def write_adapter_manifest(
    input_dir: Path,
    csv_path: Path,
    categories: list[str],
    *,
    configured_categories: list[str] | None = None,
    include_hash: bool = False,
) -> Path:
    relative = csv_path.relative_to(input_dir).as_posix()
    video_record: dict[str, object] = {
        "identity": csv_path.relative_to(input_dir).with_suffix("").as_posix(),
        "output": relative,
        "status": "completed",
        "validation": {
            "rows": sum(1 for _ in read_rows(csv_path)),
            "columns": len(categories) + 2,
            "categories": categories,
        },
    }
    if include_hash:
        video_record["output_sha256"] = file_sha256(csv_path)
    payload = {
        "schema_version": 1,
        "settings": {
            "categories": configured_categories if configured_categories is not None else categories,
            "value_type": "total",
            "analyser": "simple",
            "combination": "merge",
        },
        "summary": {"total": 1, "completed": 1, "skipped": 0, "failed": 0},
        "videos": [video_record],
        "dictionaries": [],
        "rocksteady_jar": {},
    }
    manifest = input_dir / "_manifests" / "rocksteady_run_manifest.json"
    manifest.parent.mkdir(parents=True, exist_ok=True)
    manifest.write_text(json.dumps(payload), encoding="utf-8")
    return manifest


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


class TextSegmentPostprocessingTests(unittest.TestCase):
    def test_whisper_index_ignores_recognized_pipeline_metadata_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "selected"
            video = "001_France_Test_Speaker_20250101"
            transcript = write_whisper_json(root, video)
            (root / ".text_pipeline_owner.json").write_text("{}", encoding="utf-8")
            (root / "selection_manifest.json").write_text("{}", encoding="utf-8")
            (root / "_manifests").mkdir()
            (root / "_manifests" / "transcription_run_manifest.json").write_text(
                "{}", encoding="utf-8"
            )

            index = build_whisper_index(root)

            self.assertEqual(list(index.values()), [transcript])

            (root / "unexpected_manifest.json").write_text("{}", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "Invalid Whisper JSON layout"):
                build_whisper_index(root)

    def test_language_variant_selects_separate_default_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            original_input = root / "parse_output" / "original"
            english_input = root / "parse_output" / "eng"
            output_base = root / "postprocessing_output"

            self.assertEqual(detect_text_variant(original_input), "original")
            self.assertEqual(detect_text_variant(english_input), "eng")
            self.assertEqual(
                infer_default_whisper_root(original_input),
                Path("processing/text_analysis/output/current/transcripts/original"),
            )
            self.assertEqual(
                infer_default_whisper_root(english_input),
                Path("processing/text_analysis/output/current/transcripts/eng"),
            )
            with patch("analysis.text_pipeline.postprocess.default_output_root", return_value=output_base):
                self.assertEqual(
                    resolve_segment_output_folder(original_input, None),
                    (output_base / "text" / "text_output" / "original").resolve(),
                )
                self.assertEqual(
                    resolve_segment_output_folder(english_input, None),
                    (output_base / "text" / "text_output" / "eng").resolve(),
                )

    def test_procurement_speaker_video_layout_needs_no_country(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            speaker = "Research Speaker"
            video = "Interviews_[abc123]"
            csv_path = input_dir / speaker / f"{video}.csv"
            write_rocksteady_csv(csv_path)
            write_whisper_json(
                whisper_root,
                video,
                country="",
                speaker=speaker,
            )

            self.assertEqual(
                validate_text_asset_path(csv_path, input_dir, asset_label="RockSteady CSV"),
                ("", speaker),
            )
            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                write_graphs=False,
            )
            video_rows = read_rows(output_dir / "video_level_summary.csv")
            self.assertEqual(video_rows[0]["country"], "")
            self.assertEqual(video_rows[0]["speaker"], speaker)
            self.assertTrue(
                (
                    output_dir
                    / "segment_counts"
                    / speaker
                    / f"{video}_segment_counts.csv"
                ).is_file()
            )

    def test_country_speaker_layout_matches_whisper_json_by_speaker(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            video_stem = "001_France_Research_Speaker_20260302"

            write_rocksteady_csv(
                input_dir / "France" / "Research Speaker" / f"{video_stem}.csv"
            )
            write_whisper_json(
                whisper_root,
                video_stem,
                speaker="Research Speaker",
            )

            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                write_graphs=False,
            )

            video_rows = read_rows(output_dir / "video_level_summary.csv")
            self.assertEqual(video_rows[0]["speaker"], "Research Speaker")
            self.assertEqual(video_rows[0]["country"], "France")
            self.assertEqual(video_rows[0]["person"], "Research_Speaker")

    def test_country_folder_must_match_canonical_video_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            video_stem = "001_France_Research_Speaker_20260302"
            write_rocksteady_csv(
                input_dir / "Italy" / "Research Speaker" / f"{video_stem}.csv"
            )
            write_whisper_json(whisper_root, video_stem, speaker="Research Speaker")

            with self.assertRaisesRegex(ValueError, "Country mismatch"):
                analyse_text_segments_folder(
                    input_dir,
                    output_root=root / "output",
                    whisper_root=whisper_root,
                    write_graphs=False,
                )

    def test_speaker_folder_must_match_canonical_video_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            video_stem = "001_France_Research_Speaker_20260302"
            write_rocksteady_csv(
                input_dir / "France" / "Other Speaker" / f"{video_stem}.csv"
            )
            write_whisper_json(whisper_root, video_stem, speaker="Research Speaker")

            with self.assertRaisesRegex(ValueError, "Speaker mismatch"):
                analyse_text_segments_folder(
                    input_dir,
                    output_root=root / "output",
                    whisper_root=whisper_root,
                    write_graphs=False,
                )

    def test_same_speaker_name_in_two_countries_has_distinct_speaker_ids(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            for country in ("France", "UK"):
                video_stem = f"001_{country}_Alex_Martin_20250101"
                write_rocksteady_csv(
                    input_dir / country / "Alex Martin" / f"{video_stem}.csv"
                )
                write_whisper_json(
                    whisper_root,
                    video_stem,
                    country=country,
                    speaker="Alex Martin",
                )

            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                write_graphs=False,
            )

            rows = read_rows(output_dir / "speaker_level_summary.csv")
            self.assertEqual({row["speaker_id"] for row in rows}, {"France/Alex Martin", "UK/Alex Martin"})
            self.assertTrue(
                (output_dir / "segment_counts" / "France" / "Alex Martin" / "001_France_Alex_Martin_20250101_segment_counts.csv").exists()
            )
            self.assertTrue(
                (output_dir / "segment_counts" / "UK" / "Alex Martin" / "001_UK_Alex_Martin_20250101_segment_counts.csv").exists()
            )

    def test_segment_directory_replacement_removes_stale_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            video_stem = "001 - Test Speaker - 2025-01-01"
            video_dir = root / video_stem
            replace_segment_directory(
                video_dir,
                video_stem,
                [(1, "old one"), (2, "old two"), (3, "stale text")],
            )
            stale = video_dir / f"{video_stem}__segment_000003.txt"
            self.assertTrue(stale.is_file())

            replace_segment_directory(
                video_dir,
                video_stem,
                [(1, "new one"), (2, "new two")],
            )

            names = sorted(path.name for path in video_dir.glob("*.txt"))
            self.assertEqual(
                names,
                [
                    f"{video_stem}__segment_000001.txt",
                    f"{video_stem}__segment_000002.txt",
                ],
            )
            self.assertFalse(stale.exists())

    def test_segment_sample_counts_and_indexes(self) -> None:
        self.assertEqual(DEFAULT_SEGMENT_SAMPLE_COUNTS, (3, 5, 7, 9))
        self.assertEqual(normalize_segment_sample_counts((9, 3, 5, 3)), (3, 5, 9))
        self.assertEqual(parse_segment_sample_counts("3,7,11"), (3, 7, 11))
        self.assertEqual(parse_segment_sample_counts("none"), ())
        self.assertEqual(evenly_spaced_indexes(9, 5), [0, 2, 4, 6, 8])
        self.assertEqual(evenly_spaced_indexes(4, 5), [])
        with self.assertRaisesRegex(ValueError, "odd integers"):
            normalize_segment_sample_counts((4,))

    def test_custom_segment_samples_write_only_requested_coarse_timelines(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            video_stem = "001_France_Test_Speaker_20250101"
            write_valid_rocksteady_csv(
                input_dir / "France" / "Test_Speaker" / f"{video_stem}.csv",
                segment_count=9,
            )
            write_whisper_json(whisper_root, video_stem, segment_count=9)

            result = analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                segment_sample_counts=(3, 7),
            )

            graph_dir = output_dir / "graphs" / "videos" / "France" / "Test_Speaker" / video_stem
            self.assertTrue((graph_dir / "segment_timeline_counts_full.svg").exists())
            sampled_three = graph_dir / "segment_timeline_counts_sampled_3.svg"
            sampled_seven = graph_dir / "segment_timeline_counts_sampled_7.svg"
            self.assertTrue(sampled_three.exists())
            self.assertTrue(sampled_seven.exists())
            self.assertFalse((graph_dir / "segment_timeline_counts_sampled_5.svg").exists())
            sampled_three_svg = sampled_three.read_text(encoding="utf-8")
            self.assertIn("first and last retained", sampled_three_svg)
            for segment_id in (1, 5, 9):
                self.assertIn(f'text-anchor="middle">{segment_id}</text>', sampled_three_svg)
            self.assertIn("Segment timeline sample counts: 3,7", (output_dir / "run_log.txt").read_text(encoding="utf-8"))
            self.assertIn(sampled_three, result.graph_paths)

    def test_small_graph_percentages_keep_visible_precision(self) -> None:
        self.assertEqual(format_graph_number(1.8), "1.8")
        self.assertEqual(format_graph_number(0.18), "0.18")

    def test_parse_dash_video_name(self) -> None:
        self.assertEqual(
            parse_video_name("001 - Example Speaker - 2025-11-07"),
            {
                "video_order": "001",
                "date": "2025-11-07",
                "country": "",
                "person": "Example Speaker",
            },
        )

    def test_parse_canonical_country_person_date_video_name(self) -> None:
        self.assertEqual(
            parse_video_name("001_France_Research-Lead_Speaker_20251123"),
            {
                "video_order": "001",
                "date": "20251123",
                "country": "France",
                "person": "Research-Lead_Speaker",
            },
        )
        self.assertEqual(
            parse_video_name("003_UK_Other_Speaker_202601"),
            {
                "video_order": "003",
                "date": "202601",
                "country": "UK",
                "person": "Other_Speaker",
            },
        )

    def test_segment_postprocessing_writes_professor_tables_summaries_and_graphs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            video_stems = (
                "001_France_Test_Speaker_20250101",
                "002_France_Test_Speaker_20250201",
            )
            for video_stem in video_stems:
                write_rocksteady_csv(input_dir / "France" / "Test_Speaker" / f"{video_stem}.csv")
                write_whisper_json(whisper_root, video_stem)
            result = analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
            )

            self.assertEqual(result.csv_count, 2)
            self.assertEqual(result.segment_count, 6)
            count_path = (
                output_dir
                / "segment_counts"
                / "France"
                / "Test_Speaker"
                / "001_France_Test_Speaker_20250101_segment_counts.csv"
            )
            relative_path = (
                output_dir
                / "segment_relative"
                / "France"
                / "Test_Speaker"
                / "001_France_Test_Speaker_20250101_segment_relative.csv"
            )
            enriched_path = (
                output_dir
                / "segment_level"
                / "France"
                / "Test_Speaker"
                / "001_France_Test_Speaker_20250101_segments_enriched.csv"
            )
            self.assertTrue(count_path.exists())
            self.assertTrue(relative_path.exists())
            self.assertTrue(enriched_path.exists())
            self.assertTrue((output_dir / "video_level_summary.csv").exists())
            self.assertTrue((output_dir / "speaker_level_summary.csv").exists())
            self.assertTrue((output_dir / "descriptor_statistics_by_video.csv").exists())
            comparison_dir = (
                output_dir
                / "video_mean_comparisons"
                / "France"
                / "Test_Speaker"
                / "combined"
            )
            self.assertTrue((comparison_dir / "video_means.csv").exists())
            self.assertTrue(
                (comparison_dir / "permutation_test_results.csv").exists()
            )
            self.assertFalse((output_dir / "video_distribution_comparisons").exists())
            descriptor_workbook_path = output_dir / "descriptor_statistics_by_video.xlsx"
            self.assertTrue(descriptor_workbook_path.exists())
            descriptor_workbook = load_workbook(descriptor_workbook_path)
            self.assertEqual(len(descriptor_workbook.sheetnames), 2)
            descriptor_sheet = descriptor_workbook[descriptor_workbook.sheetnames[0]]
            self.assertEqual(descriptor_sheet["A1"].value, "Video")
            self.assertTrue(descriptor_sheet.sheet_view.showGridLines)
            self.assertIsNone(descriptor_sheet["A4"].fill.fill_type)
            self.assertIsNone(descriptor_sheet["A5"].fill.fill_type)
            self.assertIsNone(descriptor_sheet["A5"].border.bottom.style)
            self.assertEqual(descriptor_sheet["A5"].value, "Mean")
            self.assertEqual(descriptor_sheet["B5"].data_type, "n")
            self.assertEqual(descriptor_sheet["B5"].number_format, "0.######")
            self.assertEqual(descriptor_sheet["B17"].number_format, "#,##0")
            self.assertEqual(descriptor_sheet["A6"].value, "Standard Error")
            self.assertEqual(descriptor_sheet["A18"].value, "Confidence Level (95.0%)")
            descriptor_workbook.close()
            readable_root = output_dir / "readable"
            self.assertTrue((readable_root / "README.md").exists())
            readable_video = read_rows(readable_root / "video_level_summary.csv")[0]
            self.assertEqual(readable_video["Positive Sentiment (% of terms)"], "20")
            self.assertNotIn("positive_proportion", readable_video)
            readable_segment = read_rows(
                readable_root
                / "segment_level"
                / "France"
                / "Test_Speaker"
                / "001_France_Test_Speaker_20250101_segments.csv"
            )[0]
            self.assertEqual(readable_segment["Transcript text"], "Test segment 1 text")
            self.assertEqual(readable_segment["Positive Sentiment count"], "3")
            self.assertEqual(readable_segment["Positive Sentiment (% of terms)"], "30")
            self.assertTrue(
                list((readable_root / "graphs" / "videos").rglob("*.svg"))
            )
            self.assertTrue(
                list((readable_root / "graphs" / "speakers").rglob("*.svg"))
            )
            self.assertTrue(
                (
                    output_dir
                    / "graphs"
                    / "videos"
                    / "France"
                    / "Test_Speaker"
                    / "001_France_Test_Speaker_20250101"
                    / "segment_timeline_counts_full.svg"
                ).exists()
            )
            self.assertTrue(
                (output_dir / "graphs" / "speakers" / "France" / "Test_Speaker" / "video_proportions.svg").exists()
            )
            self.assertTrue(
                (output_dir / "graphs" / "speakers" / "France" / "Test_Speaker" / "video_valence.svg").exists()
            )
            self.assertTrue(
                (output_dir / "graphs" / "summary" / "France" / "speaker_summary_proportions.svg").exists()
            )
            self.assertTrue(
                (output_dir / "graphs" / "summary" / "France" / "speaker_summary_valence.svg").exists()
            )
            self.assertFalse(
                (output_dir / "graphs" / "summary" / "speaker_summary_proportions.svg").exists()
            )
            country_summary_graph = (
                output_dir / "graphs" / "summary" / "France" / "speaker_summary_proportions.svg"
            ).read_text(encoding="utf-8")
            self.assertIn("France speaker summary percentages", country_summary_graph)

            count_rows = read_rows(count_path)
            relative_rows = read_rows(relative_path)
            self.assertEqual(count_rows[0]["segment_text"], "Test segment 1 text")
            self.assertEqual(count_rows[0]["start_sec"], "0")
            self.assertEqual(count_rows[0]["end_sec"], "2")
            self.assertEqual(count_rows[0]["strong_count"], "2")
            self.assertEqual(count_rows[0]["weak_count"], "1")
            self.assertNotIn("moral_count", count_rows[0])
            self.assertEqual(count_rows[0]["valence_score"], "1")
            self.assertEqual(relative_rows[0]["positive_proportion"], "0.3")
            self.assertEqual(relative_rows[0]["strong_proportion"], "0.2")
            self.assertNotIn("moral_proportion", relative_rows[0])
            self.assertEqual(relative_rows[1]["valence_score"], "-1")
            self.assertEqual(relative_rows[2]["valence_score"], "")

            video_rows = read_rows(output_dir / "video_level_summary.csv")
            first_video = next(row for row in video_rows if row["video"] == video_stems[0])
            self.assertEqual(first_video["positive_proportion"], "0.2")
            self.assertEqual(first_video["valence_score"], "0.5")
            self.assertNotIn("moral_available", first_video)
            self.assertNotIn("moral_total", first_video)
            self.assertEqual(first_video["whisper_language"], "en")

            descriptor_rows = read_rows(output_dir / "descriptor_statistics_by_video.csv")
            descriptor_names = {row["descriptor"] for row in descriptor_rows}
            self.assertTrue(
                {
                    "positive_count",
                    "positive_proportion",
                    "has_positive",
                    "strong_count",
                    "weak_proportion",
                    "valence_score",
                }.issubset(descriptor_names)
            )
            self.assertTrue(
                {
                    "start_sec",
                    "end_sec",
                    "duration_sec",
                    "whisper_word_count",
                    "valid_segment",
                    "positive_available",
                    "moral_available",
                    "rocksteady_terms",
                }.isdisjoint(descriptor_names)
            )

            speaker_graph = (
                output_dir / "graphs" / "speakers" / "France" / "Test_Speaker" / "video_proportions.svg"
            ).read_text(encoding="utf-8")
            self.assertIn("20%", speaker_graph)
            self.assertIn("Strong", speaker_graph)
            self.assertIn("Weak", speaker_graph)

            histogram = (
                output_dir
                / "graphs"
                / "videos"
                / "France"
                / "Test_Speaker"
                / "001_France_Test_Speaker_20250101"
                / "segment_count_histograms.svg"
            ).read_text(encoding="utf-8")
            self.assertIn("2 valid segments (Terms &gt; 0)", histogram)

            run_log = (output_dir / "run_log.txt").read_text(encoding="utf-8")
            self.assertIn("Expected RockSteady mode: Total", run_log)
            self.assertIn("Excluded zero-Term segments: 2", run_log)
            self.assertIn(str(whisper_root), run_log)

    def test_moral_is_populated_when_rocksteady_exports_it(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            video_stem = "001_France_Test_Speaker_20250101"
            write_rocksteady_csv(
                input_dir / "France" / "Test_Speaker" / f"{video_stem}.csv",
                include_moral=True,
            )
            write_whisper_json(whisper_root, video_stem)

            analyse_text_segments_folder(
                input_dir,
                output_root=root / "output",
                whisper_root=whisper_root,
                write_graphs=False,
            )

            count_rows = read_rows(
                root
                / "output"
                / "segment_counts"
                / "France"
                / "Test_Speaker"
                / f"{video_stem}_segment_counts.csv"
            )
            summary = read_rows(root / "output" / "video_level_summary.csv")[0]
            self.assertEqual(count_rows[0]["moral_count"], "1")
            self.assertEqual(summary["moral_available"], "1")
            self.assertEqual(summary["moral_total"], "1")
            self.assertEqual(summary["moral_proportion"], "0.066667")

    def test_whisper_segment_count_mismatch_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            video_stem = "001_France_Test_Speaker_20250101"
            write_rocksteady_csv(input_dir / "France" / "Test_Speaker" / f"{video_stem}.csv")
            write_whisper_json(whisper_root, video_stem, segment_count=2)

            with self.assertRaisesRegex(ValueError, "Segment count mismatch"):
                analyse_text_segments_folder(
                    input_dir,
                    output_root=root / "output",
                    whisper_root=whisper_root,
                )

    def test_reconcile_alignment_ignores_extra_id_and_marks_missing_id(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            video_stem = "001_France_Test_Speaker_20250101"
            csv_path = input_dir / "France" / "Test_Speaker" / f"{video_stem}.csv"
            write_valid_rocksteady_csv(csv_path, segment_count=2)
            with csv_path.open("a", encoding="utf-8", newline="") as handle:
                csv.writer(handle).writerow(
                    [
                        f"{video_stem}__segment_000004",
                        "10",
                        "1",
                        "0",
                        "0",
                        "0",
                        "1",
                        "1",
                        "0",
                    ]
                )
            write_whisper_json(whisper_root, video_stem, segment_count=3)

            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                write_graphs=False,
                segment_alignment="reconcile",
            )

            count_rows = read_rows(
                output_dir
                / "segment_counts"
                / "France"
                / "Test_Speaker"
                / f"{video_stem}_segment_counts.csv"
            )
            self.assertEqual(len(count_rows), 3)
            self.assertEqual(count_rows[2]["segment_id"], "3")
            self.assertEqual(count_rows[2]["rocksteady_row_available"], "0")
            self.assertEqual(count_rows[2]["rocksteady_terms"], "")
            self.assertEqual(count_rows[2]["positive_count"], "")

            summary = read_rows(output_dir / "video_level_summary.csv")[0]
            self.assertEqual(summary["csv_rows"], "3")
            self.assertEqual(summary["whisper_segments"], "3")
            self.assertEqual(summary["rocksteady_matched_segments"], "2")
            self.assertEqual(summary["rocksteady_missing_segments"], "1")
            self.assertEqual(summary["rocksteady_ignored_extra_rows"], "1")

            audit = read_rows(output_dir / "segment_alignment_audit.csv")[0]
            self.assertEqual(audit["alignment_status"], "reconciled")
            self.assertEqual(audit["missing_segment_ids"], "3")
            self.assertEqual(audit["ignored_segment_ids"], "4")

    def test_bilingual_text_language_matches_rocksteady_input_language(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            video_stem = "001_France_Test_Speaker_20250101"
            write_valid_rocksteady_csv(
                input_dir / "France" / "Test_Speaker" / f"{video_stem}.csv",
                segment_count=1,
            )
            whisper_path = whisper_root / "France" / "Test Speaker" / f"{video_stem}.json"
            whisper_path.parent.mkdir(parents=True)
            whisper_path.write_text(
                json.dumps(
                    {
                        "language": "it",
                        "task": "bilingual",
                        "segments": [
                            {
                                "id": 0,
                                "start": 0.0,
                                "end": 2.0,
                                "text_original": "Testo italiano",
                                "text_en": "English test text",
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )

            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                write_graphs=False,
                text_language="en",
            )

            rows = read_rows(
                output_dir
                / "segment_counts"
                / "France"
                / "Test_Speaker"
                / f"{video_stem}_segment_counts.csv"
            )
            self.assertEqual(rows[0]["segment_text"], "English test text")
            self.assertEqual(rows[0]["whisper_language"], "it")
            self.assertEqual(rows[0]["text_language"], "en")
            self.assertEqual(rows[0]["whisper_word_count"], "3")

    def test_empty_input_preserves_previous_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            output_dir = root / "output"
            input_dir.mkdir()
            output_dir.mkdir()
            previous = output_dir / "previous.csv"
            previous.write_text("keep me", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "No RockSteady segment CSV"):
                analyse_text_segments_folder(input_dir, output_root=output_dir)

            self.assertEqual(previous.read_text(encoding="utf-8"), "keep me")

    def test_failed_generation_preserves_previous_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            output_dir = root / "output"
            video_stem = "001_France_Test_Speaker_20250101"
            write_rocksteady_csv(input_dir / "France" / "Test_Speaker" / f"{video_stem}.csv")
            write_whisper_json(whisper_root, video_stem)
            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                write_graphs=False,
                run_id="previous-valid-run",
            )
            previous_manifest = (output_dir / "output_manifest.json").read_bytes()

            with patch("analysis.text_pipeline.postprocess.write_all_graphs", side_effect=RuntimeError("graph failure")):
                with self.assertRaisesRegex(RuntimeError, "graph failure"):
                    analyse_text_segments_folder(
                        input_dir,
                        output_root=output_dir,
                        whisper_root=whisper_root,
                    )

            self.assertEqual(
                (output_dir / "output_manifest.json").read_bytes(),
                previous_manifest,
            )
            self.assertEqual(
                json.loads(previous_manifest)["run_id"],
                "previous-valid-run",
            )

    def test_missing_required_column_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            path = input_dir / "France" / "Test_Speaker" / "001_France_Test_Speaker_20250101.csv"
            path.parent.mkdir(parents=True)
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["Title", "Terms", "Positiv"])
                writer.writerow(["001_France_Test_Speaker_20250101__segment_000001", "10", "1"])
            write_whisper_json(whisper_root, path.stem, segment_count=1)

            with self.assertRaisesRegex(ValueError, "missing required Total-mode columns"):
                analyse_text_segments_folder(
                    input_dir,
                    output_root=root / "output",
                    whisper_root=whisper_root,
                )

    def test_percentage_like_values_are_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            path = input_dir / "France" / "Test_Speaker" / "001_France_Test_Speaker_20250101.csv"
            path.parent.mkdir(parents=True)
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(
                    ["Title", "Terms", "Active", "Negativ", "Research Theme", "Passive", "Positiv", "Strong", "Weak"]
                )
                writer.writerow(
                    ["001_France_Test_Speaker_20250101__segment_000001", "9", "11.11", "0", "0", "0", "1", "1", "1"]
                )
            write_whisper_json(whisper_root, path.stem, segment_count=1)

            with self.assertRaisesRegex(ValueError, "Expected a non-negative whole-number count"):
                analyse_text_segments_folder(
                    input_dir,
                    output_root=root / "output",
                    whisper_root=whisper_root,
                )

    def test_category_count_exceeding_terms_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            path = input_dir / "France" / "Test_Speaker" / "001_France_Test_Speaker_20250101.csv"
            path.parent.mkdir(parents=True)
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(
                    ["Title", "Terms", "Active", "Negativ", "Research Theme", "Passive", "Positiv", "Strong", "Weak"]
                )
                writer.writerow(
                    ["001_France_Test_Speaker_20250101__segment_000001", "4", "5", "0", "0", "0", "1", "1", "1"]
                )
            write_whisper_json(whisper_root, path.stem, segment_count=1)

            with self.assertRaisesRegex(ValueError, "exceeds Terms=4"):
                analyse_text_segments_folder(
                    input_dir,
                    output_root=root / "output",
                    whisper_root=whisper_root,
                )

    def test_malformed_title_is_rejected_in_error_and_reconcile_modes(self) -> None:
        for policy in ("error", "reconcile"):
            for bad_title in (
                "wrong_video__segment_000001",
                "001_France_Test_Speaker_20250101__segment_1",
            ):
                with (
                    self.subTest(policy=policy, bad_title=bad_title),
                    tempfile.TemporaryDirectory() as temp_dir,
                ):
                    root = Path(temp_dir)
                    input_dir = root / "parse_output"
                    whisper_root = root / "transcribe_output"
                    video = "001_France_Test_Speaker_20250101"
                    path = input_dir / "France" / "Test_Speaker" / f"{video}.csv"
                    write_rocksteady_csv(path)
                    rows = list(csv.reader(path.read_text(encoding="utf-8").splitlines()))
                    rows[1][0] = bad_title
                    with path.open("w", encoding="utf-8", newline="") as handle:
                        csv.writer(handle).writerows(rows)
                    write_whisper_json(whisper_root, video)

                    with self.assertRaisesRegex(
                        ValueError, "Invalid RockSteady segment identity"
                    ):
                        analyse_text_segments_folder(
                            input_dir,
                            output_root=root / "output",
                            whisper_root=whisper_root,
                            segment_alignment=policy,
                        )

    def test_prepare_mapping_handles_skipped_empty_whisper_segment(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            prepare_root = root / "prepare_output"
            output_dir = root / "output"
            video = "001_France_Test_Speaker_20250101"
            csv_path = input_dir / "France" / "Test_Speaker" / f"{video}.csv"
            write_valid_rocksteady_csv(csv_path, segment_count=1)
            whisper_path = whisper_root / "France" / "Test_Speaker" / f"{video}.json"
            whisper_path.parent.mkdir(parents=True)
            whisper_path.write_text(
                json.dumps(
                    {
                        "language": "en",
                        "segments": [
                            {"id": 41, "start": 0.0, "end": 1.0, "text": ""},
                            {"id": 42, "start": 1.0, "end": 3.0, "text": "mapped text"},
                        ],
                    }
                ),
                encoding="utf-8",
            )
            write_prepare_mapping(
                prepare_root,
                country="France",
                speaker="Test_Speaker",
                video=video,
                segments=[PreparedSegment(1, 1, 42, "mapped text")],
            )

            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                prepare_root=prepare_root,
                write_graphs=False,
            )

            row = read_rows(
                output_dir
                / "segment_counts"
                / "France"
                / "Test_Speaker"
                / f"{video}_segment_counts.csv"
            )[0]
            self.assertEqual(row["analysis_segment_id"], "1")
            self.assertEqual(row["source_segment_index"], "1")
            self.assertEqual(row["source_segment_id"], "42")
            self.assertEqual(row["segment_text"], "mapped text")

    def test_general_language_empty_selection_discovers_all_45_categories(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            prepare_root = root / "prepare_output"
            output_dir = root / "output"
            video = "001_France_Test_Speaker_20250101"
            categories = [
                "Active",
                "Negativ",
                "Research Theme",
                "Passive",
                "Positiv",
                "Strong",
                "Weak",
                *(f"General Category {index:02d}" for index in range(38)),
            ]
            self.assertEqual(len(categories), 45)
            csv_path = input_dir / "France" / "Test Speaker" / f"{video}.csv"
            write_custom_rocksteady_csv(csv_path, categories)
            write_whisper_json(
                whisper_root,
                video,
                country="France",
                speaker="Test Speaker",
                segment_count=1,
            )
            write_prepare_mapping(
                prepare_root,
                country="France",
                speaker="Test Speaker",
                video=video,
                segments=[PreparedSegment(1, 0, 1, "Test segment 1 text")],
            )
            write_adapter_manifest(
                input_dir,
                csv_path,
                categories,
                configured_categories=[],
            )

            analyse_text_segments_folder(
                input_dir,
                output_root=output_dir,
                whisper_root=whisper_root,
                prepare_root=prepare_root,
                write_graphs=False,
            )

            manifest = json.loads((output_dir / "output_manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(len(manifest["categories"]), 45)
            self.assertEqual(
                len({category["color"] for category in manifest["categories"]}),
                45,
            )
            self.assertEqual(
                manifest["upstream_provenance"]["status"],
                "verified_inventory_only",
            )
            self.assertEqual(
                manifest["upstream_provenance"]["details"]["category_selection"],
                "all-dictionary-categories",
            )
            self.assertEqual(manifest["videos"][0]["alignment_mapping"]["status"], "verified")
            count_row = read_rows(
                output_dir
                / "segment_counts"
                / "France"
                / "Test Speaker"
                / f"{video}_segment_counts.csv"
            )[0]
            self.assertIn("general_category_37_count", count_row)
            self.assertNotIn("moral_count", count_row)
            report = (output_dir / "POSTPROCESSING_REPORT_EN.md").read_text(encoding="utf-8")
            self.assertIn("General Category 37", report)

    def test_construct_alignment_uses_explicit_bounded_lexical_balances(self) -> None:
        source = {
            "country": "UK",
            "speaker": "Test Speaker",
            "speaker_id": "UK/Test Speaker",
            "video": "001_UK_Test_Speaker_20250101",
            "segment_id": "1",
            "start_sec": "0",
            "end_sec": "2",
            "duration_sec": "2",
            "segment_text": "example",
            "positive_available": "1",
            "positive_proportion": "0.2",
            "negative_available": "1",
            "negative_proportion": "0.1",
            "active_available": "1",
            "active_proportion": "0.3",
            "passive_available": "1",
            "passive_proportion": "0.1",
            "strong_available": "1",
            "strong_proportion": "0.2",
            "weak_available": "1",
            "weak_proportion": "0.1",
            "power_available": "1",
            "power_proportion": "0.1",
            "affiliation_available": "1",
            "affiliation_proportion": "0.15",
            "hostile_available": "1",
            "hostile_proportion": "0.05",
        }

        row = construct_row(source, level="segment", variant="extra")

        self.assertEqual(len(CONSTRUCTS), 5)
        self.assertAlmostEqual(float(row["activation_balance_score"]), 0.5)
        self.assertAlmostEqual(float(row["dominance_power_balance_score"]), 0.5)
        self.assertAlmostEqual(float(row["affiliation_social_balance_score"]), 0.5)
        self.assertEqual(row["dominance_power_available"], 1)
        self.assertEqual(row["affiliation_social_evidence"], 1)

    def test_verified_upstream_requires_prepare_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "parse_output"
            whisper_root = root / "transcribe_output"
            video = "001_France_Test_Speaker_20250101"
            categories = ["Active", "Negativ", "Research Theme", "Passive", "Positiv", "Strong", "Weak"]
            csv_path = input_dir / "France" / "Test_Speaker" / f"{video}.csv"
            write_custom_rocksteady_csv(csv_path, categories)
            write_whisper_json(whisper_root, video, segment_count=1)
            write_adapter_manifest(input_dir, csv_path, categories)

            with self.assertRaisesRegex(ValueError, "require a preparation mapping"):
                analyse_text_segments_folder(
                    input_dir,
                    output_root=root / "output",
                    whisper_root=whisper_root,
                    prepare_root=root / "missing_prepare",
                    write_graphs=False,
                )

    def test_45_category_graph_wraps_legend_and_keeps_bars_in_viewbox(self) -> None:
        categories = [
            Category(
                key=f"category_{index:02d}",
                display=f"General Category {index:02d}",
                source_names=(f"General Category {index:02d}",),
                color=f"#{index * 7919 % 0xFFFFFF:06x}",
            )
            for index in range(45)
        ]
        svg = render_grouped_bar_svg(
            title="Forty-five category check",
            x_labels=["0", "1", "2", "3+"],
            series=[
                (category.display, category.color, [1.0, 2.0, 3.0, 4.0])
                for category in categories
            ],
            y_label="segments",
            width=1050,
            height=500,
        )
        viewbox = re.search(r'viewBox="0 0 (\d+) (\d+)"', svg)
        self.assertIsNotNone(viewbox)
        assert viewbox is not None
        viewbox_width = float(viewbox.group(1))
        legend_points = [
            (float(x), float(y))
            for x, y in re.findall(r'<text class="legend" x="([\d.]+)" y="([\d.]+)"', svg)
        ]
        self.assertEqual(len(legend_points), 45)
        self.assertGreater(len({y for _, y in legend_points}), 1)
        self.assertTrue(all(0 <= x < viewbox_width for x, _ in legend_points))
        plot_top_match = re.search(r'<rect class="plot-bg" x="[\d.]+" y="([\d.]+)"', svg)
        self.assertIsNotNone(plot_top_match)
        assert plot_top_match is not None
        self.assertGreater(float(plot_top_match.group(1)), max(y for _, y in legend_points))
        bars = [
            (float(x), float(width))
            for x, width in re.findall(
                r'<rect class="bar" x="([\d.]+)" y="[\d.]+" width="([\d.]+)"',
                svg,
            )
        ]
        self.assertEqual(len(bars), 45 * 4)
        self.assertTrue(all(x + width <= viewbox_width + 0.01 for x, width in bars))

    def test_report_quick_table_has_exactly_seven_cells(self) -> None:
        lines = render_speaker_quick_table(
            [
                SpeakerQuickRow(
                    speaker_id="France/Test",
                    speaker="Test",
                    videos=1,
                    valid_segments=2,
                    terms=10,
                    positive_proportion=0.2,
                    negative_proportion=0.1,
                    valence=0.333333,
                )
            ],
            language="en",
        )
        for line in lines:
            self.assertEqual(len(line.strip().strip("|").split("|")), 7)

    def test_constant_descriptor_kurtosis_is_blank(self) -> None:
        self.assertEqual(describe([2.0, 2.0, 2.0])["kurtosis"], "")

    def test_descriptor_statistics_include_excel_descriptive_fields(self) -> None:
        stats = describe([1.0, 2.0, 2.0, 3.0])
        self.assertEqual(stats["mode"], 2.0)
        self.assertAlmostEqual(stats["sample_variance"], 2.0 / 3.0)
        self.assertAlmostEqual(stats["standard_error"], stats["stddev"] / 2.0)
        self.assertAlmostEqual(stats["skewness"], 0.0)
        self.assertEqual(stats["range"], 2.0)
        self.assertEqual(stats["sum"], 8.0)
        self.assertGreater(stats["confidence_level_95"], 0.0)

    def test_fallback_publication_is_atomic_and_preserves_previous_on_failure(self) -> None:
        from processing.io_utils import publish_directory as real_publish_directory

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            output = root / "result"
            output.mkdir()
            (output / "old.txt").write_text("old", encoding="utf-8")
            staging = root / ".staging"
            staging.mkdir()
            (staging / "new.txt").write_text("new", encoding="utf-8")
            calls = 0

            def fail_first_publish(source: Path, target: Path) -> None:
                nonlocal calls
                calls += 1
                if calls == 1:
                    raise PermissionError("simulated Windows rename lock")
                real_publish_directory(source, target)

            with patch(
                "analysis.text_pipeline.transaction.publish_directory",
                side_effect=fail_first_publish,
            ):
                replace_output_dir(staging, output)
            self.assertFalse((output / "old.txt").exists())
            self.assertEqual((output / "new.txt").read_text(encoding="utf-8"), "new")

            failed_staging = root / ".failed_staging"
            failed_staging.mkdir()
            (failed_staging / "later.txt").write_text("later", encoding="utf-8")
            with patch(
                "analysis.text_pipeline.transaction.publish_directory",
                side_effect=[PermissionError("locked"), RuntimeError("blocked")],
            ):
                with self.assertRaisesRegex(RuntimeError, "validated-copy publication also failed"):
                    replace_output_dir(failed_staging, output)
            self.assertEqual((output / "new.txt").read_text(encoding="utf-8"), "new")
            self.assertFalse((output / "later.txt").exists())

    def test_derived_manifest_verifies_inventory_digest_and_csv_hash(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "derived"
            source_root = root / "source"
            relative = Path("France/Test_Speaker/001_France_Test_Speaker_20250101.csv")
            categories = ["Positiv"]
            csv_path = input_dir / relative
            source_path = source_root / relative
            write_custom_rocksteady_csv(csv_path, categories)
            write_custom_rocksteady_csv(source_path, categories)
            item = {
                "identity": relative.with_suffix("").as_posix(),
                "source": relative.as_posix(),
                "output": relative.as_posix(),
                "source_sha256": file_sha256(source_path),
                "output_sha256": file_sha256(csv_path),
                "rows": 1,
                "status": "completed",
            }
            manifest = {
                "schema_version": "2.0",
                "kind": "derived-rocksteady-category-view",
                "status": "completed",
                "source_root": str(source_root),
                "categories": categories,
                "inventory_sha256": inventory_digest([item]),
                "summary": {"total": 1, "completed": 1, "failed": 0},
                "files": [item],
            }
            (input_dir / "derived_view_manifest.json").write_text(
                json.dumps(manifest), encoding="utf-8"
            )

            provenance = inspect_upstream_provenance(input_dir, [csv_path])
            self.assertEqual(provenance.status, "verified_sha256")
            csv_path.write_text(csv_path.read_text(encoding="utf-8") + " ", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "CSV hash differs"):
                inspect_upstream_provenance(input_dir, [csv_path])

    def test_modern_adapter_manifest_verifies_video_inventory_and_hash(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "rocksteady"
            relative = Path("France/Test_Speaker/001_France_Test_Speaker_20250101.csv")
            categories = ["Active", "Negativ", "Research Theme", "Passive", "Positiv", "Strong", "Weak"]
            csv_path = input_dir / relative
            write_custom_rocksteady_csv(csv_path, categories)
            manifest_path = write_adapter_manifest(
                input_dir,
                csv_path,
                categories,
                include_hash=True,
            )
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest.update(
                {
                    "schema_version": "2.0",
                    "kind": "rocksteady-analysis-batch",
                    "status": "completed",
                    "upstream_inventory_sha256": None,
                    "inventory_sha256": inventory_digest(manifest["videos"]),
                }
            )
            manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

            provenance = inspect_upstream_provenance(input_dir, [csv_path])
            self.assertEqual(provenance.status, "verified_sha256")
            self.assertTrue(provenance.verified)
            manifest["videos"][0]["validation"]["rows"] = 2
            manifest["inventory_sha256"] = inventory_digest(manifest["videos"])
            manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "row count differs"):
                inspect_upstream_provenance(input_dir, [csv_path])


if __name__ == "__main__":
    unittest.main()
