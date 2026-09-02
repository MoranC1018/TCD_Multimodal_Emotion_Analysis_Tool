from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import openpyxl
import pytest

from analysis.combined_summary import (
    AUDIO_METRICS,
    VIDEO_METRICS,
    CombinedSource,
    InputError,
    SpeakerGroupDefinition,
    TextConstructSummary,
    build_combined_workbook,
    resolve_speaker,
)
from analysis.inference import add_probability_mirrors
from analysis.metadata import load_source_metadata
from analysis.native_face import NATIVE_FACE_METRICS
from analysis.profile import AnalysisProfile, ManualGroup, ProfileMember
from analysis.video import CanonicalVideoResult, DetectedVideoSource
from analysis.video_contract import VIDEO_NORMALIZATION_VERSION
from analysis.workflow import ModalityRequest, WorkflowError, WorkflowRequest, run_workflow


def _write_report(
    path: Path,
    sources: list[str],
    base: float = 10.0,
    *,
    metrics: tuple[str, ...] = AUDIO_METRICS,
    imotions_metadata: bool = False,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for metric_index, metric in enumerate(metrics):
            writer.writerow([metric])
            writer.writerow(
                [
                    "classification",
                    "FEA(Affectiva AFFDEX)" if imotions_metadata else "core",
                    "category",
                    "emotion",
                    "unit",
                    "score",
                ]
            )
            writer.writerow(["metric", *sources])
            writer.writerow(["count", *([10] * len(sources))])
            writer.writerow(["missing", *([0] * len(sources))])
            writer.writerow(["mean", *[base + metric_index + offset for offset in range(len(sources))]])
            writer.writerow(["stddev", *([1] * len(sources))])
            writer.writerow(["kurtosis", *([0] * len(sources))])
            writer.writerow([])


def _write_manifest(root: Path, count: int, *, speakers: list[str] | None = None) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    speakers = speakers or ["Researcher Alpha"] * count
    sources = []
    for index in range(1, count + 1):
        source_id = f"source-{index:04d}"
        speaker = speakers[index - 1]
        sources.append(
            {
                "source_id": source_id,
                "catalog_row": index + 1,
                "link": f"https://www.youtube.com/watch?v={index:011d}",
                "resolved_link": f"https://www.youtube.com/watch?v={index:011d}",
                "source_kind": "youtube",
                "speaker": speaker,
                "speaker_display": speaker,
                "selected": True,
                "status": "selected",
                "system_metadata": {
                    "title": "=Research formula" if index == 1 else f"Interview {index:02d}",
                    "duration_seconds": 60,
                    "youtube_language": "English",
                },
                "user_metadata": {
                    "Country": "Ireland" if index % 2 else "Japan",
                    "Wave": "First" if index <= count // 2 else "Second",
                },
                "output_mapping": {
                    "video_directory": str(root / speaker / f"{source_id}_Interview_{index:02d}")
                },
                "youtube": {
                    "video_id": f"{index:011d}",
                    "url": f"https://www.youtube.com/watch?v={index:011d}",
                },
            }
        )
    manifest = root / "source_manifest.json"
    manifest.write_text(
        json.dumps(
            {
                "format_version": 1,
                "catalog": {
                    "metadata_headers": ["Country", "Wave"],
                    "metadata_export_headers": {"Country": "Country", "Wave": "Wave"},
                },
                "sources": sources,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    (root / "source_metadata.csv").write_text(
        "SourceID,Country,Wave\n"
        + "\n".join(
            f"source-{index:04d},{'Ireland' if index % 2 else 'Japan'},"
            f"{'First' if index <= count // 2 else 'Second'}"
            for index in range(1, count + 1)
        )
        + "\n",
        encoding="utf-8",
    )
    return manifest


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_profile_workbook_orders_more_than_twelve_sources_and_keeps_metric_contract(
    tmp_path: Path,
) -> None:
    manifest = _write_manifest(tmp_path / "run", 14)
    metadata = load_source_metadata(manifest)
    report = tmp_path / "reports" / "Researcher Alpha" / "descriptive_statistics.csv"
    _write_report(report, [f"source-{index:04d}" for index in range(1, 15)])
    profile = AnalysisProfile(
        source_manifest=manifest,
        source_manifest_sha256=metadata.manifest_sha256,
        sort_fields=("Country", "Wave"),
        automatic_group_field="Country",
    )
    sidecar_hashes = (_digest(manifest), _digest(metadata.metadata_path))

    result = build_combined_workbook(
        {
            "audio": (
                CombinedSource("audio", "researcheralpha", "Researcher Alpha", report),
            )
        },
        tmp_path / "profiled.xlsx",
        analysis_profile=profile,
    )

    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
    audio = book["Audio"]
    cells = result.source_cells["Audio|Anger"]
    assert len(cells.speaker_cells) == 14
    assert cells.speaker_ids[:5] == (
        "source-0001",
        "source-0003",
        "source-0005",
        "source-0007",
        "source-0009",
    )
    assert [audio[cell].value for cell in cells.speaker_cells[:3]] == [10, 12, 14]
    assert cells.speaker_observations[0] == (10,)
    assert cells.speaker_observation_labels[0] == ("source-0001",)
    assert audio[cells.speaker_cells[0][0] + "1"].value == "'=Research formula"
    assert audio[cells.overall].value.startswith("=AVERAGE(")
    guide = book["Measure Guide"]
    assert guide.max_row == 35
    assert {guide.cell(row, 1).value for row in range(2, guide.max_row + 1)} == {
        "Emotions",
        "Sentiment",
        "Valence",
        "Dimensions",
        "Comparison",
    }
    assert (_digest(manifest), _digest(metadata.metadata_path)) == sidecar_hashes


def test_profile_video_provider_keeps_dynamic_order_filters_and_one_sheet(
    tmp_path: Path,
) -> None:
    manifest = _write_manifest(tmp_path / "run", 14)
    metadata = load_source_metadata(manifest)
    report = tmp_path / "reports" / "Researcher Alpha" / "descriptive_statistics.csv"
    _write_report(
        report,
        [f"source-{index:04d}" for index in range(1, 15)],
        metrics=NATIVE_FACE_METRICS,
    )
    detected = DetectedVideoSource(
        "pyfeat_native_face",
        manifest.parent.resolve(),
        "import",
        ("Verified imported Py-Feat reports.",),
    )
    canonical = CanonicalVideoResult(
        "pyfeat_native_face",
        tuple(f"source-{index:04d}" for index in range(1, 15)),
        tuple({metric: None for metric in VIDEO_METRICS} for _ in range(14)),
        detected.evidence,
        detected.warnings,
        VIDEO_NORMALIZATION_VERSION,
    )
    provenance_builder = getattr(canonical, "output_provenance", None)
    assert provenance_builder is not None
    source_hashes = (_digest(manifest), _digest(metadata.metadata_path))
    profile = AnalysisProfile(
        source_manifest=manifest,
        source_manifest_sha256=metadata.manifest_sha256,
        sort_fields=("Country", "Wave"),
        automatic_group_field="Country",
        metadata_filters=(("Country", ("Ireland",)),),
    )

    result = build_combined_workbook(
        {
            "native_face": (
                CombinedSource(
                    "native_face",
                    "researcheralpha",
                    "Researcher Alpha",
                    report,
                    video_provenance=provenance_builder(detected),
                ),
            )
        },
        tmp_path / "profiled-video.xlsx",
        analysis_profile=profile,
    )

    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
    assert result.quantitative_sheets == ("Video",)
    assert not any("native face" in name.casefold() for name in book.sheetnames)
    cells = result.source_cells["Video|Anger"]
    assert cells.speaker_ids == tuple(f"source-{index:04d}" for index in range(1, 15, 2))
    assert book["Video"][cells.overall].value.startswith("=AVERAGE(")
    engagement = result.source_cells["Video|Engagement"]
    assert all(book["Video"][coordinate].value is None for coordinate in engagement.speaker_cells)
    assert book["Video"][engagement.overall].value is None
    assert result.video_manifest_payload["sources"][0]["resolved_provider"] == "pyfeat_native_face"
    assert (_digest(manifest), _digest(metadata.metadata_path)) == source_hashes


def test_default_workbook_has_no_twelve_speaker_or_four_group_limit(tmp_path: Path) -> None:
    sources = []
    for index in range(1, 14):
        speaker = f"Researcher {index:02d}"
        report = tmp_path / "reports" / speaker / "descriptive_statistics.csv"
        _write_report(report, ["001_First", "002_Second", "003_Third", "004_Fourth", "005_Fifth"], index)
        sources.append(
            CombinedSource("audio", f"researcher{index:02d}", speaker, report)
        )

    result = build_combined_workbook({"audio": tuple(sources)}, tmp_path / "thirteen.xlsx")

    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
    cells = result.source_cells["Audio|Anger"]
    assert len(cells.speaker_cells) == 13
    assert sum(len(group[2]) for group in cells.speaker_groups) == 13
    assert book["Audio"].max_column >= 19


def test_workflow_reuses_source_sidecars_for_two_profiles_and_archives_each_output_profile(
    tmp_path: Path,
) -> None:
    manifest = _write_manifest(tmp_path / "source-run", 14)
    metadata = load_source_metadata(manifest)
    imported = manifest.parent / "imported"
    report = (
        imported
        / "emotion"
        / "Researcher Alpha"
        / "combined"
        / "other_findings"
        / "descriptive_statistics.csv"
    )
    _write_report(report, [f"source-{index:04d}" for index in range(1, 15)])
    output = tmp_path / "analysis-output"
    source_hashes = (_digest(manifest), _digest(metadata.metadata_path))

    first_profile = AnalysisProfile(
        source_manifest=manifest,
        source_manifest_sha256=metadata.manifest_sha256,
        sort_fields=("Country",),
        automatic_group_field="Country",
    )
    second_profile = AnalysisProfile(
        source_manifest=manifest,
        source_manifest_sha256=metadata.manifest_sha256,
        sort_fields=("Wave", "Country"),
        automatic_group_field="Wave",
    )
    common = {
        "output_root": output,
        "modalities": (ModalityRequest("audio", "import", imported),),
        "speaker_groups": (),
        "include_construct_comparison": False,
        "include_probability_sheets": False,
    }

    first = run_workflow(WorkflowRequest(**common, analysis_profile=first_profile))
    first_bytes = (output / "analysis_profile.json").read_bytes()
    second = run_workflow(WorkflowRequest(**common, analysis_profile=second_profile))

    assert first.workbook_path is not None
    assert second.workbook_path is not None
    assert json.loads((output / "analysis_profile.json").read_text(encoding="utf-8"))[
        "automatic_group_field"
    ] == "Wave"
    archived_profiles = list((output / "combined_analysis_history").rglob("analysis_profile.json"))
    assert len(archived_profiles) == 1
    assert archived_profiles[0].read_bytes() == first_bytes
    archived_manifests = list(
        (output / "combined_analysis_history").rglob("combined_analysis_manifest.json")
    )
    assert len(archived_manifests) == 1
    archived_manifest = json.loads(archived_manifests[0].read_text(encoding="utf-8"))
    assert archived_manifest["analysis_profile_path"] == str(archived_profiles[0].resolve())
    assert archived_manifest["archive"]["analysis_profile_sha256"] == hashlib.sha256(
        first_bytes
    ).hexdigest()
    assert (_digest(manifest), _digest(metadata.metadata_path)) == source_hashes


def test_workflow_rejects_a_profile_from_another_source_run(tmp_path: Path) -> None:
    profile_manifest = _write_manifest(tmp_path / "profile-run", 1)
    selected_manifest = _write_manifest(tmp_path / "selected-run", 1)
    profile_metadata = load_source_metadata(profile_manifest)
    report = (
        selected_manifest.parent
        / "imported"
        / "emotion"
        / "Researcher Alpha"
        / "combined"
        / "other_findings"
        / "descriptive_statistics.csv"
    )
    _write_report(report, ["source-0001"])

    request = WorkflowRequest(
        output_root=tmp_path / "output",
        modalities=(ModalityRequest("audio", "import", selected_manifest.parent / "imported"),),
        speaker_groups=(),
        include_probability_sheets=False,
        analysis_profile=AnalysisProfile(
            profile_manifest,
            profile_metadata.manifest_sha256,
        ),
    )

    with pytest.raises(WorkflowError, match="associated with the selected modality"):
        run_workflow(request)


def test_profile_manifest_is_authoritative_for_audio_sidecars_and_sidecarless_legacy_imports(
    tmp_path: Path,
) -> None:
    manifest = _write_manifest(
        tmp_path / "procurement-run",
        1,
        speakers=["Researcher Alpha"],
    )
    metadata = load_source_metadata(manifest)
    audio_root = manifest.parent / "analysis-input" / "audio"
    audio_report = (
        audio_root
        / "emotion"
        / "Researcher Alpha"
        / "combined"
        / "other_findings"
        / "descriptive_statistics.csv"
    )
    _write_report(audio_report, ["source-0001"])

    video_root = tmp_path / "ordinary-imotions-export"
    video_report = (
        video_root
        / "emotion"
        / "Researcher Alpha"
        / "combined"
        / "other_findings"
        / "descriptive_statistics.csv"
    )
    _write_report(
        video_report,
        ["source-0001_Interview_01"],
        metrics=VIDEO_METRICS,
        imotions_metadata=True,
    )

    text_root = tmp_path / "ordinary-text-export"
    text_summary = text_root / "multimodal" / "speaker_level_summary.csv"
    text_summary.parent.mkdir(parents=True)
    with text_summary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "Country",
                "Speaker",
                "Speaker ID",
                "Videos",
                "Valid segments",
                "RockSteady terms",
                "Positive Sentiment",
                "Negative Sentiment",
                "Arousal / Activation",
                "Dominance / Power",
                "Affiliation / Social orientation",
            ]
        )
        writer.writerow(
            ["Ireland", "Researcher Alpha", "lab/researcher-alpha", 1, 2, 3, 0.2, 0.1, 0.3, 0.4, 0.5]
        )

    result = run_workflow(
        WorkflowRequest(
            output_root=tmp_path / "analysis-output",
            modalities=(
                ModalityRequest("audio", "import", audio_root),
                ModalityRequest("video", "import", video_root),
                ModalityRequest("text", "import", text_root),
            ),
            speaker_groups=(),
            include_probability_sheets=False,
            analysis_profile=AnalysisProfile(manifest, metadata.manifest_sha256),
        )
    )

    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
    assert {"Audio", "Video", "Text sentiment"}.issubset(book.sheetnames)
    assert result.workbook_path.is_file()


@pytest.mark.parametrize("include_imotions", (False, True))
def test_explicit_profile_manifest_supports_all_sidecarless_legacy_imports(
    tmp_path: Path,
    include_imotions: bool,
) -> None:
    manifest = _write_manifest(
        tmp_path / "procurement-run",
        1,
        speakers=["Researcher Alpha"],
    )
    metadata = load_source_metadata(manifest)
    text_root = tmp_path / "ordinary-text-export"
    text_summary = text_root / "multimodal" / "speaker_level_summary.csv"
    text_summary.parent.mkdir(parents=True)
    with text_summary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "Country",
                "Speaker",
                "Speaker ID",
                "Videos",
                "Valid segments",
                "RockSteady terms",
                "Positive Sentiment",
                "Negative Sentiment",
                "Arousal / Activation",
                "Dominance / Power",
                "Affiliation / Social orientation",
            ]
        )
        writer.writerow(
            ["Ireland", "Researcher Alpha", "lab/researcher-alpha", 1, 2, 3, 0.2, 0.1, 0.3, 0.4, 0.5]
        )
    modalities = [ModalityRequest("text", "import", text_root)]
    if include_imotions:
        video_root = tmp_path / "ordinary-imotions-export"
        video_report = (
            video_root
            / "emotion"
            / "Researcher Alpha"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        _write_report(
            video_report,
            ["source-0001_Interview_01"],
            metrics=VIDEO_METRICS,
            imotions_metadata=True,
        )
        modalities.insert(0, ModalityRequest("video", "import", video_root))

    result = run_workflow(
        WorkflowRequest(
            output_root=tmp_path / "analysis-output",
            modalities=tuple(modalities),
            speaker_groups=(),
            include_probability_sheets=False,
            analysis_profile=AnalysisProfile(manifest, metadata.manifest_sha256),
        )
    )

    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
    assert "Text sentiment" in book.sheetnames
    assert ("Video" in book.sheetnames) is include_imotions


def test_workflow_preflight_rejects_text_profile_that_splits_one_speaker(
    tmp_path: Path,
) -> None:
    manifest = _write_manifest(
        tmp_path / "procurement-run",
        2,
        speakers=["Researcher Alpha", "Researcher Alpha"],
    )
    metadata = load_source_metadata(manifest)
    text_root = tmp_path / "ordinary-text-results"
    text_root.mkdir()
    output_root = tmp_path / "analysis-output"
    profile = AnalysisProfile(
        manifest,
        metadata.manifest_sha256,
        automatic_group_field="Country",
    )

    with pytest.raises(WorkflowError, match="Text is speaker-level|same output group"):
        run_workflow(
            WorkflowRequest(
                output_root=output_root,
                modalities=(ModalityRequest("text", "import", text_root),),
                speaker_groups=(),
                analysis_profile=profile,
            )
        )

    assert not output_root.exists()


def test_workbook_and_inference_accept_more_than_five_videos_for_one_speaker(
    tmp_path: Path,
) -> None:
    report = tmp_path / "reports" / "Researcher Alpha" / "descriptive_statistics.csv"
    source_labels = [f"source-{index:04d}" for index in range(1, 8)]
    _write_report(report, source_labels)

    result = build_combined_workbook(
        {
            "audio": (
                CombinedSource("audio", "researcheralpha", "Researcher Alpha", report),
            )
        },
        tmp_path / "seven-videos.xlsx",
    )
    add_probability_mirrors(result.workbook_path, result.source_cells)

    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
    cells = result.source_cells["Audio|Anger"]
    assert len(cells.speaker_observations[0]) == 7
    assert cells.speaker_observation_labels[0] == tuple(source_labels)
    audio_labels = [
        book["Audio"].cell(row, 3).value
        for row in range(1, book["Audio"].max_row + 1)
    ]
    assert "6th" in audio_labels
    assert "7th" in audio_labels
    inputs = book["Inference Inputs"]
    assert inputs.cell(1, 15).value == "Value 7"
    assert inputs.cell(2, 15).value == 16
    outline_text = " ".join(
        str(cell.value)
        for row in book["Probability Outline"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "speaker: up to 7 per-source means" in outline_text


def test_construct_comparison_keeps_every_member_of_a_large_group(tmp_path: Path) -> None:
    sources = []
    speakers = [f"Researcher {index:02d}" for index in range(1, 6)]
    for index, speaker in enumerate(speakers, start=1):
        report = tmp_path / "reports" / speaker / "descriptive_statistics.csv"
        _write_report(report, ["001", "002", "003", "004", "005"], index)
        sources.append(CombinedSource("audio", speaker, speaker, report))

    result = build_combined_workbook(
        {"audio": tuple(sources)},
        tmp_path / "large-group.xlsx",
        speaker_groups=(SpeakerGroupDefinition("all", "All researchers", tuple(speakers)),),
        include_construct_comparison=True,
    )

    comparison = openpyxl.load_workbook(result.workbook_path, data_only=False)[
        "Construct Comparison"
    ]
    assert [comparison.cell(5, column).value for column in (1, 9, 17, 25, 33)] == speakers
    assert comparison.max_column >= 40


def test_profile_maps_registry_speaker_ids_back_to_manifest_identity(tmp_path: Path) -> None:
    manifest = _write_manifest(tmp_path / "run", 1, speakers=["Andy Burnham"])
    metadata = load_source_metadata(manifest)
    report = tmp_path / "reports" / "Andy Burnham" / "descriptive_statistics.csv"
    _write_report(report, ["source-0001"])

    result = build_combined_workbook(
        {"audio": (CombinedSource("audio", "andy_burnham", "Andy Burnham", report),)},
        tmp_path / "known-speaker.xlsx",
        analysis_profile=AnalysisProfile(manifest, metadata.manifest_sha256),
    )

    assert result.source_cells["Audio|Anger"].speaker_ids == ("source-0001",)


def test_profile_maps_imported_text_summary_to_source_identity(tmp_path: Path) -> None:
    manifest = _write_manifest(tmp_path / "run", 1, speakers=["Andy Burnham"])
    metadata = load_source_metadata(manifest)
    summary_path = tmp_path / "text" / "speaker_level_summary.csv"
    summary_path.parent.mkdir(parents=True)
    summary_path.write_text("summary\n", encoding="utf-8")
    summary = TextConstructSummary(
        speaker_id="andy_burnham",
        display_name="Andy Burnham",
        country="Ireland",
        constructs={
            "Positive Sentiment": 0.2,
            "Negative Sentiment": 0.1,
            "Arousal / Activation": 0.3,
            "Dominance / Power": 0.4,
            "Affiliation / Social orientation": 0.5,
        },
        source_path=summary_path,
    )

    result = build_combined_workbook(
        {},
        tmp_path / "known-text-speaker.xlsx",
        analysis_profile=AnalysisProfile(manifest, metadata.manifest_sha256),
        text_summaries=(summary,),
    )

    text_sheet = openpyxl.load_workbook(result.workbook_path, data_only=False)["Text sentiment"]
    assert text_sheet["D2"].value == 20.0


def test_profile_rejects_a_modality_report_with_no_manifest_speaker(tmp_path: Path) -> None:
    manifest = _write_manifest(tmp_path / "run", 1, speakers=["Researcher Alpha"])
    metadata = load_source_metadata(manifest)
    report = tmp_path / "reports" / "Unrelated Researcher" / "descriptive_statistics.csv"
    _write_report(report, ["source-0001"])

    with pytest.raises(InputError, match="does not match any selected manifest speaker"):
        build_combined_workbook(
            {
                "audio": (
                    CombinedSource(
                        "audio",
                        "unrelatedresearcher",
                        "Unrelated Researcher",
                        report,
                    ),
                )
            },
            tmp_path / "unmatched.xlsx",
            analysis_profile=AnalysisProfile(manifest, metadata.manifest_sha256),
        )


def test_profile_rejects_missing_visible_modality_sources(tmp_path: Path) -> None:
    manifest = _write_manifest(
        tmp_path / "run",
        2,
        speakers=["Researcher Alpha", "Researcher Alpha"],
    )
    metadata = load_source_metadata(manifest)
    report = tmp_path / "reports" / "Researcher Alpha" / "descriptive_statistics.csv"
    _write_report(report, ["source-0001"])

    with pytest.raises(InputError, match="missing profiled source.*source-0002"):
        build_combined_workbook(
            {
                "audio": (
                    CombinedSource(
                        "audio",
                        "researcheralpha",
                        "Researcher Alpha",
                        report,
                    ),
                )
            },
            tmp_path / "missing.xlsx",
            analysis_profile=AnalysisProfile(manifest, metadata.manifest_sha256),
        )


def test_speaker_identity_does_not_use_substring_aliases() -> None:
    speaker = resolve_speaker("Researcher Burnham")

    assert speaker.speaker_id == "researcherburnham"
    assert speaker.display_name == "Researcher Burnham"


def test_profile_keeps_text_at_one_observation_per_speaker(tmp_path: Path) -> None:
    manifest = _write_manifest(
        tmp_path / "run",
        3,
        speakers=["Speaker A", "Speaker A", "Speaker B"],
    )
    metadata = load_source_metadata(manifest)
    summary_path = tmp_path / "text" / "speaker_level_summary.csv"
    summary_path.parent.mkdir(parents=True)
    summary_path.write_text("summary\n", encoding="utf-8")

    def summary(speaker: str, value: float) -> TextConstructSummary:
        return TextConstructSummary(
            speaker_id=speaker.casefold().replace(" ", ""),
            display_name=speaker,
            country="Research cohort",
            constructs={construct: value for construct in (
                "Positive Sentiment",
                "Negative Sentiment",
                "Arousal / Activation",
                "Dominance / Power",
                "Affiliation / Social orientation",
            )},
            source_path=summary_path,
        )

    result = build_combined_workbook(
        {},
        tmp_path / "speaker-grain-text.xlsx",
        analysis_profile=AnalysisProfile(manifest, metadata.manifest_sha256),
        text_summaries=(summary("Speaker A", 1.0), summary("Speaker B", 0.0)),
    )

    text = openpyxl.load_workbook(result.workbook_path, data_only=False)["Text sentiment"]
    assert [text["D1"].value, text["E1"].value] == ["Speaker A", "Speaker B"]
    assert [text["D2"].value, text["E2"].value] == [100.0, 0.0]
    assert text["S2"].value == "=AVERAGE(D2,E2)"


def test_profile_rejects_splitting_one_speaker_level_text_result_across_groups(
    tmp_path: Path,
) -> None:
    manifest = _write_manifest(
        tmp_path / "run",
        2,
        speakers=["Speaker A", "Speaker A"],
    )
    metadata = load_source_metadata(manifest)
    summary_path = tmp_path / "text" / "speaker_level_summary.csv"
    summary_path.parent.mkdir(parents=True)
    summary_path.write_text("summary\n", encoding="utf-8")
    summary = TextConstructSummary(
        speaker_id="speakera",
        display_name="Speaker A",
        country="Research cohort",
        constructs={
            construct: 0.5
            for construct in (
                "Positive Sentiment",
                "Negative Sentiment",
                "Arousal / Activation",
                "Dominance / Power",
                "Affiliation / Social orientation",
            )
        },
        source_path=summary_path,
    )
    profile = AnalysisProfile(
        manifest,
        metadata.manifest_sha256,
        manual_groups=(
            ManualGroup("first", "First", (ProfileMember("source", "source-0001"),)),
            ManualGroup("second", "Second", (ProfileMember("source", "source-0002"),)),
        ),
    )

    with pytest.raises(InputError, match="cannot be split across Analysis groups"):
        build_combined_workbook(
            {},
            tmp_path / "ambiguous-text.xlsx",
            analysis_profile=profile,
            text_summaries=(summary,),
        )
