"""Native Face and SourceID-grain Text contracts for Analysis."""

from __future__ import annotations

import csv
import hashlib
import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from analysis.combined_summary import TEXT_CONSTRUCTS, build_combined_workbook
from analysis.native_face import (
    NATIVE_FACE_METRICS,
    analyse_native_face_folder,
    read_native_face_export,
    read_native_face_folder,
)
from analysis.profile import AnalysisProfile, ManualGroup, ProfileMember
from analysis.text_results import TextResultsError, discover_text_results
from processing.face_analysis.outputs import AU_NAMES, artifact_metadata


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _source_context(root: Path, source_id: str = "source-0001") -> dict[str, object]:
    suffix = "one" if source_id == "source-0001" else "two"
    title = "First" if source_id == "source-0001" else "Second"
    return {
        "source_id": source_id,
        "speaker": "speaker-a",
        "speaker_display": "Speaker A",
        "source_kind": "youtube",
        "resolved_link": f"https://www.youtube.com/watch?v={source_id[-4:]}example",
        "user_metadata": {"Country": "Ireland", "Gender": "F"},
        "system_metadata": {"title": title},
        "output_mapping": {"video_directory": str(root / "media" / suffix)},
        "run_root": str(root),
        "catalog_sha256": "a" * 64,
    }


def _write_face_video(root: Path) -> Path:
    video = root / "Speaker A" / "source-0001"
    video.mkdir(parents=True)
    core = video / "face_core.csv"
    header = [
        "media_id", "frame_index", "timestamp_seconds", "face_detected", "face_count",
        "face_index", "is_primary_face", "FaceRectX", "FaceRectY", "FaceRectWidth",
        "FaceRectHeight", "FaceScore", *AU_NAMES, "Neutral", "Happy", "Sad",
        "Surprise", "Fear", "Disgust", "Anger", "valence", "arousal",
    ]
    filler = ["0"] * len(AU_NAMES)
    with core.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        # Explicit no-face and non-primary observations must not enter Analysis.
        writer.writerow(["m1", 0, 0, "false", 0, "", "false", "", "", "", "", "", *filler, "", "", "", "", "", "", "", "", ""])
        writer.writerow(["m1", 1, 1, "true", 2, 1, "false", 0, 0, 1, 1, .6, *filler, .9, .9, .9, .9, .9, .9, .9, .9, .9])
        writer.writerow(["m1", 1, 1, "true", 2, 0, "true", 0, 0, 1, 1, .95, *filler, .1, .2, .3, .4, .5, .6, .7, -.25, .75])
    context = _source_context(root)
    context_bytes = json.dumps(
        context, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    manifest = {
        "schema_version": "1.0",
        "status": "completed",
        "media_id": "m1",
        "input": {"path": str(root / "input.mp4")},
        "source": {
            "source_id": "source-0001",
            "speaker": "speaker-a",
            "speaker_display": "Speaker A",
            "catalog_sha256": "a" * 64,
            "user_metadata": context["user_metadata"],
            "system_metadata": context["system_metadata"],
            "output_mapping": context["output_mapping"],
            "source_context_sha256": hashlib.sha256(context_bytes).hexdigest(),
            "source_context": context,
            "content": {"sha256": "c" * 64, "size_bytes": 10},
        },
        "output_contract_version": "1.0",
        "outputs": {"core": artifact_metadata(core, "core")},
    }
    (video / "video_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    return core


def _write_source_sidecars(root: Path) -> Path:
    manifest = {
        "format_version": 1,
        "catalog": {
            "sha256": "a" * 64,
            "metadata_headers": ["Country", "Gender"],
            "metadata_export_headers": {"Country": "Country", "Gender": "Gender"},
        },
        "sources": [
            {
                **_source_context(root, "source-0001"), "selected": True,
            },
            {
                **_source_context(root, "source-0002"), "selected": True,
            },
        ],
    }
    path = root / "source_manifest.json"
    path.write_text(json.dumps(manifest), encoding="utf-8")
    with (root / "source_metadata.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["SourceID", "Country", "Gender"])
        writer.writeheader()
        writer.writerows([
            {"SourceID": "source-0001", "Country": "Ireland", "Gender": "F"},
            {"SourceID": "source-0002", "Country": "Ireland", "Gender": "F"},
        ])
    return path


def _write_face_run_manifest(root: Path) -> Path:
    path = root / "run_manifest.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "status": "completed",
                "catalog_sha256": "a" * 64,
                "processed_source_ids": ["source-0001"],
            }
        ),
        encoding="utf-8",
    )
    return path


def _write_native_text(root: Path) -> Path:
    multimodal = root / "postprocessing" / "multimodal"
    multimodal.mkdir(parents=True)
    for name in ("source_manifest.json", "source_metadata.csv"):
        (root / "postprocessing" / name).write_bytes((root / name).read_bytes())
    summary = multimodal / "video_level_summary.csv"
    headers = (
        "Country", "Speaker", "Speaker ID", "Video", "Source ID", "Date",
        "Valid segments", "RockSteady terms", "Positive Sentiment",
        "Negative Sentiment", "Text Valence", "Arousal / Activation",
        "Dominance / Power", "Affiliation / Social orientation",
    )
    with summary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        # Deliberately false child valences: the importer must recompute them.
        writer.writerow(["Ireland", "Speaker A", "speaker-a", "First", "source-0001", "", 2, 10, .9, .1, -1, .2, .3, .4])
        writer.writerow(["Ireland", "Speaker A", "speaker-a", "Second", "source-0002", "", 2, 10, .1, .4, 1, .3, .4, .5])
    contract = multimodal / "alignment_contract.json"
    contract.write_text(json.dumps({"schema_version": "1.0", "kind": "transcript-multimodal-alignment", "rows": {"video": 2}}), encoding="utf-8")
    batch = {
        "schema_version": "1.0", "kind": "text-postprocessing-selected-extra-pair", "status": "completed",
        "source_binding": {
            "kind": "catalog-source-sidecars",
            "catalog_sha256": "a" * 64,
            "source_manifest": "source_manifest.json",
            "source_manifest_sha256": _sha(root / "postprocessing" / "source_manifest.json"),
            "source_metadata": "source_metadata.csv",
            "source_metadata_sha256": _sha(root / "postprocessing" / "source_metadata.csv"),
            "source_contexts": [
                {
                    "source_id": source_id,
                    "sha256": hashlib.sha256(
                        json.dumps(
                            _source_context(root, source_id),
                            ensure_ascii=False,
                            sort_keys=True,
                            separators=(",", ":"),
                        ).encode("utf-8")
                    ).hexdigest(),
                    "context": _source_context(root, source_id),
                }
                for source_id in ("source-0001", "source-0002")
            ],
        },
        "multimodal": {
            "path": "multimodal", "contract": "multimodal/alignment_contract.json",
            "contract_sha256": _sha(contract), "video_summary": "multimodal/video_level_summary.csv",
            "video_summary_sha256": _sha(summary), "source_ids": ["source-0001", "source-0002"],
        },
    }
    (root / "postprocessing" / "text_postprocessing_batch_manifest.json").write_text(
        json.dumps(batch), encoding="utf-8"
    )
    return summary


def test_native_face_reader_filters_primary_rows_scales_values_and_names_provider(tmp_path: Path) -> None:
    core = _write_face_video(tmp_path)

    export = read_native_face_export(core)

    assert len(export.rows) == 1
    row = export.rows[0]
    assert row["Joy"] == "20"
    assert row["Sadness"] == "30"
    assert row["Neutral"] == "10"
    assert row["Anger"] == "70"
    assert row["Valence"] == "-25"
    assert row["Arousal"] == "75"
    assert row["Contempt"] == row["Confusion"] == ""
    assert tuple(metric for metric in NATIVE_FACE_METRICS if metric in export.info) == NATIVE_FACE_METRICS
    assert {export.info[name].provided_by for name in NATIVE_FACE_METRICS} == {"Py-Feat / Native Face"}


def test_native_face_reader_rejects_a_linked_input_root(tmp_path: Path) -> None:
    real_root = tmp_path / "real-native-face"
    real_root.mkdir()
    _write_face_video(real_root)
    linked_root = tmp_path / "linked-native-face"
    try:
        linked_root.symlink_to(real_root, target_is_directory=True)
    except OSError as exc:
        if sys.platform != "win32":
            pytest.skip(f"directory links are unavailable: {exc}")
        completed = subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(linked_root), str(real_root)],
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            pytest.skip(f"directory links are unavailable: {completed.stderr.strip()}")

    with pytest.raises(ValueError, match="symbolic link, junction, or reparse"):
        read_native_face_folder(linked_root)


def test_native_face_emotion_report_keeps_arousal_as_arousal(tmp_path: Path) -> None:
    _write_source_sidecars(tmp_path)
    _write_face_video(tmp_path)
    _write_face_run_manifest(tmp_path)
    before = {
        path.relative_to(tmp_path).as_posix(): path.read_bytes()
        for path in tmp_path.rglob("*")
        if path.is_file()
    }

    result = analyse_native_face_folder(
        tmp_path,
        output_root=tmp_path.parent / f"{tmp_path.name}-analysis",
        write_graphs=False,
    )

    combined = next(
        path
        for path in result.domain_output_dirs["emotion"].rglob("descriptive_statistics.csv")
        if path.parent.parent.name == "combined"
    )
    rows = list(csv.reader(combined.open("r", encoding="utf-8-sig", newline="")))
    section_names = tuple(row[0] for row in rows if len(row) == 1 and row[0])
    assert "Arousal" in section_names
    assert "Engagement" not in section_names
    arousal_index = next(index for index, row in enumerate(rows) if row == ["Arousal"])
    arousal_mean = next(row for row in rows[arousal_index + 1 :] if row and row[0] == "mean")
    assert [float(value) for value in arousal_mean[1:]] == [75.0]
    assert {
        path.relative_to(tmp_path).as_posix(): path.read_bytes()
        for path in tmp_path.rglob("*")
        if path.is_file()
    } == before


def test_native_face_catalog_evidence_requires_the_root_run_manifest(tmp_path: Path) -> None:
    _write_source_sidecars(tmp_path)
    _write_face_video(tmp_path)

    with pytest.raises(ValueError, match="run manifest"):
        analyse_native_face_folder(
            tmp_path, output_root=tmp_path.parent / f"{tmp_path.name}-analysis-missing-run"
        )


def test_native_face_embedded_context_must_match_the_root_sidecars(tmp_path: Path) -> None:
    _write_source_sidecars(tmp_path)
    core = _write_face_video(tmp_path)
    _write_face_run_manifest(tmp_path)
    manifest_path = core.with_name("video_manifest.json")
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    payload["source"]["source_context"]["speaker"] = "tampered-speaker"
    context_bytes = json.dumps(
        payload["source"]["source_context"],
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    payload["source"]["source_context_sha256"] = hashlib.sha256(context_bytes).hexdigest()
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(ValueError, match="speaker"):
        analyse_native_face_folder(
            tmp_path, output_root=tmp_path.parent / f"{tmp_path.name}-analysis-tampered"
        )


@pytest.mark.parametrize(
    ("field", "forged"),
    [
        ("speaker", "forged-speaker"),
        ("speaker_display", "Forged Display"),
        ("system_metadata", {"title": "Forged Title"}),
        ("user_metadata", {"Country": "Elsewhere", "Gender": "F"}),
        ("output_mapping", {"video_directory": "C:/forged"}),
    ],
)
def test_native_face_outer_identity_must_match_validated_embedded_context(
    tmp_path: Path,
    field: str,
    forged: object,
) -> None:
    _write_source_sidecars(tmp_path)
    core = _write_face_video(tmp_path)
    _write_face_run_manifest(tmp_path)
    manifest_path = core.with_name("video_manifest.json")
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    payload["source"][field] = forged
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(ValueError, match=field):
        analyse_native_face_folder(
            tmp_path,
            output_root=tmp_path.parent / f"{tmp_path.name}-analysis-forged-{field}",
        )


def test_native_text_is_preferred_recomputed_and_profile_splits_are_allowed(tmp_path: Path) -> None:
    manifest = _write_source_sidecars(tmp_path)
    summary = _write_native_text(tmp_path)
    before = {path: path.read_bytes() for path in (manifest, summary)}

    discovery = discover_text_results(tmp_path)

    assert discovery.grain == "source"
    assert [item.source_ids for item in discovery.summaries] == [
        ("source-0001",), ("source-0002",),
    ]
    assert discovery.summaries[0].constructs["Text Valence"] == pytest.approx(.8)
    assert discovery.summaries[1].constructs["Text Valence"] == pytest.approx(-.6)
    profile = AnalysisProfile(
        manifest.resolve(), _sha(manifest),
        manual_groups=(
            ManualGroup("first", "First", (ProfileMember("source", "source-0001"),)),
            ManualGroup("second", "Second", (ProfileMember("source", "source-0002"),)),
        ),
    )
    result = build_combined_workbook(
        {}, tmp_path / "analysis.xlsx", analysis_profile=profile,
        text_summaries=discovery.summaries,
    )
    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
    text = book["Text sentiment"]
    valence_row = TEXT_CONSTRUCTS.index("Text Valence") + 2
    assert text.cell(valence_row, 4).value == pytest.approx(80.0)
    assert text.cell(valence_row, 8).value == pytest.approx(-60.0)
    # Overall is recomputed from selected positive/negative values, not .1 child mean.
    assert "S2" in text.cell(valence_row, 19).value
    assert "S3" in text.cell(valence_row, 19).value
    assert "AVERAGE" not in text.cell(valence_row, 19).value
    assert {path: path.read_bytes() for path in before} == before


def test_native_text_hash_tampering_is_rejected(tmp_path: Path) -> None:
    _write_source_sidecars(tmp_path)
    summary = _write_native_text(tmp_path)
    summary.write_text(summary.read_text(encoding="utf-8") + "\n", encoding="utf-8")

    with pytest.raises(TextResultsError, match="hash"):
        discover_text_results(tmp_path)


def test_native_text_rejects_discovered_summary_that_is_not_manifest_bound(
    tmp_path: Path,
) -> None:
    """Break caught: Analysis verified one CSV but parsed another compatible CSV."""

    _write_source_sidecars(tmp_path)
    discovered = _write_native_text(tmp_path)
    bound = discovered.with_name("bound.csv")
    discovered.replace(bound)

    batch_path = tmp_path / "postprocessing" / "text_postprocessing_batch_manifest.json"
    batch = json.loads(batch_path.read_text(encoding="utf-8"))
    batch["multimodal"]["video_summary"] = "multimodal/bound.csv"
    batch["multimodal"]["video_summary_sha256"] = _sha(bound)
    batch_path.write_text(json.dumps(batch), encoding="utf-8")

    rows = list(csv.reader(bound.read_text(encoding="utf-8").splitlines()))
    positive = rows[0].index("Positive Sentiment")
    negative = rows[0].index("Negative Sentiment")
    rows[1][positive] = "0.2"
    rows[1][negative] = "0.8"
    with discovered.open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle).writerows(rows)

    with pytest.raises(TextResultsError, match="manifest-bound video summary"):
        discover_text_results(tmp_path)


def test_native_text_parses_the_same_snapshot_that_was_manifest_verified(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Break caught: the summary path changed after its manifest hash check."""

    _write_source_sidecars(tmp_path)
    summary = _write_native_text(tmp_path)
    summary_resolved = summary.resolve()
    original_read_bytes = Path.read_bytes
    replaced = False

    def read_then_replace(path: Path) -> bytes:
        nonlocal replaced
        snapshot = original_read_bytes(path)
        if path.resolve() == summary_resolved and not replaced:
            rows = list(csv.reader(snapshot.decode("utf-8").splitlines()))
            positive = rows[0].index("Positive Sentiment")
            negative = rows[0].index("Negative Sentiment")
            rows[1][positive] = "0.2"
            rows[1][negative] = "0.8"
            with path.open("w", encoding="utf-8", newline="") as handle:
                csv.writer(handle).writerows(rows)
            replaced = True
        return snapshot

    monkeypatch.setattr(Path, "read_bytes", read_then_replace)

    discovery = discover_text_results(tmp_path)

    assert replaced
    assert discovery.summaries[0].constructs["Text Valence"] == pytest.approx(0.8)


def test_native_text_rejects_replacement_sidecars_with_the_same_source_ids(tmp_path: Path) -> None:
    _write_source_sidecars(tmp_path)
    _write_native_text(tmp_path)
    manifest = tmp_path / "postprocessing" / "source_manifest.json"
    replacement = json.loads(manifest.read_text(encoding="utf-8"))
    replacement["catalog"]["sha256"] = "b" * 64
    for source in replacement["sources"]:
        source["catalog_sha256"] = "b" * 64
    manifest.write_text(json.dumps(replacement), encoding="utf-8")

    with pytest.raises(TextResultsError, match="source manifest hash"):
        discover_text_results(tmp_path)


def test_native_text_rejects_nested_alternate_source_sidecar_pair(tmp_path: Path) -> None:
    """Break caught: Analysis accepted sidecars below, rather than at, the pair root."""

    _write_source_sidecars(tmp_path)
    _write_native_text(tmp_path)
    pair_root = tmp_path / "postprocessing"
    alternate = pair_root / "alternate-sidecars"
    alternate.mkdir()
    alternate_manifest = alternate / "source_manifest.json"
    alternate_manifest.write_text(
        json.dumps(
            json.loads((pair_root / "source_manifest.json").read_text(encoding="utf-8")),
            indent=2,
        ),
        encoding="utf-8",
    )
    alternate_metadata = alternate / "source_metadata.csv"
    alternate_metadata.write_bytes((pair_root / "source_metadata.csv").read_bytes())

    batch_path = pair_root / "text_postprocessing_batch_manifest.json"
    batch = json.loads(batch_path.read_text(encoding="utf-8"))
    binding = batch["source_binding"]
    binding["source_manifest"] = "alternate-sidecars/source_manifest.json"
    binding["source_manifest_sha256"] = _sha(alternate_manifest)
    binding["source_metadata"] = "alternate-sidecars/source_metadata.csv"
    binding["source_metadata_sha256"] = _sha(alternate_metadata)
    batch_path.write_text(json.dumps(batch), encoding="utf-8")

    with pytest.raises(TextResultsError, match="explicit run-root pair"):
        discover_text_results(tmp_path)


def test_native_text_rejects_context_contract_that_does_not_match_sidecars(tmp_path: Path) -> None:
    _write_source_sidecars(tmp_path)
    _write_native_text(tmp_path)
    batch_path = tmp_path / "postprocessing" / "text_postprocessing_batch_manifest.json"
    batch = json.loads(batch_path.read_text(encoding="utf-8"))
    context_record = batch["source_binding"]["source_contexts"][0]
    context_record["context"]["speaker"] = "forged-speaker"
    context_record["sha256"] = hashlib.sha256(
        json.dumps(
            context_record["context"],
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    batch_path.write_text(json.dumps(batch), encoding="utf-8")

    with pytest.raises(TextResultsError, match="source context speaker"):
        discover_text_results(tmp_path)
