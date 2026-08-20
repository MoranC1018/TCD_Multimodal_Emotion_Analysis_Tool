import csv
import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from typing import Mapping
from unittest.mock import patch

import openpyxl
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import get_column_letter, range_boundaries

from analysis.combined_summary import (
    AUDIO_LAYOUT,
    AUDIO_METRICS,
    VIDEO_LAYOUT,
    VIDEO_METRICS,
    SpeakerGroupDefinition,
)


EXCEL_ERROR_LITERALS = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
    "#SPILL!",
    "#CALC!",
    "#FIELD!",
    "#BLOCKED!",
    "#UNKNOWN!",
    "#CONNECT!",
}


def _cell_name(node: tuple[str, str]) -> str:
    return f"{node[0]}!{node[1]}"


def _formula_targets(
    book: openpyxl.Workbook,
    owner: tuple[str, str],
    formula: str,
) -> set[tuple[str, str]]:
    """Resolve every worksheet cell referenced by one generated formula."""

    targets: set[tuple[str, str]] = set()
    for token in Tokenizer(formula).items:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        reference = token.value
        if "[" in reference or "]" in reference:
            raise AssertionError(
                f"External workbook reference in {_cell_name(owner)}: {reference}"
            )
        if "!" in reference:
            sheet_token, cell_range = reference.rsplit("!", 1)
            if sheet_token.startswith("'"):
                if not sheet_token.endswith("'"):
                    raise AssertionError(
                        f"Malformed quoted sheet reference in {_cell_name(owner)}: {reference}"
                    )
                sheet_name = sheet_token[1:-1].replace("''", "'")
            else:
                sheet_name = sheet_token
        else:
            sheet_name = owner[0]
            cell_range = reference
        if sheet_name not in book.sheetnames:
            raise AssertionError(
                f"Missing worksheet referenced by {_cell_name(owner)}: {sheet_name}"
            )
        try:
            min_column, min_row, max_column, max_row = range_boundaries(cell_range)
        except ValueError as exc:
            raise AssertionError(
                f"Unsupported cell reference in {_cell_name(owner)}: {reference}"
            ) from exc
        if None in (min_column, min_row, max_column, max_row):
            raise AssertionError(
                f"Reference must identify cells in {_cell_name(owner)}: {reference}"
            )
        target_sheet = book[sheet_name]
        used_max_row = target_sheet.max_row
        used_max_column = target_sheet.max_column
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                coordinate = f"{get_column_letter(column)}{row}"
                if (
                    target_sheet[coordinate].value is None
                    and (row > used_max_row or column > used_max_column)
                ):
                    raise AssertionError(
                        f"Missing target referenced by {_cell_name(owner)}: "
                        f"{sheet_name}!{coordinate}"
                    )
                targets.add((sheet_name, coordinate))
    return targets


def assert_workbook_formula_safety(book: openpyxl.Workbook) -> None:
    """Reject broken references, Excel errors, and any formula dependency cycle."""

    formulas: dict[tuple[str, str], str] = {}
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                normalized = value.strip().upper()
                if cell.data_type == "e" or normalized in EXCEL_ERROR_LITERALS:
                    raise AssertionError(
                        f"Excel error literal in {sheet.title}!{cell.coordinate}: {value}"
                    )
                if value.startswith("="):
                    error = next(
                        (literal for literal in EXCEL_ERROR_LITERALS if literal in normalized),
                        None,
                    )
                    if error is not None:
                        raise AssertionError(
                            f"Excel error token in {sheet.title}!{cell.coordinate}: {error}"
                        )
                    formulas[(sheet.title, cell.coordinate)] = value

    dependencies: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for owner, formula in formulas.items():
        targets = _formula_targets(book, owner, formula)
        dependencies[owner] = {target for target in targets if target in formulas}

    states: dict[tuple[str, str], int] = {}
    stack: list[tuple[str, str]] = []

    def visit(node: tuple[str, str]) -> None:
        state = states.get(node, 0)
        if state == 2:
            return
        if state == 1:
            cycle_start = stack.index(node)
            cycle = stack[cycle_start:] + [node]
            raise AssertionError(
                "Formula dependency cycle: "
                + " -> ".join(_cell_name(item) for item in cycle)
            )
        states[node] = 1
        stack.append(node)
        for dependency in dependencies.get(node, ()):
            visit(dependency)
        stack.pop()
        states[node] = 2

    for formula_cell in dependencies:
        visit(formula_cell)


def populated_coordinates(sheet: object) -> set[str]:
    """Return every coordinate containing a value across the full used range."""

    return {
        cell.coordinate
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }


def expected_shared_cell_roles(
    modality: str,
    label_columns: tuple[int, ...],
    speaker_columns: tuple[int, ...],
) -> dict[str, str]:
    """Describe the sample-style shared modality worksheet."""

    metrics = AUDIO_METRICS if modality == "Audio" else VIDEO_METRICS
    overall_column = 19
    roles: dict[str, str] = {}

    def add(row: int, columns: tuple[int, ...], role: str = "source-link") -> None:
        for column in columns:
            roles[f"{get_column_letter(column)}{row}"] = role

    add(1, (*label_columns, *speaker_columns, overall_column))
    for row in range(2, len(metrics) + 2):
        add(row, label_columns)
        add(row, (*speaker_columns, overall_column), "probability")

    layout = AUDIO_LAYOUT if modality == "Audio" else VIDEO_LAYOUT
    add(layout.count_heading, (2,))
    for row in range(layout.count_start, layout.count_end + 1):
        add(row, (3, *speaker_columns))
    add(layout.detail_heading, (2,))
    for metric_index in range(len(metrics)):
        start_row = layout.detail_start + metric_index * 5
        for row in range(start_row, start_row + 5):
            add(row, (3, *speaker_columns))
        add(start_row, (2,))
    if layout.kurtosis_heading is not None:
        add(layout.kurtosis_heading, (2,))
        assert layout.kurtosis_start is not None and layout.kurtosis_end is not None
        for row in range(layout.kurtosis_start, layout.kurtosis_end + 1):
            add(row, (2, *speaker_columns))
    return roles


def _formula_reference(sheet_name: str, coordinate: str) -> str:
    quoted = sheet_name.replace("'", "''")
    return f"'{quoted}'!{coordinate}"


def expected_inference_formulas(
    references: tuple[str, ...],
    row: int,
    family_first_row: int,
    family_last_row: int,
    settings_row: int,
) -> dict[str, str]:
    """Return the exact live helper formulas required for an inference row."""

    values = ",".join(references)
    p_values = f"$J${family_first_row}:$J${family_last_row}"
    candidates = f"$O${family_first_row}:$O${family_last_row}"
    return {
        "B": f"=COUNT({values})",
        "C": f"={_formula_reference('Inference Settings', f'D{settings_row}')}",
        "D": f'=IFERROR(AVERAGE({values}),"")',
        "E": f'=IF(B{row}<2,"",_xlfn.STDEV.S({values}))',
        "F": f'=IF(OR(E{row}="",E{row}=0),"",E{row}/SQRT(B{row}))',
        "G": f'=IF(OR(F{row}="",F{row}=0),"",(D{row}-C{row})/F{row})',
        "H": f'=IF(OR(B{row}<2,E{row}=0),"",D{row}-_xlfn.T.INV.2T(0.05,B{row}-1)*F{row})',
        "I": f'=IF(OR(B{row}<2,E{row}=0),"",D{row}+_xlfn.T.INV.2T(0.05,B{row}-1)*F{row})',
        "J": f'=IF(OR(B{row}<2,E{row}=0),"",_xlfn.T.DIST.2T(ABS(G{row}),B{row}-1))',
        "K": f'=IF(J{row}="","",_xlfn.MINIFS({candidates},{p_values},">="&J{row}))',
        "L": f'=IF(OR(E{row}="",E{row}=0),"",(D{row}-C{row})/E{row})',
        "M": f'=IF(OR(B{row}<2,E{row}=0),"",_xlfn.T.DIST(G{row},B{row}-1,TRUE))',
        "O": f'=IF(J{row}="","",MIN(1,J{row}*COUNT({p_values})/COUNTIF({p_values},"<="&J{row})))',
    }


def write_sectioned_report(
    path: Path,
    metrics: tuple[str, ...],
    mean_by_metric: Mapping[str, float] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    for metric in metrics:
        means = (
            [str(mean_by_metric[metric])] * 5
            if mean_by_metric is not None
            else ["1", "2", "3", "4", "5"]
        )
        lines.extend(
            [
                metric,
                "classification,core,category,emotion,unit,score",
                "metric,001,002,003,004,005",
                "count,10,10,10,10,10",
                "missing,0,0,0,0,0",
                f"mean,{','.join(means)}",
                "stddev,1,1,1,1,1",
                "kurtosis,0,0,0,0,0",
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_text_construct_summary(path: Path, speaker: str = "Andy Burnham") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "Country", "Speaker", "Speaker ID", "Videos", "Valid segments",
                "RockSteady terms", "Positive valence", "Negative valence",
                "Arousal / Activation", "Dominance / Power",
                "Affiliation / Social orientation",
            ]
        )
        writer.writerow(
            ["UK", speaker, f"UK/{speaker}", 5, 20, 100, 0.2, 0.1, 0.3, 0.4, 0.5]
        )
    return path


def tree_hash(root: Path) -> str:
    digest = hashlib.sha256()
    for path in sorted(root.rglob("*")):
        digest.update(path.relative_to(root).as_posix().encode("utf-8"))
        if path.is_file():
            digest.update(path.read_bytes())
    return digest.hexdigest()


def write_full_imotions_csv(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(VIDEO_METRICS)
    metadata = [
        ["#INFO"],
        ["#Study name", "Workflow test"],
        ["#METADATA"],
        ["#Category", "Timestamp", *("FEA(Emotions)" for _ in columns)],
        ["#Description", "Timestamp", *columns],
        ["#Unit", "Millisecond", *("Index" for _ in columns)],
        ["#Group", "", *("Emotion" for _ in columns)],
        ["#Display name", "", *columns],
        [
            "#Channel identifier",
            "Timestamp",
            *(f"FEA_{column.replace(' ', '_')}" for column in columns),
        ],
        ["#DATA"],
        ["Row", "Timestamp", *columns],
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(metadata)
        writer.writerow([1, 0, *range(1, len(columns) + 1)])
        writer.writerow([2, 40, *range(2, len(columns) + 2)])


def write_compact_audio_csv(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(
            [
                ["#INFO"],
                ["#SpeakerName", "Andy Burnham"],
                ["#VideoTitle", "Workflow test"],
                ["#YoutubeID", "workflow"],
                ["#DATA"],
                [
                    "WindowIndex", "StartSeconds", "EndSeconds", "PredictedEmotion",
                    "EmotionConfidence", "Anger", "Contempt", "Disgust", "Fear", "Happiness",
                    "Neutral", "Sadness", "Surprise", "Other", "Arousal", "Dominance", "Valence",
                ],
                [
                    "1", "0", "10", "Neutral", "0.7", "0.1", "0.02", "0.03", "0.04",
                    "0.2", "0.7", "0.05", "0.06", "0.01", "0.55", "0.45", "0.75",
                ],
            ]
        )


class WorkflowTests(unittest.TestCase):
    def setUp(self) -> None:
        from analysis.workflow import ModalityRequest, WorkflowRequest

        self.ModalityRequest = ModalityRequest
        self.WorkflowRequest = WorkflowRequest
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.output_root = self.root / "workflow-output"
        self.audio_import_root = self.root / "imported-audio-reports"
        self.video_import_root = self.root / "imported-video-reports"
        self._write_reports(self.audio_import_root, "audio")
        self._write_reports(self.video_import_root, "video")
        self.groups = (
            SpeakerGroupDefinition("focus", "Focus", ("Andy Burnham",)),
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_generated_combined_manifest_is_ignored_in_custom_repository_output_folders(self) -> None:
        repository = Path(__file__).resolve().parents[2]
        candidate = repository / "reports" / "custom-run" / "combined_analysis_manifest.json"

        completed = subprocess.run(
            ["git", "check-ignore", "--no-index", str(candidate)],
            cwd=repository,
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
        )

        self.assertEqual(completed.returncode, 0, completed.stderr)

    def _write_reports(self, root: Path, modality: str) -> Path:
        metrics = AUDIO_METRICS if modality == "audio" else VIDEO_METRICS
        report = root / "emotion" / "Andy Burnham" / "combined" / "other_findings" / "descriptive_statistics.csv"
        write_sectioned_report(report, metrics)
        return root

    def _request(
        self,
        *methods: str,
        groups: tuple[SpeakerGroupDefinition, ...] | None = None,
        reverse: bool = False,
    ):
        modality_names = ("imotions", "audio")[: len(methods)]
        if reverse:
            modality_names = tuple(reversed(modality_names))
        modalities = tuple(
            self.ModalityRequest(
                name,
                method,
                (
                    self.video_import_root if name == "imotions" else self.audio_import_root
                ) if method == "import" else self.root / f"{name}-input",
            )
            for name, method in zip(modality_names, methods)
        )
        for modality in modalities:
            if modality.source_method == "run":
                modality.source_path.mkdir()
        return self.WorkflowRequest(
            output_root=self.output_root,
            modalities=modalities,
            speaker_groups=self.groups if groups is None else groups,
            write_combined_workbook=True,
        )

    def _analysis_result(self, root: Path):
        from analysis.histograms import AnalysisResult

        return AnalysisResult(input_dir=root, output_dir=root, other_findings_dir=root)

    def _assert_archived_pair(
        self,
        policy: Mapping[str, object],
        expected_workbook_hash: str,
    ) -> dict[str, object]:
        archive_directory = Path(str(policy["archive_directory"]))
        archived_workbook = Path(str(policy["archived_previous_workbook"]))
        archived_manifest = Path(str(policy["archived_previous_manifest"]))
        self.assertEqual(archived_workbook.parent, archive_directory)
        self.assertEqual(archived_manifest.parent, archive_directory)
        actual_hash = hashlib.sha256(archived_workbook.read_bytes()).hexdigest()
        self.assertEqual(actual_hash, expected_workbook_hash)
        self.assertEqual(policy["archived_previous_workbook_sha256"], expected_workbook_hash)

        payload = json.loads(archived_manifest.read_text(encoding="utf-8"))
        self.assertEqual(Path(payload["workbook_path"]), archived_workbook.resolve())
        archived_metadata = payload["archive"]
        self.assertEqual(
            Path(archived_metadata["archive_directory"]),
            archive_directory.resolve(),
        )
        self.assertEqual(archived_metadata["workbook_sha256"], expected_workbook_hash)
        return payload

    def _write_complete_fixed_output(self, *, include_profile: bool = True) -> tuple[bytes, bytes, bytes]:
        self.output_root.mkdir(parents=True, exist_ok=True)
        workbook = self.output_root / "combined_analysis.xlsx"
        profile = self.output_root / "analysis_profile.json"
        manifest = self.output_root / "combined_analysis_manifest.json"
        workbook.write_bytes(b"previous workbook")
        if include_profile:
            profile.write_bytes(b'{"profile":"previous"}\n')
        payload = {
            "status": "complete",
            "started_at": "2026-08-20T10:00:00Z",
            "workbook_path": str(workbook.resolve()),
            **(
                {"analysis_profile_path": str(profile.resolve())}
                if include_profile
                else {}
            ),
        }
        manifest.write_text(json.dumps(payload), encoding="utf-8")
        return workbook.read_bytes(), profile.read_bytes() if include_profile else b"", manifest.read_bytes()

    def test_archive_preflight_preserves_fixed_output_when_profile_copy_fails(self) -> None:
        from analysis.workflow import _archive_fixed_outputs

        before = self._write_complete_fixed_output()
        real_copy = shutil.copyfile

        def fail_open_profile(source: Path, destination: Path) -> None:
            if Path(source).name == "analysis_profile.json":
                raise PermissionError("profile is open")
            real_copy(source, destination)

        with patch("analysis.workflow._copy_archive_file", side_effect=fail_open_profile):
            with self.assertRaisesRegex(PermissionError, "profile is open"):
                _archive_fixed_outputs(self.output_root, "2026-08-20T11:00:00Z")

        self.assertEqual((self.output_root / "combined_analysis.xlsx").read_bytes(), before[0])
        self.assertEqual((self.output_root / "analysis_profile.json").read_bytes(), before[1])
        self.assertEqual((self.output_root / "combined_analysis_manifest.json").read_bytes(), before[2])
        history = self.output_root / "combined_analysis_history"
        self.assertFalse(history.exists() and any(history.iterdir()))

    @unittest.skipUnless(sys.platform == "win32", "Windows open-file regression")
    def test_archive_preflight_preserves_fixed_output_when_profile_is_open(self) -> None:
        import ctypes
        from ctypes import wintypes

        from analysis.workflow import _archive_fixed_outputs

        before = self._write_complete_fixed_output()
        profile = self.output_root / "analysis_profile.json"
        create_file = ctypes.windll.kernel32.CreateFileW
        create_file.argtypes = (
            wintypes.LPCWSTR,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.LPVOID,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.HANDLE,
        )
        create_file.restype = wintypes.HANDLE
        handle = create_file(str(profile), 0x80000000, 0, None, 3, 0x80, None)
        invalid_handle = ctypes.c_void_p(-1).value
        if handle == invalid_handle:
            self.skipTest("Could not open the profile with delete/read sharing disabled")
        try:
            with self.assertRaises(OSError):
                _archive_fixed_outputs(self.output_root, "2026-08-20T11:00:00Z")
        finally:
            ctypes.windll.kernel32.CloseHandle(handle)

        self.assertEqual((self.output_root / "combined_analysis.xlsx").read_bytes(), before[0])
        self.assertEqual(profile.read_bytes(), before[1])
        self.assertEqual((self.output_root / "combined_analysis_manifest.json").read_bytes(), before[2])
        history = self.output_root / "combined_analysis_history"
        self.assertFalse(history.exists() and any(history.iterdir()))

    @unittest.skipUnless(sys.platform == "win32", "Windows open-file regression")
    def test_archive_cleanup_rolls_back_when_workbook_disallows_delete_sharing(self) -> None:
        import ctypes
        from ctypes import wintypes

        from analysis.workflow import _archive_fixed_outputs

        before = self._write_complete_fixed_output()
        workbook = self.output_root / "combined_analysis.xlsx"
        create_file = ctypes.windll.kernel32.CreateFileW
        create_file.argtypes = (
            wintypes.LPCWSTR,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.LPVOID,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.HANDLE,
        )
        create_file.restype = wintypes.HANDLE
        handle = create_file(str(workbook), 0x80000000, 0x1, None, 3, 0x80, None)
        invalid_handle = ctypes.c_void_p(-1).value
        if handle == invalid_handle:
            self.skipTest("Could not open the workbook with read sharing but delete sharing disabled")
        try:
            with self.assertRaises(OSError):
                _archive_fixed_outputs(self.output_root, "2026-08-20T11:00:00Z")
        finally:
            ctypes.windll.kernel32.CloseHandle(handle)

        self.assertEqual(workbook.read_bytes(), before[0])
        self.assertEqual((self.output_root / "analysis_profile.json").read_bytes(), before[1])
        self.assertEqual((self.output_root / "combined_analysis_manifest.json").read_bytes(), before[2])
        history = self.output_root / "combined_analysis_history"
        self.assertFalse(history.exists() and any(history.iterdir()))

    @unittest.skipUnless(sys.platform == "win32", "Windows open-file regression")
    def test_archive_cleanup_rolls_back_after_workbook_move_when_profile_disallows_delete_sharing(
        self,
    ) -> None:
        import ctypes
        from ctypes import wintypes

        from analysis.workflow import _archive_fixed_outputs

        before = self._write_complete_fixed_output()
        profile = self.output_root / "analysis_profile.json"
        create_file = ctypes.windll.kernel32.CreateFileW
        create_file.argtypes = (
            wintypes.LPCWSTR,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.LPVOID,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.HANDLE,
        )
        create_file.restype = wintypes.HANDLE
        handle = create_file(str(profile), 0x80000000, 0x1, None, 3, 0x80, None)
        invalid_handle = ctypes.c_void_p(-1).value
        if handle == invalid_handle:
            self.skipTest("Could not open the profile with read sharing but delete sharing disabled")
        try:
            with self.assertRaises(OSError):
                _archive_fixed_outputs(self.output_root, "2026-08-20T11:00:00Z")
        finally:
            ctypes.windll.kernel32.CloseHandle(handle)

        self.assertEqual((self.output_root / "combined_analysis.xlsx").read_bytes(), before[0])
        self.assertEqual(profile.read_bytes(), before[1])
        self.assertEqual((self.output_root / "combined_analysis_manifest.json").read_bytes(), before[2])
        history = self.output_root / "combined_analysis_history"
        self.assertFalse(history.exists() and any(history.iterdir()))

    @unittest.skipUnless(sys.platform == "win32", "Windows junction regression")
    def test_archive_preflight_rejects_a_history_junction_redirect(self) -> None:
        from analysis.workflow import WorkflowError, _archive_fixed_outputs

        self._write_complete_fixed_output()
        redirect = self.root / "redirect-target"
        redirect.mkdir()
        history = self.output_root / "combined_analysis_history"
        completed = subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(history), str(redirect)],
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            self.skipTest(f"Could not create test junction: {completed.stderr.strip()}")

        with self.assertRaisesRegex(WorkflowError, "reparse|junction"):
            _archive_fixed_outputs(self.output_root, "2026-08-20T11:00:00Z")
        self.assertEqual(tuple(redirect.iterdir()), ())

    @unittest.skipUnless(sys.platform == "win32", "Windows junction regression")
    def test_archive_preflight_rejects_a_history_junction_without_fixed_outputs(self) -> None:
        from analysis.workflow import WorkflowError, _archive_fixed_outputs

        self.output_root.mkdir(parents=True)
        redirect = self.root / "empty-redirect-target"
        redirect.mkdir()
        history = self.output_root / "combined_analysis_history"
        completed = subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(history), str(redirect)],
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            self.skipTest(f"Could not create test junction: {completed.stderr.strip()}")

        with self.assertRaisesRegex(WorkflowError, "reparse|junction"):
            _archive_fixed_outputs(self.output_root, "2026-08-20T11:00:00Z")
        self.assertEqual(tuple(redirect.iterdir()), ())

    @unittest.skipUnless(sys.platform == "win32", "Windows junction regression")
    def test_workflow_rejects_an_output_root_junction_before_writing(self) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        redirect = self.root / "output-redirect-target"
        redirect.mkdir()
        completed = subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(self.output_root), str(redirect)],
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            self.skipTest(f"Could not create test junction: {completed.stderr.strip()}")

        with self.assertRaisesRegex(WorkflowError, "reparse|junction"):
            run_workflow(self._request("import"))

        self.assertEqual(tuple(redirect.iterdir()), ())

    @unittest.skipUnless(sys.platform == "win32", "Windows junction regression")
    def test_workflow_rejects_an_output_parent_junction_before_creating_the_root(self) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        redirect = self.root / "output-parent-redirect-target"
        redirect.mkdir()
        parent_junction = self.root / "output-parent-junction"
        completed = subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(parent_junction), str(redirect)],
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            self.skipTest(f"Could not create test junction: {completed.stderr.strip()}")
        self.output_root = parent_junction / "workflow-output"

        with self.assertRaisesRegex(WorkflowError, "reparse|junction"):
            run_workflow(self._request("import"))

        self.assertEqual(tuple(redirect.iterdir()), ())

    @unittest.skipUnless(sys.platform == "win32", "Windows junction regression")
    def test_failed_workbook_archive_rechecks_the_preflighted_history_boundary(self) -> None:
        from analysis.workflow import (
            WorkflowError,
            _archive_failed_workbook,
            _archive_fixed_outputs,
        )

        self.output_root.mkdir(parents=True)
        history = self.output_root / "combined_analysis_history"
        _archive_fixed_outputs(self.output_root, "2026-08-20T11:00:00Z")
        redirect = self.root / "late-redirect-target"
        redirect.mkdir()
        completed = subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(history), str(redirect)],
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            self.skipTest(f"Could not create test junction: {completed.stderr.strip()}")
        workbook = self.output_root / "combined_analysis.xlsx"
        workbook.write_bytes(b"partial workbook")

        with self.assertRaisesRegex(WorkflowError, "reparse|junction"):
            _archive_failed_workbook(
                self.output_root,
                "2026-08-20T11:00:00Z",
                history,
            )

        self.assertEqual(workbook.read_bytes(), b"partial workbook")
        self.assertEqual(tuple(redirect.iterdir()), ())

    def test_failed_workflow_can_be_retried_without_manual_cleanup(self) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        request = self._request("import", "import")

        def fail_after_partial_output(*args, **kwargs):
            destination = Path(args[1])
            destination.write_bytes(b"partial workbook")
            raise RuntimeError("deliberate first-run failure")

        with patch("analysis.workflow.build_combined_workbook", side_effect=fail_after_partial_output):
            with self.assertRaisesRegex(WorkflowError, "deliberate first-run failure"):
                run_workflow(request)

        failed_manifest = json.loads(
            (self.output_root / "combined_analysis_manifest.json").read_text(encoding="utf-8")
        )
        self.assertEqual(failed_manifest["status"], "failed")

        result = run_workflow(request)

        self.assertTrue(result.workbook_path and result.workbook_path.is_file())
        complete_manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(complete_manifest["status"], "complete")
        quarantined = list(
            (self.output_root / "combined_analysis_history").rglob(
                "combined_analysis_manifest.json"
            )
        )
        self.assertTrue(
            any(json.loads(path.read_text(encoding="utf-8"))["status"] == "failed" for path in quarantined)
        )

    def test_failed_fixed_manifest_profile_and_partial_workbook_are_quarantined_together(
        self,
    ) -> None:
        from analysis.workflow import _archive_fixed_outputs

        before = self._write_complete_fixed_output()
        manifest = self.output_root / "combined_analysis_manifest.json"
        failed_payload = json.loads(manifest.read_text(encoding="utf-8"))
        failed_payload["status"] = "failed"
        manifest.write_text(json.dumps(failed_payload), encoding="utf-8")

        policy = _archive_fixed_outputs(
            self.output_root,
            "2026-08-20T11:00:00Z",
        )

        quarantine = Path(str(policy["quarantined_failed_directory"]))
        self.assertEqual(
            (quarantine / "combined_analysis.xlsx").read_bytes(),
            before[0],
        )
        self.assertEqual(
            (quarantine / "analysis_profile.json").read_bytes(),
            before[1],
        )
        self.assertEqual(
            json.loads((quarantine / "combined_analysis_manifest.json").read_text(encoding="utf-8"))[
                "status"
            ],
            "failed",
        )
        self.assertFalse((self.output_root / "combined_analysis.xlsx").exists())
        self.assertFalse((self.output_root / "analysis_profile.json").exists())
        self.assertFalse(manifest.exists())

    @patch("analysis.workflow.add_probability_mirrors")
    @patch("analysis.workflow.analyse_audio_folder")
    @patch("analysis.workflow.analyse_imotions_folder")
    def test_runs_video_then_audio_then_combined_with_submitted_groups(self, run_video, run_audio, add_mirrors) -> None:
        from analysis.workflow import run_workflow

        events: list[str] = []
        video_root = self._write_reports(self.output_root / "video", "video")
        audio_root = self._write_reports(self.output_root / "audio", "audio")
        run_video.side_effect = lambda *args, **kwargs: (events.append("video analyser"), self._analysis_result(video_root))[1]
        run_audio.side_effect = lambda *args, **kwargs: (events.append("audio analyser"), self._analysis_result(audio_root))[1]

        result = run_workflow(self._request("run", "run", reverse=True), progress=events.append)

        self.assertEqual(events[:2], ["Starting Video / iMotions analysis", "video analyser"])
        self.assertEqual(events[2], f"Completed Video / iMotions analysis: {video_root.resolve()}")
        self.assertEqual(events[3], "Starting Audio analysis")
        self.assertTrue(result.workbook_path.is_file())
        self.assertEqual(run_video.call_args.kwargs["output_root"], self.output_root / "video")
        self.assertEqual(run_audio.call_args.kwargs["output_root"], self.output_root / "audio")
        self.assertTrue(add_mirrors.called)
        self.assertEqual(result.modality_roots, {"video": video_root.resolve(), "audio": audio_root.resolve()})
        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["status"], "complete")
        self.assertEqual(manifest["request"]["speaker_groups"][0]["id"], "focus")

    def test_imported_reports_are_not_modified(self) -> None:
        from analysis.workflow import run_workflow

        before = (tree_hash(self.video_import_root), tree_hash(self.audio_import_root))
        result = run_workflow(self._request("import", "import"))

        self.assertEqual((tree_hash(self.video_import_root), tree_hash(self.audio_import_root)), before)
        self.assertTrue(result.workbook_path.is_file())

    def test_legacy_text_sentiment_results_populate_combined_workbook_without_running_text_code(self) -> None:
        from analysis.workflow import run_workflow

        text_root = self.root / "imported-text-results"
        summary_path = write_text_construct_summary(
            text_root / "text_output" / "multimodal" / "speaker_level_summary.csv"
        )
        before = tree_hash(text_root)
        request = self.WorkflowRequest(
            output_root=self.root / "text-import-output",
            modalities=(
                self.ModalityRequest("audio", "import", self.audio_import_root),
                self.ModalityRequest("text", "import", text_root),
            ),
            speaker_groups=self.groups,
        )

        result = run_workflow(request)

        self.assertEqual(tree_hash(text_root), before)
        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(book["Text sentiment"]["D2"].value, 0.2)
        positive_row = next(
            row
            for row in range(1, book["Construct Comparison"].max_row + 1)
            if book["Construct Comparison"].cell(row, 1).value
            == "Sentiment: Positive Sentiment"
        )
        self.assertIn(
            "'Text sentiment'!D2",
            book["Construct Comparison"].cell(positive_row, 4).value,
        )
        self.assertNotIn("Text sentiment Prob", book.sheetnames)
        self.assertNotIn("Measure Guide Prob", book.sheetnames)
        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["modality_roots"]["text"], str(text_root.resolve()))
        self.assertTrue(
            any(
                entry["modality"] == "text" and entry["path"] == str(summary_path.resolve())
                for entry in manifest["accepted_reports"]
            )
        )

    def test_combined_report_options_can_disable_comparison_and_probability_outputs(self) -> None:
        from analysis.workflow import run_workflow

        request = self.WorkflowRequest(
            output_root=self.root / "report-options-output",
            modalities=(self.ModalityRequest("audio", "import", self.audio_import_root),),
            speaker_groups=self.groups,
            include_construct_comparison=False,
            include_probability_sheets=False,
            confidence_level=0.90,
            headline_policy="equal",
        )

        result = run_workflow(request)

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertNotIn("Construct Comparison", book.sheetnames)
        self.assertFalse(any(name.endswith(" Prob") for name in book.sheetnames))
        self.assertNotIn("Inference Details", book.sheetnames)
        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        self.assertFalse(manifest["request"]["include_construct_comparison"])
        self.assertFalse(manifest["request"]["include_probability_sheets"])
        self.assertEqual(manifest["request"]["confidence_level"], 0.90)
        self.assertEqual(manifest["request"]["headline_policy"], "equal")

    def test_successful_rerun_archives_a_self_contained_matching_workbook_manifest_pair(self) -> None:
        from analysis.workflow import run_workflow

        first_request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(self.ModalityRequest("audio", "import", self.audio_import_root),),
            speaker_groups=self.groups,
            default_reference=0.0,
        )
        first = run_workflow(first_request)
        first_workbook_hash = hashlib.sha256(first.workbook_path.read_bytes()).hexdigest()

        second_request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=first_request.modalities,
            speaker_groups=self.groups,
            default_reference=1.0,
        )
        second = run_workflow(second_request)

        current_manifest = json.loads(second.manifest_path.read_text(encoding="utf-8"))
        archived_manifest = self._assert_archived_pair(
            current_manifest["stale_artifact_policy"],
            first_workbook_hash,
        )
        self.assertEqual(archived_manifest["request"]["default_reference"], 0.0)
        self.assertEqual(current_manifest["request"]["default_reference"], 1.0)
        current_hash = hashlib.sha256(second.workbook_path.read_bytes()).hexdigest()
        self.assertNotEqual(current_hash, first_workbook_hash)

    def test_manifest_records_versioned_accepted_and_rejected_discovery_provenance(self) -> None:
        from analysis.workflow import run_workflow

        rejected = (
            self.audio_import_root
            / "emotion"
            / "cache-copy"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(rejected, AUDIO_METRICS)

        result = run_workflow(
            self.WorkflowRequest(
                output_root=self.output_root,
                modalities=(self.ModalityRequest("audio", "import", self.audio_import_root),),
                speaker_groups=self.groups,
            )
        )

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        self.assertRegex(manifest["software"]["version"], r"^\d+\.\d+\.\d+$")
        self.assertIn("git_revision", manifest["software"])
        accepted = manifest["accepted_reports"][0]
        self.assertEqual(accepted["modality"], "audio")
        self.assertEqual(accepted["normalized_speaker"], "andyburnham")
        self.assertEqual(accepted["display_speaker"], "Andy Burnham")
        self.assertEqual(accepted["reason"], "accepted speaker-level combined report")
        rejected_entry = next(item for item in manifest["rejected_reports"] if item["path"] == str(rejected.resolve()))
        self.assertEqual(rejected_entry["modality"], "audio")
        self.assertEqual(rejected_entry["normalized_speaker"], "cachecopy")
        self.assertEqual(rejected_entry["display_speaker"], "cache-copy")
        self.assertIn("generated or temporary", rejected_entry["reason"])

    def test_failed_discovery_manifest_preserves_zero_source_rejection(self) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        analysis_root = self.root / "rejected-audio-analysis"
        candidate = (
            analysis_root
            / "emotion"
            / "cache-copy"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(candidate, AUDIO_METRICS)
        before = tree_hash(analysis_root)
        request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(self.ModalityRequest("audio", "import", analysis_root),),
            speaker_groups=self.groups,
        )

        with self.assertRaisesRegex(WorkflowError, "no speaker-level combined reports"):
            run_workflow(request)

        self.assertEqual(tree_hash(analysis_root), before)
        manifest_path = self.output_root / "combined_analysis_manifest.json"
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["status"], "failed")
        self.assertEqual(manifest["accepted_reports"], [])
        rejected = {entry["path"]: entry["reason"] for entry in manifest["rejected_reports"]}
        self.assertIn(str(candidate.resolve()), rejected)
        self.assertIn("generated or temporary", rejected[str(candidate.resolve())])

    def test_failed_discovery_manifest_preserves_invalid_and_ambiguous_candidates(self) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        scenarios: list[tuple[str, Path, tuple[Path, ...]]] = []

        invalid_root = self.root / "invalid-audio-analysis"
        invalid = (
            invalid_root
            / "emotion"
            / "Andy Burnham"
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
        )
        write_sectioned_report(invalid, ("Anger",))
        scenarios.append(("invalid", invalid_root, (invalid,)))

        ambiguous_root = self.root / "ambiguous-audio-analysis"
        ambiguous_paths = tuple(
            ambiguous_root
            / "emotion"
            / speaker
            / "combined"
            / "other_findings"
            / "descriptive_statistics.csv"
            for speaker in ("Andy-Burnham", "Andy Burnham")
        )
        for candidate in ambiguous_paths:
            write_sectioned_report(candidate, AUDIO_METRICS)
        scenarios.append(("ambiguous", ambiguous_root, ambiguous_paths))

        for label, analysis_root, candidates in scenarios:
            output_root = self.root / f"{label}-workflow-output"
            input_root = self.root / f"{label}-audio-input"
            input_root.mkdir()
            request = self.WorkflowRequest(
                output_root=output_root,
                modalities=(self.ModalityRequest("audio", "run", input_root),),
                speaker_groups=self.groups,
            )
            before = tree_hash(analysis_root)

            with self.subTest(label=label), patch(
                "analysis.workflow.analyse_audio_folder",
                return_value=self._analysis_result(analysis_root),
            ), self.assertRaises(WorkflowError):
                run_workflow(request)

            self.assertEqual(tree_hash(analysis_root), before)
            manifest_path = output_root / "combined_analysis_manifest.json"
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            rejected = {entry["path"]: entry["reason"] for entry in manifest["rejected_reports"]}
            for candidate in candidates:
                self.assertIn(str(candidate.resolve()), rejected)
                self.assertIn(label, rejected[str(candidate.resolve())].casefold())

    def test_fresh_imotions_run_reaches_combined_workbook(self) -> None:
        from analysis.workflow import run_workflow

        input_dir = self.root / "iMotions_Run" / "Andy Burnham" / "Sensor Data"
        for index in range(1, 6):
            write_full_imotions_csv(input_dir / f"{index:03}_Workflow.csv")
        request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(self.ModalityRequest("imotions", "run", input_dir.parent.parent, write_graphs=False),),
            speaker_groups=self.groups,
        )

        result = run_workflow(request)

        self.assertTrue(result.workbook_path.is_file())
        report = (
            result.modality_roots["video"] / "emotion" / "iMotions_Run" / "Andy_Burnham"
            / "combined" / "other_findings" / "descriptive_statistics.csv"
        )
        self.assertTrue(report.is_file())
        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["modality_roots"]["video"], str((self.output_root / "video").resolve()))

    def test_fresh_audio_run_reaches_combined_workbook(self) -> None:
        from analysis.workflow import run_workflow

        input_dir = self.root / "audio_run"
        for index in range(1, 6):
            write_compact_audio_csv(
                input_dir / "Andy Burnham" / f"Workflow_Test_{index}_[workflow{index}]" / "audio_analysis.csv"
            )
        request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(self.ModalityRequest("audio", "run", input_dir, write_graphs=False),),
            speaker_groups=self.groups,
        )

        result = run_workflow(request)

        self.assertTrue(result.workbook_path.is_file())
        report = (
            result.modality_roots["audio"] / "emotion" / "audio_run" / "Andy_Burnham"
            / "combined" / "other_findings" / "descriptive_statistics.csv"
        )
        self.assertTrue(report.is_file())

    def test_rejects_invalid_requests_before_any_output(self) -> None:
        from analysis.workflow import ModalityRequest, WorkflowError, run_workflow

        request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(ModalityRequest("text", "run", self.video_import_root),),
            speaker_groups=self.groups,
        )

        with self.assertRaisesRegex(WorkflowError, "import-only"):
            run_workflow(request)

        self.assertFalse(self.output_root.exists())

    def test_combined_workbook_requires_at_least_one_group_but_analysis_only_does_not(self) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        combined_request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(self.ModalityRequest("audio", "import", self.audio_import_root),),
            speaker_groups=(),
            write_combined_workbook=True,
        )
        with self.assertRaisesRegex(WorkflowError, "speaker group"):
            run_workflow(combined_request)
        self.assertFalse(self.output_root.exists())

        analysis_only = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(self.ModalityRequest("audio", "import", self.audio_import_root),),
            speaker_groups=(),
            write_combined_workbook=False,
        )
        result = run_workflow(analysis_only)
        self.assertIsNone(result.workbook_path)
        self.assertFalse((self.output_root / "combined_analysis.xlsx").exists())

    def test_cli_rejects_omitted_or_explicitly_empty_groups_for_combined_output(self) -> None:
        repository = Path(__file__).resolve().parents[2]
        for group_arguments in ((), ("--speaker-groups-json", "[]")):
            output = self.root / ("cli-omitted" if not group_arguments else "cli-empty")
            completed = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "analysis.workflow",
                    "--output-root",
                    str(output),
                    "--audio-source",
                    str(self.audio_import_root),
                    "--audio-method",
                    "import",
                    *group_arguments,
                ],
                cwd=repository,
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
            with self.subTest(group_arguments=group_arguments):
                self.assertNotEqual(completed.returncode, 0)
                self.assertIn("speaker group", completed.stderr + completed.stdout)
                self.assertFalse((output / "combined_analysis.xlsx").exists())

    def test_cli_payload_errors_are_one_line_sanitized_and_do_not_echo_values(self) -> None:
        repository = Path(__file__).resolve().parents[2]
        cases = (
            (
                "malformed-json",
                ("--reference-overrides-json", "{authorization=secret-hf_bad"),
                "--reference-overrides-json must be valid JSON",
            ),
            (
                "malformed-groups",
                (
                    "--speaker-groups-json",
                    json.dumps({"authorization": "secret-hf_bad"}),
                ),
                "--speaker-groups-json must contain a list",
            ),
            (
                "nonnumeric-override",
                (
                    "--reference-overrides-json",
                    json.dumps({"Audio - Focus": "authorization=secret-hf_bad"}),
                ),
                "Reference override values must be finite numbers",
            ),
        )

        for label, arguments, expected in cases:
            completed = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "analysis.workflow",
                    "--output-root",
                    str(self.root / f"cli-{label}"),
                    *arguments,
                ],
                cwd=repository,
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
            output = (completed.stderr + completed.stdout).strip()
            with self.subTest(label=label):
                self.assertEqual(completed.returncode, 1)
                self.assertEqual(output, f"WorkflowError: {expected}")
                self.assertNotIn("Traceback", output)
                self.assertNotIn("authorization=secret", output)
                self.assertNotIn("hf_bad", output)

    @patch("analysis.workflow.analyse_audio_folder", side_effect=RuntimeError("secret-token=hf_123 audio broke"))
    def test_failed_rerun_invalidates_stale_workbook_and_records_sanitized_stage(self, run_audio) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        successful = run_workflow(self._request("import"))
        old_workbook_hash = hashlib.sha256(successful.workbook_path.read_bytes()).hexdigest()
        fresh_audio = self.root / "fresh-audio"
        fresh_audio.mkdir()
        failed_request = self.WorkflowRequest(
            output_root=self.output_root,
            modalities=(self.ModalityRequest("audio", "run", fresh_audio),),
            speaker_groups=self.groups,
        )

        with self.assertRaisesRegex(WorkflowError, "Audio analysis failed"):
            run_workflow(failed_request)

        self.assertFalse((self.output_root / "combined_analysis.xlsx").exists())
        manifest = json.loads((self.output_root / "combined_analysis_manifest.json").read_text(encoding="utf-8"))
        self.assertEqual(manifest["status"], "failed")
        self.assertEqual(manifest["failed_stage"], "Audio analysis")
        self.assertIn("audio broke", manifest["error"])
        self.assertNotIn("hf_123", manifest["error"])
        policy = manifest["stale_artifact_policy"]
        self.assertEqual(policy["policy"], "archive_fixed_outputs_before_run")
        archived_manifest = self._assert_archived_pair(policy, old_workbook_hash)
        self.assertEqual(archived_manifest["status"], "complete")

    @patch("analysis.workflow.analyse_audio_folder", side_effect=RuntimeError("audio broke"))
    @patch("analysis.workflow.analyse_imotions_folder")
    def test_failure_preserves_completed_output_writes_failed_manifest_and_skips_workbook(self, run_video, run_audio) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        video_root = self._write_reports(self.output_root / "video", "video")
        run_video.return_value = self._analysis_result(video_root)

        with self.assertRaisesRegex(WorkflowError, "Audio analysis failed"):
            run_workflow(self._request("run", "run"))

        self.assertTrue((video_root / "emotion" / "Andy Burnham" / "combined" / "other_findings" / "descriptive_statistics.csv").is_file())
        self.assertFalse((self.output_root / "combined_analysis.xlsx").exists())
        manifest = json.loads((self.output_root / "combined_analysis_manifest.json").read_text(encoding="utf-8"))
        self.assertEqual(manifest["status"], "failed")
        self.assertEqual(manifest["modality_roots"], {"video": str(video_root.resolve())})
        self.assertEqual(manifest["failed_stage"], "Audio analysis")
        self.assertIn("audio broke", manifest["error"])


class ImportedWorkflowWorkbookIntegrationTests(unittest.TestCase):
    """Exercise the complete imported-report coordinator and workbook contract."""

    def setUp(self) -> None:
        from analysis.workflow import ModalityRequest, WorkflowRequest

        self.ModalityRequest = ModalityRequest
        self.WorkflowRequest = WorkflowRequest
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.output_root = self.root / "workflow-output"
        self.audio_root = self.root / "imported-audio"
        self.video_root = self.root / "imported-video"
        self.speaker_bases = {
            "Andy Burnham": 10.0,
            "Marine Le Pen": 20.0,
            "Nigel Farage": 30.0,
        }
        for metrics, root in (
            (AUDIO_METRICS, self.audio_root),
            (VIDEO_METRICS, self.video_root),
        ):
            for speaker, base in self.speaker_bases.items():
                values = {metric: base + index for index, metric in enumerate(metrics)}
                if speaker == "Marine Le Pen":
                    # The pair has identical Anger means, exercising zero variance.
                    values["Anger"] = self.speaker_bases["Andy Burnham"]
                report = (
                    root
                    / "emotion"
                    / speaker
                    / "combined"
                    / "other_findings"
                    / "descriptive_statistics.csv"
                )
                write_sectioned_report(report, metrics, values)

        self.groups = (
            SpeakerGroupDefinition("pair", "Pair", ("Marine Le Pen", "Andy Burnham")),
            SpeakerGroupDefinition("singleton", "Singleton", ("Nigel Farage",)),
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    @staticmethod
    def _formula_link(sheet_name: str, coordinate: str) -> str:
        return f"={_formula_reference(sheet_name, coordinate)}"

    def _assert_no_formula_errors_or_obvious_cycles(self, book: openpyxl.Workbook) -> None:
        assert_workbook_formula_safety(book)

    def test_formula_safety_rejects_missing_sheet_and_target_references(self) -> None:
        book = openpyxl.Workbook()
        book.active["A1"] = "='Missing Sheet'!B2"

        with self.assertRaises(AssertionError):
            self._assert_no_formula_errors_or_obvious_cycles(book)

        book.active["A1"] = "=B2"
        with self.assertRaises(AssertionError):
            self._assert_no_formula_errors_or_obvious_cycles(book)

    def test_formula_safety_rejects_multi_cell_and_cross_sheet_cycles(self) -> None:
        book = openpyxl.Workbook()
        first = book.active
        first.title = "First Sheet"
        second = book.create_sheet("Second Sheet")
        first["A1"] = "='Second Sheet'!B2"
        second["B2"] = "='First Sheet'!C3"
        first["C3"] = "=A1"

        with self.assertRaises(AssertionError):
            self._assert_no_formula_errors_or_obvious_cycles(book)

    def test_formula_safety_rejects_direct_cycles_and_excel_error_literals(self) -> None:
        book = openpyxl.Workbook()
        book.active["A1"] = "=A1"
        with self.assertRaisesRegex(AssertionError, "dependency cycle"):
            self._assert_no_formula_errors_or_obvious_cycles(book)

        book.active["A1"] = "#DIV/0!"
        with self.assertRaisesRegex(AssertionError, "Excel error literal"):
            self._assert_no_formula_errors_or_obvious_cycles(book)

        book.active["A1"] = "=#REF!"
        with self.assertRaisesRegex(AssertionError, "Excel error token"):
            self._assert_no_formula_errors_or_obvious_cycles(book)

    def test_formula_safety_accepts_quoted_cross_sheet_ranges(self) -> None:
        book = openpyxl.Workbook()
        source = book.active
        source.title = "O'Brien"
        source["A1"] = 1
        source["A2"] = 2
        summary = book.create_sheet("Summary")
        summary["B2"] = "=SUM('O''Brien'!A1:A2)"

        self._assert_no_formula_errors_or_obvious_cycles(book)

    def _assert_mirror_contract(
        self,
        book: openpyxl.Workbook,
    ) -> None:
        details = book["Inference Details"]
        detail_rows = {
            details.cell(row, 1).value: row
            for row in range(2, details.max_row + 1)
        }
        speaker_targets = ((4, "marinelepen"), (5, "andyburnham"), (8, "nigelfarage"))
        for modality, metrics in (("Audio", AUDIO_METRICS), ("Video", VIDEO_METRICS)):
            source = book[modality]
            mirror = book[f"{modality} Prob"]
            expected_roles = expected_shared_cell_roles(modality, (2, 7), (4, 5, 8))
            expected_coordinates = set(expected_roles)
            self.assertEqual(populated_coordinates(source), expected_coordinates)
            self.assertEqual(populated_coordinates(mirror), expected_coordinates)

            probability_keys: dict[str, str] = {}
            for row, metric in enumerate(metrics, start=2):
                for column, speaker_id in speaker_targets:
                    probability_keys[f"{get_column_letter(column)}{row}"] = (
                        f"{modality}|{metric}|{speaker_id}"
                    )
                probability_keys[f"S{row}"] = f"{modality}|{metric}|overall"

            for coordinate, role in expected_roles.items():
                if role == "probability":
                    detail_row = detail_rows[probability_keys[coordinate]]
                    self.assertEqual(
                        mirror[coordinate].value,
                        self._formula_link("Inference Details", f"M{detail_row}"),
                    )
                    self.assertEqual(mirror[coordinate].number_format, "0.00%")
                else:
                    self.assertEqual(
                        mirror[coordinate].value,
                        self._formula_link(modality, coordinate),
                    )

    def _assert_static_sheet_contract(self, book: openpyxl.Workbook) -> None:
        for static_name in (
            "Domain Def Text",
            "Domain Def Speech",
            "Measure Guide",
            "Text sentiment",
        ):
            self.assertNotIn(f"{static_name} Prob", book.sheetnames)
            values = [
                cell.value
                for row in book[static_name].iter_rows()
                for cell in row
                if cell.value is not None
            ]
            self.assertFalse(
                any(isinstance(value, str) and value.startswith("=") for value in values),
                f"Static sheet contains a formula: {static_name}",
            )
            if static_name == "Text sentiment":
                self.assertEqual(values, [], "Text sentiment must remain entirely empty")
            else:
                self.assertTrue(values, f"Static definition sheet is unexpectedly empty: {static_name}")

    def _assert_inference_formula_contract(
        self,
        book: openpyxl.Workbook,
    ) -> None:
        details = book["Inference Details"]
        inputs = book["Inference Inputs"]
        input_rows = {
            inputs.cell(row, 1).value: row
            for row in range(2, inputs.max_row + 1)
        }
        expected_keys: list[str] = []
        speaker_targets = ("marinelepen", "andyburnham", "nigelfarage")
        configurations = (("Audio", AUDIO_METRICS), ("Video", VIDEO_METRICS))
        for source_name, metrics in configurations:
            for metric in metrics:
                expected_keys.extend(f"{source_name}|{metric}|{speaker_id}" for speaker_id in speaker_targets)
                expected_keys.append(f"{source_name}|{metric}|overall")
        actual_keys = [details.cell(row, 1).value for row in range(2, details.max_row + 1)]
        self.assertEqual(actual_keys, expected_keys)
        detail_rows = {key: row for row, key in enumerate(expected_keys, start=2)}

        intended_formula_columns = set("BCDEFGHIJKLMO")
        family_start = 2
        settings_start = 3
        for source_name, metrics in configurations:
            family_count = len(metrics) * 4
            family_first_row = family_start
            family_last_row = family_start + family_count - 1
            for source_row, metric in enumerate(metrics, start=2):
                settings_row = settings_start + source_row - 2
                for speaker_id in speaker_targets:
                    target_key = f"{source_name}|{metric}|{speaker_id}"
                    detail_row = detail_rows[target_key]
                    input_row = input_rows[target_key]
                    references = tuple(
                        _formula_reference("Inference Inputs", f"{column}{input_row}")
                        for column in ("C", "E", "G", "I", "K")
                    )
                    self._assert_detail_formula_row(
                        details, detail_row, references, family_first_row, family_last_row,
                        settings_row, intended_formula_columns, f"{source_name} {metric} {speaker_id}",
                    )
                overall_key = f"{source_name}|{metric}|overall"
                overall_row = detail_rows[overall_key]
                references = tuple(
                    _formula_reference(source_name, f"{column}{source_row}")
                    for column in ("D", "E", "H")
                )
                self._assert_detail_formula_row(
                    details, overall_row, references, family_first_row, family_last_row,
                    settings_row, intended_formula_columns, f"{source_name} {metric} overall",
                )
            family_start = family_last_row + 1
            settings_start += len(metrics)

    def _assert_detail_formula_row(
        self,
        details: object,
        detail_row: int,
        references: tuple[str, ...],
        family_first_row: int,
        family_last_row: int,
        settings_row: int,
        intended_formula_columns: set[str],
        context: str,
    ) -> None:
        expected = expected_inference_formulas(
            references,
            detail_row,
            family_first_row,
            family_last_row,
            settings_row,
        )
        actual_formula_columns = {
            details.cell(detail_row, column).column_letter
            for column in range(1, details.max_column + 1)
            if isinstance(details.cell(detail_row, column).value, str)
            and details.cell(detail_row, column).value.startswith("=")
        }
        self.assertEqual(actual_formula_columns, intended_formula_columns)
        for column, formula in expected.items():
            self.assertEqual(
                details[f"{column}{detail_row}"].value,
                formula,
                f"Unexpected {column}-column helper for {context}",
            )

    def _run_imported_workflow(self, output_root: Path | None = None):
        from analysis.workflow import run_workflow

        request = self.WorkflowRequest(
            output_root=output_root or self.output_root,
            modalities=(
                self.ModalityRequest("imotions", "import", self.video_root),
                self.ModalityRequest("audio", "import", self.audio_root),
            ),
            speaker_groups=self.groups,
            write_combined_workbook=True,
            default_reference=0.0,
        )
        return run_workflow(request)

    def test_manifest_audits_metric_and_sheet_reference_resolution(self) -> None:
        from analysis.workflow import run_workflow

        result = run_workflow(
            self.WorkflowRequest(
                output_root=self.root / "reference-audit-output",
                modalities=(self.ModalityRequest("audio", "import", self.audio_root),),
                speaker_groups=self.groups,
                reference_overrides={
                    "Audio|Anger": 3.0,
                    "Audio": 4.0,
                },
            )
        )

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        by_source = {entry["matched_source"]: entry for entry in manifest["reference_resolutions"]}
        self.assertEqual(by_source["Audio|Anger"]["original_key"], "Audio|Anger")
        self.assertEqual(by_source["Audio|Anger"]["matched_scope"], "metric")
        self.assertEqual(by_source["Audio|Anger"]["resolved_reference"], 3.0)
        singleton = next(
            entry
            for entry in manifest["reference_resolutions"]
            if entry["matched_scope"] == "sheet" and entry["matched_source"] == "Audio"
        )
        self.assertEqual(singleton["original_key"], "Audio")
        self.assertEqual(singleton["resolved_reference"], 4.0)

    def test_stale_or_misspelled_reference_override_fails_the_workflow(self) -> None:
        from analysis.workflow import WorkflowError, run_workflow

        request = self.WorkflowRequest(
            output_root=self.root / "bad-reference-output",
            modalities=(self.ModalityRequest("audio", "import", self.audio_root),),
            speaker_groups=self.groups,
            reference_overrides={"Audio:arousal": 2.0},
        )

        with self.assertRaisesRegex(WorkflowError, "Unknown reference override.*Audio:arousal"):
            run_workflow(request)

    def test_mirror_contract_rejects_extra_probability_cell(self) -> None:
        result = self._run_imported_workflow(self.root / "mirror-mutation-output")
        book = openpyxl.load_workbook(result.workbook_path, data_only=False)

        book["Audio Prob"]["Z99"] = "='Inference Details'!M2"
        with self.assertRaises(AssertionError):
            self._assert_mirror_contract(book)

    def test_inference_contract_rejects_frozen_helper_formula(self) -> None:
        result = self._run_imported_workflow(self.root / "helper-mutation-output")
        book = openpyxl.load_workbook(result.workbook_path, data_only=False)

        book["Inference Details"]["D2"] = 123.0
        with self.assertRaises(AssertionError):
            self._assert_inference_formula_contract(book)

    def test_cli_imports_both_modalities_and_records_the_exact_normalized_request(self) -> None:
        cli_output = self.root / "cli-output"
        overrides = {
            "Audio|Joy": 5.5,
            "Video": -1.25,
        }
        group_payload = [
            {"id": group.group_id, "name": group.name, "speaker_ids": list(group.speaker_ids)}
            for group in self.groups
        ]
        completed = subprocess.run(
            [
                sys.executable,
                "-m",
                "analysis.workflow",
                "--output-root",
                str(cli_output),
                "--imotions-source",
                str(self.video_root),
                "--imotions-method",
                "import",
                "--audio-source",
                str(self.audio_root),
                "--audio-method",
                "import",
                "--default-reference",
                "1.25",
                "--reference-overrides-json",
                json.dumps(overrides),
                "--speaker-groups-json",
                json.dumps(group_payload),
                "--no-graphs",
                "--logscale",
                "--include-landmarks",
                "--include-timing",
                "--exclude-geometry",
            ],
            cwd=Path(__file__).resolve().parents[2],
            check=False,
            capture_output=True,
            text=True,
            timeout=60,
        )
        self.assertEqual(
            completed.returncode,
            0,
            f"CLI workflow failed:\nSTDOUT:\n{completed.stdout}\nSTDERR:\n{completed.stderr}",
        )

        workbook_path = cli_output / "combined_analysis.xlsx"
        manifest_path = cli_output / "combined_analysis_manifest.json"
        self.assertTrue(workbook_path.is_file())
        self.assertTrue(manifest_path.is_file())
        self.assertEqual(
            sorted(path.name for path in cli_output.iterdir()),
            ["combined_analysis.xlsx", "combined_analysis_manifest.json"],
        )

        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        modality_options = {
            "write_graphs": False,
            "include_logscale": True,
            "include_landmarks": True,
            "include_timing": True,
            "exclude_geometry": True,
        }
        self.assertEqual(
            manifest["request"],
            {
                "output_root": str(cli_output.resolve()),
                "modalities": [
                    {
                        "name": "imotions",
                        "source_method": "import",
                        "source_path": str(self.video_root.resolve()),
                        **modality_options,
                    },
                    {
                        "name": "audio",
                        "source_method": "import",
                        "source_path": str(self.audio_root.resolve()),
                        **modality_options,
                    },
                ],
                "speaker_groups": group_payload,
                "write_combined_workbook": True,
                "include_construct_comparison": True,
                "include_probability_sheets": True,
                "confidence_level": 0.95,
                "headline_policy": "weighted",
                "default_reference": 1.25,
                "reference_overrides": overrides,
            },
        )
        expected_reports = [
            {
                "modality": modality,
                "normalized_speaker": speaker_id,
                "display_speaker": speaker,
                "path": str(
                    (
                        root
                        / "emotion"
                        / speaker
                        / "combined"
                        / "other_findings"
                        / "descriptive_statistics.csv"
                    ).resolve()
                ),
                "reason": "accepted speaker-level combined report",
            }
            for root, modality in ((self.video_root, "video"), (self.audio_root, "audio"))
            for speaker, speaker_id in (
                ("Andy Burnham", "andyburnham"),
                ("Marine Le Pen", "marinelepen"),
                ("Nigel Farage", "nigelfarage"),
            )
        ]
        self.assertEqual(manifest["accepted_reports"], expected_reports)
        accepted_paths = [entry["path"] for entry in manifest["accepted_reports"]]
        self.assertEqual(len(accepted_paths), len(set(accepted_paths)))
        self.assertEqual(manifest["rejected_reports"], [])
        self.assertEqual(manifest["warnings"], [])
        self.assertEqual(manifest["status"], "complete")
        self.assertEqual(manifest["schema_version"], 2)
        self.assertRegex(manifest["software"]["version"], r"^\d+\.\d+\.\d+$")
        self.assertIn("git_revision", manifest["software"])
        self.assertEqual(manifest["workbook_path"], str(workbook_path.resolve()))
        self.assertEqual(
            manifest["modality_roots"],
            {
                "video": str(self.video_root.resolve()),
                "audio": str(self.audio_root.resolve()),
            },
        )

    def test_imported_modalities_generate_auditable_grouped_probability_workbook(self) -> None:
        result = self._run_imported_workflow()

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["status"], "complete")
        self.assertEqual(manifest["workbook_path"], str(result.workbook_path))
        self.assertEqual(len(manifest["accepted_reports"]), 6)
        self.assertEqual(
            manifest["request"]["speaker_groups"],
            [
                {"id": "pair", "name": "Pair", "speaker_ids": ["Marine Le Pen", "Andy Burnham"]},
                {"id": "singleton", "name": "Singleton", "speaker_ids": ["Nigel Farage"]},
            ],
        )
        self.assertEqual(
            manifest["modality_roots"],
            {"video": str(self.video_root.resolve()), "audio": str(self.audio_root.resolve())},
        )

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        quantitative_sheets = ("Audio", "Video")
        expected_sheet_order = [
            "Audio",
            "Audio Prob",
            "Domain Def Text",
            "Video",
            "Video Prob",
            "Domain Def Speech",
            "Measure Guide",
            "Text sentiment",
            "Construct Comparison",
            "Probability Outline",
            "Inference Settings",
            "Inference Details",
            "Inference Inputs",
        ]
        self.assertEqual(book.sheetnames, expected_sheet_order)
        self.assertTrue(book.calculation.fullCalcOnLoad)
        self.assertTrue(book.calculation.forceFullCalc)
        self.assertEqual(book.calculation.calcMode, "auto")

        # Every quantitative sheet has one adjacent mirror; static/empty sheets do not.
        for source_name in quantitative_sheets:
            source_index = book.sheetnames.index(source_name)
            self.assertEqual(book.sheetnames[source_index + 1], f"{source_name} Prob")
        probability_sheets = [name for name in book.sheetnames if name.endswith(" Prob")]
        self.assertEqual(probability_sheets, [f"{name} Prob" for name in quantitative_sheets])
        self._assert_static_sheet_contract(book)

        details = book["Inference Details"]
        settings = book["Inference Settings"]
        inputs = book["Inference Inputs"]
        self.assertEqual(details.sheet_state, "visible")
        self.assertEqual(settings.sheet_state, "visible")
        self.assertEqual(inputs.sheet_state, "hidden")
        detail_rows = {
            details.cell(row, 1).value: row
            for row in range(2, details.max_row + 1)
        }

        for modality, metrics in (("Audio", AUDIO_METRICS), ("Video", VIDEO_METRICS)):
            source = book[modality]
            self.assertEqual(
                [source[coordinate].value for coordinate in ("B1", "D1", "E1", "G1", "H1", "S1")],
                ["Pair", "Marine Le Pen", "Andy Burnham", "Singleton", "Nigel Farage", "Overall"],
            )

            for metric_index, metric in enumerate(metrics, start=2):
                expected_marine = 10.0 if metric == "Anger" else 20.0 + metric_index - 2
                expected_andy = 10.0 + metric_index - 2
                expected_nigel = 30.0 + metric_index - 2
                self.assertEqual(source.cell(metric_index, 4).value, expected_marine)
                self.assertEqual(source.cell(metric_index, 5).value, expected_andy)
                self.assertEqual(source.cell(metric_index, 8).value, expected_nigel)

            marine_anger_row = detail_rows[f"{modality}|Anger|marinelepen"]
            overall_anger_row = detail_rows[f"{modality}|Anger|overall"]
            overall_joy_row = detail_rows[f"{modality}|Joy|overall"]
            self.assertIn("'Inference Inputs'!C", details[f"B{marine_anger_row}"].value)
            self.assertIn(f"E{marine_anger_row}=0", details[f"M{marine_anger_row}"].value)
            self.assertIn(f"'{modality}'!D2", details[f"B{overall_anger_row}"].value)
            self.assertIn(f"'{modality}'!H2", details[f"B{overall_anger_row}"].value)
            self.assertIn("T.DIST", details[f"M{overall_joy_row}"].value)

        self._assert_mirror_contract(book)
        self._assert_inference_formula_contract(book)
        self._assert_no_formula_errors_or_obvious_cycles(book)


if __name__ == "__main__":
    unittest.main()
