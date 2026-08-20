import csv
import math
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from openpyxl.styles import Font

from analysis.combined_summary import (
    VIDEO_METRICS,
    CombinedMetricCells,
    CombinedSource,
    SpeakerGroupDefinition,
    build_combined_workbook,
    protected_manual_discovery_directories,
)
from analysis.inference import (
    add_probability_mirrors,
    benjamini_hochberg,
    calculate_inference,
)


class InferenceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.workbook_path = self.root / "combined-summary.xlsx"
        self._write_workbook("Video")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _write_workbook(self, sheet_name: str) -> None:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = sheet_name
        for coordinate, value in {
            "A1": "Metric",
            "D1": "Speaker A",
            "E1": "Speaker B",
            "F1": "Speaker C",
            "S1": "Overall",
            "A2": "Anger",
            "D2": 10.0,
            "E2": 20.0,
            "F2": 30.0,
            "S2": "=AVERAGE(D2:F2)",
            "A3": "Joy",
            "D3": 5.0,
            "E3": 15.0,
            "F3": 25.0,
            "S3": "=AVERAGE(D3:F3)",
        }.items():
            sheet[coordinate] = value
        sheet["D2"].font = Font(name="Aptos", bold=True)
        sheet.row_dimensions[2].height = 22
        sheet.column_dimensions["D"].width = 17
        sheet.freeze_panes = "A2"
        sheet.merge_cells("A4:B4")
        sheet.print_title_rows = "1:1"
        book.save(self.workbook_path)

    def _write_video_report(self, speaker: str, base: float) -> Path:
        path = self.root / speaker / "descriptive_statistics.csv"
        path.parent.mkdir(parents=True, exist_ok=True)
        sources = ["001_First", "002_Second", "003_Third", "004_Fourth", "005_Fifth"]
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for index, metric in enumerate(VIDEO_METRICS):
                writer.writerow([metric])
                writer.writerow(["classification", "core", "category", "emotion", "unit", "score"])
                writer.writerow(["metric", *sources])
                writer.writerow(["count", 10, 10, 10, 10, 10])
                writer.writerow(["missing", 0, 0, 0, 0, 0])
                writer.writerow(["mean", *[base + index + offset for offset in range(5)]])
                writer.writerow(["stddev", 1, 1, 1, 1, 1])
                writer.writerow(["kurtosis", 0, 0, 0, 0, 0])
                writer.writerow([])
        return path

    @staticmethod
    def video_cells(sheet: str = "Video") -> dict[str, CombinedMetricCells]:
        return {
            "Video|Anger": CombinedMetricCells(
                sheet=sheet,
                metric="Anger",
                overall="S2",
                speaker_cells=("D2", "E2", "F2"),
                speaker_ids=("speaker_a", "speaker_b", "speaker_c"),
            )
        }

    @staticmethod
    def two_metric_cells() -> dict[str, CombinedMetricCells]:
        cells = InferenceTests.video_cells()
        cells["Video|Joy"] = CombinedMetricCells(
            sheet="Video",
            metric="Joy",
            overall="S3",
            speaker_cells=("D3", "E3", "F3"),
            speaker_ids=("speaker_a", "speaker_b", "speaker_c"),
        )
        return cells

    def test_probability_uses_speaker_means_as_observations(self) -> None:
        result = calculate_inference([10.0, 20.0, 30.0], reference=0.0)

        self.assertEqual(result.n, 3)
        self.assertAlmostEqual(result.mean, 20.0)
        self.assertAlmostEqual(result.standard_error, 10.0 / math.sqrt(3.0))
        self.assertGreater(result.probability_above, 0.95)

    def test_confidence_level_is_adjustable_and_95_percent_remains_default(self) -> None:
        default = calculate_inference([10.0, 20.0, 30.0], reference=0.0)
        ninety = calculate_inference(
            [10.0, 20.0, 30.0],
            reference=0.0,
            confidence_level=0.90,
        )

        self.assertLess(ninety.ci_high - ninety.ci_low, default.ci_high - default.ci_low)
        with self.assertRaises(ValueError):
            calculate_inference([10.0, 20.0], reference=0.0, confidence_level=1.0)

    def test_workbook_formulas_and_outline_record_selected_confidence_level(self) -> None:
        add_probability_mirrors(
            self.workbook_path,
            self.video_cells(),
            confidence_level=0.90,
        )

        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        self.assertIn("0.1", book["Inference Details"]["H2"].value)
        self.assertEqual(book["Inference Details"]["H1"].value, "90% CI low")
        self.assertEqual(book["Probability Outline"]["B7"].value, "90% two-sided Student t interval.")

    def test_benjamini_hochberg_returns_monotonic_values_in_input_order(self) -> None:
        self.assertEqual(benjamini_hochberg([0.01, 0.04, None, 0.03]), [0.03, 0.04, None, 0.04])

    def test_probability_mirror_replaces_only_overall_cell(self) -> None:
        result = add_probability_mirrors(self.workbook_path, self.video_cells())

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        mirror = book["Video Prob"]
        self.assertEqual(mirror["D2"].value, "='Video'!D2")
        self.assertEqual(mirror["S2"].value, "='Inference Details'!M2")
        self.assertIn("T.DIST", book["Inference Details"]["M2"].value)
        self.assertEqual(mirror["S2"].number_format, "0.00%")
        self.assertEqual(result.rows[0].n, 3)
        self.assertEqual(result.rows[0].excluded_speakers, ())

    def test_mirror_preserves_sheet_presentation_and_is_adjacent(self) -> None:
        add_probability_mirrors(self.workbook_path, self.video_cells())

        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        source = book["Video"]
        mirror = book["Video Prob"]
        self.assertEqual(book.sheetnames[:2], ["Video", "Video Prob"])
        self.assertEqual(mirror["D2"].font.name, source["D2"].font.name)
        self.assertEqual(mirror["D2"].font.bold, source["D2"].font.bold)
        self.assertEqual(mirror.row_dimensions[2].height, 22)
        self.assertEqual(mirror.column_dimensions["D"].width, 17)
        self.assertEqual(mirror.freeze_panes, "A2")
        self.assertIn("A4:B4", {str(item) for item in mirror.merged_cells.ranges})
        self.assertEqual(mirror.print_title_rows, "$1:$1")

    def test_inference_details_use_formula_links_to_speaker_cells(self) -> None:
        add_probability_mirrors(self.workbook_path, self.video_cells(), reference_overrides={"Video|Anger": 5.0})

        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        details = book["Inference Details"]
        settings = book["Inference Settings"]
        self.assertIn("'Video'!D2", details["B2"].value)
        self.assertEqual(details["C2"].value, "='Inference Settings'!D3")
        self.assertEqual(settings["C2"].value, "Default reference")
        self.assertEqual(settings["D3"].value, 5.0)
        self.assertTrue(book.calculation.fullCalcOnLoad)
        self.assertTrue(book.calculation.forceFullCalc)

    def test_modern_excel_functions_use_the_ooxml_xlfn_namespace(self) -> None:
        add_probability_mirrors(self.workbook_path, self.two_metric_cells())

        details = openpyxl.load_workbook(self.workbook_path, data_only=False)["Inference Details"]
        self.assertIn("_xlfn.T.INV.2T", details["H2"].value)
        self.assertIn("_xlfn.T.INV.2T", details["I2"].value)
        self.assertIn("_xlfn.T.DIST.2T", details["J2"].value)
        self.assertIn("_xlfn.MINIFS", details["K2"].value)
        self.assertIn("_xlfn.T.DIST", details["M2"].value)

    def test_live_formulas_ignore_blank_speakers_without_coercing_them_to_zero(self) -> None:
        book = openpyxl.load_workbook(self.workbook_path)
        book["Video"]["D2"] = 10.0
        book["Video"]["E2"] = None
        book["Video"]["F2"] = 30.0
        book.save(self.workbook_path)

        result = add_probability_mirrors(self.workbook_path, self.video_cells())

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        details = book["Inference Details"]
        self.assertEqual(details["B2"].value, "=COUNT('Video'!D2,'Video'!E2,'Video'!F2)")
        self.assertEqual(details["D2"].value, '=IFERROR(AVERAGE(\'Video\'!D2,\'Video\'!E2,\'Video\'!F2),"")')
        self.assertEqual(
            details["E2"].value,
            '=IF(B2<2,"",_xlfn.STDEV.S(\'Video\'!D2,\'Video\'!E2,\'Video\'!F2))',
        )
        self.assertNotIn("IF(ISNUMBER", details["E2"].value)
        self.assertNotIn("IFERROR('Video'!", details["D2"].value)
        self.assertEqual(result.rows[0].n, 2)
        self.assertEqual(result.rows[0].mean, 20.0)
        self.assertAlmostEqual(result.rows[0].standard_deviation, math.sqrt(200.0))
        self.assertEqual(result.rows[0].excluded_speakers, ("speaker_b",))

    def test_live_formulas_ignore_text_speakers_without_coercing_them_to_zero(self) -> None:
        book = openpyxl.load_workbook(self.workbook_path)
        book["Video"]["D2"] = 10.0
        book["Video"]["E2"] = "not available"
        book["Video"]["F2"] = 30.0
        book.save(self.workbook_path)

        result = add_probability_mirrors(self.workbook_path, self.video_cells())

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        details = book["Inference Details"]
        self.assertEqual(
            details["E2"].value,
            '=IF(B2<2,"",_xlfn.STDEV.S(\'Video\'!D2,\'Video\'!E2,\'Video\'!F2))',
        )
        self.assertEqual(result.rows[0].n, 2)
        self.assertEqual(result.rows[0].mean, 20.0)
        self.assertAlmostEqual(result.rows[0].standard_deviation, math.sqrt(200.0))
        self.assertEqual(result.rows[0].excluded_speakers, ("speaker_b",))

    def test_formula_empty_speaker_cell_is_excluded_by_python_and_live_excel_formula(self) -> None:
        book = openpyxl.load_workbook(self.workbook_path)
        book["Video"]["D2"] = 10.0
        book["Video"]["E2"] = '=""'
        book["Video"]["F2"] = 30.0
        book.save(self.workbook_path)

        result = add_probability_mirrors(self.workbook_path, self.video_cells())

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        details = book["Inference Details"]
        self.assertEqual(
            details["E2"].value,
            '=IF(B2<2,"",_xlfn.STDEV.S(\'Video\'!D2,\'Video\'!E2,\'Video\'!F2))',
        )
        self.assertEqual(result.rows[0].n, 2)
        self.assertEqual(result.rows[0].mean, 20.0)
        self.assertAlmostEqual(result.rows[0].standard_deviation, math.sqrt(200.0))
        self.assertEqual(result.rows[0].excluded_speakers, ("speaker_b",))

    def test_python_inference_ignores_numeric_text_just_like_live_excel_formulas(self) -> None:
        book = openpyxl.load_workbook(self.workbook_path)
        book["Video"]["D2"] = 10.0
        book["Video"]["E2"] = "20"
        book["Video"]["F2"] = 30.0
        book.save(self.workbook_path)

        result = add_probability_mirrors(self.workbook_path, self.video_cells())

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(
            book["Inference Details"]["E2"].value,
            '=IF(B2<2,"",_xlfn.STDEV.S(\'Video\'!D2,\'Video\'!E2,\'Video\'!F2))',
        )
        self.assertEqual(result.rows[0].n, 2)
        self.assertEqual(result.rows[0].mean, 20.0)
        self.assertAlmostEqual(result.rows[0].standard_deviation, math.sqrt(200.0))
        self.assertEqual(result.rows[0].excluded_speakers, ("speaker_b",))

    def test_q_values_are_live_bh_formulas_over_the_bounded_p_value_family(self) -> None:
        add_probability_mirrors(self.workbook_path, self.two_metric_cells())

        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        details = book["Inference Details"]
        for coordinate in ("K2", "K3"):
            self.assertIsInstance(details[coordinate].value, str)
            self.assertTrue(details[coordinate].value.startswith("="))
            self.assertIn("$J$2:$J$3", details[coordinate].value)
            self.assertIn("_xlfn.MINIFS($O$2:$O$3", details[coordinate].value)
        self.assertEqual(details["O1"].value, "BH adjusted p-value")
        self.assertIn("$J$2:$J$3", details["O2"].value)
        self.assertIn("COUNTIF($J$2:$J$3,\"<=\"&J2)", details["O2"].value)

    def test_sheet_override_applies_when_the_metric_has_no_override(self) -> None:
        result = add_probability_mirrors(self.workbook_path, self.video_cells(), reference_overrides={"Video": 7.0})

        self.assertEqual(result.rows[0].reference, 7.0)
        self.assertEqual(result.reference_resolutions[0].original_key, "Video")
        self.assertEqual(result.reference_resolutions[0].matched_scope, "sheet")
        self.assertEqual(result.reference_resolutions[0].matched_source, "Video")
        self.assertEqual(result.reference_resolutions[0].resolved_reference, 7.0)

    def test_metric_override_is_audited_and_unknown_keys_are_rejected(self) -> None:
        result = add_probability_mirrors(
            self.workbook_path,
            self.video_cells(),
            reference_overrides={"Video|Anger": 5.0},
        )

        resolution = result.reference_resolutions[0]
        self.assertEqual(
            (
                resolution.original_key,
                resolution.matched_scope,
                resolution.matched_source,
                resolution.resolved_reference,
            ),
            ("Video|Anger", "metric", "Video|Anger", 5.0),
        )
        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        settings = book["Inference Settings"]
        self.assertEqual(
            [settings.cell(1, column).value for column in range(1, 5)],
            ["Original key", "Matched scope", "Matched source", "Resolved reference"],
        )
        self.assertEqual(
            [settings.cell(3, column).value for column in range(1, 5)],
            ["Video|Anger", "metric", "Video|Anger", 5.0],
        )

        self._write_workbook("Video")
        with self.assertRaisesRegex(ValueError, "Unknown reference override.*Video:anger"):
            add_probability_mirrors(
                self.workbook_path,
                self.video_cells(),
                reference_overrides={"Video:anger": 5.0},
            )

    def test_protected_reference_workbooks_are_rejected_before_opening(self) -> None:
        reference_root = self.root / "Statistics_Manual_Discovery"
        reference_root.mkdir()
        checkout_anchor = self.root / "checkout" / "analysis" / "test_anchor.py"
        checkout_anchor.parent.mkdir(parents=True)
        discovered = protected_manual_discovery_directories(checkout_anchor)
        self.assertEqual(discovered, (reference_root.resolve(),))
        reference_path = reference_root / "reference.xlsx"

        with patch(
            "analysis.inference.protected_manual_discovery_directories",
            return_value=discovered,
        ):
            with self.assertRaisesRegex(ValueError, "protected reference"):
                add_probability_mirrors(reference_path, self.video_cells())

    def test_empty_singleton_and_zero_variance_rules_are_explicit(self) -> None:
        empty = calculate_inference([], reference=0.0)
        singleton = calculate_inference([10.0], reference=0.0)
        above = calculate_inference([10.0, 10.0], reference=0.0)
        below = calculate_inference([-10.0, -10.0], reference=0.0)
        equal = calculate_inference([0.0, 0.0], reference=0.0)

        self.assertEqual((empty.n, empty.mean, empty.probability_above), (0, None, None))
        self.assertEqual((singleton.n, singleton.standard_error, singleton.probability_above), (1, None, None))
        for row in (above, below, equal):
            self.assertEqual(row.standard_deviation, 0.0)
            self.assertIsNone(row.standard_error)
            self.assertIsNone(row.ci_low)
            self.assertIsNone(row.ci_high)
            self.assertIsNone(row.p_value)
            self.assertIsNone(row.q_value)
            self.assertIsNone(row.effect_size)
            self.assertIsNone(row.probability_above)

    def test_zero_variance_workbook_formulas_leave_inference_blank(self) -> None:
        book = openpyxl.load_workbook(self.workbook_path)
        for coordinate in ("D2", "E2", "F2"):
            book["Video"][coordinate] = 10.0
        book.save(self.workbook_path)

        add_probability_mirrors(self.workbook_path, self.video_cells())

        details = openpyxl.load_workbook(self.workbook_path, data_only=False)["Inference Details"]
        for coordinate in ("F2", "H2", "I2", "J2", "L2", "M2"):
            self.assertIn("E2=0", details[coordinate].value)
        self.assertIn('F2=""', details["G2"].value)
        self.assertIn('J2=""', details["K2"].value)
        self.assertIn('J2=""', details["O2"].value)

    def test_non_finite_observations_are_excluded(self) -> None:
        result = calculate_inference([10.0, math.nan, math.inf], reference=0.0)

        self.assertEqual(result.n, 1)
        self.assertEqual(result.mean, 10.0)

    def test_mirror_name_is_limited_and_disambiguated(self) -> None:
        long_name = "ABCDEFGHIJKLMNOPQRSTUVWXYZ12345"
        self._write_workbook(long_name)

        add_probability_mirrors(self.workbook_path, self.video_cells(long_name))

        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        mirror_name = book.sheetnames[1]
        self.assertLessEqual(len(mirror_name), 31)
        self.assertTrue(mirror_name.endswith(" Prob"))

        self._write_workbook("Video")
        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        book.create_sheet("Video Prob")
        book.save(self.workbook_path)
        add_probability_mirrors(self.workbook_path, self.video_cells())
        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        self.assertIn("Video Prob 2", book.sheetnames)

    def test_missing_sheet_is_rejected(self) -> None:
        cells = self.video_cells("Missing")

        with self.assertRaisesRegex(ValueError, "Missing"):
            add_probability_mirrors(self.workbook_path, cells)

    def test_formula_quotes_apostrophes_in_sheet_names(self) -> None:
        self._write_workbook("O'Brien")

        add_probability_mirrors(self.workbook_path, self.video_cells("O'Brien"))

        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        mirror_name = book.sheetnames[1]
        self.assertEqual(book[mirror_name]["D2"].value, "='O''Brien'!D2")
        self.assertIn("'O''Brien'!D2", book["Inference Details"]["B2"].value)

    def test_group_workbook_gets_one_probability_mirror_with_speaker_and_overall_probabilities(self) -> None:
        andy = self._write_video_report("andy", 10)
        marine = self._write_video_report("marine", 20)
        result = build_combined_workbook(
            {
                "video": (
                    CombinedSource("video", "andy", "Andy Burnham", andy),
                    CombinedSource("video", "marine", "Marine Le Pen", marine),
                )
            },
            self.workbook_path,
            speaker_groups=(
                SpeakerGroupDefinition("uk", "United Kingdom", ("Andy Burnham",)),
                SpeakerGroupDefinition("fr", "France", ("Marine Le Pen",)),
            ),
        )

        inference = add_probability_mirrors(result.workbook_path, result.source_cells)

        book = openpyxl.load_workbook(result.workbook_path, data_only=False)
        self.assertEqual(
            inference.probability_sheets,
            ("Video Prob",),
        )
        self.assertEqual(
            book.sheetnames,
            [
                "Domain Def Text",
                "Video",
                "Video Prob",
                "Domain Def Speech",
                "Measure Guide",
                "Text sentiment",
                "Probability Outline",
                "Inference Settings",
                "Inference Details",
                "Inference Inputs",
            ],
        )
        details = book["Inference Details"]
        detail_rows = {
            details[f"A{row}"].value: row
            for row in range(2, details.max_row + 1)
        }
        uk_speaker_row = detail_rows["Video|Anger|andyburnham"]
        overall_row = detail_rows["Video|Anger|overall"]
        france_speaker_row = detail_rows["Video|Anger|marinelepen"]
        self.assertEqual(
            book["Video Prob"]["D2"].value,
            f"='Inference Details'!M{uk_speaker_row}",
        )
        self.assertEqual(
            book["Video Prob"]["S2"].value,
            f"='Inference Details'!M{overall_row}",
        )
        self.assertEqual(
            book["Video Prob"]["H2"].value,
            f"='Inference Details'!M{france_speaker_row}",
        )
        self.assertEqual(book["Inference Inputs"].sheet_state, "hidden")
        self.assertIn("'Inference Inputs'!", details[f"B{uk_speaker_row}"].value)
        outline = book["Probability Outline"]
        outline_text = " ".join(
            str(cell.value)
            for row in outline.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertIn("per-source means", outline_text)
        self.assertIn("speaker means", outline_text)
        self.assertIn("Default reference", outline_text)
        self.assertIn("bounded nonnegative", outline_text)
        self.assertIn("United Kingdom", outline_text)
        self.assertIn("Andy Burnham", outline_text)
        self.assertIn("France", outline_text)
        self.assertIn("Marine Le Pen", outline_text)
        self.assertIn(f"$J$2:$J${details.max_row}", details["K2"].value)

    def test_probability_outline_neutralizes_group_metadata_after_second_save(self) -> None:
        andy = self._write_video_report("andy", 10)
        result = build_combined_workbook(
            {"video": (CombinedSource("video", "andy", "Andy Burnham", andy),)},
            self.workbook_path,
            speaker_groups=(
                SpeakerGroupDefinition("=group-id", "@cohort", ("Andy Burnham",)),
            ),
        )

        add_probability_mirrors(result.workbook_path, result.source_cells)

        outline = openpyxl.load_workbook(
            result.workbook_path, data_only=False
        )["Probability Outline"]
        group_row = next(
            row
            for row in range(1, outline.max_row + 1)
            if outline.cell(row, 2).value in {"=group-id", "'=group-id"}
        )
        self.assertEqual(outline.cell(group_row, 2).value, "'=group-id")
        self.assertEqual(outline.cell(group_row, 3).value, "'@cohort")
        self.assertEqual(outline.cell(group_row, 2).data_type, "s")
        self.assertEqual(outline.cell(group_row, 3).data_type, "s")

    def test_interleaved_source_cells_are_grouped_by_sheet_in_details_and_fdr_formulas(self) -> None:
        book = openpyxl.load_workbook(self.workbook_path)
        audio = book.create_sheet("Audio")
        for coordinate, value in {
            "D2": 10.0,
            "E2": 20.0,
            "F2": 30.0,
            "D3": 5.0,
            "E3": 15.0,
            "F3": 25.0,
        }.items():
            audio[coordinate] = value
        book.save(self.workbook_path)
        audio_anger = CombinedMetricCells("Audio", "Anger", "S2", ("D2", "E2", "F2"), ("a", "b", "c"))
        audio_joy = CombinedMetricCells("Audio", "Joy", "S3", ("D3", "E3", "F3"), ("a", "b", "c"))
        video_anger = self.video_cells()["Video|Anger"]
        video_joy = self.two_metric_cells()["Video|Joy"]
        interleaved = {
            "Audio|Anger": audio_anger,
            "Video|Anger": video_anger,
            "Audio|Joy": audio_joy,
            "Video|Joy": video_joy,
        }

        add_probability_mirrors(self.workbook_path, interleaved)

        book = openpyxl.load_workbook(self.workbook_path, data_only=False)
        details = book["Inference Details"]
        self.assertEqual(
            [details[f"A{row}"].value for row in range(2, 6)],
            ["Audio|Anger", "Audio|Joy", "Video|Anger", "Video|Joy"],
        )
        for coordinate in ("K2", "K3", "O2", "O3"):
            self.assertIn("$J$2:$J$3", details[coordinate].value)
        for coordinate in ("K4", "K5", "O4", "O5"):
            self.assertIn("$J$4:$J$5", details[coordinate].value)


if __name__ == "__main__":
    unittest.main()
