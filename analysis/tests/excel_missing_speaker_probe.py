"""Recalculate missing-speaker formulas in desktop Excel.

This probe is intentionally separate from pytest because Excel automation is
Windows-only and is not available in every development or CI environment.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
if str(REPOSITORY_ROOT) not in sys.path:
    sys.path.insert(0, str(REPOSITORY_ROOT))

from analysis.combined_summary import CombinedMetricCells  # noqa: E402
from analysis.inference import add_probability_mirrors  # noqa: E402


PROBE_CASES = (
    ("Blank", None),
    ("Formula empty", '=""'),
    ("Text", "not available"),
    ("Numeric-looking text", "20"),
)
EXPECTED_COUNT = 2
EXPECTED_MEAN = 20.0
EXPECTED_STANDARD_DEVIATION = math.sqrt(200.0)


def _build_probe_workbook(path: Path) -> None:
    """Create representative source cells and add production inference formulas."""

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Video"
    sheet.append(("Metric", "", "", "Speaker A", "Speaker B", "Speaker C"))

    source_cells: dict[str, CombinedMetricCells] = {}
    for row, (case_name, middle_value) in enumerate(PROBE_CASES, start=2):
        sheet.cell(row, 1, case_name)
        sheet.cell(row, 4, 10.0)
        sheet.cell(row, 5, middle_value)
        sheet.cell(row, 6, 30.0)
        sheet.cell(row, 19, f"=AVERAGE(D{row},E{row},F{row})")
        metric_key = f"Video|{case_name}"
        source_cells[metric_key] = CombinedMetricCells(
            sheet="Video",
            metric=case_name,
            overall=f"S{row}",
            speaker_cells=(f"D{row}", f"E{row}", f"F{row}"),
            speaker_ids=("speaker_a", "speaker_b", "speaker_c"),
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    add_probability_mirrors(path, source_cells)


def _recalculate_with_excel(path: Path) -> None:
    """Open, fully recalculate, and save the workbook through Excel COM."""

    powershell = r"""
$ErrorActionPreference = 'Stop'
$excel = $null
$workbook = $null
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($env:MEAP_EXCEL_PROBE_PATH)
    $excel.CalculateFullRebuild()
    $workbook.Save()
}
finally {
    if ($null -ne $workbook) {
        $workbook.Close($false)
        [void][Runtime.InteropServices.Marshal]::ReleaseComObject($workbook)
    }
    if ($null -ne $excel) {
        $excel.Quit()
        [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""
    environment = os.environ.copy()
    environment["MEAP_EXCEL_PROBE_PATH"] = str(path.resolve())
    completed = subprocess.run(
        ["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", powershell],
        check=False,
        capture_output=True,
        text=True,
        env=environment,
    )
    if completed.returncode != 0:
        message = completed.stderr.strip().splitlines()
        detail = message[-1] if message else "desktop Excel automation failed"
        raise RuntimeError(detail)


def _read_and_validate_results(path: Path) -> list[dict[str, float | int | str]]:
    """Read Excel's cached values and enforce the blank/text exclusion contract."""

    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    details = book["Inference Details"]
    results: list[dict[str, float | int | str]] = []
    try:
        for row, (case_name, _) in enumerate(PROBE_CASES, start=2):
            count = details.cell(row, 2).value
            mean = details.cell(row, 4).value
            standard_deviation = details.cell(row, 5).value
            if count != EXPECTED_COUNT:
                raise AssertionError(f"{case_name}: expected n=2, received {count!r}")
            if not isinstance(mean, (int, float)) or not math.isclose(
                float(mean), EXPECTED_MEAN, rel_tol=0.0, abs_tol=1e-12
            ):
                raise AssertionError(f"{case_name}: expected mean=20, received {mean!r}")
            if not isinstance(standard_deviation, (int, float)) or not math.isclose(
                float(standard_deviation),
                EXPECTED_STANDARD_DEVIATION,
                rel_tol=0.0,
                abs_tol=1e-12,
            ):
                raise AssertionError(
                    f"{case_name}: expected SD=sqrt(200), received {standard_deviation!r}"
                )
            results.append(
                {
                    "case": case_name,
                    "count": int(count),
                    "mean": float(mean),
                    "sample_standard_deviation": float(standard_deviation),
                }
            )
    finally:
        book.close()
    return results


def run_probe(output_path: Path) -> list[dict[str, float | int | str]]:
    """Build, recalculate, and validate one probe workbook."""

    _build_probe_workbook(output_path)
    _recalculate_with_excel(output_path)
    return _read_and_validate_results(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional path at which to retain the recalculated probe workbook.",
    )
    args = parser.parse_args()

    if args.output is not None:
        results = run_probe(args.output.expanduser().resolve())
    else:
        with tempfile.TemporaryDirectory(prefix="meap-excel-probe-") as directory:
            results = run_probe(Path(directory) / "missing-speaker-probe.xlsx")

    print(json.dumps({"status": "passed", "cases": results}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
