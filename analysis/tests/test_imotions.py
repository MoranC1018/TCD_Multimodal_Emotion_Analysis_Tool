import csv
import re
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from analysis.imotions import (
    ImotionsRowSequence,
    analyse_imotions_folder,
    default_imotions_root,
    default_output_root,
    read_imotions_csv,
    resolve_input_folder,
    resolve_output_folder,
)
from analysis.histograms import (
    HistogramTable,
    ParsedExport,
    analyse_grouped_parsed_exports,
    clean_legacy_flat_outputs,
    ensure_disjoint_paths,
    read_output_owner,
    resolve_typed_output_folder,
    write_descriptive_statistics_csv,
    write_dict_rows,
    write_output_owner,
    write_sectioned_csv,
    write_statistical_csvs,
    write_xlsx_workbook,
)


LONG_DECIMAL_PATTERN = re.compile(r"(?<![\w.])-?\d+\.\d{3,}(?![\w.])")


HEADER = [
    "Row",
    "Timestamp",
    "Anger",
    "Valence",
    "Brow Furrow",
    "Eye Closure",
    "Mouth Open",
    "Smile",
    "Pitch",
]


def write_imotions_csv(path: Path, rows: list[list[object]]) -> None:
    """Write a compact iMotions-style export for output-layout tests."""

    metadata_rows = [
        ["#INFO"],
        ["#Study name", "Unit test"],
        ["#METADATA"],
        ["#Category", "Timestamp", "FEA(Emotions)", "FEA(Emotions)", "FEA(Action units)", "FEA(Action units)", "FEA(Action units)", "FEA(Action units)", "FEA(Head rotation)"],
        ["#Description", "Timestamp", "Anger evidence", "Valence score", "Brow movement", "Eye closure", "Mouth opening", "Smile evidence", "Head pitch"],
        ["#Unit", "Millisecond", "Index", "Index", "Index", "Index", "Index", "Index", "Degrees"],
        ["#Group", "", "Emotion", "Emotion", "Facial Expression", "Facial Expression", "Facial Expression", "Facial Expression", "Behavioral"],
        ["#Display name", "", "Anger", "Valence", "Brow Furrow", "Eye Closure", "Mouth Open", "Smile", "Pitch"],
        ["#Channel identifier", "Timestamp", "FEA_Emotion_Anger", "FEA_Emotion_Valence", "FEA_Action_Unit_Brow_Furrow", "FEA_Action_Unit_Eye_Closure", "FEA_Action_Unit_Mouth_Open", "FEA_Action_Unit_Smile", "FEA_Head_Rotation_Pitch"],
        ["#DATA"],
        HEADER,
    ]

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(metadata_rows)
        writer.writerows(rows)
        writer.writerow([])
        writer.writerow(["Summary rows after the blank line must not be counted"])
        writer.writerow([999, 999, 99, 99, 99, 99, 99, 99, 99])


class IMotionsAnalysisTests(unittest.TestCase):
    def test_histogram_csv_and_xlsx_neutralize_hostile_source_labels(self) -> None:
        hostile = ["=1+1", "+cmd", "-minus", "@speaker", "\ttab", "\rcarriage", "ordinary"]
        table = HistogramTable(
            classification="core",
            statistic="Anger",
            sources=hostile,
            bins=["0"],
            rows={"0": [1] * len(hostile)},
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            csv_path = root / "histograms.csv"
            xlsx_path = root / "histograms.xlsx"
            write_sectioned_csv(csv_path, {"Emotions": (table,)})
            self.assertTrue(write_xlsx_workbook(xlsx_path, {"Emotions": (table,)}))
            with csv_path.open(encoding="utf-8", newline="") as handle:
                csv_rows = list(csv.reader(handle))
            book = openpyxl.load_workbook(xlsx_path, data_only=False)

        expected = ["'=1+1", "'+cmd", "'-minus", "'@speaker", "'\ttab", "'\rcarriage", "ordinary"]
        self.assertEqual(csv_rows[2][2:-1], expected)
        header_cells = [book["Emotions"].cell(4, column) for column in range(3, 3 + len(expected))]
        self.assertEqual([cell.value for cell in header_cells], expected)
        self.assertTrue(all(cell.data_type == "s" for cell in header_cells))
        self.assertEqual(book["Emotions"].cell(5, 3).data_type, "n")

    def test_supplementary_csv_writers_neutralize_hostile_labels(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            descriptive_path = root / "descriptive.csv"
            write_descriptive_statistics_csv(
                descriptive_path,
                ["=source"],
                [
                    {
                        "statistic": "@metric",
                        "source": "=source",
                        "classification": "+class",
                        "category": "-category",
                        "unit": "ordinary",
                        "mean": -0.42,
                    }
                ],
            )
            dict_path = root / "manifest.csv"
            write_dict_rows(
                dict_path,
                ({"=hostile-header": "ordinary", "label": "=formula", "measurement": -0.42},),
            )
            table = HistogramTable(
                classification="emotion_0_to_100",
                statistic="@metric",
                sources=["+source-a", "-source-b"],
                bins=["0", "5", "10"],
                rows={"0": [5, 2], "5": [1, 4], "10": [3, 1]},
            )
            write_statistical_csvs(
                root,
                table.sources,
                {"emotion_0_to_100": {"metric": table}},
            )

            with descriptive_path.open(encoding="utf-8", newline="") as handle:
                descriptive_rows = list(csv.reader(handle))
            with dict_path.open(encoding="utf-8", newline="") as handle:
                manifest_reader = csv.DictReader(handle)
                manifest_row = next(manifest_reader)
                manifest_headers = manifest_reader.fieldnames
            with (root / "chi_squared_results.csv").open(
                encoding="utf-8", newline=""
            ) as handle:
                chi_rows = list(csv.reader(handle))
            with (root / "spearman_results.csv").open(
                encoding="utf-8", newline=""
            ) as handle:
                spearman_rows = list(csv.reader(handle))

        self.assertEqual(descriptive_rows[0][0], "'@metric")
        self.assertEqual(descriptive_rows[1][1], "'+class")
        self.assertEqual(descriptive_rows[2][1], "'=source")
        self.assertEqual(manifest_headers[0], "'=hostile-header")
        self.assertEqual(manifest_row["label"], "'=formula")
        self.assertEqual(float(manifest_row["measurement"]), -0.42)
        self.assertTrue(any("'+source-a" in row for row in chi_rows))
        self.assertTrue(any("'-source-b" in row for row in spearman_rows))

    def test_interrupted_report_claims_output_before_writing_partial_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "Speaker"
            input_dir.mkdir()
            source_path = input_dir / "001.csv"
            source_path.touch()
            output_root = root / "output"
            output_dir = resolve_typed_output_folder(input_dir, output_root, "emotion")
            export = ParsedExport(
                source="001",
                path=source_path,
                header=["Anger"],
                info={},
                rows=[{"Anger": "1"}],
                speaker="Speaker",
            )

            def interrupt_after_partial_output(**kwargs):
                partial_dir = kwargs["output_dir"]
                partial_dir.mkdir(parents=True, exist_ok=True)
                (partial_dir / "partial.csv").write_text("partial", encoding="utf-8")
                raise RuntimeError("interrupted")

            with patch(
                "analysis.histograms.analyse_parsed_exports",
                side_effect=interrupt_after_partial_output,
            ):
                with self.assertRaisesRegex(RuntimeError, "interrupted"):
                    analyse_grouped_parsed_exports(
                        input_dir=input_dir,
                        output_dir=output_dir,
                        exports=[export],
                        discovery_log=[],
                        write_graphs=False,
                    )

            self.assertEqual(read_output_owner(output_dir)["source"], str(input_dir.resolve()))
            self.assertEqual(resolve_typed_output_folder(input_dir, output_root, "emotion"), output_dir)
            self.assertTrue((output_dir / "001" / "partial.csv").is_file())

            clean_legacy_flat_outputs(output_dir, input_dir)

            self.assertFalse((output_dir / "001").exists())

    def test_imotions_rows_are_repeatable_and_disk_backed(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "rows.csv"
            write_imotions_csv(path, [[1, 0, 10, -25, 5, 5, 5, 5, 0]])

            export = read_imotions_csv(path)

            self.assertIsInstance(export.rows, ImotionsRowSequence)
            self.assertEqual(list(export.rows), list(export.rows))

    def test_blank_data_lines_do_not_truncate_rows_but_footer_summaries_do(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "blank-lines.csv"
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["#DATA"])
                writer.writerow(HEADER)
                writer.writerow([1, 0, 10, -25, 5, 5, 5, 5, 0])
                writer.writerow([])
                writer.writerow([2, 40, 20, -10, 5, 5, 5, 5, 0])
                writer.writerow(["Summary"])
                writer.writerow([999, 999, 99, 99, 99, 99, 99, 99, 99])

            export = read_imotions_csv(path)

            self.assertEqual(len(export.rows), 2)
            self.assertEqual([row["Row"] for row in export.rows], ["1", "2"])

    def test_header_only_imotions_export_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "empty.csv"
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["#DATA"])
                writer.writerow(HEADER)

            with self.assertRaisesRegex(ValueError, "No iMotions data rows"):
                read_imotions_csv(path)

    def test_same_named_sources_receive_distinct_owned_output_folders(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "first" / "Speaker"
            second = root / "second" / "Speaker"
            output = root / "output"
            first.mkdir(parents=True)
            second.mkdir(parents=True)
            first_output = resolve_typed_output_folder(first, output, "emotion")
            first_output.mkdir(parents=True)
            write_output_owner(first_output, first, {"combined"})

            second_output = resolve_typed_output_folder(second, output, "emotion")

            self.assertNotEqual(first_output, second_output)
            self.assertTrue(second_output.name.startswith("Speaker_"))

    def test_owned_cleanup_removes_only_declared_generated_directories(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source"
            output = root / "output"
            source.mkdir()
            generated = output / "generated"
            manual = output / "manual_notes"
            generated.mkdir(parents=True)
            manual.mkdir()
            (manual / "keep.txt").write_text("keep", encoding="utf-8")
            write_output_owner(output, source, {"generated"})

            clean_legacy_flat_outputs(output, source)

            self.assertFalse(generated.exists())
            self.assertEqual((manual / "keep.txt").read_text(encoding="utf-8"), "keep")

    def test_owned_cleanup_rejects_parent_directory_entry(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source"
            output = root / "output"
            source.mkdir()
            output.mkdir()
            write_output_owner(output, source, {".."})

            with patch("analysis.histograms.remove_generated_directory") as remove_directory:
                clean_legacy_flat_outputs(output, source)

            remove_directory.assert_not_called()
            self.assertTrue(source.is_dir())
            self.assertTrue(output.is_dir())

    def test_output_owner_rejects_oversized_control_json(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir)
            owner = output / ".meap_output_owner.json"
            owner.write_text('{"padding":"' + ("x" * (64 * 1024)) + '"}', encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "output owner JSON exceeds"):
                read_output_owner(output)

    def test_owned_cleanup_rejects_current_directory_entry(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source"
            output = root / "output"
            source.mkdir()
            output.mkdir()
            write_output_owner(output, source, {"."})

            with patch("analysis.histograms.remove_generated_directory") as remove_directory:
                clean_legacy_flat_outputs(output, source)

            remove_directory.assert_not_called()
            self.assertTrue(output.is_dir())

    def test_owned_cleanup_rejects_linked_directory_entry(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source"
            output = root / "output"
            outside = root / "outside"
            source.mkdir()
            output.mkdir()
            outside.mkdir()
            linked = output / "linked"
            try:
                linked.symlink_to(outside, target_is_directory=True)
            except OSError as exc:
                self.skipTest(f"Directory links are unavailable: {exc}")
            write_output_owner(output, source, {"linked"})

            with patch("analysis.histograms.remove_generated_directory") as remove_directory:
                clean_legacy_flat_outputs(output, source)

            remove_directory.assert_not_called()
            self.assertTrue(outside.is_dir())

    def test_analysis_input_and_output_must_not_overlap(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source"
            source.mkdir()

            with self.assertRaisesRegex(ValueError, "must not overlap"):
                ensure_disjoint_paths(source, source / "reports")

    def test_outputs_are_comparison_first_and_sectioned(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "Speaker_A"
            output_root = root / "post"
            input_dir.mkdir()

            write_imotions_csv(
                input_dir / "001_First.csv",
                [
                    [1, 0, 0, -100, 10, 20, 30, 40, -5],
                    [2, 40, 4.9, -99, 20, 30, 40, 50, -2],
                    [3, 80, 5, -1, 30, 40, 50, 60, 0],
                    [4, 120, 99.9, 0, 40, 50, 60, 70, 3],
                    [5, 160, 100, 100, 50, 60, 70, 80, 7],
                ],
            )
            write_imotions_csv(
                input_dir / "002_Second.csv",
                [
                    [1, 0, 10, -50, 50, 40, 30, 20, 6],
                    [2, 40, 15, -25, 40, 30, 20, 10, 4],
                    [3, 80, 20, 0, 30, 20, 10, 0, 2],
                    [4, 120, 25, 25, 20, 10, 0, 10, 0],
                    [5, 160, 30, 50, 10, 0, 10, 20, -2],
                ],
            )

            result = analyse_imotions_folder(input_dir, output_root=output_root, write_graphs=True)

            raw_output_dir = result.domain_output_dirs["raw"]
            histogram_csv = result.output_dir / "combined" / "histograms.csv"
            self.assertTrue(histogram_csv.exists())
            histogram_text = histogram_csv.read_text(encoding="utf-8")
            self.assertIn("Core emotions (0-100)", histogram_text)
            self.assertIn("Valence (-100 to 100)", histogram_text)
            self.assertIn("Anger\nbin_start,bin_end,001_First,002_Second,total", histogram_text)
            self.assertNotIn("Brow Furrow\nbin_start", histogram_text)
            self.assertNotIn("emotion_histograms_0_to_100.csv", [path.name for path in result.histogram_paths])
            self.assertNotIn("Pitch\nbin_start", histogram_text)

            self.assertTrue((result.output_dir / "001_First" / "histograms.csv").exists())
            self.assertTrue((result.output_dir / "002_Second" / "histograms.csv").exists())
            self.assertTrue((raw_output_dir / "001_First" / "histograms.csv").exists())
            self.assertTrue((raw_output_dir / "002_Second" / "histograms.csv").exists())
            raw_histogram_text = (raw_output_dir / "combined" / "histograms.csv").read_text(encoding="utf-8")
            self.assertIn("Other 0-100 findings", raw_histogram_text)
            self.assertIn("Brow Furrow\nbin_start,bin_end,001_First,002_Second,total", raw_histogram_text)
            self.assertNotIn("Anger\nbin_start", raw_histogram_text)

            descriptor_text = (result.output_dir / "combined" / "other_findings" / "descriptive_statistics.csv").read_text(encoding="utf-8")
            self.assertIn("Anger", descriptor_text)
            self.assertIn("metric,001_First,002_Second", descriptor_text)
            self.assertIn("mean,", descriptor_text)
            self.assertIn("kurtosis,", descriptor_text)
            self.assert_no_long_decimal_values(descriptor_text)

            raw_descriptor_text = (raw_output_dir / "combined" / "other_findings" / "descriptive_statistics.csv").read_text(encoding="utf-8")
            self.assertIn("Brow Furrow", raw_descriptor_text)
            self.assertIn("Pitch", raw_descriptor_text)
            self.assertNotIn("Anger", raw_descriptor_text)

            chi_text = (result.output_dir / "combined" / "chi_squared_results.csv").read_text(encoding="utf-8")
            self.assertIn("Pearson chi-square: Anger", chi_text)
            self.assertIn("Pairwise X-squared matrix", chi_text)
            self.assertIn("source,001_First,002_Second", chi_text)
            self.assertIn("Observed vs expected", chi_text)
            self.assert_no_long_decimal_values(chi_text)

            spearman_text = (result.output_dir / "combined" / "spearman_results.csv").read_text(encoding="utf-8")
            self.assertIn("Spearman rank correlation: Anger", spearman_text)
            self.assertIn("Spearman rho matrix", spearman_text)
            self.assert_no_long_decimal_values(spearman_text)

            region_text = (raw_output_dir / "combined" / "other_findings" / "facial_region_correlations.csv").read_text(encoding="utf-8")
            self.assertIn("region_pair,metric,001_First,002_Second", region_text)
            self.assertIn("mouth_lips_jaw", region_text)
            self.assert_no_long_decimal_values(region_text)

            graph_paths = list((result.output_dir / "combined" / "other_findings" / "histogram_graphs").glob("*.svg"))
            self.assertTrue(any("anger" in path.name for path in graph_paths))
            self.assertFalse((result.output_dir / "combined" / "other_findings" / "logscale_histograms.csv").exists())

    def test_logscale_outputs_are_opt_in(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "Speaker_A"
            input_dir.mkdir()
            write_imotions_csv(
                input_dir / "001_First.csv",
                [
                    [1, 0, 0, -100, 10, 20, 30, 40, -5],
                    [2, 40, 25, 0, 20, 30, 40, 50, -2],
                    [3, 80, 100, 100, 30, 40, 50, 60, 0],
                ],
            )

            result = analyse_imotions_folder(
                input_dir,
                output_root=root / "output",
                write_graphs=True,
                include_logscale=True,
            )

            logscale_csv = result.output_dir / "combined" / "other_findings" / "logscale_histograms.csv"
            self.assertTrue(logscale_csv.exists())
            logscale_text = logscale_csv.read_text(encoding="utf-8")
            self.assertIn("scale,log10(count + 1)", logscale_text)
            self.assertIn("Anger", logscale_text)
            self.assert_no_long_decimal_values(logscale_text)

            logscale_graphs = result.output_dir / "combined" / "other_findings" / "logscale_histogram_graphs"
            self.assertTrue(logscale_graphs.exists())
            self.assertTrue(any("anger" in path.name for path in logscale_graphs.glob("*.svg")))

    def test_low_range_imotions_core_emotions_are_scaled_to_0_to_100(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "Speaker_A"
            input_dir.mkdir()
            write_imotions_csv(
                input_dir / "001_Low.csv",
                [
                    [1, 0, 0.1, -25, 5, 5, 5, 5, 0],
                    [2, 40, 0.2, -10, 5, 5, 5, 5, 0],
                ],
            )

            result = analyse_imotions_folder(input_dir, output_root=root / "output", write_graphs=False)

            histogram_text = (result.output_dir / "combined" / "histograms.csv").read_text(encoding="utf-8")
            self.assertIn("Anger\nbin_start,bin_end,001_Low,total\n0,5,0,0\n5,10,0,0\n10,15,1,1", histogram_text)
            self.assertIn("20,25,1,1", histogram_text)

    def test_descriptive_statistics_include_excess_kurtosis(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "Speaker_A"
            input_dir.mkdir()
            write_imotions_csv(
                input_dir / "001_Kurtosis.csv",
                [
                    [1, 0, 0, -25, 5, 5, 5, 5, 0],
                    [2, 40, 0, -10, 5, 5, 5, 5, 0],
                    [3, 80, 10, 10, 5, 5, 5, 5, 0],
                    [4, 120, 10, 25, 5, 5, 5, 5, 0],
                ],
            )

            result = analyse_imotions_folder(input_dir, output_root=root / "output", write_graphs=False)
            descriptor_text = (result.output_dir / "combined" / "other_findings" / "descriptive_statistics.csv").read_text(encoding="utf-8")

        self.assertIn("kurtosis,-2", descriptor_text)

    def test_default_output_uses_analysis_output_folder(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "iMotions_Output" / "Speaker_A"
            input_dir.mkdir(parents=True)
            write_imotions_csv(input_dir / "001_Test.csv", [[1, 0, 10, -25, 5, 5, 5, 5, 0]])

            result = analyse_imotions_folder(input_dir, output_root=root / "output", write_graphs=False)

            self.assertEqual(result.output_dir, (root / "output" / "emotion" / "Speaker_A").resolve())
            self.assertEqual(result.domain_output_dirs["raw"], (root / "output" / "raw" / "Speaker_A").resolve())
            self.assertTrue((root / "output" / "emotion" / "Speaker_A" / "001_Test" / "histograms.csv").exists())
            self.assertTrue((root / "output" / "emotion" / "Speaker_A" / "combined" / "other_findings" / "descriptive_statistics.csv").exists())
            self.assertTrue((root / "output" / "raw" / "Speaker_A" / "001_Test" / "histograms.csv").exists())

    def test_parent_input_with_speaker_subfolders_keeps_face_speakers_separate(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "iMotions_Output" / "DemoDay"
            output_root = root / "output"
            speaker_a_dir = input_dir / "Speaker A" / "Sensor Data"
            speaker_b_dir = input_dir / "Speaker B" / "Sensor Data"
            speaker_a_dir.mkdir(parents=True)
            speaker_b_dir.mkdir(parents=True)
            write_imotions_csv(speaker_a_dir / "001_First.csv", [[1, 0, 10, -25, 5, 5, 5, 5, 0]])
            write_imotions_csv(speaker_a_dir / "002_Second.csv", [[1, 0, 20, -10, 5, 5, 5, 5, 0]])
            write_imotions_csv(speaker_b_dir / "001_First.csv", [[1, 0, 30, 25, 5, 5, 5, 5, 0]])

            result = analyse_imotions_folder(input_dir, output_root=output_root, write_graphs=False)

            run_dir = output_root / "emotion" / "DemoDay"
            raw_run_dir = output_root / "raw" / "DemoDay"
            speaker_a_output = run_dir / "Speaker_A"
            speaker_b_output = run_dir / "Speaker_B"
            self.assertEqual(result.output_dir, run_dir.resolve())
            self.assertEqual(result.domain_output_dirs["raw"], raw_run_dir.resolve())
            self.assertTrue((speaker_a_output / "001_First" / "histograms.csv").exists())
            self.assertTrue((speaker_a_output / "002_Second" / "histograms.csv").exists())
            self.assertTrue((speaker_a_output / "combined" / "histograms.csv").exists())
            self.assertTrue((speaker_b_output / "001_First" / "histograms.csv").exists())
            self.assertTrue((speaker_b_output / "combined" / "histograms.csv").exists())

            speaker_a_histograms = (speaker_a_output / "combined" / "histograms.csv").read_text(encoding="utf-8")
            speaker_b_histograms = (speaker_b_output / "combined" / "histograms.csv").read_text(encoding="utf-8")
            self.assertIn("Anger\nbin_start,bin_end,001_First,002_Second,total", speaker_a_histograms)
            self.assertIn("Anger\nbin_start,bin_end,001_First,total", speaker_b_histograms)
            self.assertNotIn("002_Second", speaker_b_histograms)
            self.assertTrue((raw_run_dir / "Speaker_A" / "001_First" / "histograms.csv").exists())

    def test_single_speaker_sensor_data_folder_uses_parent_speaker_name(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "Speaker C"
            sensor_dir = input_dir / "Sensor Data"
            sensor_dir.mkdir(parents=True)
            write_imotions_csv(
                sensor_dir / "001_First.csv",
                [[1, 0, 10, -25, 5, 5, 5, 5, 0]],
            )

            result = analyse_imotions_folder(
                input_dir,
                output_root=root / "output",
                write_graphs=False,
            )

            self.assertEqual(result.output_dir.name, "Speaker_C")
            self.assertTrue(
                (result.output_dir / "combined" / "other_findings" / "descriptive_statistics.csv").is_file()
            )
            self.assertFalse((result.output_dir / "Sensor_Data").exists())

    def test_bare_folder_name_can_resolve_inside_imotions_output(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            imotions_root = Path(temp_dir) / "iMotions_Output"
            input_dir = imotions_root / "Speaker_A"
            input_dir.mkdir(parents=True)

            resolved = resolve_input_folder("Speaker_A", imotions_root=imotions_root)

            self.assertEqual(resolved, input_dir.resolve())

    def test_default_analysis_folders_are_package_local(self):
        self.assertEqual(default_imotions_root(), Path("analysis/iMotions_Output").resolve())
        self.assertEqual(default_output_root(), Path("analysis/output").resolve())
        self.assertEqual(
            resolve_output_folder(Path("C:/example/iMotions_Output/Speaker_A")),
            Path("analysis/output/Speaker_A").resolve(),
        )

    def assert_no_long_decimal_values(self, text: str) -> None:
        matches = LONG_DECIMAL_PATTERN.findall(text)
        self.assertEqual(matches, [], f"Found values with more than two decimal places: {matches[:5]}")


if __name__ == "__main__":
    unittest.main()
