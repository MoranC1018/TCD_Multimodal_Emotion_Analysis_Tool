"""Speaker-level one-sample inference and workbook probability mirrors."""

from __future__ import annotations

import math
import statistics
from copy import copy
from dataclasses import dataclass, replace
from numbers import Real
from pathlib import Path
from typing import Mapping, Sequence

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from analysis.combined_summary import (
    CombinedMetricCells,
    protected_manual_discovery_directories,
)
from analysis.histograms import student_t_two_tailed_p
from spreadsheet_safety import neutralize_spreadsheet_row, neutralize_spreadsheet_value


@dataclass(frozen=True)
class InferenceRow:
    metric_key: str
    n: int
    reference: float
    mean: float | None
    standard_deviation: float | None
    standard_error: float | None
    ci_low: float | None
    ci_high: float | None
    p_value: float | None
    q_value: float | None
    effect_size: float | None
    probability_above: float | None
    excluded_speakers: tuple[str, ...]


@dataclass(frozen=True)
class InferenceResult:
    workbook_path: Path
    probability_sheets: tuple[str, ...]
    rows: tuple[InferenceRow, ...]
    reference_resolutions: tuple[ReferenceResolution, ...]


@dataclass(frozen=True)
class ReferenceResolution:
    """Explain how one metric's reference value was selected."""

    original_key: str
    matched_scope: str
    matched_source: str
    resolved_reference: float


def _detail_headers(confidence_level: float) -> tuple[str, ...]:
    percentage = f"{confidence_level * 100:g}%"
    return (
        "Metric",
        "n",
        "Reference",
        "Mean",
        "Sample standard deviation",
        "Standard error",
        "t statistic",
        f"{percentage} CI low",
        f"{percentage} CI high",
        "p-value",
        "q-value",
        "Effect size",
        "Probability above reference",
        "Excluded observations",
        "BH adjusted p-value",
    )


def student_t_probability(t_stat: float, df: int) -> float:
    """Return the Student-t CDF at ``t_stat`` for positive integer ``df``."""

    if df < 1:
        raise ValueError("Student-t degrees of freedom must be positive")
    if math.isnan(t_stat):
        raise ValueError("Student-t statistic must be finite or infinite")
    if math.isinf(t_stat):
        return 1.0 if t_stat > 0 else 0.0
    if t_stat == 0:
        return 0.5
    tail = student_t_two_tailed_p(abs(t_stat), df) / 2.0
    return 1.0 - tail if t_stat > 0 else tail


def _student_t_quantile(probability: float, df: int) -> float:
    if not 0.5 < probability < 1.0:
        raise ValueError("Student-t quantile probability must be between 0.5 and 1")
    lower = 0.0
    upper = 1.0
    while student_t_probability(upper, df) < probability:
        upper *= 2.0
    for _ in range(80):
        midpoint = (lower + upper) / 2.0
        if student_t_probability(midpoint, df) < probability:
            lower = midpoint
        else:
            upper = midpoint
    return (lower + upper) / 2.0


def calculate_inference(
    values: Sequence[float],
    reference: float,
    *,
    confidence_level: float = 0.95,
) -> InferenceRow:
    """Calculate one-sample speaker-level inference without row-level pooling."""

    reference = float(reference)
    if not math.isfinite(reference):
        raise ValueError("Reference must be finite")
    confidence_level = float(confidence_level)
    if not math.isfinite(confidence_level) or not 0.0 < confidence_level < 1.0:
        raise ValueError("Confidence level must be between 0 and 1")
    observations: list[float] = []
    for value in values:
        number = _finite_excel_number(value)
        if number is not None:
            observations.append(number)
    n = len(observations)
    if n == 0:
        return InferenceRow("", 0, reference, None, None, None, None, None, None, None, None, None, ())
    mean = statistics.fmean(observations)
    if n == 1:
        return InferenceRow("", 1, reference, mean, None, None, None, None, None, None, None, None, ())

    standard_deviation = statistics.stdev(observations)
    if standard_deviation == 0.0:
        return InferenceRow(
            "", n, reference, mean, standard_deviation, None, None, None,
            None, None, None, None, (),
        )

    standard_error = standard_deviation / math.sqrt(n)
    t_stat = (mean - reference) / standard_error
    df = n - 1
    probability_above = student_t_probability(t_stat, df)
    p_value = student_t_two_tailed_p(abs(t_stat), df)
    critical_value = _student_t_quantile(0.5 + confidence_level / 2.0, df)
    effect_size = (mean - reference) / standard_deviation
    return InferenceRow(
        "", n, reference, mean, standard_deviation, standard_error,
        mean - critical_value * standard_error,
        mean + critical_value * standard_error,
        p_value, None, effect_size, probability_above, (),
    )


def benjamini_hochberg(p_values: Sequence[float | None]) -> list[float | None]:
    """Return monotonic FDR-adjusted q-values in original order."""

    indexed = [(index, value) for index, value in enumerate(p_values) if value is not None]
    for _, value in indexed:
        if not math.isfinite(value) or value < 0.0 or value > 1.0:
            raise ValueError("p-values must be finite values between 0 and 1")
    ordered = sorted(indexed, key=lambda item: item[1])
    adjusted: list[tuple[int, float]] = []
    running = 1.0
    total = len(ordered)
    for rank, (index, value) in reversed(list(enumerate(ordered, start=1))):
        running = min(running, value * total / rank)
        adjusted.append((index, running))
    results: list[float | None] = [None] * len(p_values)
    for index, value in adjusted:
        results[index] = min(1.0, value)
    return results


def _quote_sheet_name(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def _reference_formula(sheet_name: str, coordinate: str) -> str:
    return f"={_quote_sheet_name(sheet_name)}!{coordinate}"


def _mirror_name(book: Workbook, source_name: str) -> str:
    counter = 1
    while True:
        suffix = " Prob" if counter == 1 else f" Prob {counter}"
        candidate = source_name[: 31 - len(suffix)] + suffix
        if candidate not in book.sheetnames:
            return candidate
        counter += 1


def _copy_sheet_presentation(source: Worksheet, mirror: Worksheet) -> None:
    for row in source.iter_rows():
        for cell in row:
            target = mirror[cell.coordinate]
            target._style = copy(cell._style)
            if cell.has_style:
                target.number_format = cell.number_format
            if cell.value is not None:
                target.value = _reference_formula(source.title, cell.coordinate)
    for key, dimension in source.row_dimensions.items():
        mirror.row_dimensions[key] = copy(dimension)
    for key, dimension in source.column_dimensions.items():
        mirror.column_dimensions[key] = copy(dimension)
    mirror.sheet_format = copy(source.sheet_format)
    mirror.sheet_properties = copy(source.sheet_properties)
    mirror.freeze_panes = source.freeze_panes
    mirror.auto_filter = copy(source.auto_filter)
    mirror.page_margins = copy(source.page_margins)
    mirror.page_setup = copy(source.page_setup)
    mirror.print_options = copy(source.print_options)
    mirror.print_title_rows = source.print_title_rows
    mirror.print_title_cols = source.print_title_cols
    for merged_range in source.merged_cells.ranges:
        mirror.merge_cells(str(merged_range))


def _worksheet_values(sheet: Worksheet, cells: CombinedMetricCells) -> tuple[list[float], tuple[str, ...]]:
    values: list[float] = []
    excluded: list[str] = []
    for coordinate, speaker_id in zip(cells.speaker_cells, cells.speaker_ids):
        number = _finite_excel_number(sheet[coordinate].value)
        if number is not None:
            values.append(number)
        else:
            excluded.append(speaker_id)
    return values, tuple(excluded)


def _finite_excel_number(value: object) -> float | None:
    """Match Excel reference semantics: text and logical cells are not numbers."""

    if isinstance(value, bool) or not isinstance(value, Real):
        return None
    number = float(value)
    return number if math.isfinite(number) else None


def _reference_for(
    metric_key: str,
    cells: CombinedMetricCells,
    default_reference: float,
    overrides: Mapping[str, float],
) -> ReferenceResolution:
    if metric_key in overrides:
        original_key = metric_key
        matched_scope = "metric"
        matched_source = metric_key
        value = overrides[metric_key]
    elif cells.sheet in overrides:
        original_key = cells.sheet
        matched_scope = "sheet"
        matched_source = cells.sheet
        value = overrides[cells.sheet]
    else:
        original_key = "<default>"
        matched_scope = "default"
        matched_source = "Default reference"
        value = default_reference
    value = float(value)
    if not math.isfinite(value):
        raise ValueError(f"Reference for {metric_key} must be finite")
    return ReferenceResolution(original_key, matched_scope, matched_source, value)


def _details_formula(
    references: Sequence[str],
    row: int,
    column: str,
    confidence_level: float,
) -> str:
    values = ",".join(references)
    if column == "B":
        return f"=COUNT({values})"
    if column == "D":
        return f'=IFERROR(AVERAGE({values}),"")'
    if column == "E":
        # STDEV.S keeps blank and text source references out of the sample.
        # OOXML requires the _xlfn prefix for Excel to recognize this function.
        return f'=IF(B{row}<2,"",_xlfn.STDEV.S({values}))'
    if column == "F":
        return f'=IF(OR(E{row}="",E{row}=0),"",E{row}/SQRT(B{row}))'
    if column == "G":
        return f'=IF(OR(F{row}="",F{row}=0),"",(D{row}-C{row})/F{row})'
    if column == "H":
        alpha = f"{1.0 - confidence_level:g}"
        return f'=IF(OR(B{row}<2,E{row}=0),"",D{row}-_xlfn.T.INV.2T({alpha},B{row}-1)*F{row})'
    if column == "I":
        alpha = f"{1.0 - confidence_level:g}"
        return f'=IF(OR(B{row}<2,E{row}=0),"",D{row}+_xlfn.T.INV.2T({alpha},B{row}-1)*F{row})'
    if column == "J":
        return f'=IF(OR(B{row}<2,E{row}=0),"",_xlfn.T.DIST.2T(ABS(G{row}),B{row}-1))'
    if column == "L":
        return f'=IF(OR(E{row}="",E{row}=0),"",(D{row}-C{row})/E{row})'
    if column == "M":
        return f'=IF(OR(B{row}<2,E{row}=0),"",_xlfn.T.DIST(G{row},B{row}-1,TRUE))'
    raise ValueError(f"No inference formula for column {column}")


def _bh_candidate_formula(row: int, first_row: int, last_row: int) -> str:
    p_values = f"$J${first_row}:$J${last_row}"
    # Ties use their final rank so equal p-values share one conservative candidate.
    return f'=IF(J{row}="","",MIN(1,J{row}*COUNT({p_values})/COUNTIF({p_values},"<="&J{row})))'


def _bh_q_formula(row: int, first_row: int, last_row: int) -> str:
    p_values = f"$J${first_row}:$J${last_row}"
    candidates = f"$O${first_row}:$O${last_row}"
    return f'=IF(J{row}="","",_xlfn.MINIFS({candidates},{p_values},">="&J{row}))'


def _protected_destination(path: Path) -> bool:
    destination = path.resolve()
    for protected_directory in protected_manual_discovery_directories():
        try:
            destination.relative_to(protected_directory)
        except ValueError:
            continue
        return True
    return False


def _write_probability_outline(
    book: Workbook,
    records: Sequence[tuple[str, CombinedMetricCells]],
    default_reference: float,
    overrides: Mapping[str, float],
    confidence_level: float,
) -> Worksheet:
    """Write a literal, human-readable account of every probability input rule."""

    outline = book.create_sheet("Probability Outline")
    rows = (
        ("Probability calculation outline", None),
        ("Result scope", "What is used"),
        (
            "Speaker probability",
            "Available per-source means for that participant and metric.",
        ),
        (
            "Overall probability",
            "All selected speaker means shown on the source sheet; each speaker has equal weight.",
        ),
        (
            "Displayed percentage",
            "Directional Student t CDF score T(df)(t), equivalent to one minus the one-sided p-value under the null model. It is not a Bayesian posterior probability.",
        ),
        (
            "Calculation",
            "t = (sample mean - reference) / (sample standard deviation / SQRT(n)); df = n - 1; displayed score = T.DIST(t, df, TRUE).",
        ),
        (
            "Confidence interval",
            f"{confidence_level * 100:g}% two-sided Student t interval.",
        ),
        (
            "Default reference",
            f"{default_reference:g}. Exact metric or modality overrides are listed in Inference Settings.",
        ),
        (
            "Reference caution",
            "For bounded nonnegative scores, reference 0 only tests whether values exceed the scale floor and may be uninformative; set a justified metric override when appropriate.",
        ),
        ("Minimum sample", "At least two numeric observations; otherwise the result is blank."),
        ("Zero variance", "Inference is blank because a Student t uncertainty estimate is undefined."),
        ("Missing data", "Blank and non-numeric observations are excluded; they are not converted to zero."),
        (
            "Multiple comparisons",
            "Benjamini-Hochberg q-values use one family per modality sheet across its displayed speaker and overall tests.",
        ),
        (
            "Interpretation",
            "This directional t-based score is not classifier confidence, a Bayesian posterior probability, or a guaranteed replication rate.",
        ),
    )
    for row in rows:
        outline.append(neutralize_spreadsheet_row(row))
    outline["A1"].font = Font(bold=True, size=14)
    outline["A2"].font = Font(bold=True)
    outline["B2"].font = Font(bold=True)

    outline.append(())
    outline.append(
        neutralize_spreadsheet_row(
            ("Modality", "Group id", "Group", "Linked speakers", "Expected observations")
        )
    )
    display_names: dict[str, str] = {}
    for _metric_key, cells in records:
        for speaker_id, display_name in zip(
            cells.speaker_ids, cells.speaker_display_names
        ):
            display_names.setdefault(speaker_id, display_name)
    modalities = tuple(dict.fromkeys(cells.sheet for _, cells in records))
    group_metadata = records[0][1].speaker_groups if records else ()
    for modality in modalities:
        modality_cells = tuple(cells for _, cells in records if cells.sheet == modality)
        total_speakers = max((len(cells.speaker_ids) for cells in modality_cells), default=0)
        observation_counts: dict[str, int] = {}
        for cells in modality_cells:
            for speaker_id, observations in zip(cells.speaker_ids, cells.speaker_observations):
                observation_counts[speaker_id] = max(
                    observation_counts.get(speaker_id, 0), len(observations)
                )
        for group_id, group_name, speaker_ids in group_metadata:
            expected_sources = max(
                (observation_counts.get(speaker_id, 0) for speaker_id in speaker_ids),
                default=0,
            )
            outline.append(
                neutralize_spreadsheet_row(
                    (
                        modality,
                        group_id,
                        group_name,
                        ", ".join(
                            display_names.get(speaker_id, speaker_id)
                            for speaker_id in speaker_ids
                        ),
                        f"speaker: up to {expected_sources} per-source means; "
                        f"overall: up to {total_speakers} speaker means",
                    )
                )
            )

    outline.append(())
    outline.append(
        neutralize_spreadsheet_row(
            ("Resolved reference by source metric", "Reference", "Matched scope", "Matched source")
        )
    )
    for metric_key, cells in records:
        resolution = _reference_for(metric_key, cells, default_reference, overrides)
        outline.append(
            neutralize_spreadsheet_row(
                (
                    metric_key,
                    resolution.resolved_reference,
                    resolution.matched_scope,
                    resolution.matched_source,
                )
            )
        )

    for row in outline.iter_rows():
        if row[0].value in {"Modality", "Resolved reference by source metric"}:
            for cell in row:
                cell.font = Font(bold=True)
    outline.column_dimensions["A"].width = 28
    outline.column_dimensions["B"].width = 72
    outline.column_dimensions["C"].width = 24
    outline.column_dimensions["D"].width = 52
    outline.column_dimensions["E"].width = 56
    for row in outline.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    outline.freeze_panes = "A2"
    return outline


def add_probability_mirrors(
    workbook_path: str | Path,
    source_cells: Mapping[str, CombinedMetricCells],
    *,
    default_reference: float = 0.0,
    reference_overrides: Mapping[str, float] | None = None,
    confidence_level: float = 0.95,
) -> InferenceResult:
    """Add settings, details, and adjacent probability mirrors in place."""

    destination = Path(workbook_path).expanduser().resolve()
    if _protected_destination(destination):
        raise ValueError(f"Refusing to modify protected reference workbook: {destination}")
    if not destination.is_file():
        raise ValueError(f"Workbook does not exist: {destination}")
    default_reference = float(default_reference)
    if not math.isfinite(default_reference):
        raise ValueError("Default reference must be finite")
    confidence_level = float(confidence_level)
    if not math.isfinite(confidence_level) or not 0.0 < confidence_level < 1.0:
        raise ValueError("Confidence level must be between 0 and 1")
    overrides = dict(reference_overrides or {})
    book = openpyxl.load_workbook(destination, data_only=False)
    if "Inference Settings" in book.sheetnames or "Inference Details" in book.sheetnames:
        raise ValueError("Workbook already contains inference sheets")

    records = list(source_cells.items())
    for metric_key, cells in records:
        if len(cells.speaker_cells) != len(cells.speaker_ids):
            raise ValueError(f"Speaker cells and ids differ for {metric_key}")
        if cells.sheet not in book.sheetnames:
            raise ValueError(f"Source sheet is missing for {metric_key}: {cells.sheet}")
    for _, value in overrides.items():
        if not math.isfinite(float(value)):
            raise ValueError("Reference overrides must be finite")
    accepted_override_keys = {metric_key for metric_key, _ in records}
    accepted_override_keys.update(cells.sheet for _, cells in records)
    unknown_overrides = sorted(set(overrides) - accepted_override_keys)
    if unknown_overrides:
        valid_examples = ", ".join(sorted(accepted_override_keys)[:4])
        raise ValueError(
            "Unknown reference override key(s): "
            f"{', '.join(unknown_overrides)}. Use an exact generated metric key or quantitative "
            f"sheet title, for example: {valid_examples}"
        )
    records_by_sheet: dict[str, list[tuple[str, CombinedMetricCells]]] = {}
    for record in records:
        records_by_sheet.setdefault(record[1].sheet, []).append(record)
    ordered_records = [record for sheet_records in records_by_sheet.values() for record in sheet_records]

    _write_probability_outline(
        book,
        ordered_records,
        default_reference,
        overrides,
        confidence_level,
    )
    settings = book.create_sheet("Inference Settings")
    settings.append(
        neutralize_spreadsheet_row(
            ("Original key", "Matched scope", "Matched source", "Resolved reference")
        )
    )
    settings.append(
        neutralize_spreadsheet_row(
            ("<default>", "default", "Default reference", default_reference)
        )
    )
    details = book.create_sheet("Inference Details")
    details.append(neutralize_spreadsheet_row(_detail_headers(confidence_level)))
    inputs = book.create_sheet("Inference Inputs")
    max_observations = max(
        (
            len(observations)
            for _metric_key, cells in ordered_records
            for observations in cells.speaker_observations
        ),
        default=5,
    )
    input_headers: list[str] = ["Probability target"]
    for index in range(1, max(5, max_observations) + 1):
        input_headers.extend((f"Source {index}", f"Value {index}"))
    inputs.append(
        neutralize_spreadsheet_row(tuple(input_headers))
    )

    reference_resolutions: list[ReferenceResolution] = []
    target_records: list[
        tuple[str, CombinedMetricCells, str, tuple[str, ...], tuple[float, ...], tuple[str, ...], int]
    ] = []
    for metric_key, cells in ordered_records:
        source = book[cells.sheet]
        resolution = _reference_for(metric_key, cells, default_reference, overrides)
        reference_resolutions.append(resolution)
        settings.append(
            neutralize_spreadsheet_row(
                (
                    resolution.original_key,
                    resolution.matched_scope,
                    resolution.matched_source,
                    resolution.resolved_reference,
                )
            )
        )
        settings_row = settings.max_row

        if cells.speaker_observations:
            if len(cells.speaker_observations) != len(cells.speaker_cells):
                raise ValueError(f"Speaker observations and cells differ for {metric_key}")
            for coordinate, speaker_id, observations in zip(
                cells.speaker_cells, cells.speaker_ids, cells.speaker_observations
            ):
                target_key = f"{metric_key}|{speaker_id}"
                observation_index = cells.speaker_ids.index(speaker_id)
                labels = (
                    cells.speaker_observation_labels[observation_index]
                    if cells.speaker_observation_labels
                    else tuple(f"{index:03d}" for index in range(1, len(observations) + 1))
                )
                if len(labels) != len(observations):
                    raise ValueError(f"Speaker observation labels and values differ for {target_key}")
                input_values: list[object] = [target_key]
                for label, value in zip(labels, observations):
                    input_values.extend((label, value))
                inputs.append(neutralize_spreadsheet_row(tuple(input_values)))
                input_row = inputs.max_row
                references = tuple(
                    _reference_formula(inputs.title, inputs.cell(input_row, column).coordinate)[1:]
                    for column in range(3, 3 + len(observations) * 2, 2)
                )
                values = tuple(
                    number
                    for value in observations
                    if (number := _finite_excel_number(value)) is not None
                )
                excluded = tuple(
                    label
                    for label, value in zip(labels, observations)
                    if _finite_excel_number(value) is None
                )
                target_records.append(
                    (target_key, cells, coordinate, references, values, excluded, settings_row)
                )

        values, excluded = _worksheet_values(source, cells)
        overall_key = f"{metric_key}|overall" if cells.speaker_observations else metric_key
        overall_references = tuple(
            _reference_formula(cells.sheet, coordinate)[1:] for coordinate in cells.speaker_cells
        )
        target_records.append(
            (
                overall_key,
                cells,
                cells.overall,
                overall_references,
                tuple(values),
                excluded,
                settings_row,
            )
        )

    rows: list[InferenceRow] = []
    for detail_row, (
        target_key,
        cells,
        _,
        references,
        values,
        excluded,
        settings_row,
    ) in enumerate(target_records, start=2):
        reference = float(settings[f"D{settings_row}"].value)
        row = replace(
            calculate_inference(
                values,
                reference,
                confidence_level=confidence_level,
            ),
            metric_key=target_key,
            excluded_speakers=excluded,
        )
        rows.append(row)
        details.cell(detail_row, 1, neutralize_spreadsheet_value(target_key))
        details.cell(
            detail_row,
            2,
            _details_formula(references, detail_row, "B", confidence_level),
        )
        details.cell(detail_row, 3, _reference_formula(settings.title, f"D{settings_row}"))
        for column in ("D", "E", "F", "G", "H", "I", "J", "L", "M"):
            details[f"{column}{detail_row}"] = _details_formula(
                references,
                detail_row,
                column,
                confidence_level,
            )
        details[f"N{detail_row}"] = neutralize_spreadsheet_value(", ".join(excluded))

    target_indexes_by_sheet: dict[str, list[int]] = {}
    for index, (_, cells, *_rest) in enumerate(target_records):
        target_indexes_by_sheet.setdefault(cells.sheet, []).append(index)
    q_values: list[float | None] = [None] * len(rows)
    for indexes in target_indexes_by_sheet.values():
        family_q_values = benjamini_hochberg([rows[index].p_value for index in indexes])
        for index, q_value in zip(indexes, family_q_values):
            q_values[index] = q_value
    rows = [replace(row, q_value=q_values[index]) for index, row in enumerate(rows)]
    for detail_row, row in enumerate(rows, start=2):
        cells = target_records[detail_row - 2][1]
        family_indexes = target_indexes_by_sheet[cells.sheet]
        first_detail_row = family_indexes[0] + 2
        last_detail_row = family_indexes[-1] + 2
        details[f"O{detail_row}"] = _bh_candidate_formula(detail_row, first_detail_row, last_detail_row)
        details[f"K{detail_row}"] = _bh_q_formula(detail_row, first_detail_row, last_detail_row)
        for column in "CDEFGHIJKLM":
            details[f"{column}{detail_row}"].number_format = "0.00%" if column in {"J", "K", "M"} else "0.00"
        details[f"O{detail_row}"].number_format = "0.00%"
    for column, width in {"A": 56, "B": 10, "C": 12, "D": 12, "E": 24, "F": 14, "G": 12, "H": 14, "I": 14, "J": 12, "K": 12, "L": 12, "M": 28, "N": 30, "O": 20}.items():
        details.column_dimensions[column].width = width
    settings.column_dimensions["A"].width = 32
    settings.column_dimensions["B"].width = 16
    settings.column_dimensions["C"].width = 34
    settings.column_dimensions["D"].width = 20
    inputs.column_dimensions["A"].width = 56
    for column in range(2, len(input_headers) + 1):
        inputs.column_dimensions[get_column_letter(column)].width = 16
    inputs.sheet_state = "hidden"

    probability_sheets: list[str] = []
    mirrored_sources: dict[str, Worksheet] = {}
    for _, cells, *_rest in target_records:
        if cells.sheet in mirrored_sources:
            continue
        source = book[cells.sheet]
        mirror = book.create_sheet(_mirror_name(book, source.title), book.index(source) + 1)
        _copy_sheet_presentation(source, mirror)
        mirrored_sources[cells.sheet] = mirror
        probability_sheets.append(mirror.title)
    for detail_row, (_, cells, destination_cell, *_rest) in enumerate(target_records, start=2):
        mirror = mirrored_sources[cells.sheet]
        mirror[destination_cell] = _reference_formula(details.title, f"M{detail_row}")
        mirror[destination_cell].number_format = "0.00%"

    book.calculation.calcMode = "auto"
    book.calculation.fullCalcOnLoad = True
    book.calculation.forceFullCalc = True
    book.save(destination)
    return InferenceResult(
        destination,
        tuple(probability_sheets),
        tuple(rows),
        tuple(reference_resolutions),
    )
