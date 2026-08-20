"""Build the descriptive workbook used by the combined Analysis workflow."""

from __future__ import annotations

import base64
import csv
import json
import math
import re
import statistics
import unicodedata
import zlib
from dataclasses import dataclass
from pathlib import Path
from typing import Literal, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from analysis.metadata import (
    ManifestSource,
    SourceMetadata,
    load_source_metadata,
    map_report_source_ids,
    resolve_analysis_profile,
)
from analysis.profile import AnalysisProfile
from spreadsheet_safety import neutralize_spreadsheet_value


EXPECTED_VIDEO_COUNT = 5
AUDIO_EMOTIONS = (
    "Anger", "Contempt", "Disgust", "Fear", "Joy", "Sadness", "Surprise", "Neutral", "Other",
)
AUDIO_VALENCE = ("Valence",)
AUDIO_DIMENSIONS = ("Arousal", "Dominance")
AUDIO_METRICS = (*AUDIO_EMOTIONS, *AUDIO_VALENCE, *AUDIO_DIMENSIONS)
AUDIO_REQUIRED_METRICS = (
    "Anger", "Joy", "Sadness", "Neutral", "Arousal", "Dominance", "Valence",
)
VIDEO_EMOTIONS = (
    "Anger", "Contempt", "Disgust", "Fear", "Joy", "Sadness", "Surprise", "Neutral", "Confusion",
)
VIDEO_SENTIMENT = ("Sentimentality",)
VIDEO_VALENCE = ("Valence", "Adaptive Valence")
VIDEO_DIMENSIONS = ("Engagement", "Adaptive Engagement")
VIDEO_METRICS = (*VIDEO_EMOTIONS, *VIDEO_SENTIMENT, *VIDEO_VALENCE, *VIDEO_DIMENSIONS)
VIDEO_DETAIL_METRICS = VIDEO_METRICS
VIDEO_KURTOSIS_METRICS = VIDEO_METRICS
TEXT_SENTIMENT = (
    "Positive Sentiment",
    "Negative Sentiment",
)
TEXT_DIMENSIONS = (
    "Arousal / Activation",
    "Dominance / Power",
    "Affiliation / Social orientation",
)
TEXT_CONSTRUCTS = (*TEXT_SENTIMENT, *TEXT_DIMENSIONS)
ORDINAL_LABELS = ("1st", "2nd", "3rd", "4th", "5th")
GROUP_LABEL_COLUMNS = (2, 7, 11, 15)
GROUP_SPEAKER_COLUMNS = ((4, 5, 6), (8, 9, 10), (12, 13, 14), (16, 17, 18))
GROUP_OVERALL_COLUMN = 19


@dataclass(frozen=True)
class MeasureLayout:
    headline_start: int
    headline_end: int
    count_heading: int
    count_start: int
    count_end: int
    detail_heading: int
    detail_start: int
    detail_end: int
    kurtosis_heading: int | None
    kurtosis_start: int | None
    kurtosis_end: int | None
    max_row: int


def _measure_layout(
    metrics: Sequence[str],
    *,
    count_gap: int,
    count_rows: int,
    count_starts_at_heading: bool,
    source_count: int = EXPECTED_VIDEO_COUNT,
    kurtosis_metrics: Sequence[str] = (),
) -> MeasureLayout:
    headline_start = 2
    headline_end = headline_start + len(metrics) - 1
    count_heading = headline_end + count_gap + 1
    count_start = count_heading if count_starts_at_heading else count_heading + 1
    count_end = count_start + count_rows - 1
    detail_heading = count_end + 1
    detail_start = detail_heading + 1
    detail_end = detail_start + len(metrics) * source_count - 1
    kurtosis_heading = detail_end + 1 if kurtosis_metrics else None
    kurtosis_start = kurtosis_heading + 1 if kurtosis_heading is not None else None
    kurtosis_end = (
        kurtosis_start + len(kurtosis_metrics) - 1
        if kurtosis_start is not None
        else None
    )
    return MeasureLayout(
        headline_start,
        headline_end,
        count_heading,
        count_start,
        count_end,
        detail_heading,
        detail_start,
        detail_end,
        kurtosis_heading,
        kurtosis_start,
        kurtosis_end,
        kurtosis_end or detail_end,
    )


AUDIO_LAYOUT = _measure_layout(
    AUDIO_METRICS,
    count_gap=0,
    count_rows=8,
    count_starts_at_heading=False,
)
VIDEO_LAYOUT = _measure_layout(
    VIDEO_METRICS,
    count_gap=1,
    count_rows=9,
    count_starts_at_heading=True,
    kurtosis_metrics=VIDEO_KURTOSIS_METRICS,
)

STATIC_DEFINITION_SHEETS = """
eNqVWdtu40YS/ZWGX9YGFI0p6zpvjpyxx2MnxjgXIMhLi2xRvSbZDC+WNUGAINiH3ef9iv2ORX4kX7KnqropyvYskCepi83u6rqcOtX85ejC5doW6sKs1bfmqTl6+8vReXT09qgxVV4rPElNYSqd4e/Pra1MdTQ4Oj/DhGtdufqP3/VWfdDxp90f/6gfrPrzt3+rb9ZrG1u8sbIurXS52al15XLVbIy6c5mtN+re/D1XukhYdoMl6P91Wzc2Nur4zt6fqFJXzW5Im42x2XlSWV2oHzFvZar0L+/zUX8yOa82wWr3n3a5K9SV++N3ty2s/svLYUw/o9PJKS86w6LfukTv1OxtdKbubkk4h/CHjW6UrXmR2qaFxR66wCHdmmWxbkzqKmtqVTeVK9KB2hr9MFA6buyjGfBupa5rDMgZdxub2VLdN64wf6vVpffNe+8blVi85wpd7UiD6JT0Otzm3m/zA29z3tvmbr8NqfZi7ePL9ye9DZSujNKwY8rTsXpqC8yvTa4LOFI1Oq1VYh5N5kqTqNUuaH89lAPwrrHLMqPTFqqtXYVh0ZiiwSOd7WpbD9V3RWYfYD1IbU6Peoc5NsN0OIA7aotz8Hpfm1Tj/8mAtKoPjk4KW1o+gTqNw6OyaSGrm12Gx3CJeSorAysgOGit7ig/tzrDDqYeqMI1yuSOjZCpR50ZuJODIBqJtW116Gq4n4LZxm2mq2ynKpOZR6xLhq6bNrGyd+loixiLVhvTQON4oGCZxFT1BkaLXZ63BZ7Tzj4u8KjVpK3sT4kSDb2HWTL1/hdRsMWOrYysr9XWVQlFJ6IUskeze8vvUegiHk2RNhsWLCAo3Vayf0RhFZuKlmg40EaEGGxaOEIkZAzdNhscxEsINBITWwqyAlZmIaXjV086L8kDtoizNjGsw2gWtly3GQtIqcTlCLKiYQEphZiJjZ9xdso7rG1hG8MCUmttq5wHpFHdwmjWCYiNxToIEjGDW+MA7DudsQ5nZD/YZW0T8vKAZaTX3i8iY9UMQSa0YwexmBREUsFS/siCZ6TV+6Ln8Lo0Jt5QdGm1sekGybauzM8tNt1RaHj3iZY2KIqVHQCrQTBvbbNR3U7IRRVMr1kQIkr2J5uPhowAPJ75KCHBPka8M0KMPNuN7TOmAxJe0dn4yBNyQVt0wSFCOvBjmxGerIAAQTxiq5WUj8G6EwqSzOasuDfjhPwE63iZoPj0c3EzmXudeMAhUunUZhwQU44Q164aHzLTqFtbgmo64kiGe5HzEkXT53HifdMgl3VZGl1BAd57SlrtMYTTGuEDa8hJpmTpuCKv2zoXkYRODWnZhHceCWRzseiUjpDYOm73a7InMlVWbpWZXKJqRic596Gkct0lOf4XO/GtqL8FChPo2Qy1zlgEVggPneFIIeeUlRKFFfUDItISmJOeK8uQXxnd1BRqfc8Cy0QbirGzoa8vLAlRJqJ9nFVmjVQiPBb1EAxA1yJli844heIuFubkwNw9GqoEIqFzxxtdpBJBc3EgVOHQFxlFFZWzVCJvTh5NdoXO4QbWd/7ZeJqTg1atzRIeLdiDxofG4pQBJt1w7CwirxwPGADjjTV+fOZxgwekQFPpogZ+MTotps+CjIoVYT5KkY+uBWkSQ0UNhNpjBj9a8M44sPaF6eBxdMqIDfeyncWeUjNOSc0rnCD4pY5dZQLE2CKhioNKrcRc8WEdktLpV6ITjIeBRohs7p0euMXn0MVvFNweMXNBGpJOAS+iiOwrEYsHAKK29GERce3daksreAnZuzJEbveyMctig0U72YRrnQcmOUo0+1w0RFwJZVmp+accWgl4hIwjdntsykbGo25LP+HMayoj0gh+cVsZkg2DrfpFycVx28FMNDoADUwlhqjT4G0ujKjujW1aHw6gq2mrfYpEXCiRd5kRnxO8uLaqvSfPPJHZqXrj2ixhzrOicCzWbR3KDWEDeHKeayligaU+OpD4oerRzh1hR8elKt1QgQPnKPrv+63JAB9NxtHFvGdlmq0xxTO2LJPnXc4kRA0FHylxQLRSk1GUhfrSeIpEJ38H0DLiXbHnmA58b4i/xOSlL70kEKmljPdgFo3Jiz8d/QAvEYxWhhFCwYhKCMDwpyNZ6MznlyzTG2GRMS/ClnJt0TAY5szLAqYiucJCNPnGbdGq3DqQD0GgZSeXFSdBLdAgw8WaVHIlEgXHq7vFJgdaTQ60mvIaS+TSJ9hU1e0aAI2VUoK9bNetQfM8h1jKqJ/6Y3LOPSo5k15iNrU3MNxYdMXpgNpQXLm28VVJzC0eYl7x09E3OEzJZJ9gEZUY3E70wRRy0ZJILAL5lR1XXek53EY08NuIR+6MQ2ywjrZoKeB19oC53VZ0tvd56cBymOG/JHNdAyM8PIQdClWFmkMhH8j9AE6vQSKA1lWX7itQuXD2aQfRfWZ/dojbIhR8y7jyIfzXBqW8V/pl0uS1SQegI/OkYHP30oDaFsojRu3VWrwkx9GMDgqcAXH1gqjH+tUz3hzNRl1dkuYvC5DUr1Gz8WtJO6OIu9k3SFyKODL5ybe7kt1x+Z540toTwIhpyHLvBXlhdmjhgXpu2RlDDfxauLjav0fSkI3du52g/z4Za4lNYc6DCirrLF7dnx13/JJ0nUiPRqb+EnVHtz2VWHolPUQ/HZkC3XTEWlw8n3QA2muTEQ6m17f764BoMT2tue2MFjP/zwcEMi1cEnUoH2e0+brrV6mEaJWTaZAs9S7euMyl7CGsjhrBFSO0nzOm6fxSbnRNTTo1u1TKUKSB9qiogCIJovn85WzfofsJi/0EsBukG+AwoWiiRYgR9cpH3VLygoZKpxlLN1W5tg41imneFVgykIZvGXZqoxGyKypTW2QD1uQyaQnm8rJtAil7frmBV2KNqbIIwvQRb/MVRYUeoiH4D8zIMTb0mHqfei0mUjXpWgEK40A7gv2AKzJn5pOIH9G9A7eWfUDaY1gHTXKHsqZXnl2kHFxp4AzUMRwch46eMU/v3d1Q1HjfBF95OGEK291TgFkSAJIjNBnmaccAcdBM82LdFQNZO7CcodxSnHY1u4MuJdgkN0+HfHhAjRH5vk+8D+xz0EWPTiNf/8I9Ui2tDIEhL3vY8YbkHVA87Tvc2i826hG/wLprzhhOxj3Vc+RX7bk6hYA0mT1I8SuOuRzmAI7A2Naa2sYvepdcwQXhIu0cpZ6W1uqrIoWpOp+dJ1q2/14uvE5Y+ugoWIho9BcpbcP1L1HdjdCJzxNybAhzrWBKgqm8zRAdLpGOlpCajvS5qy9/OoL4D2bXM7nIQ8Px/PbyrTpXS05G7HheYiP0ZuSlpc/J8y4nX9xTDtSFeaL3lkN10RblBt681dANAaTuMQnO3MiV3IUurMnU7VB9k9rs0TIRtnW4AkYiFUlAA5Tah8OW+pVL137WyX25rSjd9uBiKvC05OXFqdiDsur5tavBrGyojgHp0xNh6uhYsJgudqTICyUy82SxPjAxbikkpPdDHGVMVvmWoAP+TuGB75z2JOv/3DoPesfr7mh1WWah07TFK0f8ddD/fnHPzKv7grFpmrJ+++bNdrsdUisMw+uhad/Q/y9MW6G6hW8ZS6HfVEG5Tpo1tWfLs155f7PE1tzkXZD4o2/6L2lwLuUBw6sz7hxLw/3pa+mESdc8SVAQww+HS9zQ8CJUHwhupXQ/UF05WIe/jLyr/LQvaXRtdPHFTRur2//+hxJ1wzssx7whPNRScOq4YvEFiRHIQAh1gx7dkPCShN8b9CJoStUxEjwnhprpJ5MQ8bja9xvX/GWmSPlW+MN430Tc9P7f9v7zx5f3DbxH+k549xbhax1CwlQrR7qKGIzTqY+m+GRJUZJdWlelVqtb8JKCpJeTnntIs2fjaxq/QwUk7Xotzs2kO8Ft/xV0VkFPbmb4Gw8pOn35AYoUnb7yJYmUnX7m4xipPA1E79g8xRYxQnGfVgAUuHTT1lbTNzC283Sv8TXrUwE16SjTV3W+6c2/fX0Kf6b6ji7HE/UBWQmKQ8cj6Y86fpCvWvQVj043Y9+CJIBiFhtNMy9I9rVFZ63e6QpxSKqR7F4n/kL/w6yz7c3+7+3+7yXf7AB066ZqY4riK5bIh7KOwNU8lbK4V4iOu3xiA/Hjd1zTug83PWYQbglV/7vEJV8U9fKVx3cbwEngo28O2SmX2kDVL6PnyXnFknts4AsxQMJlA9VrjKhkh28nl5FQ7pZZz6VcNO1P+Oe//gkSBTZpwbheHIvfmO/15+nMFouXii76ivJMMGi+hkqClke//vo/7GTNdA==
"""


class InputError(ValueError):
    """Raised when input reports cannot safely produce a workbook."""


@dataclass(frozen=True)
class CombinedSource:
    modality: Literal["audio", "video"]
    speaker_key: str
    display_name: str
    report_path: Path


@dataclass(frozen=True)
class SpeakerGroupDefinition:
    group_id: str
    name: str
    speaker_ids: tuple[str, ...]


@dataclass(frozen=True)
class TextConstructSummary:
    """Imported speaker-level transcript constructs used only by the workbook."""

    speaker_id: str
    display_name: str
    country: str
    constructs: Mapping[str, float | None]
    source_path: Path


@dataclass(frozen=True)
class CombinedMetricCells:
    sheet: str
    metric: str
    overall: str
    speaker_cells: tuple[str, ...]
    speaker_ids: tuple[str, ...]
    speaker_observations: tuple[tuple[float | None, ...], ...] = ()
    speaker_observation_labels: tuple[tuple[str, ...], ...] = ()
    speaker_groups: tuple[tuple[str, str, tuple[str, ...]], ...] = ()
    speaker_display_names: tuple[str, ...] = ()


@dataclass(frozen=True)
class CombinedWorkbookResult:
    workbook_path: Path
    quantitative_sheets: tuple[str, ...]
    source_cells: dict[str, CombinedMetricCells]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class DiscoveryEntry:
    """Auditable decision for one discovered report candidate."""

    modality: Literal["audio", "video", "text"]
    normalized_speaker: str
    display_speaker: str
    path: Path
    reason: str


@dataclass(frozen=True)
class CombinedDiscoveryResult:
    sources: tuple[CombinedSource, ...]
    accepted: tuple[DiscoveryEntry, ...]
    rejected: tuple[DiscoveryEntry, ...]
    errors: tuple[str, ...] = ()


@dataclass(frozen=True)
class Speaker:
    speaker_id: str
    display_name: str
    workbook_header: str
    country: str
    column: str
    aliases: tuple[str, ...]


@dataclass(frozen=True)
class _ResolvedSpeakerGroup:
    group_id: str
    name: str
    speakers: tuple[Speaker, ...]


@dataclass(frozen=True)
class _LinkedLayout:
    label_columns: tuple[int, ...]
    speaker_columns: tuple[tuple[int, ...], ...]
    overall_column: int


@dataclass(frozen=True)
class MetricSeries:
    sources: tuple[str, ...]
    available: tuple[bool, ...]
    counts: tuple[int, ...]
    missing: tuple[int, ...]
    means: tuple[float, ...]
    stddevs: tuple[float, ...]
    kurtoses: tuple[float | None, ...]


def normalized(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", str(value))
    ascii_text = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", ascii_text.casefold())


METRIC_ALIASES = {normalized(metric): metric for metric in set(AUDIO_METRICS + VIDEO_METRICS)}
METRIC_ALIASES.update({"adaptiveval": "Adaptive Valence", "adaptiveeng": "Adaptive Engagement", "sentimental": "Sentimentality"})


def canonical_metric(value: str) -> str:
    try:
        return METRIC_ALIASES[normalized(value)]
    except KeyError as exc:
        raise InputError(f"Unknown metric name: {value!r}") from exc


def resolve_speaker(value: str) -> Speaker:
    """Create an exact, project-neutral identity for a supplied speaker."""

    display_name = " ".join(str(value).strip().replace("_", " ").split())
    key = normalized(display_name)
    if not key:
        raise InputError("Speaker name must contain letters or numbers")
    return Speaker(key, display_name, display_name, "", "", (display_name,))


def _resolve_speaker_groups(
    speaker_groups: Sequence[SpeakerGroupDefinition],
    speaker_catalog: Mapping[str, Speaker] | None = None,
) -> tuple[_ResolvedSpeakerGroup, ...]:
    speaker_catalog = speaker_catalog or {}
    resolved_groups: list[_ResolvedSpeakerGroup] = []
    group_ids: set[str] = set()
    names: set[str] = set()
    memberships: set[str] = set()
    for group in speaker_groups:
        group_id = group.group_id.strip()
        name = group.name.strip()
        if not group_id:
            raise InputError("Speaker group id must be nonblank")
        if not name:
            raise InputError("Speaker group name must be nonblank")
        if group_id in group_ids:
            raise InputError(f"Duplicate speaker group id: {group_id}")
        if name in names:
            raise InputError(f"Duplicate speaker group name: {name}")
        if not group.speaker_ids:
            raise InputError(f"Speaker group {group_id} must contain at least one speaker")
        speakers = tuple(
            speaker_catalog.get(resolved.speaker_id, resolved)
            for speaker_id in group.speaker_ids
            for resolved in (resolve_speaker(speaker_id),)
        )
        for speaker in speakers:
            if speaker.speaker_id in memberships:
                raise InputError(f"Speaker belongs to more than one group: {speaker.display_name}")
            memberships.add(speaker.speaker_id)
        group_ids.add(group_id)
        names.add(name)
        resolved_groups.append(_ResolvedSpeakerGroup(group_id, name, speakers))
    return tuple(resolved_groups)


def _default_speaker_groups(
    speaker_catalog: Mapping[str, Speaker],
) -> tuple[_ResolvedSpeakerGroup, ...]:
    """Create deterministic layout groups from the speakers actually supplied."""

    speakers = tuple(
        sorted(speaker_catalog.values(), key=lambda speaker: speaker.display_name.casefold())
    )
    chunks = tuple(
        speakers[index : index + len(GROUP_SPEAKER_COLUMNS[0])]
        for index in range(0, len(speakers), len(GROUP_SPEAKER_COLUMNS[0]))
    )
    return tuple(
        _ResolvedSpeakerGroup(
            "speakers" if len(chunks) == 1 else f"speakers-{index}",
            "Speakers" if len(chunks) == 1 else f"Speakers {index}",
            chunk,
        )
        for index, chunk in enumerate(chunks, start=1)
    )


def _parse_number(value: str, context: str) -> float:
    try:
        number = float(value.strip())
    except (AttributeError, ValueError) as exc:
        raise InputError(f"Invalid numeric value for {context}: {value!r}") from exc
    if not math.isfinite(number):
        raise InputError(f"Non-finite numeric value for {context}: {value!r}")
    return number


def _parse_count(value: str, context: str) -> int:
    number = _parse_number(value, context)
    if number < 0 or not math.isclose(number, round(number), abs_tol=1e-9):
        raise InputError(f"Count must be a non-negative integer for {context}: {value!r}")
    return int(number)


def _parse_optional_number(value: str) -> float | None:
    if not value or value.strip().casefold() in {"/", "na", "n/a", "nan"}:
        return None
    return _parse_number(value, "kurtosis")


def _source_slots(path: Path, metric: str, sources: Sequence[str]) -> tuple[int, ...]:
    """Map legacy 001-005 labels while allowing arbitrary ordered SourceIDs."""

    if not sources or len(set(sources)) != len(sources):
        raise InputError(f"{path}: {metric} must contain one or more unique sources")
    slots: list[int] = []
    for source in sources:
        match = re.match(r"^0*([1-5])(?:\D|$)", source)
        if match is None:
            return tuple(range(len(sources)))
        slot = int(match.group(1)) - 1
        if slot in slots:
            raise InputError(f"{path}: {metric} contains duplicate source ordinal {slot + 1:03d}")
        slots.append(slot)
    if slots != sorted(slots):
        raise InputError(f"{path}: required metric source order changes at {metric}")
    return tuple(slots)


def _slot_values(values: Sequence[object], slots: Sequence[int], missing: object) -> tuple[object, ...]:
    target_size = max(EXPECTED_VIDEO_COUNT, max(slots, default=-1) + 1)
    padded = [missing] * target_size
    for slot, value in zip(slots, values):
        padded[slot] = value
    return tuple(padded)


def parse_sectioned_csv(path: Path) -> dict[str, MetricSeries]:
    """Read the sectioned CSV emitted by ``write_descriptive_statistics_csv``."""

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    metrics: dict[str, MetricSeries] = {}
    row_index = 0
    while row_index < len(rows):
        row = rows[row_index]
        if not row or not row[0].strip():
            row_index += 1
            continue
        try:
            metric = canonical_metric(row[0])
        except InputError:
            row_index += 1
            continue
        section: dict[str, list[str]] = {}
        scan = row_index + 1
        while scan < len(rows) and rows[scan] and rows[scan][0].strip():
            section[rows[scan][0].strip().casefold()] = rows[scan][1:]
            scan += 1
        source_row = [value.strip() for value in section.get("metric", []) if value.strip()]
        slots = _source_slots(path, metric, source_row)
        source_count = len(source_row)

        def required(name: str) -> list[str]:
            values = section.get(name, [])
            if len(values) < source_count:
                raise InputError(f"{path}: {metric} is missing a complete {name!r} row")
            return values[:source_count]

        target_size = max(EXPECTED_VIDEO_COUNT, max(slots, default=-1) + 1)
        available = tuple(index in slots for index in range(target_size))
        sources = _slot_values(source_row, slots, "")
        counts = _slot_values(
            tuple(_parse_count(value, f"{path}:{metric}:count") for value in required("count")),
            slots,
            0,
        )
        missing_values = section.get("missing", ["0"] * source_count)[:source_count]
        missing = _slot_values(
            tuple(_parse_count(value, f"{path}:{metric}:missing") for value in missing_values),
            slots,
            0,
        )
        means = _slot_values(
            tuple(_parse_number(value, f"{path}:{metric}:mean") for value in required("mean")),
            slots,
            0.0,
        )
        stddevs = _slot_values(
            tuple(_parse_number(value, f"{path}:{metric}:stddev") for value in required("stddev")),
            slots,
            0.0,
        )
        kurtosis_values = section.get("kurtosis", [""] * source_count)[:source_count]
        kurtoses = _slot_values(
            tuple(_parse_optional_number(value) for value in kurtosis_values),
            slots,
            None,
        )

        metrics[metric] = MetricSeries(
            sources=tuple(str(value) for value in sources),
            available=available,
            counts=tuple(int(value) for value in counts),
            missing=tuple(int(value) for value in missing),
            means=tuple(float(value) for value in means),
            stddevs=tuple(float(value) for value in stddevs),
            kurtoses=tuple(value if value is None else float(value) for value in kurtoses),
        )
        row_index = scan + 1
    if not metrics:
        raise InputError(f"{path}: no recognized metric sections were found")
    return metrics


def _validate_report(path: Path, modality: str) -> dict[str, MetricSeries]:
    metrics = parse_sectioned_csv(path)
    required = AUDIO_REQUIRED_METRICS if modality == "audio" else VIDEO_METRICS
    missing = [metric for metric in required if metric not in metrics]
    if missing:
        raise InputError(f"{path}: missing required {modality} metrics: {', '.join(missing)}")
    expected_sources = metrics[required[0]].sources
    recognized = AUDIO_METRICS if modality == "audio" else VIDEO_METRICS
    for metric in recognized:
        if metric not in metrics:
            continue
        if metrics[metric].sources != expected_sources:
            raise InputError(f"{path}: recognized metric source order changes at {metric}")
    return metrics


def discover_combined_sources_audited(root: str | Path, modality: str) -> CombinedDiscoveryResult:
    """Find combined reports and retain a reason for every considered candidate."""

    if modality not in {"audio", "video"}:
        raise InputError(f"Unsupported modality: {modality!r}")
    root_path = Path(root).expanduser().resolve()
    if not root_path.exists():
        raise InputError(f"Analysis root does not exist: {root_path}")
    grouped: dict[str, list[tuple[Speaker, Path, str]]] = {}
    rejected: list[DiscoveryEntry] = []
    errors: list[str] = []
    for path in sorted(root_path.rglob("descriptive_statistics.csv"), key=lambda item: str(item).casefold()):
        relative = path.relative_to(root_path)
        if len(relative.parts) == 5:
            domain, supplied_speaker, combined, findings, filename = relative.parts
        elif len(relative.parts) == 6:
            domain, _run_name, supplied_speaker, combined, findings, filename = relative.parts
        elif len(relative.parts) == 4 and normalized(root_path.parent.name) == "emotion":
            domain = "emotion"
            supplied_speaker, combined, findings, filename = relative.parts
        else:
            supplied_speaker = relative.parts[1] if len(relative.parts) > 1 else ""
            rejected.append(
                DiscoveryEntry(
                    modality,
                    normalized(supplied_speaker),
                    supplied_speaker,
                    path.resolve(),
                    "candidate is not a speaker-level combined report",
                )
            )
            continue
        ignored = ("temp", "tmp", "cache", "debug", "raw")
        if any(marker in normalized(part) for part in relative.parts for marker in ignored):
            rejected.append(
                DiscoveryEntry(
                    modality,
                    normalized(supplied_speaker),
                    supplied_speaker,
                    path.resolve(),
                    "generated or temporary candidate was excluded",
                )
            )
            continue
        if (
            normalized(domain) != "emotion"
            or normalized(combined) != "combined"
            or normalized(findings) != "otherfindings"
            or filename.casefold() != "descriptive_statistics.csv"
        ):
            rejected.append(
                DiscoveryEntry(
                    modality,
                    normalized(supplied_speaker),
                    supplied_speaker,
                    path.resolve(),
                    "candidate is not in the required emotion/<speaker>/combined/other_findings layout",
                )
            )
            continue
        try:
            speaker = resolve_speaker(supplied_speaker)
            _validate_report(path, modality)
        except InputError as exc:
            reason = f"invalid candidate: {exc}"
            rejected.append(
                DiscoveryEntry(
                    modality,
                    normalized(supplied_speaker),
                    supplied_speaker,
                    path.resolve(),
                    reason,
                )
            )
            errors.append(str(exc))
            continue
        grouped.setdefault(speaker.speaker_id, []).append((speaker, path, supplied_speaker))
    selected: dict[str, CombinedSource] = {}
    accepted: list[DiscoveryEntry] = []
    for speaker_id, options in grouped.items():
        if len(options) > 1:
            canonical = [option for option in options if normalized(option[2]) == normalized(option[0].display_name)]
            if len(canonical) != 1:
                paths = ", ".join(str(option[1]) for option in options)
                error = f"Ambiguous duplicate reports for {speaker_id}: {paths}"
                errors.append(error)
                for option in options:
                    rejected.append(
                        DiscoveryEntry(
                            modality,
                            option[0].speaker_id,
                            option[0].display_name,
                            option[1].resolve(),
                            f"ambiguous duplicate candidate: {error}",
                        )
                    )
                continue
            selected_path = canonical[0][1]
            for option in options:
                if option[1] != selected_path:
                    rejected.append(
                        DiscoveryEntry(
                            modality,
                            option[0].speaker_id,
                            option[0].display_name,
                            option[1].resolve(),
                            f"duplicate alias report; canonical report selected at {selected_path.resolve()}",
                        )
                    )
            options = canonical
        speaker, path, supplied_speaker = options[0]
        selected[speaker_id] = CombinedSource(modality, normalized(supplied_speaker), speaker.display_name, path)
        accepted.append(
            DiscoveryEntry(
                modality,
                speaker.speaker_id,
                speaker.display_name,
                path.resolve(),
                "accepted speaker-level combined report",
            )
        )
    ordered_ids = sorted(
        selected,
        key=lambda speaker_id: (
            selected[speaker_id].display_name.casefold(),
            speaker_id,
        ),
    )
    ordered_sources = tuple(selected[speaker_id] for speaker_id in ordered_ids)
    accepted_by_path = {entry.path: entry for entry in accepted}
    ordered_accepted = tuple(accepted_by_path[source.report_path.resolve()] for source in ordered_sources)
    return CombinedDiscoveryResult(
        ordered_sources,
        ordered_accepted,
        tuple(rejected),
        tuple(errors),
    )


def discover_combined_sources(root: str | Path, modality: str) -> list[CombinedSource]:
    """Find speaker combined reports under an Analysis root or emotion-domain root."""

    discovery = discover_combined_sources_audited(root, modality)
    if discovery.errors:
        raise InputError(discovery.errors[0])
    return list(discovery.sources)


def _weighted_mean(series: MetricSeries) -> float:
    total = sum(series.counts)
    if total <= 0:
        raise InputError("Cannot calculate a speaker mean with zero observations")
    return sum(count * mean for count, mean in zip(series.counts, series.means)) / total


def _headline_mean(series: MetricSeries, policy: str) -> float:
    """Return the selected speaker headline while excluding unavailable videos."""

    if policy == "weighted":
        return _weighted_mean(series)
    if policy == "equal":
        values = [
            mean
            for mean, available in zip(series.means, series.available)
            if available
        ]
        if not values:
            raise InputError("Cannot calculate a speaker mean without available videos")
        return statistics.fmean(values)
    raise InputError(f"Unsupported headline policy: {policy!r}")


def _display_mean_sd(mean: float, stddev: float, parentheses: bool = False) -> str:
    separator = f"(+/- {stddev:.2f})" if parentheses else f"+/- {stddev:.2f}"
    return f"{mean:.2f} {separator}"


def _display_kurtosis(value: float | None) -> str:
    return "" if value is None else f"{value:.2f}".rstrip("0").rstrip(".")


def _observed_counts(series: MetricSeries) -> list[int]:
    return [count for count, available in zip(series.counts, series.available) if available]


def _report_source_count(reports: Mapping[str, Mapping[str, MetricSeries]]) -> int:
    return max(
        (
            len(series.sources)
            for report in reports.values()
            for series in report.values()
        ),
        default=EXPECTED_VIDEO_COUNT,
    )


def _ordinal_label(index: int) -> str:
    number = index + 1
    if 10 <= number % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(number % 10, "th")
    return f"{number}{suffix}"


def _append_coverage_warning(
    sheet_title: str,
    speaker: Speaker,
    report: Mapping[str, MetricSeries] | None,
    warnings: list[str],
) -> None:
    if report is None:
        warnings.append(f"Missing {sheet_title.lower()} report for {speaker.display_name}; cells are blank.")
        return
    series = next(iter(report.values()))
    available_count = sum(series.available)
    expected_count = max(EXPECTED_VIDEO_COUNT, len(series.available))
    if available_count == expected_count:
        return
    missing = ", ".join(
        f"{index:03d}" for index, available in enumerate(series.available, start=1) if not available
    )
    warnings.append(
        f"Partial {sheet_title.lower()} report for {speaker.display_name}: "
        f"{available_count} of {expected_count} source videos available; missing {missing}."
    )


def _append_optional_audio_warning(
    speaker: Speaker,
    report: Mapping[str, MetricSeries] | None,
    warnings: list[str],
) -> None:
    if report is None:
        return
    missing = [metric for metric in AUDIO_METRICS if metric not in report]
    if missing:
        warnings.append(
            f"Legacy audio report for {speaker.display_name} lacks optional emotions "
            f"{', '.join(missing)}; cells are blank."
        )


def _set_default_font(
    sheet: object, max_row: int, max_column: int, size: float, min_column: int = 1
) -> None:
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=min_column, max_col=max_column):
        for cell in row:
            cell.font = Font(name="Aptos Narrow", size=size)


def _linked_speaker_positions(
    groups: Sequence[_ResolvedSpeakerGroup],
) -> tuple[tuple[_ResolvedSpeakerGroup, Speaker, int], ...]:
    """Map participants to historical columns when possible, then extend dynamically."""

    layout = _linked_layout(groups)
    positions: list[tuple[_ResolvedSpeakerGroup, Speaker, int]] = []
    for group, columns in zip(groups, layout.speaker_columns):
        positions.extend((group, speaker, column) for speaker, column in zip(group.speakers, columns))
    return tuple(positions)


def _linked_layout(groups: Sequence[_ResolvedSpeakerGroup]) -> _LinkedLayout:
    legacy = len(groups) <= len(GROUP_SPEAKER_COLUMNS) and all(
        len(group.speakers) <= len(GROUP_SPEAKER_COLUMNS[0]) for group in groups
    )
    if legacy:
        speaker_columns = tuple(
            tuple(GROUP_SPEAKER_COLUMNS[index][: len(group.speakers)])
            for index, group in enumerate(groups)
        )
        return _LinkedLayout(
            tuple(GROUP_LABEL_COLUMNS[: len(groups)]),
            speaker_columns,
            GROUP_OVERALL_COLUMN,
        )

    label_columns: list[int] = []
    speaker_columns: list[tuple[int, ...]] = []
    next_label = 2
    for index, group in enumerate(groups):
        label_columns.append(next_label)
        first_speaker = 4 if index == 0 else next_label + 1
        columns = tuple(range(first_speaker, first_speaker + len(group.speakers)))
        speaker_columns.append(columns)
        next_label = (columns[-1] if columns else first_speaker) + 1
    overall_column = max(
        [3, *label_columns, *(column for columns in speaker_columns for column in columns)]
    ) + 1
    return _LinkedLayout(tuple(label_columns), tuple(speaker_columns), overall_column)


def _set_linked_column_widths(sheet: object, layout: _LinkedLayout) -> None:
    """Keep both historical and expanded participant blocks readable."""

    for column in dict.fromkeys((*GROUP_LABEL_COLUMNS, *layout.label_columns)):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.column_dimensions["C"].width = 10
    for columns in (*GROUP_SPEAKER_COLUMNS, *layout.speaker_columns):
        for column in columns:
            sheet.column_dimensions[get_column_letter(column)].width = 20
    sheet.column_dimensions[get_column_letter(layout.overall_column)].width = 14


def _write_linked_headlines(
    sheet: object,
    metrics: Sequence[str],
    reports: Mapping[str, Mapping[str, MetricSeries]],
    groups: Sequence[_ResolvedSpeakerGroup],
    warnings: list[str],
    headline_policy: str,
) -> dict[str, CombinedMetricCells]:
    layout = _linked_layout(groups)
    positions = _linked_speaker_positions(groups)
    group_metadata = tuple(
        (group.group_id, group.name, tuple(speaker.speaker_id for speaker in group.speakers))
        for group in groups
    )
    for group_index, group in enumerate(groups):
        label_column = layout.label_columns[group_index]
        sheet.cell(1, label_column, neutralize_spreadsheet_value(group.name))
        sheet.cell(1, label_column).font = Font(name="Aptos Narrow", size=11, bold=True)
        for row, metric in enumerate(metrics, start=2):
            sheet.cell(row, label_column, metric)
            sheet.cell(row, label_column).font = Font(name="Aptos Narrow", size=11)
    for _, speaker, column in positions:
        sheet.cell(1, column, neutralize_spreadsheet_value(speaker.workbook_header))
        sheet.cell(1, column).font = Font(name="Aptos Narrow", size=11)
    sheet.cell(1, layout.overall_column, "Overall")
    sheet.cell(1, layout.overall_column).font = Font(name="Aptos Narrow", size=11, bold=True)

    source_cells: dict[str, CombinedMetricCells] = {}
    for row, metric in enumerate(metrics, start=2):
        valid_cells: list[str] = []
        speaker_cells: list[str] = []
        speaker_ids: list[str] = []
        observations: list[tuple[float | None, ...]] = []
        observation_labels: list[tuple[str, ...]] = []
        for _, speaker, column in positions:
            coordinate = f"{get_column_letter(column)}{row}"
            speaker_cells.append(coordinate)
            speaker_ids.append(speaker.speaker_id)
            cell = sheet[coordinate]
            cell.number_format = "0.00"
            cell.font = Font(name="Aptos Narrow", size=11)
            report = reports.get(speaker.speaker_id)
            if report is None:
                observations.append((None,) * EXPECTED_VIDEO_COUNT)
                observation_labels.append(tuple(f"{index:03d}" for index in range(1, 6)))
                continue
            series = report.get(metric)
            if series is None:
                observations.append((None,) * EXPECTED_VIDEO_COUNT)
                observation_labels.append(tuple(f"{index:03d}" for index in range(1, 6)))
                continue
            cell.value = _headline_mean(series, headline_policy)
            valid_cells.append(coordinate)
            observations.append(
                tuple(mean if available else None for mean, available in zip(series.means, series.available))
            )
            observation_labels.append(
                tuple(source or f"{index:03d}" for index, source in enumerate(series.sources, start=1))
            )
        overall = f"{get_column_letter(layout.overall_column)}{row}"
        sheet[overall] = f"=AVERAGE({','.join(valid_cells)})" if valid_cells else ""
        sheet[overall].number_format = "0.00"
        source_cells[metric] = CombinedMetricCells(
            sheet.title,
            metric,
            overall,
            tuple(speaker_cells),
            tuple(speaker_ids),
            tuple(observations),
            tuple(observation_labels),
            group_metadata,
            tuple(speaker.display_name for _, speaker, _ in positions),
        )
    for _, speaker, _ in positions:
        _append_coverage_warning(sheet.title, speaker, reports.get(speaker.speaker_id), warnings)
        if sheet.title == "Audio":
            _append_optional_audio_warning(speaker, reports.get(speaker.speaker_id), warnings)
    return source_cells


def _write_linked_audio_sheet(
    sheet: object,
    reports: Mapping[str, Mapping[str, MetricSeries]],
    groups: Sequence[_ResolvedSpeakerGroup],
    warnings: list[str],
    headline_policy: str,
) -> dict[str, CombinedMetricCells]:
    positions = _linked_speaker_positions(groups)
    linked_layout = _linked_layout(groups)
    source_count = _report_source_count(reports)
    layout = (
        AUDIO_LAYOUT
        if source_count == EXPECTED_VIDEO_COUNT
        else _measure_layout(
            AUDIO_METRICS,
            count_gap=0,
            count_rows=3 + source_count,
            count_starts_at_heading=False,
            source_count=source_count,
        )
    )
    _set_default_font(sheet, layout.max_row, linked_layout.overall_column, 9, min_column=2)
    _set_linked_column_widths(sheet, linked_layout)
    source_cells = _write_linked_headlines(
        sheet, AUDIO_METRICS, reports, groups, warnings, headline_policy
    )
    sheet.cell(layout.count_heading, 2, "COUNT")
    for row, label in enumerate(
        ("Total", "Average", "Std Dev", *(_ordinal_label(index) for index in range(source_count))),
        start=layout.count_start,
    ):
        sheet[f"C{row}"] = label
    for _, speaker, column in positions:
        report = reports.get(speaker.speaker_id)
        if report is None:
            continue
        series = report[AUDIO_METRICS[0]]
        counts = _observed_counts(series)
        sheet.cell(layout.count_start, column, sum(counts))
        sheet.cell(layout.count_start + 1, column, sum(counts) / len(counts))
        if len(counts) >= 2:
            sheet.cell(layout.count_start + 2, column, statistics.stdev(counts)).number_format = "0.0"
        for source_index, (count, available) in enumerate(
            zip(series.counts, series.available), start=layout.count_start + 3
        ):
            if available:
                sheet.cell(source_index, column, count)
    sheet.cell(layout.detail_heading, 2, "Measures")
    for metric_index, metric in enumerate(AUDIO_METRICS):
        start_row = layout.detail_start + metric_index * source_count
        sheet[f"B{start_row}"] = metric
        for source_index in range(source_count):
            row = start_row + source_index
            sheet[f"C{row}"] = _ordinal_label(source_index)
            for _, speaker, column in positions:
                report = reports.get(speaker.speaker_id)
                series = report.get(metric) if report is not None else None
                if (
                    series is not None
                    and source_index < len(series.available)
                    and series.available[source_index]
                ):
                    sheet.cell(row, column, _display_mean_sd(series.means[source_index], series.stddevs[source_index], True))
    return source_cells


def _write_linked_video_sheet(
    sheet: object,
    reports: Mapping[str, Mapping[str, MetricSeries]],
    groups: Sequence[_ResolvedSpeakerGroup],
    warnings: list[str],
    headline_policy: str,
) -> dict[str, CombinedMetricCells]:
    positions = _linked_speaker_positions(groups)
    linked_layout = _linked_layout(groups)
    source_count = _report_source_count(reports)
    layout = (
        VIDEO_LAYOUT
        if source_count == EXPECTED_VIDEO_COUNT
        else _measure_layout(
            VIDEO_METRICS,
            count_gap=1,
            count_rows=4 + source_count,
            count_starts_at_heading=True,
            source_count=source_count,
            kurtosis_metrics=VIDEO_KURTOSIS_METRICS,
        )
    )
    _set_default_font(sheet, layout.max_row, linked_layout.overall_column, 9, min_column=2)
    _set_linked_column_widths(sheet, linked_layout)
    sheet.freeze_panes = "B2"
    source_cells = _write_linked_headlines(
        sheet, VIDEO_METRICS, reports, groups, warnings, headline_policy
    )
    sheet.cell(layout.count_heading, 2, "COUNT")
    for row, label in enumerate(
        (
            "Total",
            "Average",
            "Std Dev",
            "Missing to Count",
            *(_ordinal_label(index) for index in range(source_count)),
        ),
        start=layout.count_start,
    ):
        sheet[f"C{row}"] = label
    for _, speaker, column in positions:
        report = reports.get(speaker.speaker_id)
        if report is None:
            continue
        series = report[VIDEO_METRICS[0]]
        counts = _observed_counts(series)
        sheet.cell(layout.count_start, column, sum(counts))
        sheet.cell(layout.count_start + 1, column, sum(counts) / len(counts))
        if len(counts) >= 2:
            sheet.cell(layout.count_start + 2, column, statistics.stdev(counts)).number_format = "0.0"
        sheet.cell(layout.count_start + 3, column, sum(series.missing))
        for source_index, (count, available) in enumerate(
            zip(series.counts, series.available), start=layout.count_start + 4
        ):
            if available:
                sheet.cell(source_index, column, count)
    sheet.cell(layout.detail_heading, 2, "Measures")
    parenthesized = {"Anger", "Joy", "Sadness", "Contempt", "Disgust", "Fear", "Surprise"}
    for metric_index, metric in enumerate(VIDEO_DETAIL_METRICS):
        start_row = layout.detail_start + metric_index * source_count
        sheet[f"B{start_row}"] = metric
        for source_index in range(source_count):
            row = start_row + source_index
            sheet[f"C{row}"] = _ordinal_label(source_index)
            for _, speaker, column in positions:
                report = reports.get(speaker.speaker_id)
                if (
                    report is not None
                    and source_index < len(report[metric].available)
                    and report[metric].available[source_index]
                ):
                    series = report[metric]
                    sheet.cell(
                        row,
                        column,
                        _display_mean_sd(
                            series.means[source_index],
                            series.stddevs[source_index],
                            metric in parenthesized,
                        ),
                    )
    sheet.cell(layout.kurtosis_heading, 2, "Kurtosis")
    for row, metric in enumerate(VIDEO_KURTOSIS_METRICS, start=layout.kurtosis_start):
        sheet[f"B{row}"] = metric
        for _, speaker, column in positions:
            report = reports.get(speaker.speaker_id)
            if report is not None:
                series = report[metric]
                sheet.cell(row, column, "/".join(
                    _display_kurtosis(value) if available else ""
                    for value, available in zip(series.kurtoses, series.available)
                ))
    return source_cells


def _write_definition_sheets(book: Workbook) -> None:
    """Port the historical static definition sheets without analytical values."""

    compressed = base64.b64decode("".join(STATIC_DEFINITION_SHEETS.split()))
    values = json.loads(zlib.decompress(compressed).decode("utf-8"))
    text_sheet = book.create_sheet("Domain Def Text")
    speech_sheet = book.create_sheet("Domain Def Speech")
    for sheet in (text_sheet, speech_sheet):
        for coordinate, value in values[sheet.title].items():
            sheet[coordinate] = value
        max_row = 209 if sheet.title == "Domain Def Text" else 19
        max_column = 3 if sheet.title == "Domain Def Text" else 13
        _set_default_font(sheet, max_row, max_column, 11)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in (14, 45, 75, 106, 206):
        text_sheet[f"A{row}"].font = Font(name="Aptos Narrow", size=13.5, bold=True)
    for row in (136, 158, 183, 195):
        text_sheet[f"A{row}"].font = Font(name="Aptos Narrow", size=18, bold=True)
    for coordinate in ("A142", "B142", "C142", "A176", "B176"):
        text_sheet[coordinate].font = Font(name="Aptos Narrow", size=11, bold=True)
        text_sheet[coordinate].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row, height in {
        14: 18.5, 45: 18.5, 75: 18.5, 106: 18.5, 136: 24.25, 143: 59,
        144: 73.75, 145: 73.75, 146: 44.25, 158: 24.25, 176: 29.5,
        177: 59, 178: 59, 179: 103.25, 180: 29.5, 183: 24.25, 195: 24.25,
        206: 18.5,
    }.items():
        text_sheet.row_dimensions[row].height = height
    for row, height in {3: 59, 4: 44.25, 5: 29.5, 6: 73.75, 7: 29.5, 10: 44.25, 11: 103.25, 12: 59, 13: 73.75}.items():
        speech_sheet.row_dimensions[row].height = height
    _write_measure_guide(book)


def _write_measure_guide(book: Workbook) -> None:
    """Document every displayed measure from the same ordered metric contracts."""

    sheet = book.create_sheet("Measure Guide")
    headers = (
        "Section",
        "Modality",
        "Display measure",
        "Imported source label",
        "Workbook sheet",
        "Output range",
        "Transformation/meaning",
    )
    sheet.append(headers)
    cross_scale_note = "Cross-modality scales are not directly comparable without rescaling."

    def add(
        section: str,
        modality: str,
        metrics: Sequence[str],
        output_range: str,
        meaning: str,
        source_labels: Mapping[str, str] | None = None,
    ) -> None:
        source_labels = source_labels or {}
        workbook_sheet = "Text sentiment" if modality == "Text" else modality
        for metric in metrics:
            sheet.append(
                (
                    section,
                    modality,
                    metric,
                    source_labels.get(metric, metric),
                    workbook_sheet,
                    output_range,
                    f"{meaning} {cross_scale_note}",
                )
            )

    add(
        "Emotions",
        "Audio",
        AUDIO_EMOTIONS,
        "0..100",
        "Model probability transformed from source 0..1 to output 0..100.",
        {"Joy": "Happiness"},
    )
    add(
        "Valence",
        "Audio",
        AUDIO_VALENCE,
        "-100..100",
        "Model value transformed from source 0..1 to output -100..100.",
    )
    add(
        "Dimensions",
        "Audio",
        AUDIO_DIMENSIONS,
        "0..100",
        "Model probability transformed from source 0..1 to output 0..100.",
    )
    add("Emotions", "Video", VIDEO_EMOTIONS, "0..100", "Imported iMotions score.")
    add("Sentiment", "Video", VIDEO_SENTIMENT, "0..100", "Imported iMotions score.")
    add("Valence", "Video", VIDEO_VALENCE, "-100..100", "Imported iMotions signed score.")
    add("Dimensions", "Video", VIDEO_DIMENSIONS, "0..100", "Imported iMotions score.")
    add(
        "Sentiment",
        "Text",
        TEXT_SENTIMENT,
        "0..1",
        "Imported transcript sentiment score; legacy valence-named headers are aliases only.",
        {
            "Positive Sentiment": "Positive Sentiment (legacy: Positive valence)",
            "Negative Sentiment": "Negative Sentiment (legacy: Negative valence)",
        },
    )
    add(
        "Dimensions",
        "Text",
        TEXT_DIMENSIONS,
        "-1..1",
        "Imported transcript dimension score.",
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(name="Aptos", size=10, bold=True)
    for column, width in enumerate((14, 12, 30, 42, 20, 14, 78), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _source_speaker(source: CombinedSource) -> Speaker:
    speaker_id = normalized(source.speaker_key) or normalized(source.display_name)
    display_name = " ".join(source.display_name.strip().split())
    if not speaker_id or not display_name:
        raise InputError("Combined report speaker identity must be nonblank")
    return Speaker(speaker_id, display_name, display_name, "", "", (display_name,))


def _reports_for_sources(sources: Sequence[CombinedSource], modality: str) -> dict[str, Mapping[str, MetricSeries]]:
    reports: dict[str, Mapping[str, MetricSeries]] = {}
    for source in sources:
        if source.modality != modality:
            raise InputError(f"Source {source.report_path} does not match {modality} modality")
        speaker = _source_speaker(source)
        if speaker.speaker_id in reports:
            raise InputError(f"Duplicate {modality} report for {speaker.display_name}")
        reports[speaker.speaker_id] = _validate_report(source.report_path, modality)
    return reports


def _profile_workbook_groups(
    profile: AnalysisProfile,
    metadata: SourceMetadata,
) -> tuple[tuple[_ResolvedSpeakerGroup, ...], dict[str, Speaker]]:
    resolved = resolve_analysis_profile(metadata, profile)
    source_by_id = {source.source_id: source for source in metadata.sources}
    participants = {
        source_id: Speaker(
            source_id,
            source_by_id[source_id].title,
            source_by_id[source_id].title,
            str(source_by_id[source_id].user_metadata.get("Country", "")),
            "",
            (source_id, source_by_id[source_id].title),
        )
        for source_id in resolved.ordered_source_ids
    }
    groups = tuple(
        _ResolvedSpeakerGroup(
            group.group_id,
            group.name,
            tuple(participants[source_id] for source_id in group.source_ids),
        )
        for group in resolved.groups
        if group.source_ids
    )
    if not groups:
        raise InputError("The Analysis profile selects no sources for the combined workbook")
    return groups, participants


def _profile_reports(
    reports: Mapping[str, Mapping[str, MetricSeries]],
    metadata: SourceMetadata,
    required_source_ids: Sequence[str],
    modality: str,
) -> dict[str, Mapping[str, MetricSeries]]:
    """Split reports by SourceID and reject incomplete profile coverage."""

    if not reports:
        return {}

    sources_by_speaker: dict[str, list[ManifestSource]] = {}
    for source in metadata.sources:
        if source.selected:
            sources_by_speaker.setdefault(source.speaker_key, []).append(source)
    required = set(required_source_ids)
    split: dict[str, Mapping[str, MetricSeries]] = {}
    for speaker_id, report in reports.items():
        speaker_sources = sources_by_speaker.get(normalized(speaker_id), [])
        if not speaker_sources:
            raise InputError(
                f"{modality} report for {speaker_id!r} does not match any selected "
                "manifest speaker."
            )
        first_series = next(iter(report.values()))
        available_indexes = [
            index for index, available in enumerate(first_series.available) if available
        ]
        labels = tuple(first_series.sources[index] for index in available_indexes)
        try:
            source_ids = map_report_source_ids(metadata, speaker_sources[0].speaker, labels)
        except ValueError as exc:
            raise InputError(str(exc)) from exc
        for source_id, source_index in zip(source_ids, available_indexes):
            if source_id not in required:
                continue
            if source_id in split:
                raise InputError(
                    f"{modality} reports map profiled source more than once: {source_id}"
                )
            split[source_id] = {
                metric: _single_source_series(series, source_id, source_index)
                for metric, series in report.items()
            }
    missing = [source_id for source_id in required_source_ids if source_id not in split]
    if missing:
        raise InputError(
            f"{modality} reports are missing profiled source(s): {', '.join(missing)}"
        )
    return split


def _single_source_series(
    series: MetricSeries,
    source_id: str,
    source_index: int,
) -> MetricSeries:
    if source_index >= len(series.available) or not series.available[source_index]:
        raise InputError(f"Metric report is missing {source_id} at its established source position")
    return MetricSeries(
        sources=(source_id,),
        available=(True,),
        counts=(series.counts[source_index],),
        missing=(series.missing[source_index],),
        means=(series.means[source_index],),
        stddevs=(series.stddevs[source_index],),
        kurtoses=(series.kurtoses[source_index],),
    )


def _profile_text_reports(
    summaries: Mapping[str, TextConstructSummary],
    metadata: SourceMetadata,
    required_source_ids: Sequence[str],
) -> dict[str, TextConstructSummary]:
    """Match text summaries once per visible speaker, never once per source."""

    if not summaries:
        return {}

    profiled: dict[str, TextConstructSummary] = {}
    required = set(required_source_ids)
    selected_by_speaker: dict[str, list[ManifestSource]] = {}
    visible_by_speaker: dict[str, list[ManifestSource]] = {}
    for source in metadata.sources:
        if source.selected:
            selected_by_speaker.setdefault(source.speaker_key, []).append(source)
        if source.source_id in required:
            visible_by_speaker.setdefault(source.speaker_key, []).append(source)

    for speaker_id, summary in summaries.items():
        speaker_key = normalized(speaker_id)
        selected_sources = selected_by_speaker.get(speaker_key, [])
        if not selected_sources:
            raise InputError(
                f"Text summary for {summary.display_name!r} does not match any selected "
                "manifest speaker."
            )
        visible_sources = visible_by_speaker.get(speaker_key, [])
        if not visible_sources:
            continue
        source = visible_sources[0]
        profiled[speaker_key] = TextConstructSummary(
            speaker_id=speaker_key,
            display_name=source.speaker,
            country=str(source.user_metadata.get("Country", summary.country)),
            constructs=summary.constructs,
            source_path=summary.source_path,
        )
    missing = [
        sources[0].speaker
        for speaker_key, sources in visible_by_speaker.items()
        if speaker_key not in profiled
    ]
    if missing:
        raise InputError(
            f"Text summaries are missing profiled speaker(s): {', '.join(missing)}"
        )
    return profiled


def _profile_speaker_groups(
    source_groups: Sequence[_ResolvedSpeakerGroup],
    metadata: SourceMetadata,
    *,
    require_unique_membership: bool,
) -> tuple[_ResolvedSpeakerGroup, ...]:
    """Collapse SourceID groups to speaker grain for text and comparison views."""

    source_by_id = {source.source_id: source for source in metadata.sources}
    memberships: dict[str, str] = {}
    groups: list[_ResolvedSpeakerGroup] = []
    for group in source_groups:
        source_ids_by_speaker: dict[str, list[str]] = {}
        speaker_order: list[str] = []
        for participant in group.speakers:
            source = source_by_id[participant.speaker_id]
            if source.speaker_key not in source_ids_by_speaker:
                speaker_order.append(source.speaker_key)
                source_ids_by_speaker[source.speaker_key] = []
            source_ids_by_speaker[source.speaker_key].append(source.source_id)
        speakers: list[Speaker] = []
        for speaker_key in speaker_order:
            previous_group = memberships.get(speaker_key)
            if require_unique_membership and previous_group is not None:
                display_name = source_by_id[source_ids_by_speaker[speaker_key][0]].speaker
                raise InputError(
                    f"Speaker-level text for {display_name} cannot be split across "
                    f"Analysis groups {previous_group!r} and {group.name!r}."
                )
            memberships.setdefault(speaker_key, group.name)
            source_ids = tuple(source_ids_by_speaker[speaker_key])
            source = source_by_id[source_ids[0]]
            speakers.append(
                Speaker(
                    speaker_key,
                    source.speaker,
                    source.speaker,
                    str(source.user_metadata.get("Country", "")),
                    "",
                    source_ids,
                )
            )
        if speakers:
            groups.append(_ResolvedSpeakerGroup(group.group_id, group.name, tuple(speakers)))
    return tuple(groups)


def protected_manual_discovery_directories(
    anchor: str | Path | None = None,
) -> tuple[Path, ...]:
    """Find every existing protected reference directory above this checkout."""

    start = Path(anchor).expanduser().resolve() if anchor is not None else Path(__file__).resolve()
    protected: list[Path] = []
    for ancestor in (start, *start.parents):
        candidate = ancestor / "Statistics_Manual_Discovery"
        if candidate.is_dir():
            resolved = candidate.resolve()
            if resolved not in protected:
                protected.append(resolved)
    return tuple(protected)


def _validate_output_destination(
    destination: Path, sources_by_modality: Mapping[str, Sequence[CombinedSource]]
) -> None:
    for sources in sources_by_modality.values():
        for source in sources:
            if destination == source.report_path.expanduser().resolve():
                raise InputError(f"Output destination is a source report: {destination}")
    for protected_directory in protected_manual_discovery_directories():
        try:
            destination.relative_to(protected_directory)
        except ValueError:
            pass
        else:
            raise InputError(
                "Output destination is inside protected Statistics_Manual_Discovery: "
                f"{destination}"
            )


def _validated_text_summaries(
    summaries: Sequence[TextConstructSummary],
) -> dict[str, TextConstructSummary]:
    validated: dict[str, TextConstructSummary] = {}
    for summary in summaries:
        speaker_id = normalized(summary.speaker_id) or normalized(summary.display_name)
        display_name = " ".join(summary.display_name.strip().split())
        if not speaker_id or not display_name:
            raise InputError("Text summary speaker identity must be nonblank")
        speaker = Speaker(
            speaker_id,
            display_name,
            display_name,
            summary.country,
            "",
            (display_name,),
        )
        if speaker.speaker_id in validated:
            raise InputError(f"Duplicate text summary for {speaker.display_name}")
        missing = [construct for construct in TEXT_CONSTRUCTS if construct not in summary.constructs]
        if missing:
            raise InputError(
                f"Text summary for {speaker.display_name} is missing: {', '.join(missing)}"
            )
        constructs: dict[str, float | None] = {}
        for construct in TEXT_CONSTRUCTS:
            raw_value = summary.constructs[construct]
            if raw_value is None:
                constructs[construct] = None
                continue
            value = float(raw_value)
            if not math.isfinite(value):
                raise InputError(
                    f"Text summary for {speaker.display_name} contains a non-finite {construct}"
                )
            constructs[construct] = value
        validated[speaker.speaker_id] = TextConstructSummary(
            speaker_id=speaker.speaker_id,
            display_name=speaker.display_name,
            country=speaker.country,
            constructs=constructs,
            source_path=Path(summary.source_path).expanduser().resolve(),
        )
    return validated


def _write_text_construct_sheet(
    sheet: object,
    summaries: Mapping[str, TextConstructSummary],
    groups: Sequence[_ResolvedSpeakerGroup],
    warnings: list[str],
) -> dict[str, CombinedMetricCells]:
    """Write imported lexical constructs while keeping them out of inference inputs."""

    positions = _linked_speaker_positions(groups)
    layout = _linked_layout(groups)
    _set_default_font(sheet, 10, layout.overall_column, 10, min_column=2)
    _set_linked_column_widths(sheet, layout)
    group_metadata = tuple(
        (group.group_id, group.name, tuple(speaker.speaker_id for speaker in group.speakers))
        for group in groups
    )
    for group_index, group in enumerate(groups):
        label_column = layout.label_columns[group_index]
        sheet.cell(1, label_column, neutralize_spreadsheet_value(group.name)).font = Font(
            name="Aptos Narrow", size=11, bold=True
        )
        for row, construct in enumerate(TEXT_CONSTRUCTS, start=2):
            sheet.cell(row, label_column, construct)
    for _, speaker, column in positions:
        sheet.cell(1, column, neutralize_spreadsheet_value(speaker.workbook_header))
    sheet.cell(1, layout.overall_column, "Overall").font = Font(
        name="Aptos Narrow", size=11, bold=True
    )
    sheet["B8"] = (
        "Transcript lexical constructs are imported proxies on their documented text scales; "
        "they are not calibrated facial or audio emotion measurements."
    )
    sheet["B8"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(
        start_row=8,
        start_column=2,
        end_row=9,
        end_column=layout.overall_column,
    )

    text_cells: dict[str, CombinedMetricCells] = {}
    speaker_ids = tuple(speaker.speaker_id for _, speaker, _ in positions)
    for row, construct in enumerate(TEXT_CONSTRUCTS, start=2):
        valid_cells: list[str] = []
        speaker_cells: list[str] = []
        for _, speaker, column in positions:
            coordinate = f"{get_column_letter(column)}{row}"
            speaker_cells.append(coordinate)
            cell = sheet[coordinate]
            cell.number_format = "0.000"
            summary = summaries.get(speaker.speaker_id)
            value = summary.constructs.get(construct) if summary is not None else None
            if value is not None:
                cell.value = value
                valid_cells.append(coordinate)
        overall = f"{get_column_letter(layout.overall_column)}{row}"
        sheet[overall] = f"=AVERAGE({','.join(valid_cells)})" if valid_cells else ""
        sheet[overall].number_format = "0.000"
        text_cells[f"Text|{construct}"] = CombinedMetricCells(
            sheet.title,
            construct,
            overall,
            tuple(speaker_cells),
            speaker_ids,
            speaker_groups=group_metadata,
            speaker_display_names=tuple(
                speaker.display_name for _, speaker, _ in positions
            ),
        )
    for _, speaker, _ in positions:
        if speaker.speaker_id not in summaries:
            warnings.append(
                f"Missing text construct summary for {speaker.display_name}; cells are blank."
            )
    sheet.freeze_panes = "B2"
    return text_cells


_COMPARISON_ROWS = (
    *(("Emotions", metric, metric if metric in VIDEO_EMOTIONS else None,
       metric if metric in AUDIO_EMOTIONS else None, None, "EAF4EC")
      for metric in (*AUDIO_EMOTIONS, "Confusion") if metric in set(AUDIO_EMOTIONS + VIDEO_EMOTIONS)),
    ("Sentiment", "Sentimentality", "Sentimentality", None, None, "E8F0FA"),
    ("Sentiment", "Positive Sentiment", None, None, "Positive Sentiment", "E8F0FA"),
    ("Sentiment", "Negative Sentiment", None, None, "Negative Sentiment", "E8F0FA"),
    ("Valence", "Valence", "Valence", "Valence", None, "F8ECEC"),
    ("Valence", "Adaptive Valence", "Adaptive Valence", None, None, "F8ECEC"),
    ("Dimensions", "Arousal", None, "Arousal", "Arousal / Activation", "FFF6E3"),
    ("Dimensions", "Dominance", None, "Dominance", "Dominance / Power", "FFF6E3"),
    ("Dimensions", "Engagement", "Engagement", None, None, "FFF6E3"),
    ("Dimensions", "Adaptive Engagement", "Adaptive Engagement", None, None, "FFF6E3"),
    (
        "Dimensions",
        "Affiliation / Social orientation",
        None,
        None,
        "Affiliation / Social orientation",
        "FFF6E3",
    ),
)


def _speaker_metric_references(
    source_cells: Mapping[str, CombinedMetricCells],
    modality: str,
    metric: str,
    speaker: Speaker,
) -> tuple[str, ...]:
    cells = source_cells.get(f"{modality}|{metric}")
    if cells is None:
        return ()
    candidate_ids = tuple(dict.fromkeys((speaker.speaker_id, *speaker.aliases)))
    coordinates = tuple(
        cells.speaker_cells[index]
        for index, speaker_id in enumerate(cells.speaker_ids)
        if speaker_id in candidate_ids
    )
    sheet = cells.sheet.replace(chr(39), chr(39) * 2)
    return tuple(f"'{sheet}'!{coordinate}" for coordinate in coordinates)


def _comparison_metric_formula(
    source_cells: Mapping[str, CombinedMetricCells],
    modality: str,
    metric: str | None,
    speaker: Speaker,
) -> str | None:
    if metric is None:
        return None
    references = _speaker_metric_references(source_cells, modality, metric, speaker)
    if not references:
        return None
    if len(references) == 1:
        reference = references[0]
        return f'=IF(ISNUMBER({reference}),{reference},"Unavailable")'
    joined = ",".join(references)
    return f'=IF(COUNT({joined})>0,AVERAGE({joined}),"Unavailable")'


def _write_construct_comparison_sheet(
    book: Workbook,
    source_cells: Mapping[str, CombinedMetricCells],
    groups: Sequence[_ResolvedSpeakerGroup],
) -> None:
    """Create the at-a-glance construct tables using formula-linked source means."""

    sheet = book.create_sheet("Construct Comparison")
    largest_group = max((len(group.speakers) for group in groups), default=0)
    table_count = max(3, largest_group)
    table_starts = tuple(1 + (index * 5) for index in range(table_count))
    last_column = table_starts[-1] + 3
    thin_border = Border(
        left=Side(style="thin", color="B8B8B8"),
        right=Side(style="thin", color="B8B8B8"),
        top=Side(style="thin", color="B8B8B8"),
        bottom=Side(style="thin", color="B8B8B8"),
    )

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet["A1"] = "Multimodal construct comparison by speaker"
    sheet["A1"].font = Font(name="Aptos", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F1F1F")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sheet["A2"] = (
        "Video and Audio values are formula-linked to the selected speaker means. "
        "Text remains blank unless compatible text results are imported."
    )
    sheet["A2"].font = Font(name="Aptos", size=10, italic=True, color="333333")
    sheet["A2"].fill = PatternFill("solid", fgColor="ECECEC")
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 24

    group_row = 4
    for group in groups:
        sheet.merge_cells(
            start_row=group_row,
            start_column=1,
            end_row=group_row,
            end_column=last_column,
        )
        group_cell = sheet.cell(
            group_row, 1, neutralize_spreadsheet_value(group.name)
        )
        group_cell.font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
        group_cell.fill = PatternFill("solid", fgColor="2B2B2B")
        group_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[group_row].height = 24

        for table_start, speaker in zip(table_starts, group.speakers):
            title_row = group_row + 1
            header_row = group_row + 2
            first_data_row = group_row + 3
            table_end = table_start + 3
            sheet.merge_cells(
                start_row=title_row,
                start_column=table_start,
                end_row=title_row,
                end_column=table_end,
            )
            title = sheet.cell(
                title_row,
                table_start,
                neutralize_spreadsheet_value(speaker.display_name),
            )
            title.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
            title.fill = PatternFill("solid", fgColor="444444")
            title.alignment = Alignment(vertical="center")
            sheet.row_dimensions[title_row].height = 23

            for offset, header in enumerate(("Section / measure", "Video", "Audio", "Text")):
                cell = sheet.cell(header_row, table_start + offset, header)
                cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="666666")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
            sheet.row_dimensions[header_row].height = 31

            for row_offset, (
                section,
                label,
                video_metric,
                audio_metric,
                text_metric,
                fill_color,
            ) in enumerate(_COMPARISON_ROWS):
                row = first_data_row + row_offset
                sheet.cell(row, table_start, f"{section}: {label}")
                sheet.cell(row, table_start + 1).value = _comparison_metric_formula(
                    source_cells, "Video", video_metric, speaker
                )
                sheet.cell(row, table_start + 2).value = _comparison_metric_formula(
                    source_cells, "Audio", audio_metric, speaker
                )
                sheet.cell(row, table_start + 3).value = _comparison_metric_formula(
                    source_cells, "Text", text_metric, speaker
                )
                for column in range(table_start, table_end + 1):
                    cell = sheet.cell(row, column)
                    cell.font = Font(name="Aptos", size=9, bold=column == table_start)
                    cell.fill = PatternFill(
                        "solid", fgColor="FAFAFA" if column == table_end else fill_color
                    )
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = thin_border
                sheet.row_dimensions[row].height = 43

        group_row += len(_COMPARISON_ROWS) + 4

    for table_start in table_starts:
        sheet.column_dimensions[get_column_letter(table_start)].width = 22
        sheet.column_dimensions[get_column_letter(table_start + 1)].width = 27
        sheet.column_dimensions[get_column_letter(table_start + 2)].width = 25
        sheet.column_dimensions[get_column_letter(table_start + 3)].width = 15
    for table_start in table_starts[:-1]:
        sheet.column_dimensions[get_column_letter(table_start + 4)].width = 3
    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False


def build_combined_workbook(
    sources_by_modality: Mapping[str, Sequence[CombinedSource]],
    output_path: str | Path,
    *,
    headline_policy: str = "weighted",
    speaker_groups: Sequence[SpeakerGroupDefinition] | None = None,
    analysis_profile: AnalysisProfile | None = None,
    include_construct_comparison: bool = False,
    text_summaries: Sequence[TextConstructSummary] = (),
) -> CombinedWorkbookResult:
    """Build available historical quantitative sheets and static definitions."""

    if headline_policy not in {"weighted", "equal"}:
        raise InputError(f"Unsupported headline policy: {headline_policy!r}")
    unexpected = set(sources_by_modality) - {"audio", "video"}
    if unexpected:
        raise InputError(f"Unsupported modalities: {', '.join(sorted(unexpected))}")
    if speaker_groups is not None and not speaker_groups:
        raise InputError("At least one speaker group is required when groups are supplied")
    if analysis_profile is not None and speaker_groups is not None:
        raise InputError("Use either an Analysis profile or legacy speaker groups, not both")
    destination = Path(output_path).expanduser().resolve()
    _validate_output_destination(destination, sources_by_modality)
    text_reports = _validated_text_summaries(text_summaries)
    if any(destination == summary.source_path for summary in text_reports.values()):
        raise InputError(f"Output destination is a source report: {destination}")
    audio_reports = _reports_for_sources(sources_by_modality.get("audio", ()), "audio")
    video_reports = _reports_for_sources(sources_by_modality.get("video", ()), "video")
    if not audio_reports and not video_reports and not text_reports:
        raise InputError("At least one Audio, Video, or Text source is required")
    if analysis_profile is not None:
        profile_metadata = load_source_metadata(
            analysis_profile.source_manifest,
            expected_sha256=analysis_profile.source_manifest_sha256,
        )
        groups, participants = _profile_workbook_groups(analysis_profile, profile_metadata)
        required_source_ids = tuple(participants)
        audio_reports = _profile_reports(
            audio_reports, profile_metadata, required_source_ids, "Audio"
        )
        video_reports = _profile_reports(
            video_reports, profile_metadata, required_source_ids, "Video"
        )
        text_reports = _profile_text_reports(
            text_reports, profile_metadata, required_source_ids
        )
        comparison_groups = _profile_speaker_groups(
            groups,
            profile_metadata,
            require_unique_membership=bool(text_reports),
        )
        text_groups = comparison_groups
    else:
        speaker_catalog: dict[str, Speaker] = {}
        for sources in sources_by_modality.values():
            for source in sources:
                speaker = _source_speaker(source)
                speaker_catalog[speaker.speaker_id] = speaker
        for summary in text_reports.values():
            speaker_catalog.setdefault(
                summary.speaker_id,
                Speaker(
                    summary.speaker_id,
                    summary.display_name,
                    summary.display_name,
                    summary.country,
                    "",
                    (summary.display_name,),
                ),
            )
        groups = (
            _resolve_speaker_groups(speaker_groups, speaker_catalog)
            if speaker_groups is not None
            else _default_speaker_groups(speaker_catalog)
        )
        text_groups = groups
        comparison_groups = groups
    book = Workbook()
    book.remove(book.active)
    warnings: list[str] = []
    source_cells: dict[str, CombinedMetricCells] = {}
    quantitative_sheets: list[str] = []
    if audio_reports:
        audio = book.create_sheet("Audio")
        source_cells.update(
            {
                f"Audio|{metric}": cells
                for metric, cells in _write_linked_audio_sheet(
                    audio, audio_reports, groups, warnings, headline_policy
                ).items()
            }
        )
        quantitative_sheets.append("Audio")
    _write_definition_sheets(book)
    if video_reports:
        video = book.create_sheet("Video", book.index(book["Domain Def Text"]) + 1)
        source_cells.update(
            {
                f"Video|{metric}": cells
                for metric, cells in _write_linked_video_sheet(
                    video, video_reports, groups, warnings, headline_policy
                ).items()
            }
        )
        quantitative_sheets.append("Video")
    text_sheet = book.create_sheet("Text sentiment")
    text_cells = (
        _write_text_construct_sheet(text_sheet, text_reports, text_groups, warnings)
        if text_reports
        else {}
    )
    if include_construct_comparison:
        _write_construct_comparison_sheet(
            book,
            {**source_cells, **text_cells},
            comparison_groups,
        )
    book.active = 0
    book.calculation.fullCalcOnLoad = True
    book.calculation.forceFullCalc = True
    book.calculation.calcId = 191029
    destination.parent.mkdir(parents=True, exist_ok=True)
    book.save(destination)
    return CombinedWorkbookResult(destination, tuple(quantitative_sheets), source_cells, tuple(warnings))
