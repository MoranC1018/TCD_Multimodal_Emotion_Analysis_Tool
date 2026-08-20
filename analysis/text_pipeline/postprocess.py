#!/usr/bin/env python3
"""RockSteady segment-level text postprocessing.

This module is intentionally built for the current text workflow:

    Whisper JSON -> one RockSteady text file per segment -> one RockSteady
    Total-mode CSV per video.

RockSteady Total counts are kept as counts. The postprocessor then derives
word-level proportions, segment presence percentages, video summaries, speaker
summaries, descriptor statistics, and SVG graphs. This avoids treating text
word frequencies as if they were face/audio emotion intensities.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import math
import re
import shutil
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean
from typing import Iterable, Mapping, Sequence

from analysis.histograms import default_output_root, student_t_two_tailed_p
from analysis.text_pipeline.contracts import CATEGORIES, Category, discover_categories
from analysis.text_pipeline.provenance import (
    SegmentAlignmentContract,
    UpstreamProvenance,
    inspect_upstream_provenance,
    load_segment_alignment_contract,
    sha256_file,
)
from analysis.text_pipeline.readable import write_readable_tables
from analysis.text_pipeline.reports import (
    ReportCategory,
    SpeakerQuickRow,
    TextReportModel,
    write_report_files,
)
from analysis.text_pipeline.ownership import (
    assert_publishable_output,
    normalize_run_id,
    text_output_lock,
    validate_output_boundaries,
    write_output_owner,
)
from analysis.text_pipeline.transaction import replace_output_dir
from analysis.text_pipeline.distribution_comparisons import (
    DEFAULT_PERMUTATIONS,
    DEFAULT_RANDOM_SEED,
    TextSegmentObservation,
    write_text_mean_comparisons,
)
from processing.io_utils import atomic_write_json, make_staging_directory
from processing.text_analysis.selection import is_text_metadata_json
from spreadsheet_safety import SpreadsheetSafeWriter, neutralize_spreadsheet_value


COUNT_HISTOGRAM_BINS = ("0", "1", "2", "3+")
DEFAULT_SEGMENT_SAMPLE_COUNTS = (3, 5, 7, 9)
SEGMENT_ALIGNMENT_POLICIES = ("error", "reconcile")
TEXT_VARIANTS = ("original", "eng", "selected", "extra")


@dataclass
class SegmentRecord:
    country: str
    speaker: str
    speaker_id: str
    video: str
    segment_id: int
    source_segment_index: int
    source_segment_id: object
    title: str
    terms: float
    counts: dict[str, float | None]
    start_sec: float
    end_sec: float
    segment_text: str
    whisper_language: str
    text_language: str
    whisper_word_count: int
    categories: tuple[Category, ...]
    source_id: str = ""
    rocksteady_row_available: bool = True

    @property
    def duration_sec(self) -> float:
        return max(0.0, self.end_sec - self.start_sec)

    @property
    def valence_score(self) -> float | str:
        positive = self.counts.get("positive")
        negative = self.counts.get("negative")
        if positive is None or negative is None or positive + negative <= 0:
            return ""
        return (positive - negative) / (positive + negative)


@dataclass
class VideoAnalysis:
    country: str
    speaker: str
    speaker_id: str
    video: str
    path: Path
    segments: list[SegmentRecord]
    video_info: dict[str, str]
    whisper_path: Path
    categories: tuple[Category, ...]
    available_categories: set[str]
    rocksteady_row_count: int
    alignment_contract: SegmentAlignmentContract
    source_whisper_segment_count: int
    source_id: str = ""
    missing_segment_ids: tuple[int, ...] = ()
    ignored_segment_ids: tuple[int, ...] = ()
    alignment_policy: str = "error"
    summary: dict[str, object] = field(default_factory=dict)


@dataclass
class TextAnalysisResult:
    run_id: str
    input_dir: Path
    output_dir: Path
    csv_count: int
    segment_count: int
    video_summary_path: Path
    speaker_summary_path: Path
    descriptor_path: Path
    alignment_audit_path: Path
    output_manifest_path: Path
    graph_paths: list[Path]


def analyse_text_segments_folder(
    input_folder: str | Path,
    output_root: str | Path | None = None,
    *,
    whisper_root: str | Path | None = None,
    prepare_root: str | Path | None = None,
    write_graphs: bool = True,
    segment_sample_counts: Sequence[int] = DEFAULT_SEGMENT_SAMPLE_COUNTS,
    segment_alignment: str = "error",
    text_language: str = "original",
    run_id: str | None = None,
    output_variant: str | None = None,
    _reported_output_root: str | Path | None = None,
) -> TextAnalysisResult:
    """Analyse and atomically publish one RockSteady output variant.

    ``output_variant`` is normally inferred from the input/output folder.  It
    is exposed for orchestrators that build a complete selected/extra pair in
    a temporary parent.  ``_reported_output_root`` is an internal escape hatch
    used by that orchestrator so human-readable reports show the final path,
    not the hidden transaction directory.
    """

    sample_counts = normalize_segment_sample_counts(segment_sample_counts)
    alignment_policy = normalize_segment_alignment_policy(segment_alignment)
    selected_text_language = normalize_text_language(text_language)
    input_dir = resolve_segment_input_folder(input_folder)
    if not input_dir.exists() or not input_dir.is_dir():
        raise NotADirectoryError(f"Input folder does not exist: {input_dir}")
    csv_paths = discover_segment_csvs(input_dir)
    if not csv_paths:
        raise ValueError(f"No RockSteady segment CSV files found under {input_dir}")

    requested_output_dir = resolve_segment_output_folder(input_dir, output_root)
    resolved_whisper_root = resolve_whisper_root(whisper_root, input_dir=input_dir)
    resolved_prepare_root = resolve_prepare_root(prepare_root)
    output_dir = validate_output_boundaries(
        requested_output_dir,
        (input_dir, resolved_whisper_root, resolved_prepare_root),
    )
    reported_output_dir = (
        Path(_reported_output_root).expanduser().resolve()
        if _reported_output_root is not None
        else output_dir
    )
    inferred_variant = detect_text_variant(input_dir) or detect_text_variant(output_dir)
    if output_variant is not None:
        normalized_variant = str(output_variant).strip().casefold()
        if normalized_variant not in TEXT_VARIANTS:
            raise ValueError(
                f"Unknown text output variant {output_variant!r}; choose one of "
                f"{', '.join(TEXT_VARIANTS)}."
            )
        effective_variant = normalized_variant
    else:
        effective_variant = inferred_variant or "custom"
    effective_run_id = normalize_run_id(run_id)

    with text_output_lock(output_dir, scope="variant", variant=effective_variant):
        ownership = assert_publishable_output(
            output_dir,
            scope="variant",
            variant=effective_variant,
        )
        return _analyse_text_segments_folder_locked(
            input_dir,
            output_dir,
            reported_output_dir=reported_output_dir,
            resolved_whisper_root=resolved_whisper_root,
            resolved_prepare_root=resolved_prepare_root,
            prepare_root_was_explicit=prepare_root is not None,
            write_graphs=write_graphs,
            sample_counts=sample_counts,
            alignment_policy=alignment_policy,
            selected_text_language=selected_text_language,
            run_id=effective_run_id,
            output_variant=effective_variant,
            previous_ownership_state=ownership.state,
            csv_paths=csv_paths,
        )


def _analyse_text_segments_folder_locked(
    input_dir: Path,
    output_dir: Path,
    *,
    reported_output_dir: Path,
    resolved_whisper_root: Path,
    resolved_prepare_root: Path,
    prepare_root_was_explicit: bool,
    write_graphs: bool,
    sample_counts: Sequence[int],
    alignment_policy: str,
    selected_text_language: str,
    run_id: str,
    output_variant: str,
    previous_ownership_state: str,
    csv_paths: Sequence[Path],
) -> TextAnalysisResult:
    """Generate a variant while its family-level process lock is held."""

    upstream_provenance = inspect_upstream_provenance(input_dir, csv_paths)
    categories = discover_categories(
        csv_paths,
        expected_source_names=upstream_provenance.expected_categories,
    )
    whisper_index = build_whisper_index(resolved_whisper_root)
    videos = [
        read_rocksteady_segment_csv(
            path,
            input_dir,
            whisper_index,
            categories=categories,
            prepare_root=(
                resolved_prepare_root
                if prepare_root_was_explicit or upstream_provenance.verified
                else None
            ),
            upstream_verified=upstream_provenance.verified,
            segment_alignment=alignment_policy,
            text_language=selected_text_language,
        )
        for path in csv_paths
    ]

    for video in videos:
        video.summary = build_video_summary(video)

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    staging_dir = make_staging_directory(output_dir.parent, f".{output_dir.name}_")
    try:
        staging_graph_paths: list[Path] = []
        write_segment_outputs(staging_dir, videos)
        write_dict_rows(staging_dir / "video_level_summary.csv", [video.summary for video in videos])
        speaker_rows = build_speaker_summary_rows(videos)
        write_dict_rows(staging_dir / "speaker_level_summary.csv", speaker_rows)
        write_descriptor_statistics(staging_dir / "descriptor_statistics_by_video.csv", videos)
        write_segment_alignment_audit(staging_dir / "segment_alignment_audit.csv", videos)
        write_text_mean_comparisons(
            staging_dir,
            text_segment_observations(videos, categories),
            {category.key: category.display for category in categories},
        )
        readable_summary = write_readable_tables(
            staging_dir,
            variant=output_variant,
        )

        if write_graphs:
            staging_graph_paths.extend(
                write_all_graphs(
                    staging_dir,
                    videos,
                    speaker_rows,
                    segment_sample_counts=sample_counts,
                )
            )

        report_model = build_report_model(
            input_dir,
            reported_output_dir,
            resolved_whisper_root,
            resolved_prepare_root,
            videos,
            speaker_rows,
            staging_graph_paths,
            segment_sample_counts=sample_counts,
            provenance_status=upstream_provenance.status,
        )
        write_report_files(staging_dir, report_model)
        write_output_owner(
            staging_dir,
            scope="variant",
            variant=output_variant,
            run_id=run_id,
        )
        write_output_manifest(
            staging_dir,
            input_dir,
            resolved_whisper_root,
            resolved_prepare_root,
            videos,
            staging_graph_paths,
            categories=categories,
            upstream_provenance=upstream_provenance,
            write_graphs=write_graphs,
            segment_sample_counts=sample_counts,
            segment_alignment=alignment_policy,
            text_language=selected_text_language,
            run_id=run_id,
            output_variant=output_variant,
            previous_ownership_state=previous_ownership_state,
            readable_summary=readable_summary,
        )
        validate_output_boundaries(
            output_dir,
            (input_dir, resolved_whisper_root, resolved_prepare_root),
        )
        replace_output_dir(staging_dir, output_dir)
    except BaseException:
        if staging_dir.exists():
            shutil.rmtree(staging_dir, ignore_errors=True)
        raise

    graph_paths = [output_dir / path.relative_to(staging_dir) for path in staging_graph_paths]
    video_summary_path = output_dir / "video_level_summary.csv"
    speaker_summary_path = output_dir / "speaker_level_summary.csv"
    descriptor_path = output_dir / "descriptor_statistics_by_video.csv"
    alignment_audit_path = output_dir / "segment_alignment_audit.csv"
    output_manifest_path = output_dir / "output_manifest.json"

    return TextAnalysisResult(
        run_id=run_id,
        input_dir=input_dir,
        output_dir=output_dir,
        csv_count=len(videos),
        segment_count=sum(len(video.segments) for video in videos),
        video_summary_path=video_summary_path,
        speaker_summary_path=speaker_summary_path,
        descriptor_path=descriptor_path,
        alignment_audit_path=alignment_audit_path,
        output_manifest_path=output_manifest_path,
        graph_paths=graph_paths,
    )


def resolve_segment_input_folder(input_folder: str | Path) -> Path:
    candidate = Path(input_folder)
    if candidate.exists():
        return candidate.resolve()
    if not candidate.is_absolute():
        current_aliases = {
            "selected": Path("processing/text_analysis/output/current/rocksteady/core"),
            "core": Path("processing/text_analysis/output/current/rocksteady/core"),
            "extra": Path("processing/text_analysis/output/current/rocksteady/all"),
            "all": Path("processing/text_analysis/output/current/rocksteady/all"),
        }
        current = current_aliases.get(candidate.as_posix().strip("/").casefold())
        if current is not None and current.exists():
            return current.resolve()
        rooted = Path("processing/text_analysis/parse_output") / candidate
        if rooted.exists():
            return rooted.resolve()
    return candidate.resolve()


def resolve_segment_output_folder(input_dir: Path, output_root: str | Path | None) -> Path:
    if output_root:
        # Preserve the lexical caller path until path-safety validation can
        # inspect symlink/junction/reparse components before resolution.
        return Path(output_root).expanduser()
    base = default_output_root() / "text" / "text_output"
    variant = detect_text_variant(input_dir)
    return base / variant if variant else base


def detect_text_variant(input_dir: Path) -> str | None:
    name = input_dir.name.strip().lower()
    if name == "rocksteady output" and input_dir.parent.name.strip().lower() == "extra":
        return "extra"
    if name == "core":
        return "selected"
    if name == "all":
        return "extra"
    return name if name in TEXT_VARIANTS else None


def infer_default_whisper_root(input_dir: Path | None = None) -> Path:
    root = Path("processing/text_analysis/output/current/transcripts")
    variant = detect_text_variant(input_dir) if input_dir is not None else None
    if variant in {"selected", "extra"}:
        return Path("processing/text_analysis/output/current/selected_transcripts")
    return root / variant if variant else root


def resolve_whisper_root(whisper_root: str | Path | None, *, input_dir: Path | None = None) -> Path:
    if whisper_root is None:
        root = infer_default_whisper_root(input_dir)
    else:
        root = Path(whisper_root)
    root = root.resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"Whisper JSON root does not exist: {root}")
    return root


def resolve_prepare_root(prepare_root: str | Path | None) -> Path:
    root = Path(
        prepare_root
        if prepare_root is not None
        else "processing/text_analysis/output/current/prepared_segments"
    ).resolve()
    if not root.is_dir():
        # Legacy standalone inputs may not have preparation manifests.  Keep
        # the configured location for explicit provenance reporting; verified
        # inputs will fail with the exact expected per-video path later.
        return root
    return root


def normalize_segment_alignment_policy(value: str) -> str:
    policy = str(value or "error").strip().lower()
    if policy not in SEGMENT_ALIGNMENT_POLICIES:
        raise ValueError(
            f"Unknown segment alignment policy {value!r}; "
            f"choose one of {', '.join(SEGMENT_ALIGNMENT_POLICIES)}."
        )
    return policy


def normalize_text_language(value: str) -> str:
    language = str(value or "original").strip().lower()
    if not language:
        raise ValueError("Text language cannot be empty.")
    return language


def normalize_identity(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", ascii_value.lower())


def build_whisper_index(root: Path) -> dict[tuple[str, str, str], Path]:
    index: dict[tuple[str, str, str], Path] = {}
    for path in sorted(root.rglob("*.json")):
        if is_text_metadata_json(path, root):
            continue
        country, speaker = validate_text_asset_path(path, root, asset_label="Whisper JSON")
        key = (
            normalize_identity(country),
            normalize_identity(speaker),
            normalize_identity(path.stem),
        )
        if key in index:
            raise ValueError(
                f"Duplicate Whisper JSON identity for {country}/{speaker}/{path.stem}: "
                f"{index[key]} and {path}"
            )
        index[key] = path
    if not index:
        raise ValueError(f"No Whisper JSON files found under {root}")
    return index


def discover_segment_csvs(input_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in input_dir.rglob("*.csv")
        if path.is_file()
    )


def read_rocksteady_segment_csv(
    path: Path,
    input_dir: Path,
    whisper_index: Mapping[tuple[str, str, str], Path],
    *,
    categories: tuple[Category, ...] = CATEGORIES,
    prepare_root: Path | None = None,
    upstream_verified: bool = False,
    segment_alignment: str = "error",
    text_language: str = "original",
) -> VideoAnalysis:
    alignment_policy = normalize_segment_alignment_policy(segment_alignment)
    selected_text_language = normalize_text_language(text_language)
    rows = read_csv_flexible(path)
    validate_rocksteady_total_rows(path, rows, categories=categories)
    video = path.stem
    info = parse_video_name(video)
    country, speaker = validate_text_asset_path(path, input_dir, asset_label="RockSteady CSV")
    speaker_id = f"{country}/{speaker}" if country else speaker
    whisper_key = (
        normalize_identity(country),
        normalize_identity(speaker),
        normalize_identity(video),
    )
    whisper_path = whisper_index.get(whisper_key)
    if whisper_path is None:
        expected_identity = "/".join(part for part in (country, speaker, video) if part)
        raise ValueError(
            f"No matching Whisper JSON for RockSteady CSV {path}. "
            f"Expected Text identity key {expected_identity}."
        )
    whisper_data = read_whisper_json(whisper_path)
    source_id = str(whisper_data.get("source_id") or "")
    whisper_segments = whisper_data["segments"]
    if prepare_root is None and not upstream_verified:
        alignment_contract = SegmentAlignmentContract(
            status="legacy_unverified",
            manifest_path=None,
            manifest_sha256=None,
            source_indexes={
                index: index - 1 for index in range(1, len(whisper_segments) + 1)
            },
            source_segment_ids={
                index: None for index in range(1, len(whisper_segments) + 1)
            },
        )
    else:
        resolved_prepare_root = prepare_root or Path(
            "processing/text_analysis/output/current/prepared_segments"
        ).resolve()
        alignment_contract = load_segment_alignment_contract(
            resolved_prepare_root,
            country=country,
            speaker=speaker,
            video=video,
            upstream_verified=upstream_verified,
            legacy_segment_count=len(whisper_segments),
        )
    task = str(whisper_data.get("task", "transcribe"))
    source_language = str(whisper_data.get("language", ""))
    attached_text_language = resolve_attached_text_language(
        task,
        source_language,
        selected_text_language,
    )

    available_categories = {
        category.key
        for category in categories
        if category_source_value(rows[0], category) is not None
    }
    rows_by_segment_id: dict[int, tuple[str, Mapping[str, str]]] = {}
    for row_index, row in enumerate(rows, start=1):
        raw_title = row.get("Title") or row.get("title") or ""
        title_match = re.fullmatch(
            rf"{re.escape(video)}__segment_(\d{{6}})", raw_title.strip()
        )
        if title_match is None:
            raise ValueError(
                f"Invalid RockSteady segment identity at {path}:{row_index + 1}: "
                f"{raw_title!r}. Expected {video}__segment_NNNNNN."
            )
        title = raw_title.strip()
        segment_id = int(title_match.group(1))
        if segment_id in rows_by_segment_id:
            raise ValueError(f"Duplicate segment ID {segment_id} found in {path}")
        rows_by_segment_id[segment_id] = (title, row)

    expected_ids = set(alignment_contract.source_indexes)
    actual_ids = set(rows_by_segment_id)
    missing_segment_ids = tuple(sorted(expected_ids - actual_ids))
    ignored_segment_ids = tuple(sorted(actual_ids - expected_ids))
    if alignment_policy == "error" and (missing_segment_ids or ignored_segment_ids):
        raise ValueError(
            f"Segment count mismatch for {path}: RockSteady rows={len(rows)}, "
            f"prepared analysis segments={len(expected_ids)}, Whisper source segments="
            f"{len(whisper_segments)} in {whisper_path}. "
            f"Missing IDs={format_id_list(missing_segment_ids)}; "
            f"extra IDs={format_id_list(ignored_segment_ids)}."
        )

    segments: list[SegmentRecord] = []
    for segment_id in sorted(expected_ids):
        source_segment_index = alignment_contract.source_indexes[segment_id]
        if source_segment_index >= len(whisper_segments):
            raise ValueError(
                f"Prepare mapping source_segment_index {source_segment_index} is outside "
                f"the {len(whisper_segments)} Whisper segments in {whisper_path}"
            )
        whisper_segment = whisper_segments[source_segment_index]
        if not isinstance(whisper_segment, Mapping):
            raise ValueError(
                f"Whisper segment at index {source_segment_index} is not an object: {whisper_path}"
            )
        expected_source_id = alignment_contract.source_segment_ids.get(segment_id)
        actual_source_id = whisper_segment.get("id")
        if expected_source_id is not None and actual_source_id != expected_source_id:
            raise ValueError(
                f"Prepare mapping source_segment_id mismatch for analysis segment {segment_id} "
                f"in {whisper_path}: expected {expected_source_id!r}, found {actual_source_id!r}"
            )
        matched = rows_by_segment_id.get(segment_id)
        if matched is None:
            title = f"{video}__segment_{segment_id:06d}"
            terms = 0.0
            counts = {category.key: None for category in categories}
            rocksteady_row_available = False
        else:
            title, row = matched
            terms = parse_number(source_value(row, ("Terms",))) or 0.0
            counts = {category.key: parse_category_count(row, category) for category in categories}
            rocksteady_row_available = True
        text = whisper_segment_text(
            whisper_segment,
            task=task,
            text_language=selected_text_language,
        )
        start_sec = require_whisper_time(
            whisper_segment.get("start"), whisper_path, source_segment_index, "start"
        )
        end_sec = require_whisper_time(
            whisper_segment.get("end"), whisper_path, source_segment_index, "end"
        )
        if end_sec < start_sec:
            raise ValueError(
                f"Whisper end time precedes start time in {whisper_path}, segment {segment_id}"
            )
        segments.append(
            SegmentRecord(
                country=country,
                speaker=speaker,
                speaker_id=speaker_id,
                video=video,
                segment_id=segment_id,
                source_segment_index=source_segment_index,
                source_segment_id=actual_source_id,
                title=title,
                terms=terms,
                counts=counts,
                start_sec=start_sec,
                end_sec=end_sec,
                segment_text=text,
                whisper_language=source_language,
                text_language=attached_text_language,
                whisper_word_count=count_whisper_words(text),
                categories=categories,
                source_id=source_id,
                rocksteady_row_available=rocksteady_row_available,
            )
        )

    return VideoAnalysis(
        country=country,
        speaker=speaker,
        speaker_id=speaker_id,
        video=video,
        path=path,
        segments=segments,
        video_info=info,
        whisper_path=whisper_path,
        categories=categories,
        available_categories=available_categories,
        rocksteady_row_count=len(rows),
        alignment_contract=alignment_contract,
        source_whisper_segment_count=len(whisper_segments),
        source_id=source_id,
        missing_segment_ids=missing_segment_ids,
        ignored_segment_ids=ignored_segment_ids,
        alignment_policy=alignment_policy,
    )


def read_whisper_json(path: Path) -> dict[str, object]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Cannot read Whisper JSON {path}: {exc}") from exc
    segments = data.get("segments")
    if not isinstance(segments, list):
        raise ValueError(f"Whisper JSON has no segment list: {path}")
    return data


def resolve_attached_text_language(task: str, source_language: str, requested: str) -> str:
    if task != "bilingual":
        return source_language
    return source_language if requested == "original" else requested


def whisper_segment_text(
    segment: Mapping[str, object],
    *,
    task: str = "transcribe",
    text_language: str = "original",
) -> str:
    if task != "bilingual":
        value = segment.get("text") or ""
        return str(value).strip()

    language = normalize_text_language(text_language)
    key = "text_original" if language == "original" else f"text_{language}"
    value = segment.get(key) or ""
    if not value and language != "en":
        value = segment.get("text_original") or ""
    return str(value).strip()


def format_id_list(values: Sequence[int], *, limit: int = 12) -> str:
    if not values:
        return "none"
    shown = ",".join(str(value) for value in values[:limit])
    if len(values) > limit:
        shown += f",... (+{len(values) - limit})"
    return shown


def require_whisper_time(
    value: object,
    path: Path,
    segment_id: int,
    field_name: str,
) -> float:
    parsed = parse_number(value)
    if parsed is None or parsed < 0:
        raise ValueError(
            f"Invalid Whisper {field_name} time in {path}, segment {segment_id}: {value!r}"
        )
    return parsed


def count_whisper_words(text: str) -> int:
    return len(re.findall(r"\b[\w'-]+\b", text, flags=re.UNICODE))


def read_csv_flexible(path: Path) -> list[dict[str, str]]:
    raw_rows = list(csv.reader(read_text(path).splitlines()))
    if not raw_rows:
        return []
    header = [cell.strip() for cell in raw_rows[0]]
    n_cols = len(header)
    data_rows: list[dict[str, str]] = []
    for raw in raw_rows[1:]:
        if not any(cell.strip() for cell in raw):
            continue
        if len(raw) == n_cols + 1:
            raw = [raw[0] + "," + raw[1]] + list(raw[2:])
        data_rows.append({header[i]: (raw[i] if i < len(raw) else "") for i in range(n_cols)})
    return data_rows


def validate_rocksteady_total_rows(
    path: Path,
    rows: Sequence[Mapping[str, str]],
    *,
    categories: Sequence[Category] = CATEGORIES,
) -> None:
    if not rows:
        raise ValueError(f"RockSteady CSV contains no segment rows: {path}")

    available = {key.lower() for key in rows[0]}
    required_groups = [("Title", ("Title",)), ("Terms", ("Terms",))]
    required_groups.extend(
        (category.display, category.source_names)
        for category in categories
        if category.required
    )
    missing = [
        label
        for label, aliases in required_groups
        if not any(alias.lower() in available for alias in aliases)
    ]
    if missing:
        raise ValueError(
            f"RockSteady CSV is missing required Total-mode columns {missing}: {path}"
        )

    for row_index, row in enumerate(rows, start=2):
        terms = require_total_number(source_value(row, ("Terms",)), path, row_index, "Terms")
        for category in categories:
            raw_value = category_source_value(row, category)
            if raw_value is None and not category.required:
                continue
            value = require_total_number(raw_value, path, row_index, category.display)
            if value > terms:
                raise ValueError(
                    f"Suspicious RockSteady value at {path}:{row_index}: "
                    f"{category.display}={format_number(value)} exceeds Terms={format_number(terms)}. "
                    "Export this video from RockSteady in Total mode, not Percentage or Z-Score mode."
                )


def require_total_number(value: object, path: Path, row_index: int, column: str) -> float:
    number = parse_number(value)
    if number is None or number < 0 or abs(number - round(number)) > 1e-9:
        raise ValueError(
            f"Invalid Total-mode value at {path}:{row_index}, column {column}: {value!r}. "
            "Expected a non-negative whole-number count from RockSteady Total mode."
        )
    return number


def category_source_value(row: Mapping[str, str], category: Category) -> object:
    return source_value(row, category.source_names)


def source_value(row: Mapping[str, str], aliases: Sequence[str]) -> object:
    for source_name in aliases:
        if source_name in row:
            return row[source_name]
    lower_map = {key.lower(): value for key, value in row.items()}
    for source_name in aliases:
        if source_name.lower() in lower_map:
            return lower_map[source_name.lower()]
    return None


def read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="latin-1", errors="replace")


def validate_text_asset_path(path: Path, root: Path, *, asset_label: str) -> tuple[str, str]:
    """Return an optional country and required speaker encoded by an asset path."""
    try:
        relative = path.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{asset_label} is outside its configured root {root}: {path}") from exc
    if len(relative.parts) == 2:
        return "", relative.parts[0]
    if len(relative.parts) != 3:
        raise ValueError(
            f"Invalid {asset_label} layout: {relative}. Expected "
            "Speaker/Video or Country/Speaker/Video."
        )

    speaker = path.parent.name
    info = parse_video_name(path.stem)
    filename_country = info.get("country", "")
    filename_person = info.get("person", "")
    canonical = bool(
        re.fullmatch(
            r"\d{3}_[^_]+_.+_\d{4,8}(?:unknown)?",
            path.stem,
        )
    )

    country = path.parent.parent.name
    if not canonical:
        raise ValueError(
            f"Invalid canonical video name for {asset_label} {relative}. Expected "
            "NNN_Country_Speaker_YYYYMMDD."
        )
    if normalize_identity(country) != normalize_identity(filename_country):
        raise ValueError(
            f"Country mismatch in {asset_label} {relative}: folder={country!r}, "
            f"filename={filename_country!r}."
        )
    if canonical and normalize_identity(speaker) != normalize_identity(filename_person):
        raise ValueError(
            f"Speaker mismatch in {asset_label} {relative}: folder={speaker!r}, "
            f"filename={filename_person.replace('_', ' ')!r}."
        )
    return country, speaker


def parse_video_name(video: str) -> dict[str, str]:
    # Canonical format shared by all countries:
    #   001_France_Research_Speaker_20260302
    # The person's name is deliberately greedy because it may contain any
    # number of underscore-separated components; the final token is the date.
    canonical_match = re.fullmatch(
        r"(?P<order>\d{3})_(?P<country>[^_]+)_(?P<person>.+)_(?P<date>\d{4,8}(?:unknown)?)",
        video,
    )
    if canonical_match:
        groups = canonical_match.groupdict()
        return {
            "video_order": groups["order"],
            "date": groups["date"],
            "country": groups["country"],
            "person": groups["person"],
        }

    # Retain compatibility with the earlier synthetic/test convention:
    #   001_20250101_Test_Speaker
    legacy_match = re.fullmatch(
        r"(?P<order>\d{3})_(?P<date>\d{4,8}(?:unknown)?)_(?P<country>[^_]+)_(?P<person>.+)",
        video,
    )
    if legacy_match:
        groups = legacy_match.groupdict()
        return {
            "video_order": groups["order"],
            "date": groups["date"],
            "country": groups["country"],
            "person": groups["person"],
        }

    dash_match = re.match(
        r"^(?P<order>\d{3})\s+-\s+(?P<person>.+)\s+-\s+"
        r"(?P<date>\d{4}-\d{2}-\d{2}|\d{8}|\d{6}unknown|\d{4}unknown)$",
        video,
    )
    if dash_match:
        groups = dash_match.groupdict()
        return {
            "video_order": groups["order"],
            "date": groups["date"],
            "country": "",
            "person": groups["person"],
        }

    return {"video_order": "", "date": "", "country": "", "person": ""}


def parse_category_count(row: Mapping[str, str], category: Category) -> float | None:
    raw_value = category_source_value(row, category)
    if raw_value is None:
        return None
    parsed = parse_number(raw_value)
    return parsed if parsed is not None else 0.0


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return number


def build_video_summary(video: VideoAnalysis) -> dict[str, object]:
    valid = valid_segments(video.segments)
    terms_total = sum(segment.terms for segment in valid)
    row: dict[str, object] = {
        "country": video.country,
        "speaker": video.speaker,
        "speaker_id": video.speaker_id,
        "video": video.video,
        "source_id": video.source_id,
        "video_order": video.video_info.get("video_order", ""),
        "date": video.video_info.get("date", ""),
        "filename_country": video.video_info.get("country", ""),
        "person": video.video_info.get("person", ""),
        "csv_rows": video.rocksteady_row_count,
        "whisper_segments": len(video.segments),
        "analysis_segments": len(video.segments),
        "source_whisper_segments": video.source_whisper_segment_count,
        "alignment_mapping_status": video.alignment_contract.status,
        "prepare_manifest": (
            str(video.alignment_contract.manifest_path)
            if video.alignment_contract.manifest_path
            else ""
        ),
        "rocksteady_matched_segments": sum(
            1 for segment in video.segments if segment.rocksteady_row_available
        ),
        "rocksteady_missing_segments": len(video.missing_segment_ids),
        "rocksteady_missing_segment_ids": format_id_list(video.missing_segment_ids),
        "rocksteady_ignored_extra_rows": len(video.ignored_segment_ids),
        "rocksteady_ignored_segment_ids": format_id_list(video.ignored_segment_ids),
        "segment_alignment_policy": video.alignment_policy,
        "segments_with_terms": len(valid),
        "segments_excluded_zero_terms": sum(
            1
            for segment in video.segments
            if segment.rocksteady_row_available and segment.terms <= 0
        ),
        "rocksteady_terms_total": terms_total,
        "whisper_word_count_total": sum(segment.whisper_word_count for segment in video.segments),
        "whisper_language": video.segments[0].whisper_language if video.segments else "",
        "text_language": video.segments[0].text_language if video.segments else "",
        "whisper_json": str(video.whisper_path),
    }

    for category in video.categories:
        available = category.key in video.available_categories
        row[f"{category.key}_available"] = int(available)
        if not available:
            row[f"{category.key}_total"] = ""
            row[f"{category.key}_proportion"] = ""
            row[f"{category.key}_segment_percent"] = ""
            row[f"mean_segment_{category.key}_count"] = ""
            row[f"median_segment_{category.key}_count"] = ""
            row[f"max_segment_{category.key}_count"] = ""
            continue
        values = [numeric_category_count(segment, category.key) for segment in valid]
        total = sum(values)
        row[f"{category.key}_total"] = total
        row[f"{category.key}_proportion"] = proportion(total, terms_total)
        row[f"{category.key}_segment_percent"] = percent(sum(1 for value in values if value > 0), len(valid))
        row[f"mean_segment_{category.key}_count"] = safe_mean(values)
        row[f"median_segment_{category.key}_count"] = percentile(sorted(values), 0.5)
        row[f"max_segment_{category.key}_count"] = max(values) if values else ""

    row["valence_score"] = (
        valence_score(
            numeric_or_zero(row.get("positive_total")),
            numeric_or_zero(row.get("negative_total")),
        )
        if {"positive", "negative"}.issubset(video.available_categories)
        else ""
    )
    return row


def build_speaker_summary_rows(videos: Sequence[VideoAnalysis]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    categories = categories_for_videos(videos)
    for speaker_id in sorted({video.speaker_id for video in videos}):
        speaker_videos = sorted(
            [video for video in videos if video.speaker_id == speaker_id],
            key=lambda item: item.video,
        )
        country = speaker_videos[0].country
        speaker = speaker_videos[0].speaker
        valid = [
            segment
            for video in speaker_videos
            for segment in video.segments
            if segment.rocksteady_row_available and segment.terms > 0
        ]
        terms_total = sum(segment.terms for segment in valid)
        row: dict[str, object] = {
            "country": country,
            "speaker": speaker,
            "speaker_id": speaker_id,
            "videos_count": len(speaker_videos),
            "csv_rows": sum(video.rocksteady_row_count for video in speaker_videos),
            "whisper_segments": sum(len(video.segments) for video in speaker_videos),
            "analysis_segments": sum(len(video.segments) for video in speaker_videos),
            "source_whisper_segments": sum(
                video.source_whisper_segment_count for video in speaker_videos
            ),
            "alignment_mapping_status": ",".join(
                sorted({video.alignment_contract.status for video in speaker_videos})
            ),
            "rocksteady_matched_segments": sum(
                1
                for video in speaker_videos
                for segment in video.segments
                if segment.rocksteady_row_available
            ),
            "rocksteady_missing_segments": sum(
                len(video.missing_segment_ids) for video in speaker_videos
            ),
            "rocksteady_ignored_extra_rows": sum(
                len(video.ignored_segment_ids) for video in speaker_videos
            ),
            "segments_with_terms": len(valid),
            "segments_excluded_zero_terms": sum(
                1
                for video in speaker_videos
                for segment in video.segments
                if segment.rocksteady_row_available and segment.terms <= 0
            ),
            "rocksteady_terms_total": terms_total,
            "whisper_word_count_total": sum(
                segment.whisper_word_count
                for video in speaker_videos
                for segment in video.segments
            ),
        }
        for category in categories:
            available_videos = sum(
                1 for video in speaker_videos if category.key in video.available_categories
            )
            complete = available_videos == len(speaker_videos)
            row[f"{category.key}_available_videos"] = available_videos
            row[f"{category.key}_complete"] = int(complete)
            if not complete:
                row[f"{category.key}_total"] = ""
                row[f"{category.key}_proportion"] = ""
                row[f"{category.key}_segment_percent"] = ""
                row[f"mean_video_{category.key}_proportion"] = ""
                row[f"median_video_{category.key}_proportion"] = ""
                continue
            segment_values = [numeric_category_count(segment, category.key) for segment in valid]
            total = sum(segment_values)
            video_proportions = numeric_summary_values(speaker_videos, f"{category.key}_proportion")
            row[f"{category.key}_total"] = total
            row[f"{category.key}_proportion"] = proportion(total, terms_total)
            row[f"{category.key}_segment_percent"] = percent(
                sum(1 for value in segment_values if value > 0), len(valid)
            )
            row[f"mean_video_{category.key}_proportion"] = safe_mean(video_proportions)
            row[f"median_video_{category.key}_proportion"] = percentile(sorted(video_proportions), 0.5)
        row["valence_score"] = (
            valence_score(
                numeric_or_zero(row.get("positive_total")),
                numeric_or_zero(row.get("negative_total")),
            )
            if row.get("positive_complete") == 1 and row.get("negative_complete") == 1
            else ""
        )
        rows.append(row)
    return rows


def numeric_summary_values(videos: Sequence[VideoAnalysis], key: str) -> list[float]:
    values: list[float] = []
    for video in videos:
        value = parse_number(video.summary.get(key, ""))
        if value is not None:
            values.append(value)
    return values


def valid_segments(segments: Sequence[SegmentRecord]) -> list[SegmentRecord]:
    return [
        segment
        for segment in segments
        if segment.rocksteady_row_available and segment.terms > 0
    ]


def text_segment_observations(
    videos: Sequence[VideoAnalysis],
    categories: Sequence[Category],
) -> list[TextSegmentObservation]:
    observations: list[TextSegmentObservation] = []
    for video in videos:
        for segment in valid_segments(video.segments):
            counts = {
                category.key: (
                    segment.counts.get(category.key)
                    if category.key in video.available_categories
                    else None
                )
                for category in categories
            }
            observations.append(
                TextSegmentObservation(
                    country=video.country,
                    speaker=video.speaker,
                    video=video.video,
                    segment_id=str(segment.segment_id),
                    terms=segment.terms,
                    category_counts=counts,
                    positive_count=counts.get("positive"),
                    negative_count=counts.get("negative"),
                )
            )
    return observations


def categories_for_videos(videos: Sequence[VideoAnalysis]) -> tuple[Category, ...]:
    if not videos:
        return ()
    categories = videos[0].categories
    expected = tuple(category.key for category in categories)
    for video in videos[1:]:
        found = tuple(category.key for category in video.categories)
        if found != expected:
            raise ValueError(
                f"Inconsistent category contract for {video.path}: expected {expected}, found {found}"
            )
    return categories


def numeric_category_count(segment: SegmentRecord, key: str) -> float:
    value = segment.counts.get(key)
    return value if value is not None else 0.0


def numeric_or_zero(value: object) -> float:
    parsed = parse_number(value)
    return parsed if parsed is not None else 0.0


def valence_score(positive: float, negative: float) -> float | str:
    denominator = positive + negative
    if denominator <= 0:
        return ""
    return (positive - negative) / denominator


def write_segment_outputs(output_dir: Path, videos: Sequence[VideoAnalysis]) -> None:
    for video in videos:
        identity_dir = Path(video.country) / video.speaker if video.country else Path(video.speaker)
        count_rows = [segment_count_row(segment) for segment in video.segments]
        relative_rows = [segment_relative_row(segment) for segment in video.segments]
        enriched_rows = [segment_enriched_row(segment) for segment in video.segments]
        write_dict_rows(
            output_dir / "segment_counts" / identity_dir / f"{video.video}_segment_counts.csv",
            count_rows,
        )
        write_dict_rows(
            output_dir / "segment_relative" / identity_dir / f"{video.video}_segment_relative.csv",
            relative_rows,
        )
        write_dict_rows(
            output_dir / "segment_level" / identity_dir / f"{video.video}_segments_enriched.csv",
            enriched_rows,
        )


def segment_identity_row(segment: SegmentRecord) -> dict[str, object]:
    return {
        "country": segment.country,
        "speaker": segment.speaker,
        "speaker_id": segment.speaker_id,
        "video": segment.video,
        "source_id": segment.source_id,
        "segment_id": segment.segment_id,
        "analysis_segment_id": segment.segment_id,
        "source_segment_index": segment.source_segment_index,
        "source_segment_id": (
            segment.source_segment_id if segment.source_segment_id is not None else ""
        ),
        "start_sec": segment.start_sec,
        "end_sec": segment.end_sec,
        "duration_sec": segment.duration_sec,
        "segment_text": segment.segment_text,
        "whisper_language": segment.whisper_language,
        "text_language": segment.text_language,
        "rocksteady_row_available": int(segment.rocksteady_row_available),
        "rocksteady_terms": segment.terms if segment.rocksteady_row_available else "",
        "whisper_word_count": segment.whisper_word_count,
        "valid_segment": int(segment.rocksteady_row_available and segment.terms > 0),
    }


def segment_count_row(segment: SegmentRecord) -> dict[str, object]:
    row = segment_identity_row(segment)
    row["valence_score"] = segment.valence_score
    for category in segment.categories:
        value = segment.counts.get(category.key)
        row[f"{category.key}_count"] = value if value is not None else ""
    return row


def segment_relative_row(segment: SegmentRecord) -> dict[str, object]:
    row = segment_identity_row(segment)
    row["valence_score"] = segment.valence_score
    for category in segment.categories:
        value = segment.counts.get(category.key)
        row[f"{category.key}_proportion"] = (
            proportion(value, segment.terms) if value is not None else ""
        )
    return row


def segment_enriched_row(segment: SegmentRecord) -> dict[str, object]:
    row = segment_identity_row(segment)
    row["rocksteady_title"] = segment.title
    for category in segment.categories:
        value = segment.counts.get(category.key)
        row[f"{category.key}_available"] = int(value is not None)
        row[f"{category.key}_count"] = value if value is not None else ""
        row[f"{category.key}_proportion"] = (
            proportion(value, segment.terms) if value is not None else ""
        )
        row[f"has_{category.key}"] = int(value > 0) if value is not None else ""
    row["valence_score"] = segment.valence_score
    return row


def write_descriptor_statistics(path: Path, videos: Sequence[VideoAnalysis]) -> None:
    rows: list[dict[str, object]] = []
    for video in videos:
        enriched_rows = [segment_enriched_row(segment) for segment in valid_segments(video.segments)]
        descriptors = descriptor_names(enriched_rows, video.categories)
        for descriptor in descriptors:
            values = [parse_number(row.get(descriptor, "")) for row in enriched_rows]
            finite = [value for value in values if value is not None]
            stats = describe(finite)
            stats.update(
                {
                    "country": video.country,
                    "speaker": video.speaker,
                    "speaker_id": video.speaker_id,
                    "video": video.video,
                    "descriptor": descriptor,
                    "missing": len(values) - len(finite),
                }
            )
            rows.append(stats)
    ordered = [
        "country",
        "speaker",
        "speaker_id",
        "video",
        "descriptor",
        "count",
        "missing",
        "mean",
        "standard_error",
        "mode",
        "stddev",
        "sample_variance",
        "skewness",
        "min",
        "q1",
        "median",
        "q3",
        "max",
        "range",
        "sum",
        "confidence_level_95",
        "kurtosis",
        "nonzero_count",
        "nonzero_percent",
    ]
    write_dict_rows(path, rows, fieldnames=ordered)
    write_descriptor_statistics_workbook(path.with_suffix(".xlsx"), rows)


def write_descriptor_statistics_workbook(
    path: Path,
    rows: Sequence[Mapping[str, object]],
) -> None:
    """Write screenshot-style descriptive-statistic tables for every video.

    Each video receives one worksheet. Every available descriptor is rendered
    as a two-column ``Statistic / Value`` table so the workbook is generated
    directly from the current postprocessing data rather than from a fixed
    professor-summary template.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    statistic_fields = (
        ("Mean", "mean"),
        ("Standard Error", "standard_error"),
        ("Median", "median"),
        ("Mode", "mode"),
        ("Standard Deviation", "stddev"),
        ("Sample Variance", "sample_variance"),
        ("Kurtosis", "kurtosis"),
        ("Skewness", "skewness"),
        ("Range", "range"),
        ("Minimum", "min"),
        ("Maximum", "max"),
        ("Sum", "sum"),
        ("Count", "count"),
        ("Confidence Level (95.0%)", "confidence_level_95"),
    )
    grouped: dict[tuple[str, str, str, str], list[Mapping[str, object]]] = {}
    for row in rows:
        key = (
            str(row.get("country", "")),
            str(row.get("speaker", "")),
            str(row.get("speaker_id", "")),
            str(row.get("video", "")),
        )
        grouped.setdefault(key, []).append(row)

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    for (country, speaker, speaker_id, video), descriptor_rows in grouped.items():
        title = unique_excel_sheet_title(video, used_titles)
        sheet = workbook.create_sheet(title)
        sheet.sheet_view.showGridLines = True
        sheet.freeze_panes = "A4"
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 20
        sheet["A1"] = "Video"
        sheet["B1"] = neutralize_spreadsheet_value(video)
        sheet["A2"] = "Country / Speaker"
        sheet["B2"] = neutralize_spreadsheet_value(f"{country} / {speaker}")
        sheet["A1"].font = sheet["A2"].font = Font(bold=True)

        output_row = 4
        for descriptor_row in descriptor_rows:
            descriptor = str(descriptor_row.get("descriptor", ""))
            sheet.cell(output_row, 1, neutralize_spreadsheet_value(descriptor))
            sheet.cell(output_row, 2, "Value")
            for column in (1, 2):
                cell = sheet.cell(output_row, column)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left")
            output_row += 1

            for label, field in statistic_fields:
                label_cell = sheet.cell(output_row, 1, label)
                raw_value = descriptor_row.get(field, "")
                numeric_value = parse_number(raw_value)
                value_cell = sheet.cell(
                    output_row,
                    2,
                    numeric_value
                    if numeric_value is not None
                    else neutralize_spreadsheet_value(raw_value),
                )
                if field == "count" or (
                    numeric_value is not None and numeric_value.is_integer()
                ):
                    value_cell.number_format = "#,##0"
                else:
                    value_cell.number_format = "0.######"
                output_row += 1
            output_row += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def unique_excel_sheet_title(video: str, used_titles: set[str]) -> str:
    invalid = re.compile(r"[\\/*?:\[\]]")
    base = invalid.sub("_", video).strip(" '") or "Video"
    candidate = base[:31]
    suffix = 2
    while candidate.casefold() in used_titles:
        tail = f"_{suffix}"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used_titles.add(candidate.casefold())
    return candidate


def write_segment_alignment_audit(path: Path, videos: Sequence[VideoAnalysis]) -> None:
    rows: list[dict[str, object]] = []
    for video in videos:
        matched = sum(
            1 for segment in video.segments if segment.rocksteady_row_available
        )
        rows.append(
            {
                "country": video.country,
                "speaker": video.speaker,
                "speaker_id": video.speaker_id,
                "video": video.video,
                "alignment_status": (
                    "exact"
                    if not video.missing_segment_ids and not video.ignored_segment_ids
                    else "reconciled"
                ),
                "alignment_policy": video.alignment_policy,
                "alignment_mapping_status": video.alignment_contract.status,
                "prepare_manifest": (
                    str(video.alignment_contract.manifest_path)
                    if video.alignment_contract.manifest_path
                    else ""
                ),
                "prepare_manifest_sha256": video.alignment_contract.manifest_sha256 or "",
                "rocksteady_csv_rows": video.rocksteady_row_count,
                "whisper_segments": len(video.segments),
                "analysis_segments": len(video.segments),
                "source_whisper_segments": video.source_whisper_segment_count,
                "matched_segments": matched,
                "missing_rocksteady_segments": len(video.missing_segment_ids),
                "missing_segment_ids": ",".join(
                    str(value) for value in video.missing_segment_ids
                ),
                "ignored_extra_rows": len(video.ignored_segment_ids),
                "ignored_segment_ids": ",".join(
                    str(value) for value in video.ignored_segment_ids
                ),
                "text_language": (
                    video.segments[0].text_language if video.segments else ""
                ),
                "rocksteady_csv": str(video.path),
                "whisper_json": str(video.whisper_path),
            }
        )
    write_dict_rows(path, rows)


def descriptor_names(
    rows: Sequence[Mapping[str, object]],
    categories: Sequence[Category] = CATEGORIES,
) -> list[str]:
    candidates: list[str] = []
    for category in categories:
        candidates.extend(
            (
                f"{category.key}_count",
                f"{category.key}_proportion",
                f"has_{category.key}",
            )
        )
    candidates.append("valence_score")
    return [
        name
        for name in candidates
        if any(parse_number(row.get(name, "")) is not None for row in rows)
    ]


def describe(values: Sequence[float]) -> dict[str, object]:
    if not values:
        return {
            "count": 0,
            "mean": "",
            "standard_error": "",
            "mode": "",
            "stddev": "",
            "sample_variance": "",
            "skewness": "",
            "min": "",
            "q1": "",
            "median": "",
            "q3": "",
            "max": "",
            "range": "",
            "sum": "",
            "confidence_level_95": "",
            "kurtosis": "",
            "nonzero_count": 0,
            "nonzero_percent": "",
        }
    ordered = sorted(values)
    avg = mean(ordered)
    sample_stddev = stddev(ordered)
    return {
        "count": len(ordered),
        "mean": avg,
        "standard_error": sample_stddev / math.sqrt(len(ordered)),
        "mode": mode(ordered),
        "stddev": sample_stddev,
        "sample_variance": sample_stddev ** 2,
        "skewness": sample_skewness(ordered, avg, sample_stddev),
        "min": ordered[0],
        "q1": percentile(ordered, 0.25),
        "median": percentile(ordered, 0.5),
        "q3": percentile(ordered, 0.75),
        "max": ordered[-1],
        "range": ordered[-1] - ordered[0],
        "sum": sum(ordered),
        "confidence_level_95": confidence_level_95(len(ordered), sample_stddev),
        "kurtosis": excess_kurtosis(ordered, avg),
        "nonzero_count": sum(1 for value in ordered if value != 0),
        "nonzero_percent": percent(sum(1 for value in ordered if value != 0), len(ordered)),
    }


def stddev(values: Sequence[float]) -> float:
    if len(values) < 2:
        return 0.0
    avg = mean(values)
    variance = sum((value - avg) ** 2 for value in values) / (len(values) - 1)
    return math.sqrt(variance)


def mode(values: Sequence[float]) -> float | str:
    frequencies = Counter(values)
    highest_frequency = max(frequencies.values(), default=0)
    if highest_frequency < 2:
        return ""
    return min(value for value, frequency in frequencies.items() if frequency == highest_frequency)


def sample_skewness(
    values: Sequence[float],
    avg: float | None = None,
    sample_stddev: float | None = None,
) -> float | str:
    count = len(values)
    if count < 3:
        return ""
    if avg is None:
        avg = mean(values)
    if sample_stddev is None:
        sample_stddev = stddev(values)
    if sample_stddev == 0:
        return ""
    standardized_cube_sum = sum(((value - avg) / sample_stddev) ** 3 for value in values)
    return count * standardized_cube_sum / ((count - 1) * (count - 2))


def confidence_level_95(count: int, sample_stddev: float) -> float | str:
    if count < 2:
        return ""
    target_p = 0.05
    lower = 0.0
    upper = 2.0
    while student_t_two_tailed_p(upper, count - 1) > target_p:
        upper *= 2.0
    for _ in range(80):
        midpoint = (lower + upper) / 2.0
        if student_t_two_tailed_p(midpoint, count - 1) > target_p:
            lower = midpoint
        else:
            upper = midpoint
    critical_t = (lower + upper) / 2.0
    return critical_t * sample_stddev / math.sqrt(count)


def percentile(sorted_values: Sequence[float], fraction: float) -> float | str:
    if not sorted_values:
        return ""
    if len(sorted_values) == 1:
        return sorted_values[0]
    index = fraction * (len(sorted_values) - 1)
    lower = math.floor(index)
    upper = math.ceil(index)
    if lower == upper:
        return sorted_values[lower]
    weight = index - lower
    return sorted_values[lower] * (1 - weight) + sorted_values[upper] * weight


def excess_kurtosis(values: Sequence[float], avg: float | None = None) -> float | str:
    if not values:
        return ""
    if avg is None:
        avg = mean(values)
    variance = sum((value - avg) ** 2 for value in values) / len(values)
    if variance == 0:
        # Kurtosis divides by variance squared and is mathematically undefined
        # for a constant descriptor.  A blank cell preserves that distinction
        # instead of presenting an invented value of zero.
        return ""
    fourth = sum((value - avg) ** 4 for value in values) / len(values)
    return fourth / (variance ** 2) - 3


def proportion(value: float, terms: float) -> float | str:
    if terms <= 0:
        return ""
    return value / terms


def percent(count: int, total: int) -> float | str:
    if total <= 0:
        return ""
    return count / total * 100


def safe_mean(values: Sequence[float]) -> float | str:
    if not values:
        return ""
    return mean(values)


def normalize_segment_sample_counts(sample_counts: Sequence[int]) -> tuple[int, ...]:
    """Validate and canonicalize requested odd timeline sample sizes."""

    normalized: set[int] = set()
    for sample_count in sample_counts:
        if isinstance(sample_count, bool) or not isinstance(sample_count, int):
            raise ValueError("Segment sample counts must be integers.")
        if sample_count < 3 or sample_count % 2 == 0:
            raise ValueError("Segment sample counts must be odd integers greater than or equal to 3.")
        normalized.add(sample_count)
    return tuple(sorted(normalized))


def parse_segment_sample_counts(value: str) -> tuple[int, ...]:
    """Parse a CLI value such as ``3,5,7,9`` or ``none``."""

    text = value.strip()
    if text.lower() in {"none", "off"}:
        return ()
    if not text:
        raise argparse.ArgumentTypeError("Provide comma-separated odd integers, for example 3,5,7,9.")
    try:
        sample_counts = tuple(int(part.strip()) for part in text.split(","))
        return normalize_segment_sample_counts(sample_counts)
    except ValueError as error:
        raise argparse.ArgumentTypeError(str(error)) from error


def evenly_spaced_indexes(total: int, sample_count: int) -> list[int]:
    """Return zero-based indexes spread across a sequence, including both ends."""

    normalize_segment_sample_counts((sample_count,))
    if total < sample_count:
        return []
    return [round(index * (total - 1) / (sample_count - 1)) for index in range(sample_count)]


def sample_segments_evenly(
    segments: Sequence[SegmentRecord],
    sample_count: int,
) -> list[SegmentRecord]:
    return [segments[index] for index in evenly_spaced_indexes(len(segments), sample_count)]


def write_all_graphs(
    output_dir: Path,
    videos: Sequence[VideoAnalysis],
    speaker_rows: Sequence[Mapping[str, object]],
    *,
    segment_sample_counts: Sequence[int] = DEFAULT_SEGMENT_SAMPLE_COUNTS,
) -> list[Path]:
    sample_counts = normalize_segment_sample_counts(segment_sample_counts)
    categories = categories_for_videos(videos)
    graph_paths: list[Path] = []
    for video in videos:
        identity_dir = Path(video.country) / video.speaker if video.country else Path(video.speaker)
        video_graph_dir = output_dir / "graphs" / "videos" / identity_dir / video.video
        graph_paths.append(
            write_svg(
                video_graph_dir / "segment_timeline_counts_full.svg",
                render_timeline_counts_svg(video),
            )
        )
        segments = valid_segments(video.segments)
        for sample_count in sample_counts:
            sampled_segments = sample_segments_evenly(segments, sample_count)
            if not sampled_segments:
                continue
            graph_paths.append(
                write_svg(
                    video_graph_dir / f"segment_timeline_counts_sampled_{sample_count}.svg",
                    render_timeline_counts_svg(
                        video,
                        sampled_segments=sampled_segments,
                        requested_sample_count=sample_count,
                    ),
                )
            )
        graph_paths.append(
            write_svg(
                video_graph_dir / "segment_count_histograms.svg",
                render_count_histograms_svg(video),
            )
        )

    speaker_graph_dir = output_dir / "graphs" / "speakers"
    for speaker_id in sorted({video.speaker_id for video in videos}):
        speaker_videos = sorted(
            [video for video in videos if video.speaker_id == speaker_id],
            key=lambda item: item.video,
        )
        speaker = speaker_videos[0].speaker
        country = speaker_videos[0].country
        identity_dir = Path(country) / speaker if country else Path(speaker)
        graph_paths.append(
            write_svg(
                speaker_graph_dir / identity_dir / "video_proportions.svg",
                render_speaker_video_proportions_svg(speaker, speaker_videos),
            )
        )
        graph_paths.append(
            write_svg(
                speaker_graph_dir / identity_dir / "video_valence.svg",
                render_speaker_video_valence_svg(speaker, speaker_videos),
            )
        )

    summary_graph_dir = output_dir / "graphs" / "summary"
    countries = sorted({str(row.get("country", "")).strip() for row in speaker_rows})
    for country in countries:
        country_rows = [
            row
            for row in speaker_rows
            if str(row.get("country", "")).strip() == country
        ]
        country_dir = summary_graph_dir / (country or "Unknown")
        graph_paths.append(
            write_svg(
                country_dir / "speaker_summary_proportions.svg",
                render_speaker_summary_proportions_svg(
                    country_rows, country=country, categories=categories
                ),
            )
        )
        graph_paths.append(
            write_svg(
                country_dir / "speaker_summary_valence.svg",
                render_speaker_summary_valence_svg(country_rows, country=country),
            )
        )
    return graph_paths


def write_svg(path: Path, text: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return path


def render_timeline_counts_svg(
    video: VideoAnalysis,
    *,
    sampled_segments: Sequence[SegmentRecord] | None = None,
    requested_sample_count: int | None = None,
) -> str:
    width = 1500
    left, right, bottom = 92, 42, 92
    categories = [
        category
        for category in video.categories
        if category.key in video.available_categories
    ]
    plot_w = width - left - right
    legend_rows = legend_row_count(categories, plot_w, item_spacing=126)
    legend_extra = max(0, legend_rows - 1) * 22
    top = 112 + legend_extra
    height = 520 + legend_extra
    plot_h = height - top - bottom
    if sampled_segments is None:
        x_labels, series_values, timeline_note, y_label = build_timeline_series(
            valid_segments(video.segments),
            categories,
        )
        title_suffix = ""
    else:
        if requested_sample_count is None or len(sampled_segments) != requested_sample_count:
            raise ValueError("A sampled timeline requires exactly the requested number of segments.")
        x_labels, series_values, _, y_label = build_timeline_series(sampled_segments, categories)
        source_segment_count = len(valid_segments(video.segments))
        timeline_note = (
            f"{requested_sample_count} evenly spaced segments selected from "
            f"{source_segment_count:,} valid segments; "
            "first and last retained"
        )
        title_suffix = f" ({requested_sample_count}-segment sample)"
    raw_max_y = max([value for values in series_values.values() for value in values] or [1])
    max_y = nice_axis_max(raw_max_y)

    parts = svg_header(width, height)
    parts.extend(
        chart_title_svg(
            left,
            f"{full_video_label(video.video)} segment timeline{title_suffix}",
            timeline_note,
        )
    )
    parts.extend(
        legend_svg(categories, left, 86, max_width=plot_w, item_spacing=126)
    )
    parts.append(plot_background_svg(left, top, plot_w, plot_h))
    parts.append(axis_svg(left, top, plot_w, plot_h, y_label=y_label))

    for tick in range(0, 5):
        y = top + plot_h - (plot_h * tick / 4)
        label = max_y * tick / 4
        parts.append(f'<line class="grid" x1="{left}" y1="{y:.2f}" x2="{left + plot_w}" y2="{y:.2f}"/>')
        parts.append(
            f'<text class="axis-label" x="{left - 8}" y="{y + 4:.2f}" '
            f'text-anchor="end">{format_number(label)}</text>'
        )

    denom = max(1, len(x_labels) - 1)
    desired_ticks = len(x_labels) if sampled_segments is not None else 7
    for index in segment_tick_indexes(len(x_labels), desired=desired_ticks):
        x = left + plot_w * index / denom
        parts.append(
            f'<line class="tick" x1="{x:.2f}" y1="{top + plot_h}" '
            f'x2="{x:.2f}" y2="{top + plot_h + 6}"/>'
        )
        parts.append(
            f'<text class="axis-label" x="{x:.2f}" y="{top + plot_h + 24}" '
            f'text-anchor="middle">{x_labels[index]}</text>'
        )
    parts.append(
        f'<text class="axis-title" x="{left + plot_w / 2:.2f}" y="{height - 36}" '
        'text-anchor="middle">Whisper segment order</text>'
    )

    for category in categories:
        points = []
        for index, value in enumerate(series_values[category.key]):
            x = left + plot_w * index / denom
            y = top + plot_h - (value / max_y * plot_h)
            points.append(f"{x:.2f},{y:.2f}")
        parts.append(
            f'<polyline class="series" points="{" ".join(points)}" '
            f'stroke="{category.color}"><title>{escape(category.display)}</title></polyline>'
        )

    parts.append(svg_footer())
    return "\n".join(parts)


def build_timeline_series(
    segments: Sequence[SegmentRecord],
    categories: Sequence[Category],
    *,
    max_points: int = 120,
) -> tuple[list[int], dict[str, list[float]], str, str]:
    if len(segments) <= max_points:
        labels = [segment.segment_id for segment in segments]
        values = {
            category.key: [numeric_category_count(segment, category.key) for segment in segments]
            for category in categories
        }
        note = f"{len(segments):,} Whisper segments; raw RockSteady Total counts per segment"
        return labels, values, note, "category word count"

    bucket_count = max_points
    labels: list[int] = []
    values = {category.key: [] for category in categories}
    for bucket_index in range(bucket_count):
        start = round(len(segments) * bucket_index / bucket_count)
        end = round(len(segments) * (bucket_index + 1) / bucket_count)
        end = max(end, start + 1)
        bucket_segments = segments[start:end]
        labels.append(round((bucket_segments[0].segment_id + bucket_segments[-1].segment_id) / 2))
        for category in categories:
            values[category.key].append(
                sum(
                    numeric_category_count(segment, category.key)
                    for segment in bucket_segments
                )
            )

    note = (
        f"{len(segments):,} Whisper segments binned into {bucket_count} equal-order groups; "
        "y-axis sums RockSteady Total counts within each group"
    )
    return labels, values, note, "category word count per group"


def render_count_histograms_svg(video: VideoAnalysis) -> str:
    categories = [
        category
        for category in video.categories
        if category.key in video.available_categories
    ]
    segments = valid_segments(video.segments)
    table = {
        category.key: count_histogram([numeric_category_count(segment, category.key) for segment in segments])
        for category in categories
    }
    return render_grouped_bar_svg(
        title=f"{full_video_label(video.video)} segment category-count histograms",
        subtitle=f"{len(segments):,} valid segments (Terms > 0): counts with 0, 1, 2, or 3+ category hits",
        x_labels=list(COUNT_HISTOGRAM_BINS),
        series=[
            (category.display, category.color, [table[category.key][label] for label in COUNT_HISTOGRAM_BINS])
            for category in categories
        ],
        y_label="segments",
        width=1050,
        height=500,
    )


def render_speaker_video_proportions_svg(speaker: str, videos: Sequence[VideoAnalysis]) -> str:
    contract_categories = categories_for_videos(videos)
    categories = [
        category
        for category in contract_categories
        if any(category.key in video.available_categories for video in videos)
    ]
    return render_grouped_bar_svg(
        title=f"{display_name(speaker)} video-level percentages",
        subtitle="RockSteady Total count / RockSteady Terms, displayed as a percentage",
        x_labels=[video_axis_label(video.video) for video in videos],
        series=[
            (
                category.display,
                category.color,
                [float_or_zero(video.summary.get(f"{category.key}_proportion")) * 100 for video in videos],
            )
            for category in categories
        ],
        y_label="dictionary hits as % of RockSteady Terms",
        width=max(1500, 300 * max(1, len(videos))),
        height=560,
        show_value_labels=True,
        value_suffix="%",
    )


def render_speaker_summary_proportions_svg(
    rows: Sequence[Mapping[str, object]],
    *,
    country: str = "",
    categories: Sequence[Category] | None = None,
) -> str:
    contract_categories = tuple(categories or CATEGORIES)
    plotted_categories = [
        category
        for category in contract_categories
        if any(parse_number(row.get(f"{category.key}_proportion")) is not None for row in rows)
    ]
    ordered_rows = sorted(rows, key=lambda row: str(row.get("speaker", "")))
    return render_grouped_bar_svg(
        title=f"{country + ' ' if country else ''}speaker summary percentages",
        subtitle="RockSteady Total count / RockSteady Terms, displayed as a percentage",
        x_labels=[speaker_axis_label(str(row.get("speaker", ""))) for row in ordered_rows],
        series=[
            (
                category.display,
                category.color,
                [float_or_zero(row.get(f"{category.key}_proportion")) * 100 for row in ordered_rows],
            )
            for category in plotted_categories
        ],
        y_label="dictionary hits as % of RockSteady Terms",
        width=max(1500, 260 * max(1, len(ordered_rows))),
        height=560,
        show_value_labels=True,
        value_suffix="%",
    )


def render_speaker_video_valence_svg(speaker: str, videos: Sequence[VideoAnalysis]) -> str:
    return render_valence_bar_svg(
        title=f"{display_name(speaker)} video-level valence",
        subtitle="Derived score: (Positive - Negative) / (Positive + Negative)",
        x_labels=[video_axis_label(video.video) for video in videos],
        values=[parse_number(video.summary.get("valence_score")) for video in videos],
        width=max(1050, 178 * max(1, len(videos))),
    )


def render_speaker_summary_valence_svg(
    rows: Sequence[Mapping[str, object]],
    *,
    country: str = "",
) -> str:
    ordered_rows = sorted(rows, key=lambda row: str(row.get("speaker", "")))
    return render_valence_bar_svg(
        title=f"{country + ' ' if country else ''}speaker summary valence",
        subtitle="Derived score: (Positive - Negative) / (Positive + Negative)",
        x_labels=[speaker_axis_label(str(row.get("speaker", ""))) for row in ordered_rows],
        values=[parse_number(row.get("valence_score")) for row in ordered_rows],
        width=max(1120, 170 * max(1, len(ordered_rows))),
    )


def render_valence_bar_svg(
    *,
    title: str,
    subtitle: str,
    x_labels: Sequence[str],
    values: Sequence[float | None],
    width: int,
    height: int = 520,
    y_label: str = "valence score (-1 to 1)",
    colors: Sequence[str] | None = None,
) -> str:
    left, right, top, bottom = 92, 42, 88, 104
    plot_w = width - left - right
    plot_h = height - top - bottom
    zero_y = top + plot_h / 2
    group_w = plot_w / max(1, len(x_labels))
    bar_w = max(18, min(70, group_w * 0.48))

    parts = svg_header(width, height)
    parts.extend(chart_title_svg(left, title, subtitle))
    parts.append(plot_background_svg(left, top, plot_w, plot_h))
    for tick_value in (-1.0, -0.5, 0.0, 0.5, 1.0):
        y = top + (1 - tick_value) / 2 * plot_h
        css_class = "axis" if tick_value == 0 else "grid"
        parts.append(
            f'<line class="{css_class}" x1="{left}" y1="{y:.2f}" '
            f'x2="{left + plot_w}" y2="{y:.2f}"/>'
        )
        parts.append(
            f'<text class="axis-label" x="{left - 8}" y="{y + 4:.2f}" '
            f'text-anchor="end">{tick_value:g}</text>'
        )
    parts.append(
        f'<text class="axis-label" x="{left}" y="{top - 12}">{escape(y_label)}</text>'
    )

    for index, label in enumerate(x_labels):
        center_x = left + group_w * index + group_w / 2
        value = values[index] if index < len(values) else None
        if value is not None:
            clamped = max(-1.0, min(1.0, value))
            value_y = top + (1 - clamped) / 2 * plot_h
            y = min(zero_y, value_y)
            bar_h = max(1.0, abs(zero_y - value_y))
            color = (
                colors[index]
                if colors is not None and index < len(colors)
                else "#16a34a" if clamped > 0 else "#dc2626" if clamped < 0 else "#64748b"
            )
            parts.append(
                f'<rect class="bar" x="{center_x - bar_w / 2:.2f}" y="{y:.2f}" '
                f'width="{bar_w:.2f}" height="{bar_h:.2f}" fill="{color}">'
                f'<title>{escape(label)}: {format_number(value)}</title></rect>'
            )
            label_y = max(top + 14, value_y - 6) if clamped >= 0 else min(top + plot_h - 4, value_y + 16)
            parts.append(
                f'<text class="value-label" x="{center_x:.2f}" y="{label_y:.2f}" '
                f'text-anchor="middle">{escape(format_number(value))}</text>'
            )
        else:
            parts.append(
                f'<text class="axis-label" x="{center_x:.2f}" y="{zero_y - 6:.2f}" '
                f'text-anchor="middle">NA</text>'
            )
        parts.append(
            multiline_text_svg(
                "x-label",
                center_x,
                top + plot_h + 24,
                label_lines(label),
                anchor="middle",
            )
        )

    parts.append(svg_footer())
    return "\n".join(parts)


def count_histogram(values: Sequence[float]) -> dict[str, int]:
    counts = {label: 0 for label in COUNT_HISTOGRAM_BINS}
    for value in values:
        if value <= 0:
            counts["0"] += 1
        elif value == 1:
            counts["1"] += 1
        elif value == 2:
            counts["2"] += 1
        else:
            counts["3+"] += 1
    return counts


def render_grouped_bar_svg(
    *,
    title: str,
    x_labels: Sequence[str],
    series: Sequence[tuple[str, str, Sequence[float]]],
    y_label: str,
    subtitle: str = "",
    width: int = 900,
    height: int = 420,
    show_value_labels: bool = False,
    value_suffix: str = "",
) -> str:
    left, right, bottom = 92, 42, 104
    series_count = max(1, len(series))
    group_count = max(1, len(x_labels))
    # Keep every category bar inside its group even for General Language's
    # 45-category output. A wide SVG is preferable to overlapping/clipped bars.
    minimum_group_width = max(
        120,
        math.ceil(
            (series_count * 4 + max(0, series_count - 1) * 4) / 0.7
        ),
    )
    width = max(width, left + right + minimum_group_width * group_count)
    legend_categories = [
        Category(name.lower(), name, tuple(), color) for name, color, _ in series
    ]
    legend_rows = legend_row_count(
        legend_categories, width - left - right, item_spacing=126
    )
    legend_extra = max(0, legend_rows - 1) * 22
    top = 112 + legend_extra
    height += legend_extra
    plot_w = width - left - right
    plot_h = height - top - bottom
    raw_max_y = max([value for _, _, values in series for value in values] or [1])
    max_y = nice_axis_max(raw_max_y * (1.14 if show_value_labels else 1.0))
    group_w = plot_w / max(1, len(x_labels))
    bar_gap = 4
    bar_w = max(4, (group_w * 0.7 - bar_gap * max(0, len(series) - 1)) / max(1, len(series)))

    parts = svg_header(width, height)
    parts.extend(chart_title_svg(left, title, subtitle))
    parts.extend(
        legend_svg(
            legend_categories,
            left,
            86,
            max_width=plot_w,
            item_spacing=126,
        )
    )
    parts.append(plot_background_svg(left, top, plot_w, plot_h))
    parts.append(axis_svg(left, top, plot_w, plot_h, y_label=y_label))
    for tick in range(0, 5):
        y = top + plot_h - (plot_h * tick / 4)
        label = max_y * tick / 4
        parts.append(f'<line class="grid" x1="{left}" y1="{y:.2f}" x2="{left + plot_w}" y2="{y:.2f}"/>')
        parts.append(
            f'<text class="axis-label" x="{left - 8}" y="{y + 4:.2f}" '
            f'text-anchor="end">{format_number(label)}</text>'
        )

    for group_index, label in enumerate(x_labels):
        group_x = left + group_w * group_index + group_w * 0.14
        for series_index, (name, color, values) in enumerate(series):
            value = values[group_index] if group_index < len(values) else 0.0
            bar_h = value / max_y * plot_h
            x = group_x + series_index * (bar_w + bar_gap)
            y = top + plot_h - bar_h
            parts.append(
                f'<rect class="bar" x="{x:.2f}" y="{y:.2f}" width="{bar_w:.2f}" '
                f'height="{bar_h:.2f}" fill="{color}">'
                f'<title>{escape(name)} {escape(label)}: '
                f'{format_graph_number(value)}{escape(value_suffix)}</title></rect>'
            )
            if show_value_labels and value > 0 and bar_w >= 10:
                value_y = max(top + 14, y - 5)
                parts.append(
                    f'<text class="value-label" x="{x + bar_w / 2:.2f}" y="{value_y:.2f}" '
                    f'text-anchor="middle">{escape(format_graph_number(value))}{escape(value_suffix)}</text>'
                )
        label_x = left + group_w * group_index + group_w / 2
        parts.append(
            multiline_text_svg(
                "x-label",
                label_x,
                top + plot_h + 24,
                label_lines(label),
                anchor="middle",
            )
        )

    parts.append(svg_footer())
    return "\n".join(parts)


def svg_header(width: int, height: int) -> list[str]:
    return [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}">',
        "<style>",
        "text{font-family:Arial,Helvetica,sans-serif;fill:#111827}.title{font-size:22px;font-weight:700}",
        ".subtitle{font-size:13px;fill:#4b5563}.axis-label,.x-label,.legend,.axis-title{font-size:13px}",
        ".value-label{font-size:12px;font-weight:700;fill:#111827}"
        ".grid{stroke:#d1d5db;stroke-width:1}.tick{stroke:#6b7280;stroke-width:1}",
        ".axis{stroke:#374151;stroke-width:1.3}"
        ".series{fill:none;stroke-width:2.1;opacity:.92}.bar{opacity:.9}"
        ".plot-bg{fill:#f8fafc;stroke:#e5e7eb}",
        "</style>",
        f'<rect width="{width}" height="{height}" fill="#ffffff"/>',
    ]


def svg_footer() -> str:
    return "</svg>"


def axis_svg(left: int, top: int, plot_w: int, plot_h: int, *, y_label: str) -> str:
    bottom = top + plot_h
    return "\n".join(
        [
            f'<line class="axis" x1="{left}" y1="{top}" x2="{left}" y2="{bottom}"/>',
            f'<line class="axis" x1="{left}" y1="{bottom}" x2="{left + plot_w}" y2="{bottom}"/>',
            f'<text class="axis-label" x="{left}" y="{top - 12}">{escape(y_label)}</text>',
        ]
    )


def legend_svg(
    categories: Sequence[Category],
    x: int,
    y: int,
    *,
    max_width: int | None = None,
    item_spacing: int = 100,
    row_spacing: int = 22,
) -> list[str]:
    """Render a width-aware legend instead of clipping large dictionaries."""

    parts: list[str] = []
    for category, cursor, row in _legend_layout(
        categories, max_width=max_width, item_spacing=item_spacing
    ):
        row_y = y + row * row_spacing
        parts.append(
            f'<rect x="{x + cursor}" y="{row_y - 11}" width="12" height="12" '
            f'fill="{category.color}"/>'
        )
        parts.append(
            f'<text class="legend" x="{x + cursor + 17}" y="{row_y}">'
            f'{escape(category.display)}</text>'
        )
    return parts


def legend_row_count(
    categories: Sequence[Category],
    max_width: int,
    *,
    item_spacing: int = 100,
) -> int:
    layout = _legend_layout(
        categories, max_width=max_width, item_spacing=item_spacing
    )
    return max((row for _, _, row in layout), default=-1) + 1


def _legend_layout(
    categories: Sequence[Category],
    *,
    max_width: int | None,
    item_spacing: int,
) -> list[tuple[Category, int, int]]:
    available_width = max_width if max_width is not None else 1_000_000
    if available_width <= 0:
        raise ValueError("Legend width must be positive")
    layout: list[tuple[Category, int, int]] = []
    cursor = 0
    row = 0
    for category in categories:
        # Conservative estimate for 13px Arial plus marker and spacing.
        cell_width = max(item_spacing, 34 + len(category.display) * 8)
        if cursor and cursor + cell_width > available_width:
            row += 1
            cursor = 0
        layout.append((category, cursor, row))
        cursor += min(cell_width, available_width)
    return layout


def short_label(label: str, limit: int = 28) -> str:
    return label if len(label) <= limit else label[: limit - 1] + "..."


def chart_title_svg(left: int, title: str, subtitle: str = "") -> list[str]:
    parts = [f'<text class="title" x="{left}" y="32">{escape(title)}</text>']
    if subtitle:
        parts.append(f'<text class="subtitle" x="{left}" y="54">{escape(subtitle)}</text>')
    return parts


def plot_background_svg(left: int, top: int, plot_w: int, plot_h: int) -> str:
    return f'<rect class="plot-bg" x="{left}" y="{top}" width="{plot_w}" height="{plot_h}"/>'


def segment_tick_indexes(count: int, *, desired: int) -> list[int]:
    if count <= 0:
        return []
    if count == 1:
        return [0]
    ticks = {0, count - 1}
    for index in range(1, desired - 1):
        ticks.add(round((count - 1) * index / (desired - 1)))
    return sorted(ticks)


def nice_axis_max(value: float) -> float:
    if value <= 0:
        return 1.0
    magnitude = 10 ** math.floor(math.log10(value))
    for multiple in (1, 1.2, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10):
        candidate = multiple * magnitude
        if candidate >= value:
            return candidate
    return 10 * magnitude


def label_lines(label: str) -> list[str]:
    lines = [line.strip() for line in str(label).split("\n") if line.strip()]
    return lines or [""]


def multiline_text_svg(class_name: str, x: float, y: float, lines: Sequence[str], *, anchor: str) -> str:
    escaped_lines = [escape(short_label(line, 26)) for line in lines]
    tspans = []
    for index, line in enumerate(escaped_lines):
        dy = "0" if index == 0 else "15"
        tspans.append(f'<tspan x="{x:.2f}" dy="{dy}">{line}</tspan>')
    return (
        f'<text class="{class_name}" x="{x:.2f}" y="{y:.2f}" '
        f'text-anchor="{anchor}">{"".join(tspans)}</text>'
    )


def display_name(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("_", " ")).strip()


def speaker_axis_label(speaker: str) -> str:
    words = display_name(speaker).split()
    if len(words) <= 1:
        return display_name(speaker)
    return "\n".join(words[:2])


def video_axis_label(video: str) -> str:
    info = parse_video_name(video)
    order = info.get("video_order") or video[:3]
    date = format_date_label(info.get("date", ""))
    return f"{order}\n{date}" if date else order


def full_video_label(video: str) -> str:
    info = parse_video_name(video)
    if not info.get("video_order"):
        return display_name(video)
    country = info.get("country", "")
    person = display_name(info.get("person", ""))
    date = format_date_label(info.get("date", ""))
    pieces = [info["video_order"], date, country, person]
    return " ".join(piece for piece in pieces if piece)


def format_date_label(date: str) -> str:
    if re.fullmatch(r"\d{8}", date):
        return f"{date[:4]}-{date[4:6]}-{date[6:]}"
    if re.fullmatch(r"\d{6}unknown", date):
        return f"{date[:4]}-{date[4:6]} unknown"
    if re.fullmatch(r"\d{4}unknown", date):
        return f"{date[:4]} unknown"
    return date


def format_graph_number(value: float) -> str:
    if abs(value) >= 100:
        return f"{value:.0f}"
    if abs(value) >= 10:
        return f"{value:.1f}".rstrip("0").rstrip(".")
    if abs(value) >= 1:
        return f"{value:.1f}".rstrip("0").rstrip(".")
    return f"{value:.2f}".rstrip("0").rstrip(".")


def float_or_zero(value: object) -> float:
    parsed = parse_number(value)
    return parsed if parsed is not None else 0.0


def escape(text: object) -> str:
    return html.escape(str(text), quote=True)


def build_report_model(
    input_dir: Path,
    output_dir: Path,
    whisper_root: Path,
    prepare_root: Path,
    videos: Sequence[VideoAnalysis],
    speaker_rows: Sequence[Mapping[str, object]],
    graph_paths: Sequence[Path],
    *,
    segment_sample_counts: Sequence[int],
    provenance_status: str,
) -> TextReportModel:
    """Create the single data model used by both reports and the run log."""

    categories = categories_for_videos(videos)
    quick_rows = tuple(
        SpeakerQuickRow(
            speaker_id=str(row.get("speaker_id", "")),
            speaker=str(row.get("speaker", "")),
            videos=row.get("videos_count", ""),
            valid_segments=row.get("segments_with_terms", ""),
            terms=row.get("rocksteady_terms_total", ""),
            positive_proportion=row.get("positive_proportion", ""),
            negative_proportion=row.get("negative_proportion", ""),
            valence=row.get("valence_score", ""),
        )
        for row in speaker_rows
    )
    descriptor_rows = sum(
        len(
            descriptor_names(
                [segment_enriched_row(segment) for segment in valid_segments(video.segments)],
                video.categories,
            )
        )
        for video in videos
    )
    return TextReportModel(
        input_dir=input_dir,
        output_dir=output_dir,
        whisper_root=whisper_root,
        prepare_root=prepare_root,
        video_count=len(videos),
        speaker_count=len(speaker_rows),
        segment_count=sum(len(video.segments) for video in videos),
        source_segment_count=sum(
            video.source_whisper_segment_count for video in videos
        ),
        rocksteady_row_count=sum(video.rocksteady_row_count for video in videos),
        matched_segment_count=sum(
            segment.rocksteady_row_available
            for video in videos
            for segment in video.segments
        ),
        missing_segment_count=sum(len(video.missing_segment_ids) for video in videos),
        ignored_segment_count=sum(len(video.ignored_segment_ids) for video in videos),
        valid_segment_count=sum(len(valid_segments(video.segments)) for video in videos),
        zero_term_segment_count=sum(
            segment.rocksteady_row_available and segment.terms <= 0
            for video in videos
            for segment in video.segments
        ),
        terms_total=sum(
            float_or_zero(video.summary.get("rocksteady_terms_total")) for video in videos
        ),
        descriptor_rows=descriptor_rows,
        graph_count=len(graph_paths),
        segment_sample_counts=tuple(segment_sample_counts),
        alignment_policies=tuple(sorted({video.alignment_policy for video in videos})),
        text_languages=tuple(
            sorted(
                {
                    segment.text_language
                    for video in videos
                    for segment in video.segments[:1]
                }
            )
        ),
        provenance_status=provenance_status,
        alignment_mapping_statuses=tuple(
            sorted({video.alignment_contract.status for video in videos})
        ),
        categories=tuple(
            ReportCategory(
                category=category,
                available_videos=sum(
                    category.key in video.available_categories for video in videos
                ),
            )
            for category in categories
        ),
        speakers=quick_rows,
    )


def write_output_manifest(
    output_dir: Path,
    input_dir: Path,
    whisper_root: Path,
    prepare_root: Path,
    videos: Sequence[VideoAnalysis],
    graph_paths: Sequence[Path],
    *,
    categories: Sequence[Category],
    upstream_provenance: UpstreamProvenance,
    write_graphs: bool,
    segment_sample_counts: Sequence[int],
    segment_alignment: str,
    text_language: str,
    run_id: str,
    output_variant: str,
    previous_ownership_state: str,
    readable_summary: Mapping[str, object],
) -> Path:
    """Describe every output family and the run that produced it."""

    segment_count = sum(len(video.segments) for video in videos)
    valid_segment_count = sum(len(valid_segments(video.segments)) for video in videos)
    missing_segment_count = sum(len(video.missing_segment_ids) for video in videos)
    ignored_segment_count = sum(len(video.ignored_segment_ids) for video in videos)
    payload = {
        "schema_version": "2.1",
        "kind": "text-postprocessing-variant",
        "run_id": run_id,
        "variant": output_variant,
        "status": "completed",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "inputs": {
            "rocksteady_csv_root": str(input_dir),
            "whisper_json_root": str(whisper_root),
            "prepare_mapping_root": str(prepare_root),
        },
        "upstream_provenance": upstream_provenance.as_dict(),
        "config": {
            "rocksteady_value_type": "total",
            "text_language": text_language,
            "segment_alignment": segment_alignment,
            "write_graphs": write_graphs,
            "segment_sample_counts": list(segment_sample_counts),
            "mean_permutation_test": {
                "permutations": DEFAULT_PERMUTATIONS,
                "random_seed": DEFAULT_RANDOM_SEED,
                "multiple_comparison_adjustment": "Holm",
            },
        },
        "publication": {
            "ownership_schema": "1.0",
            "previous_target_state": previous_ownership_state,
            "atomic_directory_replacement": True,
        },
        "summary": {
            "videos": len(videos),
            "segments": segment_count,
            "source_whisper_segments": sum(
                video.source_whisper_segment_count for video in videos
            ),
            "valid_segments": valid_segment_count,
            "missing_rocksteady_segments": missing_segment_count,
            "ignored_rocksteady_rows": ignored_segment_count,
            "graphs": len(graph_paths),
        },
        "artifacts": {
            "start_here": {
                "path": "video_level_summary.csv",
                "description": "One row per video; recommended first table for analysis.",
            },
            "speaker_summary": {
                "path": "speaker_level_summary.csv",
                "description": "One row per speaker, aggregated from video totals.",
            },
            "video_mean_comparisons": {
                "root": "video_mean_comparisons/",
                "combined_pattern": (
                    "video_mean_comparisons/<Country>/<Speaker>/combined/"
                    "{video_means.csv,permutation_test_results.csv}"
                ),
                "description": (
                    "Per-video weighted means and reproducible segment-label "
                    "permutation tests with Holm-adjusted p-values."
                ),
            },
            "descriptor_statistics": {
                "path": "descriptor_statistics_by_video.csv",
                "description": "Long-format descriptive statistics by video and descriptor.",
            },
            "descriptor_statistics_workbook": {
                "path": "descriptor_statistics_by_video.xlsx",
                "description": (
                    "One worksheet per video with screenshot-style Statistic/Value "
                    "tables for every descriptor."
                ),
            },
            "alignment_audit": {
                "path": "segment_alignment_audit.csv",
                "description": "Per-video Whisper/RockSteady matching evidence.",
            },
            "readable_tables": {
                "root": "readable/",
                "start_here": "readable/video_level_summary.csv",
                "speaker": "readable/speaker_level_summary.csv",
                "segment_pattern": (
                    "readable/segment_level/<Speaker>/<Video>_segments.csv "
                    "(legacy: <Country>/<Speaker>/<Video>_segments.csv)"
                ),
                "category_guide": "readable/category_guide.csv",
                "summary": readable_summary,
                "description": "Plain headers and percentages for human reading.",
            },
            "segment_counts": {
                "path_pattern": (
                    "segment_counts/<Speaker>/<Video>_segment_counts.csv "
                    "(legacy: <Country>/<Speaker>/<Video>_segment_counts.csv)"
                ),
                "description": "Raw RockSteady Total category counts per segment.",
            },
            "segment_relative": {
                "path_pattern": (
                    "segment_relative/<Speaker>/<Video>_segment_relative.csv "
                    "(legacy: <Country>/<Speaker>/<Video>_segment_relative.csv)"
                ),
                "description": "Category count divided by RockSteady Terms per segment.",
            },
            "segment_enriched": {
                "path_pattern": (
                    "segment_level/<Speaker>/<Video>_segments_enriched.csv "
                    "(legacy: <Country>/<Speaker>/<Video>_segments_enriched.csv)"
                ),
                "description": "Counts, proportions, text, timestamps and availability flags.",
            },
            "method_reports": ["POSTPROCESSING_REPORT.md", "POSTPROCESSING_REPORT_EN.md"],
            "run_log": "run_log.txt",
            "graphs_root": "graphs/" if write_graphs else None,
        },
        "categories": [
            {
                "key": category.key,
                "display": category.display,
                "color": category.color,
                "required": category.required,
                "source_column": category.source_column or category.source_names[0],
                "source_names": list(category.source_names),
                "available_videos": sum(
                    category.key in video.available_categories for video in videos
                ),
            }
            for category in categories
        ],
        "videos": [
            {
                "country": video.country,
                "speaker": video.speaker,
                "speaker_id": video.speaker_id,
                "video": video.video,
                "source_id": video.source_id,
                "analysis_segments": len(video.segments),
                "source_whisper_segments": video.source_whisper_segment_count,
                "valid_segments": len(valid_segments(video.segments)),
                "alignment_status": (
                    "exact"
                    if not video.missing_segment_ids and not video.ignored_segment_ids
                    else "reconciled"
                ),
                "alignment_mapping": video.alignment_contract.as_dict(),
                "rocksteady_csv": str(video.path),
                "rocksteady_csv_sha256": sha256_file(video.path),
                "whisper_json": str(video.whisper_path),
                "whisper_json_sha256": sha256_file(video.whisper_path),
            }
            for video in videos
        ],
    }
    payload["output_inventory"] = [
        {
            "path": path.relative_to(output_dir).as_posix(),
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }
        for path in sorted(output_dir.rglob("*"))
        if path.is_file() and path.name != "output_manifest.json"
    ]
    path = output_dir / "output_manifest.json"
    atomic_write_json(path, payload)
    return path


def write_dict_rows(
    path: Path,
    rows: Sequence[Mapping[str, object]],
    fieldnames: Sequence[str] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = infer_fieldnames(rows)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = SpreadsheetSafeWriter(
            csv.DictWriter(handle, fieldnames=list(fieldnames), extrasaction="ignore")
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({key: format_cell(row.get(key, "")) for key in fieldnames})


def infer_fieldnames(rows: Sequence[Mapping[str, object]]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return fields


def format_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float):
        return format_number(value)
    return value


def format_number(value: object) -> str:
    if not isinstance(value, (int, float)):
        return str(value)
    number = float(value)
    if not math.isfinite(number):
        return ""
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    if abs(number) < 1:
        return f"{number:.6f}".rstrip("0").rstrip(".")
    return f"{number:.2f}".rstrip("0").rstrip(".")


def build_single_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Postprocess one RockSteady segment-level Total CSV variant.",
        epilog=(
            "For the normal selected+extra workflow, use "
            "`python -m analysis.text_pipeline.postprocess analyse-pair --help`."
        ),
    )
    parser.add_argument(
        "input_folder",
        nargs="?",
        default=Path("processing/text_analysis/output/current/rocksteady/core"),
        type=Path,
        help=(
            "Speaker/Video or legacy Country/Speaker/Video RockSteady CSV root "
            "(default: processing/text_analysis/output/current/rocksteady/core)."
        ),
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=None,
        help="Output folder. Defaults to analysis/output/text/text_output.",
    )
    parser.add_argument(
        "--whisper-root",
        type=Path,
        default=None,
        help=(
            "Root folder containing matching Whisper JSON files. By default, "
            "core/all outputs use output/current/selected_transcripts; "
            "original/eng inputs select their matching transcript trees."
        ),
    )
    parser.add_argument(
        "--prepare-root",
        type=Path,
        default=None,
        help=(
            "Root containing Speaker/Video (or legacy Country/Speaker/Video) "
            ".prepare_manifest.json mappings. "
            "Manifest-backed pipeline input defaults to "
            "processing/text_analysis/output/current/prepared_segments; pass this "
            "option explicitly when attaching a mapping to standalone CSVs."
        ),
    )
    parser.add_argument(
        "--segment-samples",
        type=parse_segment_sample_counts,
        default=DEFAULT_SEGMENT_SAMPLE_COUNTS,
        metavar="ODD_COUNTS",
        help=(
            "Comma-separated odd sample sizes for simplified segment timelines "
            "(default: 3,5,7,9; use 'none' to keep only the full timeline)."
        ),
    )
    parser.add_argument(
        "--segment-alignment",
        choices=SEGMENT_ALIGNMENT_POLICIES,
        default="error",
        help=(
            "How to handle RockSteady/Whisper segment differences: error (default) "
            "or reconcile by Title segment ID, ignoring extra IDs and retaining "
            "explicit placeholders for missing RockSteady rows."
        ),
    )
    parser.add_argument(
        "--text-lang",
        default="original",
        help=(
            "Whisper text attached to bilingual segment outputs: original (default), "
            "en, or another available language key. Use the same language that was "
            "sent to RockSteady."
        ),
    )
    parser.add_argument("--no-graphs", action="store_true", help="Skip SVG graph generation.")
    parser.add_argument(
        "--run-id",
        default=None,
        help="Optional parent-pipeline run ID; a UUID is generated when omitted.",
    )
    return parser


def build_pair_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python -m analysis.text_pipeline.postprocess analyse-pair",
        description=(
            "Generate selected and extra text reports in one transaction. The visible "
            "output root changes only after both variants and their source identities validate."
        ),
    )
    parser.add_argument(
        "selected_input_folder",
        nargs="?",
        type=Path,
        default=Path("processing/text_analysis/output/current/rocksteady/core"),
        help="Selected/core RockSteady CSV root.",
    )
    parser.add_argument(
        "extra_input_folder",
        nargs="?",
        type=Path,
        default=Path("processing/text_analysis/output/current/rocksteady/all"),
        help="Full/extra RockSteady CSV root.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=None,
        help=(
            "Common parent published with selected/ and extra/ children. Defaults to "
            "analysis/output/text/text_output."
        ),
    )
    parser.add_argument(
        "--whisper-root",
        type=Path,
        default=None,
        help="Common matching Whisper JSON root (defaults to output/current/selected_transcripts).",
    )
    parser.add_argument(
        "--prepare-root",
        type=Path,
        default=None,
        help="Common Speaker/Video or legacy Country/Speaker/Video preparation-mapping root.",
    )
    parser.add_argument(
        "--segment-samples",
        type=parse_segment_sample_counts,
        default=DEFAULT_SEGMENT_SAMPLE_COUNTS,
        metavar="ODD_COUNTS",
        help="Comma-separated odd timeline sample sizes; use 'none' for only full timelines.",
    )
    parser.add_argument(
        "--segment-alignment",
        choices=SEGMENT_ALIGNMENT_POLICIES,
        default="error",
        help="RockSteady/Whisper mismatch policy (default: error).",
    )
    parser.add_argument(
        "--text-lang",
        default="original",
        help="Whisper text language sent to RockSteady and attached to output rows.",
    )
    parser.add_argument("--no-graphs", action="store_true", help="Skip SVG graph generation.")
    parser.add_argument(
        "--run-id",
        default=None,
        help="Optional parent-pipeline run ID; a UUID is generated when omitted.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> None:
    arguments = list(sys.argv[1:] if argv is None else argv)
    if arguments and arguments[0] == "analyse-pair":
        from analysis.text_pipeline.batch import analyse_text_segment_pair

        args = build_pair_parser().parse_args(arguments[1:])
        pair = analyse_text_segment_pair(
            args.selected_input_folder,
            args.extra_input_folder,
            output_root=args.output_root,
            whisper_root=args.whisper_root,
            prepare_root=args.prepare_root,
            write_graphs=not args.no_graphs,
            segment_sample_counts=args.segment_samples,
            segment_alignment=args.segment_alignment,
            text_language=args.text_lang,
            run_id=args.run_id,
        )
        print(f"Run ID:            {pair.run_id}")
        print(f"Output family:     {pair.output_root}")
        print(f"Video identities:  {pair.identity_count}")
        print(f"Selected summary:  {pair.selected.video_summary_path}")
        print(f"Extra summary:     {pair.extra.video_summary_path}")
        print(f"Batch manifest:    {pair.batch_manifest_path}")
        return

    parser = build_single_parser()
    args = parser.parse_args(arguments)

    result = analyse_text_segments_folder(
        args.input_folder,
        output_root=args.output_root,
        whisper_root=args.whisper_root,
        prepare_root=args.prepare_root,
        write_graphs=not args.no_graphs,
        segment_sample_counts=args.segment_samples,
        segment_alignment=args.segment_alignment,
        text_language=args.text_lang,
        run_id=args.run_id,
    )
    print(f"Run ID:            {result.run_id}")
    print(f"Input folder:      {result.input_dir}")
    print(f"Output folder:     {result.output_dir}")
    print(f"Video CSV files:   {result.csv_count}")
    print(f"Segment rows:      {result.segment_count}")
    print(f"Video summary:     {result.video_summary_path}")
    print(f"Speaker summary:   {result.speaker_summary_path}")
    print(f"Descriptor stats:  {result.descriptor_path}")
    print(f"Alignment audit:   {result.alignment_audit_path}")
    print(f"Output manifest:   {result.output_manifest_path}")
    print(f"Graphs:            {len(result.graph_paths)}")


if __name__ == "__main__":
    main()
