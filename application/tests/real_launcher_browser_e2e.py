"""Opt-in, real HTTP/browser/subprocess acceptance (never collected by pytest).

Run with a Python environment containing the project's runtime requirements::

    python application/tests/real_launcher_browser_e2e.py --output <new-directory> \
        --playwright <playwright-package> --browser <browser-executable>

Only EULA/settings path resolution is redirected into the new output directory.
No HTTP routes, command builders, process runners, or analytical functions are
mocked. Synthetic imported statistics are explicitly distinct from model output.
"""
from __future__ import annotations

import argparse
import contextlib
import csv
import hashlib
import json
import os
from pathlib import Path
import shutil
import subprocess
import sys
import threading
import traceback

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_fixtures(root: Path, video: Path | None) -> dict[str, str]:
    from analysis.combined_summary import AUDIO_METRICS

    media = root / "fixtures" / "Synthetic Speaker" / "synthetic-tone.mp4"
    media.parent.mkdir(parents=True)
    corrupt = media.with_name("invalid-video.mp4")
    corrupt.write_bytes(b"Deliberately invalid video for real subprocess failure acceptance.\n")
    if video:
        shutil.copyfile(video, media)
    else:
        subprocess.run([
            "ffmpeg", "-hide_banner", "-loglevel", "error", "-f", "lavfi",
            "-i", "testsrc2=size=320x240:rate=30:duration=12", "-f", "lavfi",
            "-i", "sine=frequency=440:sample_rate=48000:duration=12", "-c:v",
            "libx264", "-pix_fmt", "yuv420p", "-c:a", "aac", "-shortest", str(media),
        ], check=True)
    reports = root / "fixtures" / "synthetic-imported-audio-statistics"
    report = reports / "emotion" / "Synthetic Researcher" / "combined" / "other_findings" / "descriptive_statistics.csv"
    report.parent.mkdir(parents=True)
    source_ids = [f"source-{i:04d}" for i in range(1, 5)]
    with report.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for offset, metric in enumerate(AUDIO_METRICS):
            writer.writerows([
                [metric], ["classification", "core", "category", "emotion", "unit", "score"],
                ["metric", *source_ids], ["count", 2, 8, 6, 4], ["missing", 0, 0, 0, 0],
                ["mean", *[10 * i + offset for i in range(1, 5)]],
                ["stddev", 1, 2, 3, 4], ["kurtosis", 0, 0, 0, 0], [],
            ])
    sources = [{
        "source_id": source_id, "speaker": "Synthetic Researcher",
        "speaker_display": "Synthetic Researcher", "selected": True,
        "source_kind": "local", "system_metadata": {"title": f"Synthetic observation {i}"},
        "user_metadata": {"Country": "Ireland" if i % 2 else "Japan"},
        "output_mapping": {"video_directory": str(reports / source_id)},
    } for i, source_id in enumerate(source_ids, 1)]
    manifest = reports / "source_manifest.json"
    manifest.write_text(json.dumps({
        "format_version": 1,
        "catalog": {"metadata_headers": ["Country"], "metadata_export_headers": {"Country": "Country"}},
        "sources": sources,
    }, indent=2), encoding="utf-8")
    with (reports / "source_metadata.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["SourceID", "Country"])
        writer.writerows((s["source_id"], s["user_metadata"]["Country"]) for s in sources)
    (reports / "SYNTHETIC_FIXTURE.txt").write_text(
        "Fabricated acceptance-test descriptive statistics, not human observations or model inference.\n"
        "Anger means: 10,20,30,40; counts: 2,8,6,4; country order: Ireland,Japan,Ireland,Japan.\n",
        encoding="utf-8",
    )
    return {"video": str(media), "corruptVideo": str(corrupt), "reports": str(reports), "manifest": str(manifest)}


def probe(path: Path) -> dict:
    return json.loads(subprocess.check_output([
        "ffprobe", "-v", "error", "-show_streams", "-show_format", "-of", "json", str(path),
    ], text=True, encoding="utf-8"))


def validate_supplied_video(path: Path) -> None:
    """Reject unsupported fixtures before settings, server, or pipeline setup."""
    if not path.is_file():
        raise ValueError(f"--video must name an existing 12-second tone video: {path}")
    try:
        metadata = probe(path)
        duration = float(metadata["format"]["duration"])
        stream_types = {stream["codec_type"] for stream in metadata["streams"]}
    except (OSError, subprocess.CalledProcessError, ValueError, KeyError, TypeError) as exc:
        raise ValueError(f"--video could not be read by FFprobe: {path}") from exc
    if not 11.8 <= duration <= 12.2:
        raise ValueError(
            f"--video requires a 12-second fixture (tolerance 0.2 seconds); "
            f"received {duration:.3f} seconds. Omit --video to generate a supported fixture."
        )
    if not {"video", "audio"} <= stream_types:
        raise ValueError("--video requires both video and audio streams; omit it to generate a tone fixture.")


def validate_artifacts(root: Path) -> dict:
    import math
    import openpyxl
    from analysis.audio import read_audio_analysis_csv

    result = {}
    for mode, expected in (("full", 12.0), ("focus", 5.0)):
        candidates = sorted((root / mode).rglob("*imotions*.mp4"))
        assert candidates, f"{mode}: no iMotions-compatible output found"
        movie = candidates[-1]
        metadata = probe(movie)
        duration = float(metadata["format"]["duration"])
        assert abs(duration - expected) <= .2, (mode, duration, expected)
        assert {s["codec_type"] for s in metadata["streams"]} >= {"video", "audio"}
        result[mode] = {"path": str(movie), "duration_seconds": duration, "sha256": sha(movie), "streams": metadata["streams"]}
        if mode == "focus":
            # Selected source spans are 1..3 and 6..8 seconds, with a one-second
            # black/silent gap at output 2..3. Decode away from boundaries.
            raw = subprocess.check_output([
                "ffmpeg", "-v", "error", "-ss", "2.5", "-i", str(movie),
                "-frames:v", "1", "-vf", "scale=16:16,format=gray", "-f", "rawvideo", "-",
            ])
            assert raw and max(raw) <= 4, f"Focus gap frame is not black: max={max(raw)}"
            pcm = subprocess.check_output([
                "ffmpeg", "-v", "error", "-ss", "2.25", "-i", str(movie),
                "-t", "0.5", "-vn", "-ac", "1", "-ar", "8000", "-f", "s16le", "-",
            ])
            import array
            samples = array.array("h", pcm)
            assert samples and max(abs(v) for v in samples) <= 10, "Focus gap is not silent"
            result[mode]["gap"] = {"frame_max_gray": max(raw), "audio_max_abs_pcm16": max(abs(v) for v in samples)}
    audio_files = list((root / "audio").rglob("audio_analysis.csv"))
    feature_files = list((root / "audio").rglob("opensmile_features.csv"))
    assert len(audio_files) == len(feature_files) == 1
    audio = read_audio_analysis_csv(audio_files[0])
    assert len(audio.rows) >= 2
    for row in audio.rows:
        for metric in ("Anger", "Contempt", "Disgust", "Fear", "Joy", "Sadness", "Surprise", "Neutral", "Other", "Arousal", "Dominance", "Valence"):
            assert row.get(metric, "") in ("", None), (metric, row.get(metric))
    with feature_files[0].open(encoding="utf-8-sig", newline="") as handle:
        features = list(csv.DictReader(handle))
    assert len(features) == len(audio.rows)
    numeric = [float(v) for row in features for k, v in row.items() if k not in {"Row", "WindowStart", "WindowEnd", "name", "frameTime"} and v]
    assert len(numeric) > 100 and all(math.isfinite(v) for v in numeric)
    result["audio"] = {"rows": len(audio.rows), "finite_feature_values": len(numeric), "emotion_columns_blank": True}
    workbook = root / "analysis" / "combined_analysis.xlsx"
    manifest = json.loads((workbook.parent / "combined_analysis_manifest.json").read_text(encoding="utf-8"))
    assert manifest["status"] == "complete", manifest.get("status")
    book = openpyxl.load_workbook(workbook, data_only=False)
    sheet = book["Audio"]
    # Source means and metadata must preserve identity under Country ordering.
    rows = list(sheet.values)
    anger = next(row for row in rows if "Anger" in row)
    source_columns = [3, 4, 7, 8]
    means = [anger[index] for index in source_columns]
    assert means == [10, 30, 20, 40], anger
    assert [rows[0][index] for index in source_columns] == [
        "Synthetic observation 1", "Synthetic observation 3",
        "Synthetic observation 2", "Synthetic observation 4",
    ]
    assert anger[-1] == "=AVERAGE(D2,E2,H2,I2)"
    all_values = [c.value for row in sheet for c in row]
    assert "Ireland" in all_values and "Japan" in all_values
    assert "Audio (probability)" in book.sheetnames or any("prob" in n.lower() for n in book.sheetnames), book.sheetnames
    assert not any(c.data_type == "e" for s in book for row in s for c in row), "Workbook has stored Excel error cells"
    profile = json.loads((workbook.parent / "analysis_profile.json").read_text(encoding="utf-8"))
    assert profile["automatic_group_field"] == "Country"
    result["analysis"] = {"path": str(workbook), "sha256": sha(workbook), "sheets": book.sheetnames, "anger_ordered_means": means, "overall_formula_expected_value": 25, "formula_validation": "Reference membership checked; workbook not recalculated in Excel", "profile": profile}
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--playwright", required=True)
    parser.add_argument("--browser", required=True)
    parser.add_argument("--node", default="node")
    parser.add_argument("--video", type=Path, help="Optional 12-second tone video (duration tolerance 0.2 seconds).")
    args = parser.parse_args()
    root = args.output.resolve()
    if root.exists() and any(root.iterdir()):
        parser.error("--output must be absent or empty, to preserve earlier evidence")
    from application import launcher
    launcher.initialize_launcher_ffmpeg_runtime()
    if args.video:
        try:
            validate_supplied_video(args.video)
        except ValueError as exc:
            parser.error(str(exc))
    root.mkdir(parents=True, exist_ok=True)
    original_settings = launcher.backend.ui_settings_path(REPO)
    original_paths = [original_settings, launcher.backend.settings_backup_path(original_settings), launcher.backend.eula_path(REPO)]
    original_paths.extend(launcher.backend.credential_store.secret_path(original_settings, name) for name in ("youtubeApiKey", "huggingFaceToken"))
    original_fingerprints = {p: sha(p) if p.is_file() else None for p in original_paths}
    # Configuration seam only. This process owns the isolated LauncherState;
    # production _local settings, credentials, and terms files stay untouched.
    launcher.backend.eula_path = lambda _repo: root / "state" / "eula.txt"
    launcher.backend.ui_settings_path = lambda _repo: root / "state" / "ui_settings.json"
    os.environ["MEA_CREDENTIAL_STORE_ROOT"] = str(root / "state" / "credentials")
    for name in ("YOUTUBE_API_KEY", "HF_TOKEN", "HUGGINGFACE_TOKEN", "HUGGING_FACE_HUB_TOKEN"):
        os.environ.pop(name, None)
    launcher.backend.write_eula_state(REPO, True)
    launcher.backend.save_ui_settings(REPO, {"resourceLimitsEnabled": False})
    fixtures = write_fixtures(root, args.video)
    fixture_hashes = {str(p.relative_to(root)): sha(p) for p in (root / "fixtures").rglob("*") if p.is_file()}
    server = launcher.LauncherHttpServer(("127.0.0.1", 0), launcher.VideoStackUiHandler)
    url = f"{launcher.launcher_origin(server.server_address)}/?token={launcher.API_TOKEN}"
    config = {"url": url, "token": launcher.API_TOKEN, "fixtures": fixtures, "output": str(root), "playwright": args.playwright, "browser": args.browser}
    config_file = root / "browser-config.json"
    config_file.write_text(json.dumps(config), encoding="utf-8")
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    report = {"scope": "Real browser / real launcher HTTP / actual subprocesses; only settings and EULA path resolvers redirected. No route mocks. Synthetic imported analysis statistics; no model inference claim.", "python": sys.executable}
    try:
        with (root / "launcher.log").open("w", encoding="utf-8") as log, contextlib.redirect_stdout(log), contextlib.redirect_stderr(log):
            thread.start()
            completed = subprocess.run([args.node, str(Path(__file__).with_name("real_launcher_browser_harness.js")), str(config_file)], cwd=REPO, text=True, encoding="utf-8", errors="replace", capture_output=True, timeout=900)
            (root / "browser.stdout.log").write_text(completed.stdout, encoding="utf-8")
            (root / "browser.stderr.log").write_text(completed.stderr, encoding="utf-8")
            assert completed.returncode == 0, f"Browser harness failed ({completed.returncode}); see browser.stderr.log"
            report["artifacts"] = validate_artifacts(root)
            assert fixture_hashes == {str(p.relative_to(root)): sha(p) for p in (root / "fixtures").rglob("*") if p.is_file()}, "Inputs changed during processing"
            report["fixtures_unchanged"] = True
            assert original_fingerprints == {p: sha(p) if p.is_file() else None for p in original_paths}, "Original settings, EULA, or credential files changed"
            report["original_configuration_unchanged"] = True
            report["status"] = "passed"
    except Exception:
        report["status"] = "failed"
        report["error"] = traceback.format_exc()
    finally:
        launcher.terminate_active_process("Stopping isolated acceptance-test process.")
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)
        config_file.unlink(missing_ok=True)  # Ephemeral bootstrap token is not durable evidence.
        (root / "acceptance.json").write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0 if report["status"] == "passed" else 1


if __name__ == "__main__":
    raise SystemExit(main())
