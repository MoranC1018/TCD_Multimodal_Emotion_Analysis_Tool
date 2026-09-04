"""Opt-in public CLI acceptance with real subprocesses and independently read outputs.

Never collected by pytest. Run with the installed project Python environment::

    python application/tests/real_cli_e2e.py --output NEW --phase nonmodel
    python application/tests/real_cli_e2e.py --output NEW --phase models --model-video FIXTURE

Models require an existing offline cache and a 15-second face/speech fixture.
No CLI routes, builders, runners, decoders or analytical functions are mocked.
"""
from __future__ import annotations

import argparse
import array
import csv
import hashlib
import io
import json
import math
import os
from pathlib import Path
import shutil
import statistics
import subprocess
import sys
import time
import traceback

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))
from application.tests.real_launcher_browser_e2e import write_fixtures


def sha(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def save(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def csv_rows(path: Path) -> list[dict]:
    content = path.read_text(encoding="utf-8-sig")
    if "#DATA\n" in content:
        content = content.split("#DATA\n", 1)[1]
    return list(csv.DictReader(io.StringIO(content)))


def close(actual, expected, tolerance=1e-8) -> None:
    assert isinstance(actual, (int, float)) and math.isfinite(actual), actual
    assert abs(actual - expected) <= tolerance, (actual, expected, tolerance)


def only(root: Path, pattern: str) -> Path:
    matches = sorted(root.rglob(pattern))
    assert len(matches) == 1, (str(root), pattern, [str(p) for p in matches])
    return matches[0]


def ref(step: str) -> dict:
    return {"from_step": step, "output": "output_root"}


def source_snapshot() -> dict:
    """Record source identity even when acceptance runs before a local commit."""
    command = ["git", "-c", f"safe.directory={REPO.as_posix()}", "rev-parse", "HEAD"]
    result = subprocess.run(command, cwd=REPO, capture_output=True, text=True, check=True)
    sources = {}
    for directory in ("application", "procurement", "processing", "analysis"):
        for path in sorted((REPO / directory).rglob("*.py")):
            if "__pycache__" not in path.parts:
                sources[path.relative_to(REPO).as_posix()] = sha(path)
    return {"head": result.stdout.strip(), "python_sources_sha256": sources}


class Acceptance:
    def __init__(self, root: Path, phase: str, model_video: Path | None, max_ram_percent: float = 90):
        self.root, self.phase, self.model_video = root, phase, model_video
        self.results: list[dict] = []
        self.sequence = 0
        self.env = {
            **os.environ, "PYTHONUTF8": "1", "PYTHONUNBUFFERED": "1",
            "HF_HUB_OFFLINE": "1", "TRANSFORMERS_OFFLINE": "1",
            **{name: "4" for name in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS", "ONNX_NUM_THREADS", "ORT_NUM_THREADS", "MEA_NATIVE_THREADS")},
        }
        self.resources = {"nativeThreads": 4, "maxCpuCores": 4, "maxRamPercent": max_ram_percent}
        from procurement.external_tools import resolve_media_binary
        self.ffmpeg, self.ffprobe = map(str, (resolve_media_binary("ffmpeg"), resolve_media_binary("ffprobe")))
        self.env["PATH"] = str(Path(self.ffmpeg).parent) + os.pathsep + self.env.get("PATH", "")
        self.fixtures: dict[str, str] = {}

    def record(self, name: str, action) -> None:
        began = time.monotonic()
        print(f"START {name}", flush=True)
        try:
            evidence = action() or {}
            result = {"case": name, "state": "passed", "evidence": evidence}
        except Exception as exc:
            result = {"case": name, "state": "failed", "error": f"{type(exc).__name__}: {exc}", "traceback": traceback.format_exc()}
        result["seconds"] = round(time.monotonic() - began, 3)
        self.results.append(result)
        save(self.root / "acceptance.json", self.summary())
        print(f"{result['state'].upper()} {name}: {result.get('error', '')}", flush=True)

    def summary(self) -> dict:
        return {
            "state": "passed" if self.results and all(r["state"] == "passed" for r in self.results) else "failed",
            "phase": self.phase, "repo": str(REPO), "python": sys.executable, "resources": self.resources,
            "cases": self.results,
            "limits": "Synthetic operational acceptance. No population accuracy claim, all-platform claim, Excel recalculation or scientific-publication approval. Only the named phase/cases are covered.",
        }

    def invoke(self, *args, expected=0, timeout=180) -> dict:
        self.sequence += 1
        label = f"{self.sequence:03d}-{args[0]}"
        command = [sys.executable, "-m", "application.cli", *map(str, args)]
        result = subprocess.run(command, cwd=REPO, env=self.env, stdin=subprocess.DEVNULL, capture_output=True, text=True, encoding="utf-8", timeout=timeout)
        logs = self.root / "calls"
        logs.mkdir(exist_ok=True)
        (logs / f"{label}.stdout.json").write_text(result.stdout, encoding="utf-8")
        (logs / f"{label}.stderr.log").write_text(result.stderr, encoding="utf-8")
        save(logs / f"{label}.command.json", {"command": command, "returncode": result.returncode})
        payload = json.loads(result.stdout)  # Reject banners or multiple stdout values.
        assert isinstance(payload, dict), payload
        assert result.returncode == expected, {"expected": expected, "actual": result.returncode, "payload": payload, "stderr_log": str(logs / f"{label}.stderr.log")}
        return payload

    def job(self, name: str, steps: list[dict], *, folder: Path | None = None) -> Path:
        path = (folder or self.root / "jobs") / f"{name}.json"
        save(path, {"schema_version": 1, "resources": self.resources, "steps": steps})
        return path

    def run_job(self, name: str, steps: list[dict], *, expected=0, timeout=300) -> dict:
        payload = self.invoke("run", "--job", self.job(name, steps), "--run-dir", self.root / "runs" / name, expected=expected, timeout=timeout)
        if expected == 0:
            assert payload["state"] == "completed" and all(s["state"] == "completed" for s in payload["steps"]), payload
            for output in payload["outputs"].values():
                assert Path(output["output_root"]).is_dir(), output
        return payload

    def probe(self, path: Path) -> dict:
        return json.loads(subprocess.check_output([self.ffprobe, "-v", "error", "-show_streams", "-show_format", "-of", "json", str(path)], env=self.env, text=True, encoding="utf-8"))

    def validate_movie(self, root: Path, duration: float) -> dict:
        movie = only(root, "*imotions*.mp4")
        metadata = self.probe(movie)
        actual = float(metadata["format"]["duration"])
        close(actual, duration, .25)
        assert {s["codec_type"] for s in metadata["streams"]} >= {"audio", "video"}
        return {"path": str(movie), "duration_seconds": actual, "sha256": sha(movie)}

    def metadata_commands(self) -> dict:
        schema = self.invoke("schema")
        assert {"procurement", "face", "audio", "text", "analysis"} <= set(schema["x-stage-options"])
        self.invoke("schema", "--stage", "native.face")
        self.invoke("settings")
        doctor = self.invoke("doctor", "--component", "procurement", "--device", "cpu")
        assert doctor["state"] == "ready"
        scan = self.invoke("inspect", "source", self.fixtures["video"], "--no-enrich")
        assert scan["groups"] and scan["source_kind"] == "file", scan
        return {"stage_count": len(schema["x-stage-options"]), "source_scan": scan, "doctor": doctor}

    def dry_run(self) -> dict:
        directory = self.root / "runs" / "dry-run"
        job = self.job("dry-run", [
            {"id": "full", "stage": "procurement", "options": {"source_path": self.fixtures["video"], "mode": "full"}},
            {"id": "audio", "stage": "audio", "options": {"source_path": ref("full"), "include_emotions": False}},
        ])
        payload = self.invoke("run", "--job", job, "--run-dir", directory, "--dry-run")
        assert payload["state"] == "dry_run" and payload["steps"][1]["state"] == "deferred"
        assert not directory.exists()
        assert not Path(payload["steps"][0]["output_root"]).exists()
        self.invoke("validate", "--job", job, "--run-dir", directory)
        assert not directory.exists()
        return payload

    def full_audio(self) -> dict:
        result = self.run_job("full-audio", [
            {"id": "full", "stage": "procurement", "options": {"source_path": self.fixtures["video"], "mode": "full"}},
            {"id": "audio", "stage": "audio", "options": {"source_path": ref("full"), "include_emotions": False, "device": "cpu", "window_seconds": 6, "stride_seconds": 3}},
        ])
        movie = self.validate_movie(Path(result["outputs"]["full"]["output_root"]), 12)
        root = Path(result["outputs"]["audio"]["output_root"])
        rows = csv_rows(only(root, "audio_analysis.csv"))
        acoustic = csv_rows(only(root, "opensmile_features.csv"))
        assert len(rows) == len(acoustic) == 3, (len(rows), len(acoustic))
        emotion = ("Anger", "Contempt", "Disgust", "Fear", "Happiness", "Sadness", "Surprise", "Neutral", "Other", "Arousal", "Dominance", "Valence")
        assert all(row.get(metric) in (None, "") for row in rows for metric in emotion)
        assert len(acoustic[0]) == 93
        values = [float(value) for row in acoustic for key, value in row.items() if key not in {"Row", "WindowStart", "WindowEnd", "name", "frameTime"}]
        assert all(math.isfinite(value) for value in values)
        assert [(float(r["StartSeconds"]), float(r["EndSeconds"])) for r in rows] == [(0, 6), (3, 9), (6, 12)]
        return {"movie": movie, "audio_windows": len(rows), "finite_acoustic_values": len(values), "emotion_model_values_absent": True}

    def focus(self) -> dict:
        source = self.fixtures["video"]
        manifest = self.root / "fixtures" / "focus.json"
        save(manifest, {"source_path": source, "source_kind": "file", "gap_seconds": 1, "selected_segments": [
            {"source_path": source, "source_kind": "file", "speaker": "Synthetic Speaker", "start_seconds": start, "end_seconds": end}
            for start, end in ((1, 3), (6, 8))
        ]})
        result = self.run_job("focus", [{"id": "focus", "stage": "procurement", "options": {"source_path": source, "mode": "manual", "segment_manifest": str(manifest)}}])
        movie = self.validate_movie(Path(result["outputs"]["focus"]["output_root"]), 5)
        frame = subprocess.check_output([self.ffmpeg, "-v", "error", "-ss", "2.5", "-i", movie["path"], "-frames:v", "1", "-vf", "scale=16:16,format=gray", "-f", "rawvideo", "-"], env=self.env)
        pcm = subprocess.check_output([self.ffmpeg, "-v", "error", "-ss", "2.25", "-i", movie["path"], "-t", "0.5", "-vn", "-ac", "1", "-ar", "8000", "-f", "s16le", "-"], env=self.env)
        samples = array.array("h", pcm)
        assert frame and max(frame) <= 4
        assert samples and max(abs(value) for value in samples) <= 10
        return {"movie": movie, "gap_max_gray": max(frame), "gap_max_abs_pcm16": max(abs(value) for value in samples)}

    def sampling(self) -> dict:
        result = self.run_job("sampling", [{"id": "sample", "stage": "procurement", "options": {"source_path": self.fixtures["video"], "mode": "standard", "percentage": .5, "max_segment_seconds": 2}}])
        return self.validate_movie(Path(result["outputs"]["sample"]["output_root"]), 6)

    def catalog(self) -> dict:
        filename = self.root / "fixtures" / "sources.csv"
        with filename.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Link", "Speaker", "Country"])
            writer.writerow([str(Path(self.fixtures["video"]).relative_to(filename.parent)), "Synthetic Speaker", "Ireland"])
        scan = self.invoke("inspect", "source", filename, "--no-enrich")
        assert scan["catalog_sha256"] == sha(filename), scan
        result = self.run_job("catalog-full", [{"id": "catalog", "stage": "procurement", "options": {"source_path": str(filename), "mode": "full"}}])
        output = Path(result["outputs"]["catalog"]["output_root"])
        inspected = self.invoke("inspect", "catalog", output)
        assert inspected["catalog_sha256"] == sha(filename), inspected
        assert [r["source_id"] for r in inspected["sources"]] == ["source-0001"], inspected
        return {"output_root": str(output), "scan": inspected, "source_manifest_sha256": sha(output / "source_manifest.json")}

    def invalid(self) -> dict:
        checks = []
        for name, options in (("unknown", {"mystery": True}), ("bad-type", {"include_emotions": "false"}), ("bad-range", {"window_seconds": 0})):
            job = self.job(name, [{"id": "audio", "stage": "audio", "options": {"source_path": self.fixtures["video"], **options}}])
            directory = self.root / "runs" / name
            result = self.invoke("run", "--job", job, "--run-dir", directory, expected=2)
            assert result["state"] == "validation_failed" and not directory.exists()
            checks.append(result)
        return {"checks": checks}

    def corrupt_fail_stop(self) -> dict:
        result = self.run_job("corrupt", [
            {"id": "bad", "stage": "audio", "options": {"source_path": self.fixtures["corruptVideo"], "include_emotions": False, "device": "cpu"}},
            {"id": "never", "stage": "procurement", "options": {"source_path": self.fixtures["video"], "mode": "full"}},
        ], expected=3)
        assert result["state"] == "failed" and result["steps"][0]["returncode"] != 0
        assert result["steps"][1]["state"] == "pending" and "never" not in result["outputs"]
        assert not (self.root / "runs" / "corrupt" / "outputs" / "never").exists()
        log = Path(result["steps"][0]["log_path"]).read_text(encoding="utf-8")
        assert any(word in log.casefold() for word in ("ffprobe", "invalid data", "moov atom")), log[-1000:]
        return result

    def portable(self) -> dict:
        original = self.root / "portable original"
        original.mkdir()
        shutil.copyfile(self.fixtures["video"], original / "voix échantillon.mp4")
        job = self.job("travail vidéo", [{"id": "full", "stage": "procurement", "options": {"source_path": "voix échantillon.mp4", "mode": "full"}}], folder=original)
        copied = self.root / "copie déplacée avec espaces"
        shutil.copytree(original, copied)
        result = self.invoke("run", "--job", copied / job.name, "--run-dir", self.root / "runs" / "portable")
        assert result["state"] == "completed"
        command = result["steps"][0]["command"]
        assert str(copied / "voix échantillon.mp4") in command
        assert str(original / "voix échantillon.mp4") not in command
        assert sha(copied / "voix échantillon.mp4") == sha(Path(self.fixtures["video"]))
        return self.validate_movie(Path(result["outputs"]["full"]["output_root"]), 12)

    def imported_analysis(self) -> dict:
        reports = self.fixtures["reports"]
        speakers = self.invoke("inspect", "analysis-speakers", "--modality", "audio", "import", reports)
        profile = self.invoke("inspect", "analysis-profile", "--modality", "audio", "import", reports)
        assert "Country" in [r["name"] for r in profile["metadataFields"]]
        result = self.run_job("imported-analysis", [{"id": "analysis", "stage": "analysis", "options": {"modalities": [{"name": "audio", "source_method": "import", "source_path": reports}], "write_graphs": False}, "native_options": {"profile_options": {"automatic_group_field": "Country"}}}])
        output = Path(result["outputs"]["analysis"]["output_root"])
        import openpyxl
        workbook = output / "combined_analysis.xlsx"
        book = openpyxl.load_workbook(workbook, data_only=False)
        rows = list(book["Audio"].values)
        anger = next(row for row in rows if "Anger" in row)
        assert [anger[index] for index in (3, 4, 7, 8)] == [10, 30, 20, 40], anger
        assert [rows[0][index] for index in (3, 4, 7, 8)] == ["Synthetic observation 1", "Synthetic observation 3", "Synthetic observation 2", "Synthetic observation 4"]
        assert anger[-1] == "=AVERAGE(D2,E2,H2,I2)", anger
        assert not any(c.data_type == "e" for sheet in book for row in sheet for c in row)
        assert any("prob" in name.casefold() for name in book.sheetnames)
        saved_profile = json.loads((output / "analysis_profile.json").read_text())
        assert saved_profile["automatic_group_field"] == "Country"
        book.close()
        return {"workbook": str(workbook), "sha256": sha(workbook), "country_ordered_anger": [10, 30, 20, 40], "overall_expected": 25, "formula_scope": "Reference membership checked; no Excel recalculation", "speakers": speakers}

    def cancel(self) -> dict:
        import psutil
        media = self.root / "fixtures" / "cancellation-long.mp4"
        subprocess.run([self.ffmpeg, "-v", "error", "-stream_loop", "39", "-i", self.fixtures["video"], "-c", "copy", str(media)], env=self.env, check=True)
        job = self.job("cancel", [{"id": "full", "stage": "procurement", "options": {"source_path": str(media), "mode": "full"}}])
        directory = self.root / "runs" / "cancel"
        command = [sys.executable, "-m", "application.cli", "run", "--job", str(job), "--run-dir", str(directory), "--timeout", "60"]
        observed: dict[int, float] = {}
        with (self.root / "cancel.stdout.json").open("w", encoding="utf-8") as stdout, (self.root / "cancel.stderr.log").open("w", encoding="utf-8") as stderr:
            process = subprocess.Popen(command, cwd=REPO, env=self.env, stdin=subprocess.DEVNULL, stdout=stdout, stderr=stderr)
            try:
                deadline = time.monotonic() + 40
                saw_decoder = False
                while time.monotonic() < deadline and process.poll() is None:
                    try:
                        descendants = psutil.Process(process.pid).children(recursive=True)
                        for child in descendants:
                            observed[child.pid] = child.create_time()
                        saw_decoder = any("ffmpeg" in p.name().casefold() for p in descendants)
                    except psutil.Error:
                        pass
                    if saw_decoder:
                        break
                    time.sleep(.05)
                assert saw_decoder, "No live FFmpeg descendant observed before cancellation"
                status = self.invoke("status", "--run-dir", directory)
                assert status["state"] == "running" and status["steps"][0].get("child_pid")
                requested = self.invoke("cancel", "--run-dir", directory)
                assert requested["state"] == "cancel_requested"
                code = process.wait(timeout=40)
                assert code == 130, code
            finally:
                if process.poll() is None:
                    for pid, created in observed.items():
                        try:
                            child = psutil.Process(pid)
                            if abs(child.create_time() - created) < .01:
                                child.kill()
                        except psutil.Error:
                            pass
                    process.kill()
                    process.wait(timeout=10)
        result = json.loads((self.root / "cancel.stdout.json").read_text())
        assert result["state"] == "cancelled"
        survivors = []
        for pid, created in observed.items():
            try:
                child = psutil.Process(pid)
                if abs(child.create_time() - created) < .01 and child.is_running():
                    survivors.append(pid)
            except psutil.Error:
                pass
        assert not survivors, survivors
        assert self.invoke("status", "--run-dir", directory)["state"] == "cancelled"
        return {"state": result["state"], "observed_descendant_pids": list(observed), "surviving_owned_descendants": survivors, "real_ffmpeg_observed": True}

    def timeout(self) -> dict:
        job = self.job("deadline", [{"id": "audio", "stage": "audio", "options": {"source_path": self.fixtures["video"], "include_emotions": False}}])
        result = self.invoke("run", "--job", job, "--run-dir", self.root / "runs" / "deadline", "--timeout", ".1", expected=124)
        assert result["state"] == "timed_out"
        return result

    def models(self) -> dict:
        assert self.model_video is not None, "--model-video is required for model acceptance"
        metadata = self.probe(self.model_video)
        close(float(metadata["format"]["duration"]), 15, .25)
        assert {s["codec_type"] for s in metadata["streams"]} >= {"video", "audio"}
        model_source = self.root / "fixtures" / "synthetic-model.mp4"
        shutil.copyfile(self.model_video, model_source)
        # Whisper's downloader does not obey HF_HUB_OFFLINE. Prove its cache
        # is present and valid before the public CLI could request inference.
        # Importing Whisper also imports Torch. Keep that preflight in a short-
        # lived process so it does not retain RAM during actual model inference.
        url = json.loads(subprocess.check_output(
            [sys.executable, "-c", "import json, whisper; print(json.dumps(whisper._MODELS['small']))"],
            cwd=REPO, env=self.env, text=True, encoding="utf-8", timeout=120,
        ))
        cache = Path(os.getenv("XDG_CACHE_HOME", str(Path.home() / ".cache"))) / "whisper" / "small.pt"
        assert cache.is_file() and sha(cache) == url.split("/")[-2], "A verified cached small Whisper checkpoint is required"
        for component in ("face", "audio", "text"):
            result = self.invoke("doctor", "--component", component, "--device", "cpu", timeout=300)
            assert result["state"] == "ready", result
        catalog = self.root / "fixtures" / "model-sources.csv"
        with catalog.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Link", "Speaker", "Country"])
            writer.writerow([model_source.name, "Synthetic Test", "UK"])
        result = self.run_job("catalog-models-profile", [
            {"id": "catalog", "stage": "procurement", "options": {"source_path": str(catalog), "mode": "full"}},
            {"id": "face", "stage": "face", "options": {"source_path": ref("catalog"), "sample_fps": 1, "batch_size": 2, "device": "cpu"}},
            {"id": "audio", "stage": "audio", "options": {"source_path": ref("catalog"), "device": "cpu"}},
            {"id": "text", "stage": "text", "options": {"source_path": ref("catalog"), "whisper_model": "small", "whisper_device": "cpu", "whisper_language": "en", "threads": 4, "write_graphs": False}},
            {"id": "analysis", "stage": "analysis", "options": {"modalities": [
                {"name": "video", "source_method": "run", "source_path": ref("face")},
                {"name": "audio", "source_method": "run", "source_path": ref("audio")},
                {"name": "text", "source_method": "import", "source_path": ref("text")},
            ], "write_graphs": False}, "native_options": {"profile_options": {"automatic_group_field": "Country"}}},
        ], timeout=2400)
        roots = {name: Path(value["output_root"]) for name, value in result["outputs"].items()}
        for modality in ("face", "audio", "text"):
            for name in ("source_manifest.json", "source_metadata.csv"):
                assert (roots[modality] / name).read_bytes() == (roots["catalog"] / name).read_bytes()
        face = csv_rows(only(roots["face"], "face_core.csv"))
        audio = csv_rows(only(roots["audio"], "audio_analysis.csv"))
        acoustic = csv_rows(only(roots["audio"], "opensmile_features.csv"))
        assert len(face) == 15 and sum(row["face_detected"] == "True" for row in face) >= 12
        assert len(audio) == len(acoustic) == 2 and len(acoustic[0]) == 93
        assert [(float(r["StartSeconds"]), float(r["EndSeconds"])) for r in audio] == [(0, 10), (5, 15)]
        audio_manifest = json.loads(only(roots["audio"], "audio_analysis_manifest.json").read_text())
        approved_audio = {
            "tiantiaf/whisper-large-v3-msp-podcast-emotion": "b92dab65151206a603810ec8b72eb528b9dd983c",
            "superb/wav2vec2-base-superb-er": "441a7599c3b22107314dcbd9166621c5c83f2cc5",
        }
        categorical = audio_manifest["categorical_model_name"]
        assert audio_manifest["categorical_model_version"] == approved_audio[categorical]
        assert audio_manifest["dimensional_model_name"] == "audeering/wav2vec2-large-robust-12-ft-emotion-msp-dim"
        assert audio_manifest["dimensional_model_version"] == "6eba34a2485ea31cb03600241787c3a5edab8626"
        assert audio_manifest["input_sha256"] == sha(Path(audio_manifest["input_video"]))
        assert audio_manifest["source_id"] == "source-0001" and audio_manifest["source_metadata"]["Country"] == "UK"
        assert audio_manifest["model_device"] == "cpu" and not audio_manifest["emotion_models_skipped"]
        supported_audio = ["Anger", "Happiness", "Neutral", "Sadness"]
        extra_classes = ["Contempt", "Disgust", "Fear", "Surprise", "Other"]
        if categorical.startswith("superb/"):
            assert all(row[field] == "" for row in audio for field in extra_classes)
        else:
            supported_audio += extra_classes
        for row in audio:
            assert all(math.isfinite(float(row[field])) for field in supported_audio)
            close(sum(float(row[field]) for field in supported_audio), 1, 1e-5)
        face_manifest = json.loads(only(roots["face"], "video_manifest.json").read_text())
        assert face_manifest["status"] == "completed" and face_manifest["backend"]["resolved_device"] == "cpu"
        face_weights = face_manifest["backend"]["weights"]["components"]
        assert set(face_weights) == {"retinaface_r34", "arcface_r50", "face_multitask_v2"}
        for weight in face_weights.values():
            assert weight["sha256"] == sha(Path(weight["local_path"]))
            assert weight["requested_revision"] == weight["snapshot_commit"]
        transcription_manifest = json.loads(only(roots["text"], "transcription_run_manifest.json").read_text())
        assert transcription_manifest["status"] == "completed"
        whisper_provenance = transcription_manifest["whisper_provenance"]
        assert whisper_provenance
        for provenance in whisper_provenance.values():
            assert provenance["checkpoint"]["expected_sha256"] == sha(cache)
            assert provenance["checkpoint"]["requested_name"] == "small"
            assert provenance["engine"]["distribution"] == "openai-whisper" and provenance["engine"]["version"]
            assert provenance["runtime"]["torch_version"] and provenance["runtime"]["ffmpeg_version"]
            assert provenance["device"] == "cpu"
            assert all(p["language"] == "en" and p["fp16"] is False for p in provenance["decoding_passes"])
        terms = csv_rows(roots["text"] / "analysis" / "selected" / "video_level_summary.csv")[0]
        constructs = csv_rows(roots["text"] / "analysis" / "multimodal" / "video_level_summary.csv")[0]
        assert terms["source_id"] == constructs["Source ID"] == "source-0001"
        total = int(terms["rocksteady_terms_total"])
        assert total > 0
        close(float(constructs["Positive Sentiment"]), int(terms["positive_total"]) / total, 1e-6)
        close(float(constructs["Negative Sentiment"]), int(terms["negative_total"]) / total, 1e-6)
        active, passive = int(terms["active_total"]), int(terms["passive_total"])
        assert active + passive > 0
        close(float(constructs["Arousal / Activation"]), (active - passive) / (active + passive), 1e-5)
        import openpyxl
        workbook = roots["analysis"] / "combined_analysis.xlsx"
        book = openpyxl.load_workbook(workbook, data_only=False)
        checks = []
        def check(sheet, row, expected):
            actual = book[sheet].cell(row, 4).value
            close(actual, expected)
            checks.append({"cell": f"{sheet}!D{row}", "actual": actual, "expected_raw": expected})
        for row, field in ((2, "Anger"), (3, "Happiness"), (4, "Sadness"), (5, "Neutral"), (7, "Arousal"), (8, "Dominance")):
            check("Audio", row, statistics.mean(float(r[field]) * 100 for r in audio))
        check("Audio", 6, statistics.mean(float(r["Valence"]) * 200 - 100 for r in audio))
        for row, field in ((2, "Anger"), (4, "Disgust"), (5, "Fear"), (6, "Happy"), (7, "Sad"), (8, "Surprise"), (9, "Neutral")):
            check("Video", row, statistics.mean(float(r[field]) * 100 for r in face if r["face_detected"] == "True"))
        assert book["Video"]["D3"].value is None
        for row, field in ((2, "Positive Sentiment"), (3, "Negative Sentiment"), (4, "Text Valence"), (5, "Arousal / Activation"), (6, "Dominance / Power")):
            check("Text sentiment", row, float(constructs[field]) * 100)
        assert not any(cell.data_type == "e" for sheet in book for row in sheet for cell in row)
        profile = json.loads((roots["analysis"] / "analysis_profile.json").read_text())
        assert profile["automatic_group_field"] == "Country"
        assert profile["source_manifest"]["sha256"] == sha(roots["catalog"] / "source_manifest.json")
        book.close()
        return {"output_roots": {k: str(v) for k, v in roots.items()}, "face_samples": len(face), "audio_windows": len(audio), "text_terms": total, "source_sidecars_identical": True, "workbook": str(workbook), "workbook_sha256": sha(workbook), "independent_raw_checks": checks, "profile": profile, "models": {"audio_categorical": categorical, "audio_categorical_revision": audio_manifest["categorical_model_version"], "audio_supported_classes": supported_audio, "audio_model_warnings": audio_manifest["model_errors"], "audio_dimensional_revision": audio_manifest["dimensional_model_version"], "face_backend": face_manifest["backend"], "whisper_provenance": whisper_provenance}}

    def clean_speaker(self) -> dict:
        assert self.model_video is not None
        result = self.invoke("doctor", "--component", "clean-speaker", "--device", "cpu", timeout=300)
        assert result["state"] == "ready"
        result = self.run_job("clean-speaker", [{"id": "clean", "stage": "procurement", "options": {
            "source_path": str(self.model_video), "mode": "clean-speaker-beta", "beta_device": "cpu",
            "beta_identity_stills": 5, "beta_scan_fps": 1, "beta_validation_fps": 1,
            "beta_keep_debug": True, "beta_resource_guard_percent": 0,
            "beta_resource_guard_timeout_seconds": 60, "beta_isolated_video_processes": False,
            "beta_skip_completed_outputs": False, "beta_video_cooldown_seconds": 0,
            "beta_max_affinity_cores": 4, "beta_native_threads": 4,
        }}], timeout=1200)
        output = Path(result["outputs"]["clean"]["output_root"])
        summary = json.loads(only(output, "clean_speaker_beta_summary.json").read_text())
        assert summary["processed"] == 1 and summary["failed"] == summary["unusable"] == 0, summary
        assert summary["results"][0]["status"] == "ok"
        movie = Path(summary["results"][0]["output_video"])
        duration = float(self.probe(movie)["format"]["duration"])
        assert 10 <= duration <= 15.3, duration
        assert list(output.rglob("face_visibility_intervals.json")) and list(output.rglob("voice_activity_intervals.json"))
        return {"summary": summary, "movie_sha256": sha(movie), "duration_seconds": duration}

    def execute(self) -> int:
        from application import backend
        save(self.root / "source-start.json", source_snapshot())
        settings = backend.ui_settings_path(REPO)
        protected = [settings, backend.settings_backup_path(settings), backend.eula_path(REPO)]
        protected.extend(backend.credential_store.secret_path(settings, key) for key in ("youtubeApiKey", "huggingFaceToken"))
        before = {p: sha(p) if p.is_file() else None for p in protected}
        supplied_hash = sha(self.model_video) if self.model_video else None
        # Reuse the existing public synthetic-fixture helper. FFmpeg discovery
        # is inherited temporarily by generation, not persisted to user settings.
        previous_path = os.environ.get("PATH", "")
        os.environ["PATH"] = self.env["PATH"]
        try:
            self.fixtures = write_fixtures(self.root, None)
        finally:
            os.environ["PATH"] = previous_path
        fixture_hashes = {key: sha(Path(value)) for key, value in self.fixtures.items() if Path(value).is_file()}
        if self.phase in {"nonmodel", "all"}:
            for name, action in (("schema-inspect-doctor", self.metadata_commands), ("dry-run-no-output", self.dry_run), ("full-to-opensmile", self.full_audio), ("focus-decoded-gap", self.focus), ("standard-sampling", self.sampling), ("catalog-inspect", self.catalog), ("invalid-no-output", self.invalid), ("decoder-failure-stops-workflow", self.corrupt_fail_stop), ("copied-unicode-relative-job", self.portable), ("imported-country-workbook", self.imported_analysis), ("public-status-cancel-child-cleanup", self.cancel), ("workflow-timeout", self.timeout)):
                self.record(name, action)
        if self.phase in {"models", "all"}:
            self.record("catalog-models-profile-raw-oracle", self.models)
            self.record("clean-speaker-offline", self.clean_speaker)
        def preservation():
            assert before == {p: sha(p) if p.is_file() else None for p in protected}, "Desktop settings, credentials or acceptance files changed"
            assert fixture_hashes == {key: sha(Path(self.fixtures[key])) for key in fixture_hashes}
            if self.model_video:
                assert sha(self.model_video) == supplied_hash
            return {"protected_desktop_files_checked": len(protected), "synthetic_inputs_unchanged": True, "supplied_model_fixture_unchanged": bool(self.model_video)}
        self.record("input-settings-credential-preservation", preservation)
        save(self.root / "source-finish.json", source_snapshot())
        return 0 if self.summary()["state"] == "passed" else 1


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, required=True, help="New evidence directory; existing directories are refused")
    parser.add_argument("--phase", choices=("nonmodel", "models", "all"), default="nonmodel")
    parser.add_argument("--model-video", type=Path, help="Existing 15-second fixture with a visible speaking face, required for models/all")
    parser.add_argument("--max-ram-percent", type=float, default=90, help="Explicit job RAM guard, 10–95 percent; default 90")
    args = parser.parse_args()
    root = args.output.expanduser().absolute()
    if root.exists():
        parser.error("--output must be new; previous evidence is preserved")
    if args.phase != "nonmodel" and (args.model_video is None or not args.model_video.is_file()):
        parser.error("--model-video must name an existing fixture for models/all")
    if not math.isfinite(args.max_ram_percent) or not 10 <= args.max_ram_percent <= 95:
        parser.error("--max-ram-percent must be between 10 and 95")
    root.mkdir(parents=True)
    return Acceptance(root, args.phase, args.model_video.resolve() if args.model_video else None, args.max_ram_percent).execute()


if __name__ == "__main__":
    raise SystemExit(main())
